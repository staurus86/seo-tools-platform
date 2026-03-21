"""
Keyword Clusterizer router.
"""
import io
import os
from datetime import datetime
from typing import Optional, List, Dict, Any

from fastapi import APIRouter, BackgroundTasks, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, field_validator

from app.api.routers._task_store import create_task_pending, update_task_state, get_task_result

router = APIRouter(tags=["SEO Tools"])


def _parse_clusterizer_frequency(value: Any) -> Optional[float]:
    from app.tools.clusterizer.input_parser import parse_clusterizer_frequency

    return parse_clusterizer_frequency(value)


def _split_clusterizer_text_rows(raw_text: str) -> List[str]:
    from app.tools.clusterizer.input_parser import split_clusterizer_text_rows

    return split_clusterizer_text_rows(raw_text)


def _parse_clusterizer_keyword_line(raw_line: str) -> List[Dict[str, Any]]:
    from app.tools.clusterizer.input_parser import parse_clusterizer_keyword_line

    return parse_clusterizer_keyword_line(raw_line)


def _collect_clusterizer_keyword_rows(
    keywords: Optional[List[str]], keywords_text: Optional[str]
) -> List[Dict[str, Any]]:
    from app.tools.clusterizer.input_parser import collect_clusterizer_keyword_rows

    return collect_clusterizer_keyword_rows(keywords, keywords_text)


def check_keywords_clusterizer(
    *,
    keywords: Optional[List[str]] = None,
    keyword_rows: Optional[List[Dict[str, Any]]] = None,
    method: str = "jaccard",
    similarity_threshold: float = 0.35,
    min_cluster_size: int = 2,
    clustering_mode: str = "balanced",
    progress_callback=None,
) -> Dict[str, Any]:
    """Basic keyword clusterizer without SERP data."""
    from app.tools.clusterizer import run_keyword_clusterizer

    return run_keyword_clusterizer(
        keywords=keywords or [],
        keyword_rows=keyword_rows or [],
        method=method,
        similarity_threshold=similarity_threshold,
        min_cluster_size=min_cluster_size,
        clustering_mode=clustering_mode,
        progress_callback=progress_callback,
    )


class ClusterizerRequest(BaseModel):
    keywords: Optional[List[str]] = None
    keywords_text: Optional[str] = None
    method: Optional[str] = "jaccard"
    clustering_mode: Optional[str] = "balanced"
    similarity_threshold_pct: int = 35
    min_cluster_size: int = 2

    @field_validator("keywords", mode="before")
    @classmethod
    def _normalize_keywords(cls, value):
        if value is None:
            return []
        if isinstance(value, str):
            return [x.strip() for x in str(value).replace("\r", "\n").split("\n") if x.strip()]
        if isinstance(value, list):
            return [str(v).strip() for v in value if str(v).strip()]
        return []

    @field_validator("keywords_text", mode="before")
    @classmethod
    def _normalize_keywords_text(cls, value):
        if value is None:
            return ""
        return str(value).strip()

    @field_validator("clustering_mode", mode="before")
    @classmethod
    def _normalize_clustering_mode(cls, value):
        if value is None:
            return "balanced"
        return str(value).strip().lower()


@router.post("/tasks/clusterizer")
async def create_clusterizer_task(data: ClusterizerRequest, background_tasks: BackgroundTasks):
    """Keyword clustering by textual similarity without search-engine fetching."""
    from app.config import settings

    keyword_rows = _collect_clusterizer_keyword_rows(data.keywords, data.keywords_text)
    keywords = [str(item.get("keyword") or "").strip() for item in keyword_rows if str(item.get("keyword") or "").strip()]
    input_demand_total = sum(float(item.get("frequency") or 0.0) for item in keyword_rows)

    max_keywords = max(1, int(getattr(settings, "CLUSTERIZER_MAX_KEYWORDS", 2000) or 2000))
    if not keywords:
        raise HTTPException(status_code=422, detail="Добавьте хотя бы один ключ для кластеризации")
    if len(keyword_rows) > max_keywords:
        raise HTTPException(status_code=422, detail=f"Слишком много ключей: максимум {max_keywords}")

    method = str(data.method or "jaccard").strip().lower()
    if method not in {"jaccard", "overlap", "dice"}:
        method = "jaccard"
    clustering_mode = str(data.clustering_mode or "balanced").strip().lower()
    if clustering_mode not in {"strict", "balanced", "broad"}:
        clustering_mode = "balanced"
    similarity_threshold_pct = max(1, min(100, int(data.similarity_threshold_pct or 35)))
    min_cluster_size = max(1, min(50, int(data.min_cluster_size or 2)))
    similarity_threshold = similarity_threshold_pct / 100.0

    task_id = f"clusterizer-{datetime.now().timestamp()}"
    create_task_pending(
        task_id,
        "clusterizer",
        f"keywords:{len(keyword_rows)}",
        status_message="Задача поставлена в очередь",
    )

    def _run_clusterizer_task() -> None:
        try:
            update_task_state(
                task_id,
                status="RUNNING",
                progress=5,
                status_message="Подготовка ключей",
            )

            def _progress(progress: int, message: str) -> None:
                update_task_state(
                    task_id,
                    status="RUNNING",
                    progress=progress,
                    status_message=message,
                )

            result = check_keywords_clusterizer(
                keywords=[],
                keyword_rows=keyword_rows,
                method=method,
                similarity_threshold=similarity_threshold,
                min_cluster_size=min_cluster_size,
                clustering_mode=clustering_mode,
                progress_callback=_progress,
            )

            result_payload = result.get("results", {}) if isinstance(result, dict) else {}
            summary_payload = result_payload.get("summary", {}) if isinstance(result_payload, dict) else {}
            summary_payload["input_demand_total"] = round(
                float(summary_payload.get("input_demand_total") or input_demand_total), 4
            )
            if isinstance(result_payload, dict):
                result_payload["summary"] = summary_payload
            if isinstance(result, dict):
                result["results"] = result_payload

            update_task_state(
                task_id,
                status="SUCCESS",
                progress=100,
                status_message="Кластеризация завершена",
                result=result,
                error=None,
            )
        except Exception as exc:
            update_task_state(
                task_id,
                status="FAILURE",
                progress=100,
                status_message="Кластеризация завершилась с ошибкой",
                error=str(exc),
            )

    background_tasks.add_task(_run_clusterizer_task)
    return {"task_id": task_id, "status": "PENDING", "message": "Кластеризация ключей запущена"}


@router.post("/tasks/clusterizer/upload")
async def create_clusterizer_task_upload(
    background_tasks: BackgroundTasks,
    keywords_file: UploadFile = File(...),
    method: str = Form("jaccard"),
    clustering_mode: str = Form("balanced"),
    similarity_threshold_pct: int = Form(35),
    min_cluster_size: int = Form(2),
):
    """Keyword clustering from uploaded CSV or XLSX file (columns: keyword, frequency)."""
    from app.config import settings
    from app.tools.clusterizer.input_parser import parse_clusterizer_file

    filename = str(keywords_file.filename or "")
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext not in {"csv", "xlsx"}:
        raise HTTPException(status_code=422, detail="Поддерживаются только файлы .csv и .xlsx")

    file_bytes = await keywords_file.read()
    if not file_bytes:
        raise HTTPException(status_code=422, detail="Файл пустой")

    max_file_size = max(1024, int(getattr(settings, "CLUSTERIZER_MAX_FILE_SIZE_BYTES", 50 * 1024 * 1024) or (50 * 1024 * 1024)))
    if len(file_bytes) > max_file_size:
        raise HTTPException(status_code=422, detail=f"Файл превышает лимит {max_file_size} байт")

    try:
        keyword_rows = parse_clusterizer_file(file_bytes, filename)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Ошибка чтения файла: {exc}") from exc

    keywords = [str(item.get("keyword") or "").strip() for item in keyword_rows if str(item.get("keyword") or "").strip()]
    if not keywords:
        raise HTTPException(status_code=422, detail="Файл не содержит ключевых слов")

    max_keywords = max(1, int(getattr(settings, "CLUSTERIZER_MAX_KEYWORDS", 25000) or 25000))
    if len(keyword_rows) > max_keywords:
        raise HTTPException(status_code=422, detail=f"Слишком много ключей: максимум {max_keywords}")

    _method = str(method or "jaccard").strip().lower()
    if _method not in {"jaccard", "overlap", "dice"}:
        _method = "jaccard"
    _clustering_mode = str(clustering_mode or "balanced").strip().lower()
    if _clustering_mode not in {"strict", "balanced", "broad"}:
        _clustering_mode = "balanced"
    _threshold_pct = max(1, min(100, int(similarity_threshold_pct or 35)))
    _min_cluster_size = max(1, min(50, int(min_cluster_size or 2)))
    _similarity_threshold = _threshold_pct / 100.0
    input_demand_total = sum(float(item.get("frequency") or 0.0) for item in keyword_rows)

    task_id = f"clusterizer-{datetime.now().timestamp()}"
    create_task_pending(
        task_id,
        "clusterizer",
        f"file:{filename}:{len(keyword_rows)}kw",
        status_message="Задача поставлена в очередь",
    )

    def _run_upload_task() -> None:
        try:
            update_task_state(task_id, status="RUNNING", progress=5, status_message="Подготовка ключей из файла")

            def _progress(progress: int, message: str) -> None:
                update_task_state(task_id, status="RUNNING", progress=progress, status_message=message)

            result = check_keywords_clusterizer(
                keywords=[],
                keyword_rows=keyword_rows,
                method=_method,
                similarity_threshold=_similarity_threshold,
                min_cluster_size=_min_cluster_size,
                clustering_mode=_clustering_mode,
                progress_callback=_progress,
            )

            result_payload = result.get("results", {}) if isinstance(result, dict) else {}
            summary_payload = result_payload.get("summary", {}) if isinstance(result_payload, dict) else {}
            summary_payload["input_demand_total"] = round(
                float(summary_payload.get("input_demand_total") or input_demand_total), 4
            )
            if isinstance(result_payload, dict):
                result_payload["summary"] = summary_payload
            if isinstance(result, dict):
                result["results"] = result_payload

            update_task_state(
                task_id,
                status="SUCCESS",
                progress=100,
                status_message="Кластеризация завершена",
                result=result,
                error=None,
            )
        except Exception as exc:
            update_task_state(
                task_id,
                status="FAILURE",
                progress=100,
                status_message="Кластеризация завершилась с ошибкой",
                error=str(exc),
            )

    background_tasks.add_task(_run_upload_task)
    return {"task_id": task_id, "status": "PENDING", "message": f"Кластеризация {len(keyword_rows)} ключей из файла запущена"}


# ── XLSX Export ──────────────────────────────────────────────────────────────


def _build_clusterizer_xlsx(task_data: Dict[str, Any]) -> bytes:
    """Build an XLSX workbook from clusterizer task result and return bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Brand styling (matches design-system)
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill(start_color="0F4C81", end_color="0F4C81", fill_type="solid")
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    CELL_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    ALT_FILL_A = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    ALT_FILL_B = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    def _apply_header(ws, headers: List[str]) -> None:
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = CELL_BORDER
        ws.freeze_panes = "A2"

    def _apply_row_style(ws, row_idx: int, col_count: int) -> None:
        fill = ALT_FILL_A if row_idx % 2 == 0 else ALT_FILL_B
        align = Alignment(horizontal="left", vertical="center", wrap_text=False)
        for col_idx in range(1, col_count + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.border = CELL_BORDER
            cell.alignment = align

    def _auto_width(ws, col_count: int, max_rows: int = 200) -> None:
        for col_idx in range(1, col_count + 1):
            max_len = 10
            for row_idx in range(1, min(max_rows + 2, ws.max_row + 1)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    max_len = max(max_len, min(60, len(str(val)) + 2))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len

    result = task_data.get("result", {}) or {}
    results = result.get("results", {}) or {}
    clusters = results.get("clusters", []) or []
    flat_keywords = results.get("cluster_keywords_flat", []) or []
    unclustered = results.get("unclustered_keywords", []) or []
    summary = results.get("summary", {}) or {}

    wb = Workbook()

    # ── Sheet 1: Clusters ────────────────────────────────────────────────
    ws_clusters = wb.active
    ws_clusters.title = "Clusters"
    ws_clusters.sheet_properties.tabColor = "0F4C81"
    cluster_headers = [
        "Cluster ID", "Representative", "Size", "Quality",
        "Intent", "Avg Similarity", "Keywords",
    ]
    _apply_header(ws_clusters, cluster_headers)
    for row_idx, cluster in enumerate(clusters, start=2):
        keywords_joined = ", ".join(cluster.get("keywords", []))
        ws_clusters.cell(row=row_idx, column=1, value=cluster.get("cluster_id", ""))
        ws_clusters.cell(row=row_idx, column=2, value=cluster.get("representative", ""))
        ws_clusters.cell(row=row_idx, column=3, value=cluster.get("size", 0))
        ws_clusters.cell(row=row_idx, column=4, value=cluster.get("cohesion", ""))
        ws_clusters.cell(row=row_idx, column=5, value=cluster.get("intent", ""))
        ws_clusters.cell(row=row_idx, column=6, value=cluster.get("avg_similarity", 0))
        ws_clusters.cell(row=row_idx, column=7, value=keywords_joined)
        _apply_row_style(ws_clusters, row_idx, len(cluster_headers))
    _auto_width(ws_clusters, len(cluster_headers))

    # ── Sheet 2: All Keywords ────────────────────────────────────────────
    ws_keywords = wb.create_sheet("All Keywords")
    ws_keywords.sheet_properties.tabColor = "0E7490"
    kw_headers = [
        "Keyword", "Frequency", "Cluster ID",
        "Cluster Representative", "Intent",
    ]
    _apply_header(ws_keywords, kw_headers)

    # Build cluster lookup for intent
    cluster_intent_map: Dict[int, str] = {}
    for cluster in clusters:
        cluster_intent_map[int(cluster.get("cluster_id", 0))] = str(cluster.get("intent", ""))

    for row_idx, kw_row in enumerate(flat_keywords, start=2):
        cid = int(kw_row.get("cluster_id", 0))
        ws_keywords.cell(row=row_idx, column=1, value=kw_row.get("keyword", ""))
        ws_keywords.cell(row=row_idx, column=2, value=kw_row.get("demand", 0))
        ws_keywords.cell(row=row_idx, column=3, value=cid)
        ws_keywords.cell(row=row_idx, column=4, value=kw_row.get("representative", ""))
        ws_keywords.cell(row=row_idx, column=5, value=cluster_intent_map.get(cid, ""))
        _apply_row_style(ws_keywords, row_idx, len(kw_headers))
    _auto_width(ws_keywords, len(kw_headers))

    # ── Sheet 3: Singletons ──────────────────────────────────────────────
    ws_singletons = wb.create_sheet("Singletons")
    ws_singletons.sheet_properties.tabColor = "F59E0B"
    singleton_headers = ["Keyword", "Frequency"]
    _apply_header(ws_singletons, singleton_headers)

    # Get singleton data from flat_keywords (cluster size == 1)
    singleton_rows = [
        kw_row for kw_row in flat_keywords
        if int(kw_row.get("cluster_size", 0)) == 1
    ]
    # Fallback to unclustered list if flat data doesn't have cluster_size
    if not singleton_rows and unclustered:
        for row_idx, keyword in enumerate(unclustered, start=2):
            ws_singletons.cell(row=row_idx, column=1, value=keyword)
            ws_singletons.cell(row=row_idx, column=2, value="")
            _apply_row_style(ws_singletons, row_idx, len(singleton_headers))
    else:
        for row_idx, kw_row in enumerate(singleton_rows, start=2):
            ws_singletons.cell(row=row_idx, column=1, value=kw_row.get("keyword", ""))
            ws_singletons.cell(row=row_idx, column=2, value=kw_row.get("demand", 0))
            _apply_row_style(ws_singletons, row_idx, len(singleton_headers))
    _auto_width(ws_singletons, len(singleton_headers))

    # ── Sheet 4: Summary ─────────────────────────────────────────────────
    ws_summary = wb.create_sheet("Summary")
    ws_summary.sheet_properties.tabColor = "10B981"
    summary_headers = ["Metric", "Value"]
    _apply_header(ws_summary, summary_headers)

    summary_rows = [
        ("Total Keywords (input)", summary.get("keywords_input_total", 0)),
        ("Unique Keywords", summary.get("keywords_unique_total", 0)),
        ("Duplicates Removed", summary.get("duplicates_removed", 0)),
        ("Total Demand (input)", summary.get("input_demand_total", 0)),
        ("Unique Demand", summary.get("unique_demand_total", 0)),
        ("Clusters Total", summary.get("clusters_total", 0)),
        ("Primary Clusters (multi-keyword)", summary.get("primary_clusters_total", 0)),
        ("Multi-keyword Clusters", summary.get("multi_keyword_clusters", 0)),
        ("Singleton Clusters", summary.get("singleton_clusters", 0)),
        ("Biggest Cluster Size", summary.get("biggest_cluster_size", 0)),
        ("Average Cluster Size", summary.get("avg_cluster_size", 0)),
        ("Average Cluster Cohesion", summary.get("avg_cluster_cohesion", 0)),
        ("High Quality Clusters", summary.get("high_quality_clusters", 0)),
        ("Low Confidence Keywords", summary.get("low_confidence_keywords", 0)),
        ("Primary Demand Share %", summary.get("primary_demand_share_pct", 0)),
        ("Singleton Demand Share %", summary.get("singleton_demand_share_pct", 0)),
        ("Top Cluster Demand Share %", summary.get("top_cluster_demand_share_pct", 0)),
    ]
    for row_idx, (metric, value) in enumerate(summary_rows, start=2):
        ws_summary.cell(row=row_idx, column=1, value=metric)
        ws_summary.cell(row=row_idx, column=2, value=value)
        _apply_row_style(ws_summary, row_idx, len(summary_headers))
    _auto_width(ws_summary, len(summary_headers))

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@router.get("/tasks/clusterizer/{task_id}/export/xlsx")
async def export_clusterizer_xlsx(task_id: str):
    """Export clusterizer results as XLSX file."""
    task_data = get_task_result(task_id)
    if not task_data:
        raise HTTPException(status_code=404, detail="Task not found")

    status = str(task_data.get("status", "")).upper()
    if status != "SUCCESS":
        raise HTTPException(
            status_code=400,
            detail=f"Task is not completed yet (status: {status})",
        )

    xlsx_bytes = _build_clusterizer_xlsx(task_data)

    safe_id = task_id.replace("/", "_").replace("\\", "_")
    filename = f"clusterizer_{safe_id}.xlsx"

    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
