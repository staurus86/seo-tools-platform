"""
Link Profile Audit router.
"""
import io
from datetime import datetime
from typing import Any, Dict, List, Optional, Set

from fastapi import APIRouter, BackgroundTasks, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse

from app.validators import normalize_http_input as _normalize_http_input
from app.api.routers._task_store import create_task_pending, update_task_state, get_task_result

router = APIRouter(tags=["SEO Tools"])


@router.post("/tasks/link-profile-audit")
async def create_link_profile_audit(
    background_tasks: BackgroundTasks,
    our_domain: str = Form(...),
    backlink_files: List[UploadFile] = File(...),
    batch_file: Optional[UploadFile] = File(None),
    commercial_keywords: str = Form(""),
    informational_keywords: str = Form(""),
    spam_keywords: str = Form(""),
    brand_keywords: str = Form(""),
):
    """Backlink profile audit from uploaded backlink exports and one batch file."""
    from app.config import settings
    from app.tools.link_profile import run_link_profile_audit

    domain = _normalize_http_input(our_domain) if "://" in str(our_domain or "") else str(our_domain or "").strip()
    if not str(domain).strip():
        raise HTTPException(status_code=422, detail="Укажите домен проекта")
    if not backlink_files:
        raise HTTPException(status_code=422, detail="Добавьте хотя бы один файл бэклинков")
    # batch_file is optional for all-in-one XLSX packs (e.g. test links.xlsx).

    max_backlink_files = max(1, min(200, int(getattr(settings, "LINK_PROFILE_MAX_BACKLINK_FILES", 20) or 20)))
    max_file_size_bytes = max(1, int(getattr(settings, "LINK_PROFILE_MAX_FILE_SIZE_BYTES", 25 * 1024 * 1024) or (25 * 1024 * 1024)))
    max_batch_file_size_bytes = max(
        1,
        int(getattr(settings, "LINK_PROFILE_MAX_BATCH_FILE_SIZE_BYTES", 50 * 1024 * 1024) or (50 * 1024 * 1024)),
    )
    max_total_upload_bytes = max(
        max_file_size_bytes,
        int(getattr(settings, "LINK_PROFILE_MAX_TOTAL_UPLOAD_BYTES", 150 * 1024 * 1024) or (150 * 1024 * 1024)),
    )
    if len(backlink_files) > max_backlink_files:
        raise HTTPException(status_code=422, detail=f"Слишком много файлов бэклинков: максимум {max_backlink_files}")

    allowed_ext = (".csv", ".xlsx")
    backlog_payloads: List[tuple[str, bytes]] = []
    total_upload_bytes = 0
    for up in backlink_files:
        name = str(up.filename or "")
        if not name.lower().endswith(allowed_ext):
            raise HTTPException(status_code=422, detail=f"Неподдерживаемый формат файла: {name}")
        blob = await up.read()
        if not blob:
            raise HTTPException(status_code=422, detail=f"Пустой файл: {name}")
        if len(blob) > max_file_size_bytes:
            raise HTTPException(
                status_code=422,
                detail=f"Файл {name} превышает лимит {max_file_size_bytes} байт",
            )
        total_upload_bytes += len(blob)
        if total_upload_bytes > max_total_upload_bytes:
            raise HTTPException(
                status_code=422,
                detail=f"Суммарный размер файлов превышает лимит {max_total_upload_bytes} байт",
            )
        backlog_payloads.append((name, blob))

    batch_name = ""
    batch_payload = b""
    batch_upload = batch_file if getattr(batch_file, "read", None) and getattr(batch_file, "filename", None) is not None else None
    if batch_upload and str(batch_upload.filename or "").strip():
        batch_name = str(batch_upload.filename or "")
        if not batch_name.lower().endswith(allowed_ext):
            raise HTTPException(status_code=422, detail="Batch файл должен быть .csv или .xlsx")
        batch_payload = await batch_upload.read()
        if not batch_payload:
            raise HTTPException(status_code=422, detail="Batch файл пустой")
        if len(batch_payload) > max_batch_file_size_bytes:
            raise HTTPException(
                status_code=422,
                detail=f"Batch файл превышает лимит {max_batch_file_size_bytes} байт",
            )
        total_upload_bytes += len(batch_payload)
        if total_upload_bytes > max_total_upload_bytes:
            raise HTTPException(
                status_code=422,
                detail=f"Суммарный размер файлов превышает лимит {max_total_upload_bytes} байт",
            )

    task_id = f"link-profile-{datetime.now().timestamp()}"
    create_task_pending(task_id, "link_profile_audit", str(our_domain or "").strip(), status_message="Задача поставлена в очередь")

    def _run_link_profile_task() -> None:
        try:
            update_task_state(task_id, status="RUNNING", progress=10, status_message="Подготовка данных для анализа")

            def _progress(progress: int, message: str) -> None:
                update_task_state(
                    task_id,
                    status="RUNNING",
                    progress=progress,
                    status_message=message,
                )

            result = run_link_profile_audit(
                our_domain=str(our_domain or "").strip(),
                backlink_files=backlog_payloads,
                batch_file=(batch_name, batch_payload) if batch_name else None,
                commercial_keywords=commercial_keywords,
                informational_keywords=informational_keywords,
                spam_keywords=spam_keywords,
                brand_keywords=brand_keywords,
                progress_callback=_progress,
            )
            update_task_state(
                task_id,
                status="SUCCESS",
                progress=100,
                status_message="Аудит ссылочного профиля завершен",
                result=result,
                error=None,
            )
        except Exception as exc:
            update_task_state(
                task_id,
                status="FAILURE",
                progress=100,
                status_message="Аудит ссылочного профиля завершился с ошибкой",
                error=str(exc),
            )

    background_tasks.add_task(_run_link_profile_task)
    return {"task_id": task_id, "status": "PENDING", "message": "Аудит ссылочного профиля запущен"}


# ---------------------------------------------------------------------------
# Helper: extract task result or raise
# ---------------------------------------------------------------------------

def _get_completed_task(task_id: str) -> Dict[str, Any]:
    """Return task data dict; raise HTTPException if not found / not complete."""
    task_data = get_task_result(task_id)
    if not task_data:
        raise HTTPException(status_code=404, detail="Task not found")
    status = str(task_data.get("status", "")).upper()
    if status != "SUCCESS":
        raise HTTPException(status_code=400, detail=f"Task is not completed yet (status: {status})")
    return task_data


# ---------------------------------------------------------------------------
# GET /api/tasks/link-profile/{task_id}/disavow
# ---------------------------------------------------------------------------

@router.get("/tasks/link-profile/{task_id}/disavow")
async def get_disavow_file(task_id: str, threshold: float = 55.0):
    """Return Google Disavow file as text/plain download."""
    task_data = _get_completed_task(task_id)
    result = task_data.get("result", {}) or {}
    results = result.get("results", {}) or {}
    tables = results.get("tables", {}) or {}

    # Try pre-computed disavow first; if threshold differs, regenerate
    disavow = tables.get("disavow_preview", {}) or {}
    if not disavow.get("generated") or disavow.get("threshold_used") != threshold:
        risk_signals = tables.get("risk_signals", []) or []
        from app.tools.link_profile.service_v1 import _generate_disavow
        disavow = _generate_disavow(risk_signals, threshold=threshold)

    if not disavow.get("generated"):
        raise HTTPException(status_code=404, detail=disavow.get("reason", "No toxic domains found"))

    content = disavow.get("content", "")
    safe_id = task_id.replace("/", "_").replace("\\", "_")
    filename = f"disavow_{safe_id}.txt"

    return StreamingResponse(
        io.BytesIO(content.encode("utf-8")),
        media_type="text/plain; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# GET /api/tasks/link-profile/{task_id}/export/xlsx  — Full XLSX report
# ---------------------------------------------------------------------------

def _build_link_profile_export_xlsx(task_data: Dict[str, Any]) -> bytes:
    """Build a comprehensive XLSX workbook for link profile results."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    result = task_data.get("result", {}) or {}
    results = result.get("results", {}) or {}
    tables = results.get("tables", {}) or {}
    summary = results.get("summary", {}) or {}
    url = result.get("url", "")

    wb = Workbook()

    # --- Styles ---
    header_fill = PatternFill(start_color="0F4C81", end_color="0F4C81", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    zebra_fill = PatternFill(start_color="F8FBFF", end_color="F8FBFF", fill_type="solid")

    def _write_sheet(
        name: str,
        rows: List[Dict[str, Any]],
        headers: Optional[List[str]] = None,
    ) -> None:
        ws = wb.create_sheet(name[:31])
        ws.sheet_properties.tabColor = "0F4C81"
        rows = rows or []
        if not headers:
            headers = list((rows[0] or {}).keys()) if rows else ["Info"]
        # Header row
        for col_idx, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col_idx, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = header_align
            c.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = min(40, max(12, len(str(h)) + 4))
        # Data rows
        if not rows:
            ws.cell(row=2, column=1, value="No data").border = thin_border
            return
        for r_idx, item in enumerate(rows, 2):
            for c_idx, h in enumerate(headers, 1):
                val = item.get(h) if isinstance(item, dict) else None
                if val is None and isinstance(item, dict):
                    norm = {str(k).strip().lower(): v for k, v in item.items()}
                    val = norm.get(str(h).strip().lower())
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = thin_border
                if r_idx % 2 == 0:
                    cell.fill = zebra_fill

    # --- Cover ---
    ws_cover = wb.active
    ws_cover.title = "Cover"
    ws_cover.sheet_properties.tabColor = "0F4C81"
    brand_fill_cover = PatternFill(start_color="0F4C81", end_color="0F4C81", fill_type="solid")
    for col in range(1, 8):
        ws_cover.cell(row=1, column=col).fill = brand_fill_cover
    title_cell = ws_cover.cell(row=4, column=2, value="Link Profile Report")
    title_cell.font = Font(size=24, bold=True, color="0F4C81")
    ws_cover.cell(row=5, column=2, value="SEO Tools Platform").font = Font(size=12, color="94A3B8")
    if url:
        ws_cover.cell(row=7, column=2, value=f"Domain: {url}").font = Font(size=11, color="0E7490")
    ws_cover.cell(row=8, column=2, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(size=10, color="94A3B8")
    ws_cover.column_dimensions["B"].width = 60
    ws_cover.sheet_view.showGridLines = False

    # --- 1. Summary ---
    summary_rows: List[Dict[str, Any]] = []
    for k, v in summary.items():
        summary_rows.append({"Metric": k, "Value": v})
    _write_sheet("Summary", summary_rows, ["Metric", "Value"])

    # --- 2. Competitor Benchmark ---
    _write_sheet(
        "Competitor Benchmark",
        tables.get("competitor_benchmark", []) or [],
        [
            "Домен", "Domain Rating", "Total Backlinks", "Referring Domains",
            "Follow %", "Lost %", "HTTP 2xx %", "Homepage %",
            "UGC %", "Sponsored %", "Rendered %", "Quality Score",
        ],
    )

    # --- 3. Gap Donors ---
    _write_sheet(
        "Gap Donors",
        tables.get("gap_donors_priority", []) or [],
        [
            "Domain", "Competitors Covered", "Coverage %",
            "Avg DR", "Avg Traffic", "Follow %", "Lost %", "Opportunity Score",
        ],
    )

    # --- 4. Anchor Analysis ---
    _write_sheet(
        "Anchor Analysis",
        tables.get("anchor_analysis", []) or [],
        ["anchor", "lemma", "count"],
    )

    # --- 5. Risk Signals ---
    _write_sheet(
        "Risk Signals",
        tables.get("risk_signals", []) or [],
        [
            "Domain", "Links", "Avg DR", "Avg Traffic",
            "Avg External Links", "Lost %", "Nofollow %",
            "Sponsored %", "Risk Score", "Risk Level",
        ],
    )

    # --- 6. Link Velocity (if available) ---
    velocity_rows = tables.get("link_velocity_rows", []) or []
    if velocity_rows:
        _write_sheet(
            "Link Velocity",
            velocity_rows,
            ["month", "gained", "lost", "net"],
        )

    # --- 7. Disavow Preview ---
    disavow = tables.get("disavow_preview", {}) or {}
    if disavow.get("generated"):
        disavow_content = str(disavow.get("content", ""))
        disavow_lines = disavow_content.split("\n")
        disavow_rows = [{"line": line} for line in disavow_lines if line.strip()]
        ws_dis = wb.create_sheet("Disavow Preview")
        ws_dis.sheet_properties.tabColor = "0F4C81"
        ws_dis.cell(row=1, column=1, value=f"Domains: {disavow.get('domains_count', 0)}").font = Font(bold=True, color="0F4C81")
        ws_dis.cell(row=2, column=1, value=f"Threshold: {disavow.get('threshold_used', 55.0)}").font = Font(color="94A3B8")
        for r_idx, row in enumerate(disavow_rows, 4):
            ws_dis.cell(row=r_idx, column=1, value=row["line"]).border = thin_border
        ws_dis.column_dimensions["A"].width = 60
    else:
        _write_sheet("Disavow Preview", [], ["Info"])

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf.getvalue()


@router.get("/tasks/link-profile/{task_id}/export/xlsx")
async def export_link_profile_xlsx(task_id: str):
    """Export link profile audit results as XLSX file."""
    task_data = _get_completed_task(task_id)
    xlsx_bytes = _build_link_profile_export_xlsx(task_data)
    safe_id = task_id.replace("/", "_").replace("\\", "_")
    filename = f"link_profile_{safe_id}.xlsx"
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
