import csv
import io
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from statistics import mean, median
from typing import Any, Callable, Dict, List, Optional, Sequence, Tuple
from urllib.parse import urlparse

from openpyxl import load_workbook

ProgressCallback = Optional[Callable[[int, str], None]]
MAX_LINK_PROFILE_ROWS = 120000
MAX_RAW_TABLE_ROWS = 5000
MAX_RESULT_TABLE_ROWS = 10000


_MULTI_PART_TLDS = {
    "co.uk",
    "org.uk",
    "gov.uk",
    "ac.uk",
    "co.jp",
    "ne.jp",
    "or.jp",
    "com.au",
    "net.au",
    "org.au",
    "co.nz",
    "com.br",
    "com.tr",
    "co.kr",
    "com.mx",
    "co.in",
}

DEFAULT_COMMERCIAL_KEYWORDS = [
    "купить", "цена", "заказать", "промо", "скидка", "акция", "доставка", "покупка", "заказ", "оплата",
    "магазин", "товар", "онлайн", "online", "sale", "deals", "discount", "shop", "coupon", "code", "код",
    "бесплатно", "бесплатный", "cheap", "outlet", "clearance", "new", "arrivals", "promo", "promotion",
    "offer", "deal", "buy", "order", "purchase", "delivery", "save", "bargain", "price", "cart",
    "checkout", "special", "limited", "exclusive", "subscription", "подписка", "product", "item", "store",
    "market", "маркет",
]

DEFAULT_INFORMATIONAL_KEYWORDS = [
    "как", "что", "почему", "статья", "руководство", "обзор", "советы", "инструкция", "гид", "обучение",
    "уроки", "tutorial", "guide", "tips", "review", "how", "what", "why", "learn", "explained", "faq",
    "вопросы", "ответы", "информация", "info", "подробно", "анализ", "исследование", "блог", "blog",
    "article", "news", "новости", "история", "history", "факты", "facts", "причины", "решение", "solution",
    "совет", "рекомендации", "best", "top", "рейтинг", "список", "list", "объяснение", "description", "описание",
]

DEFAULT_SPAM_KEYWORDS = [
    "seo", "backlinks", "backlinking", "traffic", "boost", "ranking", "indexing", "dofollow", "black", "white",
    "grey", "linkbuilding", "links", "purchased", "buy links", "sell links", "click here", "click now",
    "free traffic", "fast results", "guaranteed ranking", "telegram", "tg", "gambling", "casino", "bet",
    "betting", "stavka", "фриспин", "freespin", "viagra", "cialis", "pharmacy", "drugs", "hacked", "взлом",
    "взломка", "взломанный", "torrent", "торрент", "pirate", "crack", "keygen", "adult", "xxx", "porn",
    "escort", "webcam", "cheap pills", "earn money", "make money", "crypto scam", "investment scam",
    "get rich", "lottery", "free gift", "win prize", "survey", "spam", "↑↑↑", "??????",
]

DEFAULT_NAVIGATIONAL_KEYWORDS = [
    "официальный сайт", "официальный", "сайт", "homepage", "home", "главная", "main page",
    "контакты", "contact", "contacts", "вход", "login", "sign in", "signin", "личный кабинет",
    "кабинет", "профиль", "account", "profile", "каталог", "catalog", "раздел", "поддержка",
    "support", "help", "faq", "about", "about us", "о компании", "menu",
]

DEFAULT_GENERIC_KEYWORDS = [
    "тут", "здесь", "подробнее", "читать", "далее", "перейти", "ссылка", "линк", "here", "click",
    "click here", "read more", "learn more", "more", "source", "website", "site", "visit",
]


def _to_registrable_domain(host: str) -> str:
    value = (host or "").strip().lower().rstrip(".")
    if not value:
        return ""
    if ":" in value:
        value = value.split(":", 1)[0]
    if value.startswith("www."):
        value = value[4:]
    if re.fullmatch(r"\d{1,3}(?:\.\d{1,3}){3}", value):
        return value
    parts = [p for p in value.split(".") if p]
    if len(parts) <= 2:
        return value
    tail2 = ".".join(parts[-2:])
    tail3 = ".".join(parts[-3:])
    if tail2 in _MULTI_PART_TLDS and len(parts) >= 3:
        return tail3
    if re.fullmatch(r"[a-z]{2}", parts[-1]) and parts[-2] in {"co", "com", "net", "org", "gov", "edu", "ac"} and len(parts) >= 3:
        return tail3
    return tail2


def _parse_keywords(raw: str) -> List[str]:
    if not raw:
        return []
    parts = re.split(r"[,;\n\r\t]+", str(raw))
    return [p.strip().lower() for p in parts if p and p.strip()]


def _merge_keywords(base: List[str], extra: List[str]) -> List[str]:
    out: List[str] = []
    seen = set()
    for token in list(base or []) + list(extra or []):
        t = str(token or "").strip().lower()
        if not t or t in seen:
            continue
        seen.add(t)
        out.append(t)
    return out


def _normalize_domain(raw: str) -> str:
    value = (raw or "").strip().lower()
    if not value:
        return ""
    if "//" not in value:
        value = f"https://{value}"
    parsed = urlparse(value)
    host = (parsed.netloc or "").lower().strip()
    return _to_registrable_domain(host)


def _extract_domain(raw: str) -> str:
    value = (raw or "").strip()
    if not value:
        return ""
    if "//" not in value:
        value = f"https://{value}"
    try:
        parsed = urlparse(value)
        host = (parsed.netloc or "").lower().strip()
        return _to_registrable_domain(host)
    except Exception:
        return ""


def _is_our_target(target_domain: str, our_domain: str) -> bool:
    if not target_domain or not our_domain:
        return False
    return target_domain == our_domain or target_domain.endswith(f".{our_domain}")


def _decode_text(payload: bytes) -> str:
    if payload.startswith(b"\xef\xbb\xbf"):
        try:
            return payload.decode("utf-8-sig")
        except Exception:
            pass
    if payload.startswith((b"\xff\xfe", b"\xfe\xff")):
        try:
            return payload.decode("utf-16")
        except Exception:
            pass

    # Heuristic: frequent NUL bytes usually indicate UTF-16 payloads.
    head = payload[:4096]
    nul_ratio = (head.count(b"\x00") / max(1, len(head)))
    if nul_ratio > 0.1:
        for enc in ("utf-16", "utf-16-le"):
            try:
                return payload.decode(enc)
            except Exception:
                continue

    for enc in ("utf-8-sig", "utf-8", "cp1251", "latin-1", "utf-16", "utf-16-le"):
        try:
            return payload.decode(enc)
        except Exception:
            continue
    return payload.decode("utf-8", errors="replace")


def _read_csv_rows(payload: bytes, max_rows: Optional[int] = None) -> List[Dict[str, Any]]:
    text = _decode_text(payload)
    stream = io.StringIO(text, newline="")
    first_line = ""
    for line in text.splitlines():
        if line.strip():
            first_line = line
            break
    # Fast path for Ahrefs exports: UTF-16 + tab-delimited with stable headers.
    delimiter_override: Optional[str] = None
    if first_line and first_line.count("\t") >= max(first_line.count(","), first_line.count(";"), first_line.count("|")):
        dialect = csv.excel
        delimiter_override = "\t"
    else:
        sample = text[:8192]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except Exception:
            dialect = csv.excel
    try:
        if delimiter_override:
            reader = csv.DictReader(stream, dialect=dialect, delimiter=delimiter_override)
        else:
            reader = csv.DictReader(stream, dialect=dialect)
        rows: List[Dict[str, Any]] = []
        for row in reader:
            rows.append(dict(row))
            if max_rows is not None and len(rows) >= max_rows:
                break
        return rows
    except csv.Error:
        # Fallback for malformed CSV where fields contain raw newlines.
        lines = text.splitlines()
        if not lines:
            return []
        header = lines[0]
        delim_candidates = [("\t", header.count("\t")), (";", header.count(";")), (",", header.count(",")), ("|", header.count("|"))]
        delim = max(delim_candidates, key=lambda x: x[1])[0]
        headers = [h.strip() for h in lines[0].split(delim)]
        rows: List[Dict[str, Any]] = []
        for line in lines[1:]:
            if not line.strip():
                continue
            parts = line.split(delim)
            if len(parts) < len(headers):
                parts = parts + [""] * (len(headers) - len(parts))
            elif len(parts) > len(headers):
                parts = parts[: len(headers) - 1] + [delim.join(parts[len(headers) - 1 :])]
            rows.append({headers[i]: parts[i] for i in range(len(headers))})
            if max_rows is not None and len(rows) >= max_rows:
                break
        return rows


def _read_xlsx_rows(payload: bytes, max_rows: Optional[int] = None) -> List[Dict[str, Any]]:
    wb = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    ws = wb.active
    iterator = ws.iter_rows(values_only=True)
    header_row = next(iterator, None)
    if not header_row:
        return []
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    rows: List[Dict[str, Any]] = []
    for values in iterator:
        item: Dict[str, Any] = {}
        for idx, key in enumerate(headers):
            if not key:
                continue
            item[key] = values[idx] if idx < len(values) else None
        rows.append(item)
        if max_rows is not None and len(rows) >= max_rows:
            break
    return rows


def _worksheet_to_rows(ws, max_rows: Optional[int] = None) -> List[Dict[str, Any]]:
    iterator = ws.iter_rows(values_only=True)
    header_row = next(iterator, None)
    if not header_row:
        return []
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    rows: List[Dict[str, Any]] = []
    for values in iterator:
        if values is None:
            continue
        item: Dict[str, Any] = {}
        non_empty = False
        for idx, key in enumerate(headers):
            if not key:
                continue
            value = values[idx] if idx < len(values) else None
            if value not in (None, ""):
                non_empty = True
            item[key] = value
        if non_empty:
            rows.append(item)
            if max_rows is not None and len(rows) >= max_rows:
                break
    return rows


def _parse_analysis_data_sections(ws) -> List[Dict[str, Any]]:
    sections: List[Dict[str, Any]] = []
    header_markers = {"домен", "competitor"}
    current_header: List[str] = []
    current_rows: List[Dict[str, Any]] = []

    def _finalize():
        nonlocal current_header, current_rows
        if current_header and current_rows:
            title = " / ".join(current_header[:3])
            sections.append({"title": title, "rows": current_rows})
        current_header = []
        current_rows = []

    iterator = ws.iter_rows(values_only=True)
    for row in iterator:
        compact = [v for v in row if v not in (None, "")]
        if not compact:
            _finalize()
            continue
        first = str(compact[0]).strip().lower()
        if first in header_markers and len(compact) >= 2:
            _finalize()
            current_header = [str(x).strip() for x in compact]
            continue
        if current_header:
            rec: Dict[str, Any] = {}
            for idx, h in enumerate(current_header):
                rec[h] = compact[idx] if idx < len(compact) else None
            current_rows.append(rec)
    _finalize()
    return sections


def _read_test_links_pack(
    payload: bytes,
    *,
    max_backlink_rows: Optional[int] = None,
    max_batch_rows: Optional[int] = None,
    max_priority_rows: Optional[int] = None,
) -> Dict[str, List[Dict[str, Any]]]:
    wb = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    sheet_map = {str(n): wb[n] for n in wb.sheetnames}
    backlinks_rows: List[Dict[str, Any]] = []
    batch_rows: List[Dict[str, Any]] = []
    priority_rows: List[Dict[str, Any]] = []
    analysis_sections: List[Dict[str, Any]] = []

    backlink_sheets = ["Ссылки с конкурентов", "Дубли без нашего сайта", "Ссылки с главных", "Ссылки с редиректов"]
    for s in backlink_sheets:
        ws = sheet_map.get(s)
        if not ws:
            continue
        remaining = None
        if max_backlink_rows is not None:
            remaining = max(0, max_backlink_rows - len(backlinks_rows))
            if remaining <= 0:
                break
        for row in _worksheet_to_rows(ws, max_rows=remaining):
            row["__sheet__"] = s
            backlinks_rows.append(row)

    for s in ("RU", "Все"):
        ws = sheet_map.get(s)
        if not ws:
            continue
        remaining = None
        if max_batch_rows is not None:
            remaining = max(0, max_batch_rows - len(batch_rows))
            if remaining <= 0:
                break
        batch_rows.extend(_worksheet_to_rows(ws, max_rows=remaining))

    ws_prio = sheet_map.get("Приоритетные доноры")
    if ws_prio:
        priority_rows.extend(_worksheet_to_rows(ws_prio, max_rows=max_priority_rows))
    ws_analysis = sheet_map.get("Анализ данных")
    if ws_analysis:
        analysis_sections.extend(_parse_analysis_data_sections(ws_analysis))

    return {
        "backlinks_rows": backlinks_rows,
        "batch_rows": batch_rows,
        "priority_rows": priority_rows,
        "analysis_sections": analysis_sections,
    }


def _read_tabular_rows(filename: str, payload: bytes, max_rows: Optional[int] = None) -> List[Dict[str, Any]]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        return _read_csv_rows(payload, max_rows=max_rows)
    if lower.endswith(".xlsx"):
        return _read_xlsx_rows(payload, max_rows=max_rows)
    raise ValueError(f"Unsupported file format: {filename}")


def _pick_value(row: Dict[str, Any], aliases: Sequence[str]) -> Any:
    norm = {str(k).strip().lower(): v for k, v in row.items()}
    for alias in aliases:
        if alias in norm:
            return norm.get(alias)
    return None


def _to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    raw = str(value).strip().replace(",", ".")
    if not raw:
        return None
    try:
        return float(raw)
    except Exception:
        return None


def _to_follow_bool(value: Any) -> Optional[bool]:
    if value is None:
        return None
    raw = str(value).strip().lower()
    if not raw:
        return None
    if raw in {"dofollow", "follow", "true", "1", "yes"}:
        return True
    if raw in {"nofollow", "no-follow", "false", "0", "no"}:
        return False
    return None


def _to_flag_bool(value: Any) -> Optional[bool]:
    if value is None:
        return None
    raw = str(value).strip().lower()
    if not raw:
        return None
    if raw in {"1", "true", "yes", "y", "on"}:
        return True
    if raw in {"0", "false", "no", "n", "off"}:
        return False
    if raw in {"nofollow", "ugc", "sponsored"}:
        return True
    return None


def _to_iso_date(value: Any) -> str:
    if value is None:
        return ""
    raw = str(value).strip()
    if not raw:
        return ""
    # Accept ISO-like formats and keep only date part for consistent aggregation.
    if len(raw) >= 10 and re.match(r"^\d{4}-\d{2}-\d{2}", raw):
        return raw[:10]
    if len(raw) >= 10 and re.match(r"^\d{2}\.\d{2}\.\d{4}", raw):
        d, m, y = raw[:10].split(".")
        return f"{y}-{m}-{d}"
    return raw[:10]


def _classify_anchor(anchor: str, keywords: Dict[str, List[str]]) -> str:
    text = re.sub(r"\s+", " ", str(anchor or "").strip().lower())
    if not text:
        return "empty"

    if re.match(r"^(https?://|www\.)", text) or re.fullmatch(r"[a-z0-9.-]+\.[a-z]{2,}(?:/[^\s]*)?", text):
        return "naked_url"

    def _has_any(items: List[str]) -> bool:
        return any(k and k in text for k in items)

    has_spam = _has_any(keywords.get("spam", []))
    has_brand = _has_any(keywords.get("brand", []))
    has_commercial = _has_any(keywords.get("commercial", []))
    has_info = _has_any(keywords.get("informational", []))
    has_nav = _has_any(keywords.get("navigational", []))
    has_generic = _has_any(keywords.get("generic", []))

    if has_spam:
        return "spam"
    if has_brand and has_commercial:
        return "brand_commercial"
    if has_brand and has_nav:
        return "brand_navigational"
    if has_brand:
        return "brand"
    if has_nav:
        return "navigational"
    if has_commercial:
        return "commercial"
    if has_info:
        return "informational"
    if has_generic or len(text) <= 3:
        return "generic"
    return "other"


def _derive_brand_keywords(domain: str) -> List[str]:
    host = (domain or "").lower().strip()
    if not host:
        return []
    host = host.split(":")[0]
    if host.startswith("www."):
        host = host[4:]
    main = host.split(".")[0]
    chunks = re.split(r"[^a-z0-9а-яё]+", main, flags=re.IGNORECASE)
    out = [host, main]
    out.extend(chunks)
    seen = set()
    result: List[str] = []
    for item in out:
        token = (item or "").strip().lower()
        if len(token) < 2 or token in seen:
            continue
        seen.add(token)
        result.append(token)
    return result


def _brand_token_from_domain(domain: str) -> str:
    host = _normalize_domain(domain)
    if not host:
        return ""
    parts = host.split(".")
    if len(parts) >= 2:
        token = parts[-2]
    else:
        token = parts[0]
    token = re.sub(r"[^a-z0-9а-яё]+", "", token.lower(), flags=re.IGNORECASE)
    return token


def _extract_zone(domain: str) -> str:
    parts = (domain or "").split(".")
    if len(parts) < 2:
        return "other"
    zone = parts[-1].lower()
    return zone if 1 < len(zone) <= 10 else "other"


def _dr_bucket(dr: Optional[float]) -> str:
    if dr is None:
        return "unknown"
    if dr < 10:
        return "0-9"
    if dr < 30:
        return "10-29"
    if dr < 50:
        return "30-49"
    if dr < 70:
        return "50-69"
    return "70+"


def _dr_bucket_decile(dr: Optional[float]) -> str:
    if dr is None:
        return "DR unknown"
    value = max(0.0, min(100.0, float(dr)))
    low = int(value // 10) * 10
    if low >= 90:
        return "DR 90-100"
    return f"DR {low}-{low + 9}"


def _anchor_words(anchor: str) -> List[str]:
    return [w.lower() for w in re.findall(r"[A-Za-zА-Яа-яЁё0-9]{3,}", anchor or "")]


def _has_redirect_301(*values: Any) -> bool:
    for value in values:
        text = str(value or "").strip().lower()
        if not text:
            continue
        if re.search(r"(?<!\d)301(?!\d)", text):
            return True
    return False


def _is_homepage_url(url: str) -> bool:
    if not url:
        return False
    try:
        parsed = urlparse(url if "://" in url else f"https://{url}")
        path = (parsed.path or "").strip()
        return path in ("", "/")
    except Exception:
        return False


def _flatten_breakdown_rows(rows: List[Dict[str, Any]], *, dim: str, value_key: str, row_limit: int = 500) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    grouped: Dict[str, Counter[str]] = defaultdict(Counter)
    for row in rows:
        comp = str(row.get("target_domain") or "")
        if not comp:
            continue
        value = str(row.get(value_key) or "").strip().lower()
        if not value:
            value = "unknown"
        grouped[comp][value] += 1
    for comp, cnt in grouped.items():
        total = sum(cnt.values()) or 1
        for v, c in cnt.items():
            out.append(
                {
                    "competitor_domain": comp,
                    dim: v,
                    "count": c,
                    "pct": round((c / total) * 100, 2),
                }
            )
    out.sort(key=lambda x: (x.get("competitor_domain", ""), -(x.get("count", 0))))
    return out[:row_limit]


def _pct(part: float, total: float) -> float:
    return round((float(part) / max(1.0, float(total))) * 100.0, 2)


def _counter_to_pct_rows(counter: Counter[str], *, value_key: str, limit: int = 50) -> List[Dict[str, Any]]:
    total = sum(counter.values()) or 1
    rows: List[Dict[str, Any]] = []
    for k, v in counter.most_common(limit):
        rows.append(
            {
                value_key: k,
                "count": int(v),
                "pct": _pct(v, total),
            }
        )
    return rows


def _normalize_backlink_row(row: Dict[str, Any]) -> Dict[str, Any]:
    page_title = _pick_value(
        row,
        ("referring page title", "page title", "title"),
    )
    source = _pick_value(
        row,
        (
            "source_url",
            "referring page url",
            "referring_url",
            "source",
            "url_from",
            "from",
            "referrer",
            "referring page",
            "source page",
            "url",
        ),
    )
    target = _pick_value(
        row,
        (
            "target_url",
            "target url",
            "target",
            "url_to",
            "to",
            "destination",
            "target page",
            "url target",
        ),
    )
    anchor = _pick_value(
        row,
        ("anchor", "anchor_text", "text", "anchor text"),
    )
    follow = _pick_value(
        row,
        ("follow", "link_type", "type", "rel", "is_dofollow", "dofollow"),
    )
    nofollow = _pick_value(
        row,
        ("nofollow",),
    )
    redirect_status_codes = _pick_value(
        row,
        ("redirect chain status codes", "redirect_status_codes", "redirect status", "status codes"),
    )
    link_type = _pick_value(
        row,
        ("type", "link type"),
    )
    platform = _pick_value(
        row,
        ("platform",),
    )
    http_code = _pick_value(
        row,
        ("referring page http code", "http code", "status"),
    )
    language = _pick_value(
        row,
        ("language", "lang"),
    )
    lost_status = _pick_value(
        row,
        ("lost status", "drop reason", "discovered status"),
    )
    drop_reason = _pick_value(
        row,
        ("drop reason",),
    )
    discovered_status = _pick_value(
        row,
        ("discovered status",),
    )
    first_seen = _pick_value(
        row,
        ("first seen",),
    )
    last_seen = _pick_value(
        row,
        ("last seen",),
    )
    lost_flag = _pick_value(
        row,
        ("lost",),
    )
    content = _pick_value(
        row,
        ("content",),
    )
    ugc_flag = _pick_value(
        row,
        ("ugc",),
    )
    sponsored_flag = _pick_value(
        row,
        ("sponsored",),
    )
    rendered_flag = _pick_value(
        row,
        ("rendered",),
    )
    raw_flag = _pick_value(
        row,
        ("raw",),
    )
    external_links_value = _pick_value(
        row,
        ("external links",),
    )
    linked_domains_value = _pick_value(
        row,
        ("linked domains",),
    )
    referring_domains_value = _pick_value(
        row,
        ("referring domains",),
    )
    page_traffic_value = _pick_value(
        row,
        ("page traffic",),
    )
    keywords_count_value = _pick_value(
        row,
        ("keywords",),
    )
    author = _pick_value(
        row,
        ("author",),
    )
    links_in_group_value = _pick_value(
        row,
        ("links in group",),
    )
    left_context = _pick_value(
        row,
        ("left context",),
    )
    right_context = _pick_value(
        row,
        ("right context",),
    )
    redirect_chain_urls = _pick_value(
        row,
        ("redirect chain urls",),
    )
    redirect_chain_status = _pick_value(
        row,
        ("redirect chain status codes",),
    )
    ur_value = _pick_value(
        row,
        ("ur", "url rating"),
    )
    domain_rating = _pick_value(
        row,
        ("dr", "domain rating", "domain_rating", "ahrefs_dr"),
    )
    traffic = _pick_value(
        row,
        ("traffic", "domain traffic", "organic_traffic", "ahrefs_traffic"),
    )

    source_url = str(source or "").strip()
    target_url = str(target or "").strip()
    follow_flag = _to_follow_bool(follow)
    nofollow_flag = _to_follow_bool(nofollow)
    lost_flag_value = _to_flag_bool(lost_flag)
    lost_status_value = str(lost_status or "").strip().lower()
    if not lost_status_value and lost_flag_value is True:
        lost_status_value = "lost"
    if follow_flag is None and nofollow_flag is not None:
        follow_flag = not nofollow_flag
    return {
        "source_url": source_url,
        "source_domain": _extract_domain(source_url),
        "source_is_homepage": _is_homepage_url(source_url),
        "target_url": target_url,
        "target_domain": _extract_domain(target_url),
        "target_is_homepage": _is_homepage_url(target_url),
        "anchor": str(anchor or "").strip(),
        "follow": follow_flag,
        "has_redirect_301": _has_redirect_301(redirect_status_codes, redirect_chain_status, http_code),
        "link_type": str(link_type or "").strip().lower(),
        "http_code": str(http_code or "").strip(),
        "language": str(language or "").strip().lower(),
        "platform": str(platform or "").strip().lower(),
        "page_title": str(page_title or "").strip(),
        "lost_status": lost_status_value,
        "drop_reason": str(drop_reason or "").strip().lower(),
        "discovered_status": str(discovered_status or "").strip().lower(),
        "first_seen": _to_iso_date(first_seen),
        "last_seen": _to_iso_date(last_seen),
        "lost_flag": lost_flag_value,
        "content": str(content or "").strip().lower(),
        "ugc": _to_flag_bool(ugc_flag),
        "sponsored": _to_flag_bool(sponsored_flag),
        "rendered": _to_flag_bool(rendered_flag),
        "raw_flag": _to_flag_bool(raw_flag),
        "external_links": _to_float(external_links_value),
        "linked_domains": _to_float(linked_domains_value),
        "referring_domains_count": _to_float(referring_domains_value),
        "page_traffic": _to_float(page_traffic_value),
        "keywords_count": _to_float(keywords_count_value),
        "author": str(author or "").strip(),
        "links_in_group": _to_float(links_in_group_value),
        "left_context": str(left_context or "").strip(),
        "right_context": str(right_context or "").strip(),
        "redirect_chain_urls": str(redirect_chain_urls or "").strip(),
        "redirect_chain_status": str(redirect_chain_status or "").strip(),
        "ur": _to_float(ur_value),
        "dr": _to_float(domain_rating),
        "traffic": _to_float(traffic),
    }


def _normalize_batch_row(row: Dict[str, Any]) -> Tuple[str, Dict[str, Optional[float]]]:
    domain = _pick_value(row, ("domain", "target", "referring_domain", "site", "host"))
    dr = _pick_value(row, ("dr", "domain rating", "domain_rating", "ahrefs_dr"))
    traffic = _pick_value(row, ("traffic", "organic / traffic", "domain traffic", "organic_traffic", "ahrefs_traffic"))
    backlinks_all = _pick_value(row, ("backlinks / all", "backlinks_all", "total backlinks"))
    backlinks_followed = _pick_value(row, ("backlinks / followed", "backlinks_followed"))
    backlinks_not_followed = _pick_value(row, ("backlinks / not followed", "backlinks_not_followed"))
    backlinks_redirects = _pick_value(row, ("backlinks / redirects", "backlinks_redirects"))
    backlinks_internal = _pick_value(row, ("backlinks / internal", "backlinks_internal"))
    ref_domains_all = _pick_value(row, ("ref. domains / all", "ref domains all", "ref_domains_all"))
    ref_domains_followed = _pick_value(row, ("ref. domains / followed", "ref_domains_followed"))
    ref_domains_not_followed = _pick_value(row, ("ref. domains / not followed", "ref_domains_not_followed"))
    organic_keywords_total = _pick_value(row, ("organic / total keywords", "organic_keywords_total"))
    organic_keywords_top3 = _pick_value(row, ("organic / keywords (top 3)", "organic_keywords_top3"))
    organic_keywords_4_10 = _pick_value(row, ("organic / keywords (4-10)", "organic_keywords_4_10"))
    organic_keywords_11_20 = _pick_value(row, ("organic / keywords (11-20)", "organic_keywords_11_20"))
    paid_keywords = _pick_value(row, ("paid / keywords", "paid_keywords"))
    paid_traffic = _pick_value(row, ("paid / traffic", "paid_traffic"))
    outgoing_domains_followed = _pick_value(row, ("outgoing domains / followed", "outgoing_domains_followed"))
    outgoing_domains_all_time = _pick_value(row, ("outgoing domains / all time", "outgoing_domains_all_time"))
    outgoing_links_followed = _pick_value(row, ("outgoing links / followed", "outgoing_links_followed"))
    outgoing_links_all_time = _pick_value(row, ("outgoing links / all time", "outgoing_links_all_time"))
    d = _normalize_domain(str(domain or ""))
    return d, {
        "dr": _to_float(dr),
        "traffic": _to_float(traffic),
        "backlinks_all": _to_float(backlinks_all),
        "backlinks_followed": _to_float(backlinks_followed),
        "backlinks_not_followed": _to_float(backlinks_not_followed),
        "backlinks_redirects": _to_float(backlinks_redirects),
        "backlinks_internal": _to_float(backlinks_internal),
        "ref_domains_all": _to_float(ref_domains_all),
        "ref_domains_followed": _to_float(ref_domains_followed),
        "ref_domains_not_followed": _to_float(ref_domains_not_followed),
        "organic_keywords_total": _to_float(organic_keywords_total),
        "organic_keywords_top3": _to_float(organic_keywords_top3),
        "organic_keywords_4_10": _to_float(organic_keywords_4_10),
        "organic_keywords_11_20": _to_float(organic_keywords_11_20),
        "paid_keywords": _to_float(paid_keywords),
        "paid_traffic": _to_float(paid_traffic),
        "outgoing_domains_followed": _to_float(outgoing_domains_followed),
        "outgoing_domains_all_time": _to_float(outgoing_domains_all_time),
        "outgoing_links_followed": _to_float(outgoing_links_followed),
        "outgoing_links_all_time": _to_float(outgoing_links_all_time),
    }


def _is_blank_value(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _batch_row_priority_key(row: Dict[str, Any]) -> Tuple[Any, ...]:
    value_aliases: List[Sequence[str]] = [
        ("url rating", "ur"),
        ("domain rating", "dr"),
        ("ahrefs rank",),
        ("organic / total keywords", "organic_keywords_total"),
        ("organic / keywords (top 3)", "organic_keywords_top3"),
        ("organic / keywords (4-10)", "organic_keywords_4_10"),
        ("organic / keywords (11-20)", "organic_keywords_11_20"),
        ("organic / keywords (21-50)", "organic_keywords_21_50"),
        ("organic / keywords (51+)", "organic_keywords_51_plus"),
        ("organic / traffic", "organic_traffic", "traffic"),
        ("organic / value", "organic_value"),
        ("organic / top countries",),
        ("ref. domains / all", "ref domains all", "ref_domains_all"),
        ("ref. domains / followed", "ref_domains_followed"),
        ("ref. domains / not followed", "ref_domains_not_followed"),
        ("ref. ips / ips", "ref_ips_ips"),
        ("ref. ips / subnets", "ref_ips_subnets"),
        ("backlinks / all", "backlinks_all", "total backlinks"),
        ("backlinks / followed", "backlinks_followed"),
        ("backlinks / not followed", "backlinks_not_followed"),
        ("backlinks / redirects", "backlinks_redirects"),
        ("backlinks / internal", "backlinks_internal"),
        ("outgoing domains / followed", "outgoing_domains_followed"),
        ("outgoing domains / all time", "outgoing_domains_all_time"),
        ("outgoing links / followed", "outgoing_links_followed"),
        ("outgoing links / all time", "outgoing_links_all_time"),
    ]
    values = [_pick_value(row, aliases) for aliases in value_aliases]
    non_empty_count = sum(1 for value in values if not _is_blank_value(value))

    numeric_count = 0
    numeric_sum = 0.0
    for value in values:
        num = _to_float(value)
        if num is None:
            continue
        numeric_count += 1
        numeric_sum += num

    backlinks_all = _to_float(_pick_value(row, ("backlinks / all", "backlinks_all", "total backlinks")))
    ref_domains_all = _to_float(_pick_value(row, ("ref. domains / all", "ref domains all", "ref_domains_all")))
    traffic = _to_float(_pick_value(row, ("organic / traffic", "organic_traffic", "traffic")))
    dr = _to_float(_pick_value(row, ("domain rating", "dr")))

    fingerprint = "|".join(
        f"{str(k).strip().lower()}={str(v).strip()}"
        for k, v in sorted((row or {}).items(), key=lambda kv: str(kv[0]).strip().lower())
    )
    return (
        non_empty_count,
        numeric_count,
        float(backlinks_all) if backlinks_all is not None else float("-inf"),
        float(ref_domains_all) if ref_domains_all is not None else float("-inf"),
        float(traffic) if traffic is not None else float("-inf"),
        float(dr) if dr is not None else float("-inf"),
        numeric_sum,
        fingerprint,
    )


def _select_preferred_batch_rows(rows: List[Dict[str, Any]]) -> Tuple[Dict[str, Dict[str, Any]], int, int]:
    preferred_rows: Dict[str, Dict[str, Any]] = {}
    preferred_keys: Dict[str, Tuple[Any, ...]] = {}
    domain_counts: Counter[str] = Counter()

    for row in rows:
        target_raw = _pick_value(row, ("target", "domain", "site", "url"))
        domain_key = _normalize_domain(str(target_raw or ""))
        if not domain_key:
            continue
        domain_counts[domain_key] += 1
        key = _batch_row_priority_key(row)
        prev_key = preferred_keys.get(domain_key)
        if prev_key is None or key > prev_key:
            preferred_rows[domain_key] = row
            preferred_keys[domain_key] = key

    duplicate_domains = sum(1 for count in domain_counts.values() if count > 1)
    duplicate_rows = sum(max(0, count - 1) for count in domain_counts.values())
    return preferred_rows, duplicate_domains, duplicate_rows


def run_link_profile_audit(
    *,
    our_domain: str,
    backlink_files: List[Tuple[str, bytes]],
    batch_file: Optional[Tuple[str, bytes]] = None,
    commercial_keywords: str = "",
    informational_keywords: str = "",
    spam_keywords: str = "",
    brand_keywords: str = "",
    progress_callback: ProgressCallback = None,
) -> Dict[str, Any]:
    domain = _normalize_domain(our_domain)
    if not domain:
        raise ValueError("Укажите корректный домен проекта")
    if not backlink_files:
        raise ValueError("Добавьте хотя бы один файл бэклинков")

    keywords = {
        "commercial": _merge_keywords(DEFAULT_COMMERCIAL_KEYWORDS, _parse_keywords(commercial_keywords)),
        "informational": _merge_keywords(DEFAULT_INFORMATIONAL_KEYWORDS, _parse_keywords(informational_keywords)),
        "spam": _merge_keywords(DEFAULT_SPAM_KEYWORDS, _parse_keywords(spam_keywords)),
        "navigational": list(DEFAULT_NAVIGATIONAL_KEYWORDS),
        "generic": list(DEFAULT_GENERIC_KEYWORDS),
        "brand": _parse_keywords(brand_keywords),
    }
    auto_brand_tokens: List[str] = _derive_brand_keywords(domain)
    for fname, _ in backlink_files:
        base = re.sub(r"-backlinks.*$", "", str(fname or ""), flags=re.IGNORECASE)
        auto_brand_tokens.extend(_derive_brand_keywords(base))
        token = _brand_token_from_domain(base)
        if token:
            auto_brand_tokens.append(token)

    if progress_callback:
        progress_callback(15, "Чтение файла batch analysis")

    batch_name = ""
    batch_rows: List[Dict[str, Any]] = []
    preferred_batch_rows_by_domain: Dict[str, Dict[str, Any]] = {}
    batch_duplicate_domains = 0
    batch_duplicate_rows = 0
    auto_batch_rows: List[Dict[str, Any]] = []
    precomputed_priority_rows: List[Dict[str, Any]] = []
    imported_analysis_sections: List[Dict[str, Any]] = []

    if batch_file:
        batch_name, batch_payload = batch_file
        batch_rows = _read_tabular_rows(batch_name, batch_payload, max_rows=50000)
    batch_metrics: Dict[str, Dict[str, Optional[float]]] = {}
    batch_targets: List[str] = []

    if progress_callback:
        progress_callback(35, "Чтение файлов бэклинков")

    normalized_rows: List[Dict[str, Any]] = []
    raw_competitor_links_rows: List[Dict[str, Any]] = []
    raw_homepage_links_rows: List[Dict[str, Any]] = []
    raw_redirect_links_rows: List[Dict[str, Any]] = []
    raw_duplicates_without_our_rows: List[Dict[str, Any]] = []
    row_limit_hit = False
    file_summaries: List[Dict[str, Any]] = []
    for filename, payload in backlink_files:
        rows: List[Dict[str, Any]] = []
        remaining_rows = max(0, MAX_LINK_PROFILE_ROWS - len(normalized_rows))
        if remaining_rows <= 0:
            row_limit_hit = True
            break
        if str(filename or "").lower().endswith(".xlsx"):
            pack = _read_test_links_pack(
                payload,
                max_backlink_rows=remaining_rows,
                max_batch_rows=50000,
                max_priority_rows=5000,
            )
            if pack.get("backlinks_rows"):
                rows = pack.get("backlinks_rows", [])
                auto_batch_rows.extend(pack.get("batch_rows", []))
                precomputed_priority_rows.extend(pack.get("priority_rows", []))
                imported_analysis_sections.extend(pack.get("analysis_sections", []))
            else:
                rows = _read_tabular_rows(filename, payload, max_rows=remaining_rows)
        else:
            rows = _read_tabular_rows(filename, payload, max_rows=remaining_rows)
        valid_rows = 0
        for row in rows:
            if len(normalized_rows) >= MAX_LINK_PROFILE_ROWS:
                row_limit_hit = True
                break
            sheet_name = str(row.get("__sheet__") or "").strip()
            if sheet_name == "Ссылки с главных":
                if len(raw_homepage_links_rows) < MAX_RAW_TABLE_ROWS:
                    raw_homepage_links_rows.append(
                        {
                            "Referring page URL": _pick_value(row, ("referring page url",)) or "",
                            "Target URL": _pick_value(row, ("target url",)) or "",
                            "Anchor": _pick_value(row, ("anchor",)) or "",
                            "Domain Rating": _pick_value(row, ("domain rating", "dr")) or "",
                            "UR": _pick_value(row, ("ur", "url rating")) or "",
                            "Domain traffic": _pick_value(row, ("domain traffic", "traffic")) or "",
                            "Nofollow": _pick_value(row, ("nofollow",)) or "",
                            "Lost status": _pick_value(row, ("lost status",)) or "",
                        }
                    )
            elif sheet_name == "Ссылки с конкурентов":
                if len(raw_competitor_links_rows) < MAX_RAW_TABLE_ROWS:
                    raw_competitor_links_rows.append(
                        {
                            "Referring page URL": _pick_value(row, ("referring page url",)) or "",
                            "Target URL": _pick_value(row, ("target url",)) or "",
                            "Anchor": _pick_value(row, ("anchor",)) or "",
                            "Domain Rating": _pick_value(row, ("domain rating", "dr")) or "",
                            "UR": _pick_value(row, ("ur", "url rating")) or "",
                            "Domain traffic": _pick_value(row, ("domain traffic", "traffic")) or "",
                            "Nofollow": _pick_value(row, ("nofollow",)) or "",
                            "Lost status": _pick_value(row, ("lost status",)) or "",
                        }
                    )
            elif sheet_name == "Ссылки с редиректов":
                if len(raw_redirect_links_rows) < MAX_RAW_TABLE_ROWS:
                    raw_redirect_links_rows.append(
                        {
                            "Referring page URL": _pick_value(row, ("referring page url",)) or "",
                            "Target URL": _pick_value(row, ("target url",)) or "",
                            "Anchor": _pick_value(row, ("anchor",)) or "",
                            "Domain Rating": _pick_value(row, ("domain rating", "dr")) or "",
                            "UR": _pick_value(row, ("ur", "url rating")) or "",
                            "Domain traffic": _pick_value(row, ("domain traffic", "traffic")) or "",
                            "Nofollow": _pick_value(row, ("nofollow",)) or "",
                            "Lost status": _pick_value(row, ("lost status",)) or "",
                        }
                    )
            elif sheet_name == "Дубли без нашего сайта":
                if len(raw_duplicates_without_our_rows) < MAX_RAW_TABLE_ROWS:
                    raw_duplicates_without_our_rows.append(
                        {
                            "Referring page URL": _pick_value(row, ("referring page url",)) or "",
                            "Target URL": _pick_value(row, ("target url",)) or "",
                            "Anchor": _pick_value(row, ("anchor",)) or "",
                            "Domain Rating": _pick_value(row, ("domain rating", "dr")) or "",
                            "UR": _pick_value(row, ("ur", "url rating")) or "",
                            "Domain traffic": _pick_value(row, ("domain traffic", "traffic")) or "",
                            "Nofollow": _pick_value(row, ("nofollow",)) or "",
                            "Lost status": _pick_value(row, ("lost status",)) or "",
                        }
                    )
            normalized = _normalize_backlink_row(row)
            if not normalized["source_domain"]:
                continue
            if not normalized["target_domain"]:
                continue
            normalized_rows.append(normalized)
            valid_rows += 1
        file_summaries.append({"file": filename, "rows": len(rows), "valid_rows": valid_rows})
        if row_limit_hit:
            break

    if (not batch_rows) and auto_batch_rows:
        batch_rows = auto_batch_rows
        batch_name = "from_test_links_pack"

    preferred_batch_rows_by_domain, batch_duplicate_domains, batch_duplicate_rows = _select_preferred_batch_rows(batch_rows)
    for d in sorted(preferred_batch_rows_by_domain.keys()):
        _, metrics = _normalize_batch_row(preferred_batch_rows_by_domain[d])
        batch_metrics[d] = metrics
        batch_targets.append(d)

    for target_domain in batch_targets:
        auto_brand_tokens.extend(_derive_brand_keywords(target_domain))
        token = _brand_token_from_domain(target_domain)
        if token:
            auto_brand_tokens.append(token)
    derived_brand_keywords = list(dict.fromkeys([x for x in auto_brand_tokens if x and len(x) >= 2]))
    brand_keywords_used = list(dict.fromkeys((keywords.get("brand") or []) + derived_brand_keywords))
    keywords["brand"] = brand_keywords_used

    if not normalized_rows:
        raise ValueError("Не удалось извлечь ссылки из входных файлов")

    if progress_callback:
        progress_callback(60, "Расчет метрик ссылочного профиля")

    source_to_targets: Dict[str, set[str]] = defaultdict(set)
    source_to_competitors: Dict[str, set[str]] = defaultdict(set)
    competitor_counter: Counter[str] = Counter()
    competitor_ref_domains: Dict[str, set[str]] = defaultdict(set)
    our_source_counter: Counter[str] = Counter()
    follow_counter: Counter[str] = Counter()
    follow_our_counter: Counter[str] = Counter()
    follow_comp_counter: Counter[str] = Counter()
    anchor_type_counter: Counter[str] = Counter()
    anchor_counter: Counter[str] = Counter()
    anchor_word_counter: Counter[str] = Counter()
    dr_bucket_counter: Counter[str] = Counter()
    zone_counter: Counter[str] = Counter()
    redirect_301_counter: Counter[str] = Counter()
    homepage_donor_counter: Counter[str] = Counter()
    lost_status_counter: Counter[str] = Counter()
    drop_reason_counter: Counter[str] = Counter()
    discovered_status_counter: Counter[str] = Counter()
    http_class_counter: Counter[str] = Counter()
    language_counter: Counter[str] = Counter()
    platform_counter: Counter[str] = Counter()
    link_type_counter: Counter[str] = Counter()
    content_counter: Counter[str] = Counter()
    attr_counter: Counter[str] = Counter()
    source_domain_stats: Dict[str, Dict[str, float]] = defaultdict(
        lambda: {
            "count": 0.0,
            "dr_sum": 0.0,
            "dr_n": 0.0,
            "ur_sum": 0.0,
            "ur_n": 0.0,
            "traffic_sum": 0.0,
            "traffic_n": 0.0,
            "nofollow_n": 0.0,
            "lost_n": 0.0,
            "http2xx_n": 0.0,
            "ugc_n": 0.0,
            "sponsored_n": 0.0,
            "rendered_n": 0.0,
            "raw_n": 0.0,
            "external_links_sum": 0.0,
            "external_links_n": 0.0,
            "page_traffic_sum": 0.0,
            "page_traffic_n": 0.0,
            "keywords_sum": 0.0,
            "keywords_n": 0.0,
        }
    )
    target_agg: Dict[str, Dict[str, Any]] = defaultdict(
        lambda: {
            "total": 0,
            "follow": 0,
            "nofollow": 0,
            "unknown": 0,
            "target_home": 0,
            "home_follow": 0,
            "home_nofollow": 0,
            "int_follow": 0,
            "int_nofollow": 0,
            "lost": 0,
            "dr_counts": Counter(),
            "dr10_counts": Counter(),
            "zone_counts": Counter(),
            "http_class": Counter(),
            "link_type": Counter(),
            "language": Counter(),
            "platform": Counter(),
            "content": Counter(),
            "anchor_type": Counter(),
            "ugc": 0,
            "sponsored": 0,
            "rendered": 0,
            "raw_flag": 0,
        }
    )
    source_follow_profile: Dict[str, Counter[str]] = defaultdict(Counter)

    dr_values: List[float] = []
    traffic_values: List[float] = []
    our_dr_values: List[float] = []
    our_traffic_values: List[float] = []
    comp_dr_values: List[float] = []
    comp_traffic_values: List[float] = []
    our_dr_unknown = 0
    comp_dr_unknown = 0

    our_links = 0
    for row in normalized_rows:
        src = row["source_domain"]
        trg = row["target_domain"]
        sd = source_domain_stats[src]
        source_to_targets[src].add(trg)
        zone_counter[_extract_zone(src)] += 1
        if row.get("has_redirect_301"):
            redirect_301_counter[src] += 1
        if row.get("source_is_homepage"):
            homepage_donor_counter[src] += 1
        if row.get("lost_status"):
            lost_status_counter[row.get("lost_status")] += 1
            sd["lost_n"] += 1
        if row.get("drop_reason"):
            drop_reason_counter[str(row.get("drop_reason"))] += 1
        if row.get("discovered_status"):
            discovered_status_counter[str(row.get("discovered_status"))] += 1
        code = str(row.get("http_code") or "")
        if code.startswith("2"):
            http_class_counter["2xx"] += 1
            sd["http2xx_n"] += 1
        elif code.startswith("3"):
            http_class_counter["3xx"] += 1
        elif code.startswith("4"):
            http_class_counter["4xx"] += 1
        elif code.startswith("5"):
            http_class_counter["5xx"] += 1
        else:
            http_class_counter["unknown"] += 1
        link_type_counter[str(row.get("link_type") or "unknown")] += 1
        language_counter[str(row.get("language") or "unknown")] += 1
        platform_counter[str(row.get("platform") or "unknown")] += 1
        content_counter[str(row.get("content") or "unknown")] += 1
        sd["count"] += 1
        if row.get("dr") is not None:
            sd["dr_sum"] += float(row.get("dr"))
            sd["dr_n"] += 1
        if row.get("ur") is not None:
            sd["ur_sum"] += float(row.get("ur"))
            sd["ur_n"] += 1
        if row.get("traffic") is not None:
            sd["traffic_sum"] += float(row.get("traffic"))
            sd["traffic_n"] += 1
        if row.get("external_links") is not None:
            sd["external_links_sum"] += float(row.get("external_links"))
            sd["external_links_n"] += 1
        if row.get("page_traffic") is not None:
            sd["page_traffic_sum"] += float(row.get("page_traffic"))
            sd["page_traffic_n"] += 1
        if row.get("keywords_count") is not None:
            sd["keywords_sum"] += float(row.get("keywords_count"))
            sd["keywords_n"] += 1
        if row.get("ugc") is True:
            sd["ugc_n"] += 1
            attr_counter["ugc_true"] += 1
        elif row.get("ugc") is False:
            attr_counter["ugc_false"] += 1
        else:
            attr_counter["ugc_unknown"] += 1
        if row.get("sponsored") is True:
            sd["sponsored_n"] += 1
            attr_counter["sponsored_true"] += 1
        elif row.get("sponsored") is False:
            attr_counter["sponsored_false"] += 1
        else:
            attr_counter["sponsored_unknown"] += 1
        if row.get("rendered") is True:
            sd["rendered_n"] += 1
            attr_counter["rendered_true"] += 1
        elif row.get("rendered") is False:
            attr_counter["rendered_false"] += 1
        else:
            attr_counter["rendered_unknown"] += 1
        if row.get("raw_flag") is True:
            sd["raw_n"] += 1
            attr_counter["raw_true"] += 1
        elif row.get("raw_flag") is False:
            attr_counter["raw_false"] += 1
        else:
            attr_counter["raw_unknown"] += 1

        if _is_our_target(trg, domain):
            our_links += 1
            our_source_counter[src] += 1
        else:
            competitor_counter[trg] += 1
            competitor_ref_domains[trg].add(src)
            source_to_competitors[src].add(trg)

        follow = row.get("follow")
        if follow is True:
            follow_counter["dofollow"] += 1
            source_follow_profile[src]["dofollow"] += 1
            if _is_our_target(trg, domain):
                follow_our_counter["dofollow"] += 1
            else:
                follow_comp_counter["dofollow"] += 1
        elif follow is False:
            follow_counter["nofollow"] += 1
            sd["nofollow_n"] += 1
            source_follow_profile[src]["nofollow"] += 1
            if _is_our_target(trg, domain):
                follow_our_counter["nofollow"] += 1
            else:
                follow_comp_counter["nofollow"] += 1
        else:
            follow_counter["unknown"] += 1
            source_follow_profile[src]["unknown"] += 1
            if _is_our_target(trg, domain):
                follow_our_counter["unknown"] += 1
            else:
                follow_comp_counter["unknown"] += 1

        anchor = row.get("anchor") or ""
        anchor_type = _classify_anchor(anchor, keywords)
        anchor_type_counter[anchor_type] += 1
        if anchor:
            anchor_counter[anchor] += 1
            for token in _anchor_words(anchor):
                anchor_word_counter[token] += 1

        row_dr = row.get("dr")
        row_traffic = row.get("traffic")
        source_metrics = batch_metrics.get(src, {})
        effective_dr = row_dr if row_dr is not None else source_metrics.get("dr")
        effective_traffic = row_traffic if row_traffic is not None else source_metrics.get("traffic")
        if effective_dr is not None:
            dr_values.append(float(effective_dr))
            dr_bucket_counter[_dr_bucket(float(effective_dr))] += 1
            if _is_our_target(trg, domain):
                our_dr_values.append(float(effective_dr))
            else:
                comp_dr_values.append(float(effective_dr))
        else:
            dr_bucket_counter["unknown"] += 1
            if _is_our_target(trg, domain):
                our_dr_unknown += 1
            else:
                comp_dr_unknown += 1
        if effective_traffic is not None:
            traffic_values.append(float(effective_traffic))
            if _is_our_target(trg, domain):
                our_traffic_values.append(float(effective_traffic))
            else:
                comp_traffic_values.append(float(effective_traffic))
        ta = target_agg[trg]
        ta["total"] += 1
        is_home = bool(row.get("target_is_homepage"))
        if is_home:
            ta["target_home"] += 1
        if follow is True:
            ta["follow"] += 1
            if is_home:
                ta["home_follow"] += 1
            else:
                ta["int_follow"] += 1
        elif follow is False:
            ta["nofollow"] += 1
            if is_home:
                ta["home_nofollow"] += 1
            else:
                ta["int_nofollow"] += 1
        else:
            ta["unknown"] += 1
        if row.get("dr") is not None:
            ta["dr_counts"][_dr_bucket(row.get("dr"))] += 1
        if effective_dr is not None:
            ta["dr10_counts"][_dr_bucket_decile(effective_dr)] += 1
        ta["zone_counts"][_extract_zone(src)] += 1
        if row.get("lost_status"):
            ta["lost"] += 1
        ta["http_class"]["2xx" if str(row.get("http_code") or "").startswith("2") else "3xx" if str(row.get("http_code") or "").startswith("3") else "4xx" if str(row.get("http_code") or "").startswith("4") else "5xx" if str(row.get("http_code") or "").startswith("5") else "unknown"] += 1
        ta["link_type"][str(row.get("link_type") or "unknown")] += 1
        ta["language"][str(row.get("language") or "unknown")] += 1
        ta["platform"][str(row.get("platform") or "unknown")] += 1
        ta["content"][str(row.get("content") or "unknown")] += 1
        ta["anchor_type"][anchor_type] += 1
        if row.get("ugc") is True:
            ta["ugc"] += 1
        if row.get("sponsored") is True:
            ta["sponsored"] += 1
        if row.get("rendered") is True:
            ta["rendered"] += 1
        if row.get("raw_flag") is True:
            ta["raw_flag"] += 1

    duplicates_with_our: List[Dict[str, Any]] = []
    duplicates_without_our: List[Dict[str, Any]] = []
    single_our: List[Dict[str, Any]] = []
    single_competitor: List[Dict[str, Any]] = []

    for src, targets in source_to_targets.items():
        has_our = any(_is_our_target(t, domain) for t in targets)
        competitor_targets = [t for t in targets if not _is_our_target(t, domain)]
        if has_our and competitor_targets:
            duplicates_with_our.append(
                {
                    "domain": src,
                    "targets_count": len(targets),
                    "competitors_count": len(competitor_targets),
                    "competitor_targets": ", ".join(sorted(competitor_targets)[:5]),
                }
            )
        elif has_our and not competitor_targets:
            single_our.append({"domain": src, "targets_count": len(targets)})
        elif (not has_our) and len(competitor_targets) == 1:
            single_competitor.append({"domain": src, "competitor": competitor_targets[0]})
        elif (not has_our) and len(competitor_targets) >= 2:
            duplicates_without_our.append(
                {
                    "domain": src,
                    "competitors_count": len(competitor_targets),
                    "competitor_targets": ", ".join(sorted(competitor_targets)[:5]),
                }
            )

    # Fallback: if source sheets are not present, build raw tables from normalized rows.
    if not raw_homepage_links_rows:
        for row in normalized_rows:
            if not row.get("source_is_homepage"):
                continue
            if len(raw_homepage_links_rows) >= MAX_RAW_TABLE_ROWS:
                break
            raw_homepage_links_rows.append(
                {
                    "Referring page URL": row.get("source_url") or "",
                    "Target URL": row.get("target_url") or "",
                    "Anchor": row.get("anchor") or "",
                    "Domain Rating": row.get("dr"),
                    "UR": row.get("ur"),
                    "Domain traffic": row.get("traffic"),
                    "Nofollow": "nofollow" if row.get("follow") is False else "dofollow" if row.get("follow") is True else "",
                    "Lost status": row.get("lost_status") or "",
                }
            )
    if not raw_redirect_links_rows:
        for row in normalized_rows:
            code = str(row.get("http_code") or "")
            if not row.get("has_redirect_301") and not code.startswith("3"):
                continue
            if len(raw_redirect_links_rows) >= MAX_RAW_TABLE_ROWS:
                break
            raw_redirect_links_rows.append(
                {
                    "Referring page URL": row.get("source_url") or "",
                    "Target URL": row.get("target_url") or "",
                    "Anchor": row.get("anchor") or "",
                    "Domain Rating": row.get("dr"),
                    "UR": row.get("ur"),
                    "Domain traffic": row.get("traffic"),
                    "Nofollow": "nofollow" if row.get("follow") is False else "dofollow" if row.get("follow") is True else "",
                    "Lost status": row.get("lost_status") or "",
                }
            )
    if not raw_duplicates_without_our_rows:
        dup_wo_domains = {str(x.get("domain") or "") for x in duplicates_without_our}
        for row in normalized_rows:
            src = str(row.get("source_domain") or "")
            trg = str(row.get("target_domain") or "")
            if src not in dup_wo_domains:
                continue
            if _is_our_target(trg, domain):
                continue
            if len(raw_duplicates_without_our_rows) >= MAX_RAW_TABLE_ROWS:
                break
            raw_duplicates_without_our_rows.append(
                {
                    "Referring page URL": row.get("source_url") or "",
                    "Target URL": row.get("target_url") or "",
                    "Anchor": row.get("anchor") or "",
                    "Domain Rating": row.get("dr"),
                    "UR": row.get("ur"),
                    "Domain traffic": row.get("traffic"),
                    "Nofollow": "nofollow" if row.get("follow") is False else "dofollow" if row.get("follow") is True else "",
                    "Lost status": row.get("lost_status") or "",
                }
            )
    if not raw_competitor_links_rows:
        for row in normalized_rows:
            trg = str(row.get("target_domain") or "")
            if _is_our_target(trg, domain):
                continue
            if row.get("source_is_homepage") or row.get("has_redirect_301"):
                continue
            if len(raw_competitor_links_rows) >= MAX_RAW_TABLE_ROWS:
                break
            raw_competitor_links_rows.append(
                {
                    "Referring page URL": row.get("source_url") or "",
                    "Target URL": row.get("target_url") or "",
                    "Anchor": row.get("anchor") or "",
                    "Domain Rating": row.get("dr"),
                    "UR": row.get("ur"),
                    "Domain traffic": row.get("traffic"),
                    "Nofollow": "nofollow" if row.get("follow") is False else "dofollow" if row.get("follow") is True else "",
                    "Lost status": row.get("lost_status") or "",
                }
            )

    per_target_metrics: Dict[str, Dict[str, Any]] = {}
    for target_domain, agg in target_agg.items():
        total = int(agg.get("total", 0))
        follow = int(agg.get("follow", 0))
        nofollow = int(agg.get("nofollow", 0))
        unknown = int(agg.get("unknown", 0))
        home = int(agg.get("target_home", 0))
        internal = total - home
        home_follow = int(agg.get("home_follow", 0))
        home_nofollow = int(agg.get("home_nofollow", 0))
        int_follow = int(agg.get("int_follow", 0))
        int_nofollow = int(agg.get("int_nofollow", 0))
        dr_counts = agg.get("dr_counts", Counter())
        zone_counts = agg.get("zone_counts", Counter())
        per_target_metrics[target_domain] = {
            "total_links": total,
            "homepage_pct": round((home / max(1, total)) * 100, 2),
            "internal_pct": round((internal / max(1, total)) * 100, 2),
            "follow_pct": round((follow / max(1, follow + nofollow + unknown)) * 100, 2),
            "nofollow_pct": round((nofollow / max(1, follow + nofollow + unknown)) * 100, 2),
            "homepage_follow_pct": round((home_follow / max(1, follow)) * 100, 2),
            "homepage_nofollow_pct": round((home_nofollow / max(1, nofollow)) * 100, 2),
            "internal_follow_pct": round((int_follow / max(1, follow)) * 100, 2),
            "internal_nofollow_pct": round((int_nofollow / max(1, nofollow)) * 100, 2),
            "lost_pct": round((int(agg.get("lost", 0)) / max(1, total)) * 100, 2),
            "ugc_pct": round((int(agg.get("ugc", 0)) / max(1, total)) * 100, 2),
            "sponsored_pct": round((int(agg.get("sponsored", 0)) / max(1, total)) * 100, 2),
            "rendered_pct": round((int(agg.get("rendered", 0)) / max(1, total)) * 100, 2),
            "raw_pct": round((int(agg.get("raw_flag", 0)) / max(1, total)) * 100, 2),
            "dr_counts": dr_counts,
            "dr10_counts": agg.get("dr10_counts", Counter()),
            "zone_counts": zone_counts,
            "http_class": agg.get("http_class", Counter()),
            "link_type": agg.get("link_type", Counter()),
            "language": agg.get("language", Counter()),
            "platform": agg.get("platform", Counter()),
            "content": agg.get("content", Counter()),
            "anchor_type": agg.get("anchor_type", Counter()),
        }

    competitors_list = [d for d in per_target_metrics.keys() if not _is_our_target(str(d), domain)]
    our_metrics = per_target_metrics.get(domain) or next((v for k, v in per_target_metrics.items() if _is_our_target(str(k), domain)), {})
    summary_domains_order: List[str] = list(dict.fromkeys([str(d) for d in competitors_list if str(d)] + [domain]))
    benchmark_rows: List[Dict[str, Any]] = []
    if competitors_list:
        keys = [
            "homepage_pct",
            "internal_pct",
            "follow_pct",
            "nofollow_pct",
            "homepage_follow_pct",
            "homepage_nofollow_pct",
            "internal_follow_pct",
            "internal_nofollow_pct",
            "lost_pct",
        ]
        for key in keys:
            vals = [float(per_target_metrics[d].get(key, 0.0)) for d in competitors_list]
            avg_v = round(mean(vals), 2) if vals else 0.0
            med_v = round(median(vals), 2) if vals else 0.0
            our_v = round(float(our_metrics.get(key, 0.0)), 2) if our_metrics else 0.0
            benchmark_rows.append(
                {
                    "metric": key,
                    "our_site_pct": our_v,
                    "competitors_avg_pct": avg_v,
                    "competitors_median_pct": med_v,
                    "delta_vs_avg": round(our_v - avg_v, 2),
                    "delta_vs_median": round(our_v - med_v, 2),
                }
            )

    report_follow_nofollow_rows: List[Dict[str, Any]] = []
    report_home_internal_rows: List[Dict[str, Any]] = []
    report_follow_split_rows: List[Dict[str, Any]] = []
    report_geo_distribution_rows: List[Dict[str, Any]] = []
    report_anchor_matrix_rows: List[Dict[str, Any]] = []
    for d in summary_domains_order:
        m = per_target_metrics.get(d, {}) or {}
        if not m:
            continue
        total_links = max(1, int(m.get("total_links", 0) or 0))
        report_follow_nofollow_rows.append(
            {
                "Конкурент": d,
                "Follow (False)": round(float(m.get("follow_pct", 0.0) or 0.0) / 100.0, 4),
                "Nofollow (True)": round(float(m.get("nofollow_pct", 0.0) or 0.0) / 100.0, 4),
            }
        )
        report_home_internal_rows.append(
            {
                "Конкурент": d,
                "Главная страница": round(float(m.get("homepage_pct", 0.0) or 0.0) / 100.0, 4),
                "Внутренние страницы": round(float(m.get("internal_pct", 0.0) or 0.0) / 100.0, 4),
            }
        )
        report_follow_split_rows.append(
            {
                "Конкурент": d,
                "Follow на главную": round(float(m.get("homepage_follow_pct", 0.0) or 0.0) / 100.0, 4),
                "Nofollow на главную": round(float(m.get("homepage_nofollow_pct", 0.0) or 0.0) / 100.0, 4),
                "Follow на внутренние": round(float(m.get("internal_follow_pct", 0.0) or 0.0) / 100.0, 4),
                "Nofollow на внутренние": round(float(m.get("internal_nofollow_pct", 0.0) or 0.0) / 100.0, 4),
            }
        )
        zc = m.get("zone_counts", Counter()) or Counter()
        report_geo_distribution_rows.append(
            {
                "Конкурент": d,
                ".ru": round(float(zc.get("ru", 0)) / total_links, 4),
                ".com": round(float(zc.get("com", 0)) / total_links, 4),
            }
        )
        ac = m.get("anchor_type", Counter()) or Counter()
        report_anchor_matrix_rows.append(
            {
                "Домен": d,
                "безанкорный": int(ac.get("empty", 0) + ac.get("naked_url", 0)),
                "брендовый": int(ac.get("brand", 0) + ac.get("brand_navigational", 0)),
                "инфо": int(ac.get("informational", 0)),
                "коммерция": int(ac.get("commercial", 0) + ac.get("brand_commercial", 0)),
                "навигационный": int(ac.get("navigational", 0) + ac.get("generic", 0) + ac.get("other", 0)),
                "спам": int(ac.get("spam", 0)),
            }
        )

    def _append_avg_median(rows: List[Dict[str, Any]], label_key: str = "Конкурент") -> List[Dict[str, Any]]:
        if not rows:
            return rows
        keys = [k for k in rows[0].keys() if k != label_key and k != "Домен"]
        label_key_eff = label_key if label_key in rows[0] else "Домен"
        avg_row: Dict[str, Any] = {label_key_eff: "Средние"}
        med_row: Dict[str, Any] = {label_key_eff: "Медиана"}
        for k in keys:
            vals = [x.get(k) for x in rows if isinstance(x.get(k), (int, float))]
            if vals:
                avg_row[k] = round(mean([float(v) for v in vals]), 4)
                med_row[k] = round(median([float(v) for v in vals]), 4)
            else:
                avg_row[k] = None
                med_row[k] = None
        return rows + [avg_row, med_row]

    report_follow_nofollow_rows = _append_avg_median(report_follow_nofollow_rows, "Конкурент")
    report_home_internal_rows = _append_avg_median(report_home_internal_rows, "Конкурент")
    report_follow_split_rows = _append_avg_median(report_follow_split_rows, "Конкурент")
    report_geo_distribution_rows = _append_avg_median(report_geo_distribution_rows, "Конкурент")
    report_anchor_matrix_rows = _append_avg_median(report_anchor_matrix_rows, "Домен")

    link_types_by_competitor = _flatten_breakdown_rows(normalized_rows, dim="link_type", value_key="link_type", row_limit=3000)
    http_codes_by_competitor = _flatten_breakdown_rows(normalized_rows, dim="http_code", value_key="http_code", row_limit=3000)
    languages_by_competitor = _flatten_breakdown_rows(normalized_rows, dim="language", value_key="language", row_limit=3000)

    our_ref_domains = set(our_source_counter.keys())
    our_site_rows: List[Dict[str, Any]] = []
    for src, links_count in our_source_counter.most_common(50):
        metrics = batch_metrics.get(src, {})
        our_site_rows.append(
            {
                "ref_domain": src,
                "links_to_our_site": links_count,
                "batch_dr": metrics.get("dr"),
                "batch_traffic": metrics.get("traffic"),
            }
        )

    competitor_rows: List[Dict[str, Any]] = []
    for competitor, links_count in competitor_counter.most_common(200):
        metrics = batch_metrics.get(competitor, {})
        comp_refs = competitor_ref_domains.get(competitor, set())
        shared = len(comp_refs & our_ref_domains)
        backlinks_all = float(metrics.get("backlinks_all") or 0.0)
        backlinks_followed = float(metrics.get("backlinks_followed") or 0.0)
        backlinks_not_followed = float(metrics.get("backlinks_not_followed") or 0.0)
        backlinks_redirects = float(metrics.get("backlinks_redirects") or 0.0)
        backlinks_internal = float(metrics.get("backlinks_internal") or 0.0)
        ref_domains_all = float(metrics.get("ref_domains_all") or 0.0)
        ref_domains_followed = float(metrics.get("ref_domains_followed") or 0.0)
        ref_domains_not_followed = float(metrics.get("ref_domains_not_followed") or 0.0)
        competitor_rows.append(
            {
                "competitor_domain": competitor,
                "links": links_count,
                "ref_domains": len(comp_refs),
                "shared_with_our_site": shared,
                "batch_dr": metrics.get("dr"),
                "batch_traffic": metrics.get("traffic"),
                "batch_backlinks_all": backlinks_all,
                "batch_ref_domains_all": ref_domains_all,
                "batch_backlinks_followed_pct": round((backlinks_followed / max(1.0, backlinks_all)) * 100, 2) if backlinks_all else None,
                "batch_backlinks_nofollow_pct": round((backlinks_not_followed / max(1.0, backlinks_all)) * 100, 2) if backlinks_all else None,
                "batch_backlinks_redirect_pct": round((backlinks_redirects / max(1.0, backlinks_all)) * 100, 2) if backlinks_all else None,
                "batch_backlinks_internal_pct": round((backlinks_internal / max(1.0, backlinks_all)) * 100, 2) if backlinks_all else None,
                "batch_ref_domains_followed_pct": round((ref_domains_followed / max(1.0, ref_domains_all)) * 100, 2) if ref_domains_all else None,
                "batch_ref_domains_nofollow_pct": round((ref_domains_not_followed / max(1.0, ref_domains_all)) * 100, 2) if ref_domains_all else None,
                "batch_organic_keywords_total": metrics.get("organic_keywords_total"),
            }
        )

    summary_domains_order: List[str] = []
    for item in competitor_rows:
        d = str(item.get("competitor_domain") or "").strip()
        if d and d not in summary_domains_order:
            summary_domains_order.append(d)
    if domain not in summary_domains_order:
        summary_domains_order.append(domain)

    # Build RU sheet rows directly from batch analysis data to preserve full metrics.
    ru_headers = [
        "#", "Target", "IP", "URL Rating", "Domain Rating", "Ahrefs Rank",
        "Organic / Total Keywords", "Organic / Keywords (Top 3)", "Organic / Keywords (4-10)",
        "Organic / Keywords (11-20)", "Organic / Keywords (21-50)", "Organic / Keywords (51+)",
        "Organic / Traffic", "Organic / Value", "Organic / Top Countries",
        "Ref. domains / All", "Ref. domains / Followed", "Ref. domains / Not followed",
        "Ref. IPs / IPs", "Ref. IPs / Subnets",
        "Backlinks / All", "Backlinks / Followed", "Backlinks / Not followed",
        "Backlinks / Redirects", "Backlinks / Internal",
        "Outgoing domains / Followed", "Outgoing domains / All time",
        "Outgoing links / Followed", "Outgoing links / All time",
    ]
    batch_rows_by_domain: Dict[str, Dict[str, Any]] = preferred_batch_rows_by_domain

    report_ru_rows: List[Dict[str, Any]] = []
    ru_idx = 1
    for d in summary_domains_order:
        src = batch_rows_by_domain.get(d)
        if not src:
            continue
        target_value = _pick_value(src, ("target", "domain", "site", "url")) or d
        rec = {
            "#": ru_idx,
            "Target": str(target_value),
            "IP": _pick_value(src, ("ip",)),
            "URL Rating": _pick_value(src, ("url rating", "ur")),
            "Domain Rating": _pick_value(src, ("domain rating", "dr")),
            "Ahrefs Rank": _pick_value(src, ("ahrefs rank",)),
            "Organic / Total Keywords": _pick_value(src, ("organic / total keywords", "organic_keywords_total")),
            "Organic / Keywords (Top 3)": _pick_value(src, ("organic / keywords (top 3)", "organic_keywords_top3")),
            "Organic / Keywords (4-10)": _pick_value(src, ("organic / keywords (4-10)", "organic_keywords_4_10")),
            "Organic / Keywords (11-20)": _pick_value(src, ("organic / keywords (11-20)", "organic_keywords_11_20")),
            "Organic / Keywords (21-50)": _pick_value(src, ("organic / keywords (21-50)", "organic_keywords_21_50")),
            "Organic / Keywords (51+)": _pick_value(src, ("organic / keywords (51+)", "organic_keywords_51_plus")),
            "Organic / Traffic": _pick_value(src, ("organic / traffic", "organic_traffic")),
            "Organic / Value": _pick_value(src, ("organic / value", "organic_value")),
            "Organic / Top Countries": _pick_value(src, ("organic / top countries",)),
            "Ref. domains / All": _pick_value(src, ("ref. domains / all", "ref domains all", "ref_domains_all")),
            "Ref. domains / Followed": _pick_value(src, ("ref. domains / followed", "ref_domains_followed")),
            "Ref. domains / Not followed": _pick_value(src, ("ref. domains / not followed", "ref_domains_not_followed")),
            "Ref. IPs / IPs": _pick_value(src, ("ref. ips / ips", "ref_ips_ips")),
            "Ref. IPs / Subnets": _pick_value(src, ("ref. ips / subnets", "ref_ips_subnets")),
            "Backlinks / All": _pick_value(src, ("backlinks / all", "backlinks_all", "total backlinks")),
            "Backlinks / Followed": _pick_value(src, ("backlinks / followed", "backlinks_followed")),
            "Backlinks / Not followed": _pick_value(src, ("backlinks / not followed", "backlinks_not_followed")),
            "Backlinks / Redirects": _pick_value(src, ("backlinks / redirects", "backlinks_redirects")),
            "Backlinks / Internal": _pick_value(src, ("backlinks / internal", "backlinks_internal")),
            "Outgoing domains / Followed": _pick_value(src, ("outgoing domains / followed", "outgoing_domains_followed")),
            "Outgoing domains / All time": _pick_value(src, ("outgoing domains / all time", "outgoing_domains_all_time")),
            "Outgoing links / Followed": _pick_value(src, ("outgoing links / followed", "outgoing_links_followed")),
            "Outgoing links / All time": _pick_value(src, ("outgoing links / all time", "outgoing_links_all_time")),
        }
        report_ru_rows.append(rec)
        ru_idx += 1

    report_core_metrics_rows: List[Dict[str, Any]] = []
    for d in summary_domains_order:
        bm = batch_metrics.get(d, {}) or {}
        dr_v = _to_float(bm.get("dr"))
        bl_v = _to_float(bm.get("backlinks_all"))
        rd_v = _to_float(bm.get("ref_domains_all"))
        if dr_v is None and bl_v is None and rd_v is None:
            continue
        report_core_metrics_rows.append(
            {
                "Конкурент": d,
                "Domain Rating": dr_v,
                "Total Backlinks": int(bl_v) if bl_v is not None else None,
                "Referring Domains": int(rd_v) if rd_v is not None else None,
            }
        )
    if report_core_metrics_rows:
        dr_vals = [float(x.get("Domain Rating")) for x in report_core_metrics_rows if x.get("Domain Rating") is not None]
        bl_vals = [float(x.get("Total Backlinks")) for x in report_core_metrics_rows if x.get("Total Backlinks") is not None]
        rd_vals = [float(x.get("Referring Domains")) for x in report_core_metrics_rows if x.get("Referring Domains") is not None]
        report_core_metrics_rows.append(
            {
                "Конкурент": "Средние",
                "Domain Rating": round(mean(dr_vals), 2) if dr_vals else None,
                "Total Backlinks": int(round(mean(bl_vals))) if bl_vals else None,
                "Referring Domains": int(round(mean(rd_vals))) if rd_vals else None,
            }
        )
        report_core_metrics_rows.append(
            {
                "Конкурент": "Медиана",
                "Domain Rating": round(median(dr_vals), 2) if dr_vals else None,
                "Total Backlinks": int(round(median(bl_vals))) if bl_vals else None,
                "Referring Domains": int(round(median(rd_vals))) if rd_vals else None,
            }
        )

    competitor_rank_rows: List[Dict[str, Any]] = []
    for row in competitor_rows:
        dr_v = float(row.get("batch_dr") or 0.0)
        backlinks_v = float(row.get("batch_backlinks_all") or 0.0)
        follow_v = float(row.get("batch_backlinks_followed_pct") or 0.0)
        score = round((dr_v * 0.45) + (min(backlinks_v, 200000.0) / 2000.0 * 0.3) + (follow_v * 0.25), 2)
        competitor_rank_rows.append(
            {
                "competitor_domain": row.get("competitor_domain"),
                "batch_dr": row.get("batch_dr"),
                "batch_backlinks_all": row.get("batch_backlinks_all"),
                "batch_backlinks_followed_pct": row.get("batch_backlinks_followed_pct"),
                "rank_score": score,
            }
        )
    competitor_rank_rows.sort(key=lambda x: float(x.get("rank_score") or 0.0), reverse=True)
    for idx, item in enumerate(competitor_rank_rows, start=1):
        item["rank"] = idx

    comparison_rows: List[Dict[str, Any]] = []
    our_batch_dr = _to_float((batch_metrics.get(domain) or {}).get("dr"))
    if our_batch_dr is None:
        for d_key, m in batch_metrics.items():
            if _is_our_target(str(d_key), domain):
                our_batch_dr = _to_float((m or {}).get("dr"))
                if our_batch_dr is not None:
                    break
    our_follow_pct = float(our_metrics.get("follow_pct", 0.0) or 0.0)
    our_lost_pct = float(our_metrics.get("lost_pct", 0.0) or 0.0)
    our_http2xx_pct = _pct((our_metrics.get("http_class", Counter()) or Counter()).get("2xx", 0), max(1, int((our_metrics.get("total_links", 0) or 0))))
    for row in competitor_rows:
        competitor = str(row.get("competitor_domain") or "")
        comp_refs = competitor_ref_domains.get(competitor, set())
        shared = len(comp_refs & our_ref_domains)
        comp_only = len(comp_refs - our_ref_domains)
        our_only = len(our_ref_domains - comp_refs)
        overlap_pct = round((shared / max(1, len(comp_refs))) * 100, 2)
        comp_metrics = per_target_metrics.get(competitor, {})
        comp_follow_pct = float(comp_metrics.get("follow_pct", 0.0) or 0.0)
        comp_lost_pct = float(comp_metrics.get("lost_pct", 0.0) or 0.0)
        comp_http2xx_pct = _pct((comp_metrics.get("http_class", Counter()) or Counter()).get("2xx", 0), max(1, int((comp_metrics.get("total_links", 0) or 0))))
        comp_batch_dr = _to_float((batch_metrics.get(competitor) or {}).get("dr"))
        dr_gap_pct = None
        if (our_batch_dr is not None) and (comp_batch_dr is not None):
            dr_gap_pct = _pct(comp_batch_dr - our_batch_dr, max(1.0, our_batch_dr))
        comparison_rows.append(
            {
                "competitor_domain": competitor,
                "shared_ref_domains": shared,
                "competitor_only_domains": comp_only,
                "our_only_domains": our_only,
                "overlap_pct": overlap_pct,
                "coverage_of_our_ref_domains_pct": round((shared / max(1, len(our_ref_domains))) * 100, 2),
                "donor_gap_pct": round((comp_only / max(1, len(comp_refs))) * 100, 2),
                "competitor_follow_pct": comp_follow_pct,
                "our_follow_pct": our_follow_pct,
                "follow_gap_pp": round(comp_follow_pct - our_follow_pct, 2),
                "competitor_lost_pct": comp_lost_pct,
                "our_lost_pct": our_lost_pct,
                "lost_gap_pp": round(comp_lost_pct - our_lost_pct, 2),
                "competitor_http_2xx_pct": comp_http2xx_pct,
                "our_http_2xx_pct": our_http2xx_pct,
                "http_2xx_gap_pp": round(comp_http2xx_pct - our_http2xx_pct, 2),
                "our_batch_dr": our_batch_dr,
                "competitor_batch_dr": comp_batch_dr,
                "dr_gap_pct_vs_our": dr_gap_pct,
            }
        )

    competitor_quality_rows: List[Dict[str, Any]] = []
    for row in competitor_rows:
        comp = str(row.get("competitor_domain") or "")
        comp_metrics = per_target_metrics.get(comp, {})
        total_links = int(comp_metrics.get("total_links", 0) or 0)
        http2xx_pct = _pct((comp_metrics.get("http_class", Counter()) or Counter()).get("2xx", 0), max(1, total_links))
        follow_pct = float(comp_metrics.get("follow_pct", 0.0) or 0.0)
        lost_pct = float(comp_metrics.get("lost_pct", 0.0) or 0.0)
        homepage_pct = float(comp_metrics.get("homepage_pct", 0.0) or 0.0)
        quality_score = round(
            (follow_pct * 0.35)
            + (http2xx_pct * 0.25)
            + ((100.0 - lost_pct) * 0.25)
            + ((100.0 - homepage_pct) * 0.15),
            2,
        )
        competitor_quality_rows.append(
            {
                "competitor_domain": comp,
                "links_in_dataset": total_links,
                "follow_pct": round(follow_pct, 2),
                "lost_pct": round(lost_pct, 2),
                "http_2xx_pct": round(http2xx_pct, 2),
                "homepage_target_pct": round(homepage_pct, 2),
                "quality_score_0_100": quality_score,
            }
        )
    competitor_quality_rows.sort(key=lambda x: float(x.get("quality_score_0_100", 0.0)), reverse=True)

    dr_deciles = [
        "DR 0-9",
        "DR 10-19",
        "DR 20-29",
        "DR 30-39",
        "DR 40-49",
        "DR 50-59",
        "DR 60-69",
        "DR 70-79",
        "DR 80-89",
        "DR 90-100",
    ]
    dr_distribution_matrix_rows: List[Dict[str, Any]] = []
    matrix_domains = [x.get("competitor_domain") for x in competitor_rows if x.get("competitor_domain")]
    for comp in matrix_domains:
        m = per_target_metrics.get(str(comp), {})
        total_links = max(1, int(m.get("total_links", 0) or 0))
        dr10 = m.get("dr10_counts", Counter()) or Counter()
        row: Dict[str, Any] = {"Домен": comp}
        for bucket in dr_deciles:
            row[bucket] = _pct(dr10.get(bucket, 0), total_links)
        dr_distribution_matrix_rows.append(row)
    if dr_distribution_matrix_rows:
        avg_row: Dict[str, Any] = {"Домен": "Средние"}
        med_row: Dict[str, Any] = {"Домен": "Медиана"}
        for bucket in dr_deciles:
            vals = [float(x.get(bucket, 0.0) or 0.0) for x in dr_distribution_matrix_rows]
            avg_row[bucket] = round(mean(vals), 2) if vals else 0.0
            med_row[bucket] = round(median(vals), 2) if vals else 0.0
        dr_distribution_matrix_rows.append(avg_row)
        dr_distribution_matrix_rows.append(med_row)
    our_dr10 = (our_metrics or {}).get("dr10_counts", Counter()) or Counter()
    our_total = max(1, int((our_metrics or {}).get("total_links", 0) or 0))
    our_row: Dict[str, Any] = {"Домен": domain}
    for bucket in dr_deciles:
        our_row[bucket] = _pct(our_dr10.get(bucket, 0), our_total)
    dr_distribution_matrix_rows.append(our_row)
    report_dr_distribution_rows: List[Dict[str, Any]] = []
    for row in dr_distribution_matrix_rows:
        rec: Dict[str, Any] = {"Конкурент": row.get("Домен")}
        for bucket in dr_deciles:
            value = row.get(bucket)
            if isinstance(value, (int, float)):
                rec[bucket] = round(float(value) / 100.0, 4)
            else:
                rec[bucket] = value
        report_dr_distribution_rows.append(rec)

    def _build_mix_benchmark(
        value_getter: Callable[[Dict[str, Any]], Counter[str]],
        categories: List[str],
        title_prefix: str = "",
    ) -> List[Dict[str, Any]]:
        rows: List[Dict[str, Any]] = []
        for cat in categories:
            comp_vals: List[float] = []
            for comp in competitors_list:
                m = per_target_metrics.get(comp, {})
                cnt = value_getter(m) or Counter()
                total_links = max(1, int(m.get("total_links", 0) or 0))
                comp_vals.append(_pct(cnt.get(cat, 0), total_links))
            our_cnt = value_getter(our_metrics or {}) or Counter()
            our_val = _pct(our_cnt.get(cat, 0), max(1, int((our_metrics or {}).get("total_links", 0) or 0)))
            rows.append(
                {
                    "metric": f"{title_prefix}{cat}",
                    "our_site_pct": our_val,
                    "competitors_avg_pct": round(mean(comp_vals), 2) if comp_vals else 0.0,
                    "competitors_median_pct": round(median(comp_vals), 2) if comp_vals else 0.0,
                }
            )
        return rows

    anchor_mix_benchmark_rows = _build_mix_benchmark(
        lambda m: m.get("anchor_type", Counter()) or Counter(),
        ["empty", "naked_url", "brand", "brand_commercial", "brand_navigational", "commercial", "informational", "navigational", "generic", "spam", "other"],
        "anchor:",
    )
    link_type_categories = [k for k, _ in link_type_counter.most_common(20)]
    link_type_benchmark_rows = _build_mix_benchmark(
        lambda m: m.get("link_type", Counter()) or Counter(),
        link_type_categories,
        "type:",
    )
    http_categories = ["2xx", "3xx", "4xx", "5xx", "unknown"]
    http_benchmark_rows = _build_mix_benchmark(
        lambda m: m.get("http_class", Counter()) or Counter(),
        http_categories,
        "http:",
    )
    follow_home_internal_benchmark_rows = benchmark_rows[:]

    top_anchors = [{"anchor": a, "count": c} for a, c in anchor_counter.most_common(30)]
    anchor_word_rows = [{"word": w, "count": c} for w, c in anchor_word_counter.most_common(30)]
    priority_domains = sorted(
        duplicates_with_our,
        key=lambda x: (x.get("competitors_count", 0), x.get("targets_count", 0)),
        reverse=True,
    )[:30]
    candidate_domains = set([x.get("domain") for x in duplicates_without_our if x.get("domain")] + [x.get("domain") for x in single_competitor if x.get("domain")])
    priority_score_rows: List[Dict[str, Any]] = []
    for d in candidate_domains:
        stats = source_domain_stats.get(str(d), {})
        n = max(1.0, float(stats.get("count", 0.0)))
        avg_dr = float(stats.get("dr_sum", 0.0)) / max(1.0, float(stats.get("dr_n", 0.0)))
        avg_ur = float(stats.get("ur_sum", 0.0)) / max(1.0, float(stats.get("ur_n", 0.0)))
        avg_traffic = float(stats.get("traffic_sum", 0.0)) / max(1.0, float(stats.get("traffic_n", 0.0)))
        nofollow_rate = float(stats.get("nofollow_n", 0.0)) / n
        if avg_dr <= 5:
            continue
        priority_score = (avg_dr * 0.5) + (avg_ur * 0.3) + ((avg_traffic or 0.0) * 0.2 / 1000.0)
        priority_score_rows.append(
            {
                "domain": d,
                "avg_dr": round(avg_dr, 2),
                "avg_ur": round(avg_ur, 2),
                "avg_traffic": round(avg_traffic, 2),
                "nofollow_rate": round(nofollow_rate, 4),
                "priority_score": round(priority_score, 4),
            }
        )
    strict_rows = [r for r in priority_score_rows if r["avg_dr"] > 40 and r["nofollow_rate"] < 0.5]
    if not strict_rows:
        strict_rows = [r for r in priority_score_rows if r["avg_dr"] > 30 and r["nofollow_rate"] < 0.7]
    priority_score_domains = sorted(strict_rows, key=lambda x: x["priority_score"], reverse=True)[:300]
    if precomputed_priority_rows:
        imported: List[Dict[str, Any]] = []
        for row in precomputed_priority_rows:
            d = _pick_value(row, ("domain",))
            dr = _to_float(_pick_value(row, ("domain rating", "dr")))
            ur = _to_float(_pick_value(row, ("ur", "url rating")))
            tr = _to_float(_pick_value(row, ("domain traffic", "organic / traffic", "traffic")))
            nf = _to_follow_bool(_pick_value(row, ("nofollow",)))
            ps = _to_float(_pick_value(row, ("priority score",)))
            domain_key = _normalize_domain(str(d or ""))
            if not domain_key:
                continue
            imported.append(
                {
                    "domain": domain_key,
                    "avg_dr": dr,
                    "avg_ur": ur,
                    "avg_traffic": tr,
                    "nofollow_rate": 1.0 if nf is True else 0.0 if nf is False else None,
                    "priority_score": ps,
                }
            )
        if imported:
            idx = {str(x.get("domain")): x for x in priority_score_domains}
            for row in imported:
                idx[str(row.get("domain"))] = row
            priority_score_domains = sorted(
                idx.values(),
                key=lambda x: float(x.get("priority_score") or 0.0),
                reverse=True,
            )[:500]
    opportunity_domains_rows: List[Dict[str, Any]] = []
    total_competitors = max(1, len(competitor_counter))
    for d in candidate_domains:
        stats = source_domain_stats.get(str(d), {})
        competitors_hit = len(source_to_competitors.get(str(d), set()))
        total = float(stats.get("count", 0.0))
        if total <= 0:
            continue
        avg_dr = float(stats.get("dr_sum", 0.0)) / max(1.0, float(stats.get("dr_n", 0.0)))
        avg_traffic = float(stats.get("traffic_sum", 0.0)) / max(1.0, float(stats.get("traffic_n", 0.0)))
        nofollow_rate = float(stats.get("nofollow_n", 0.0)) / total
        lost_rate = float(stats.get("lost_n", 0.0)) / total
        http2xx_rate = float(stats.get("http2xx_n", 0.0)) / total
        coverage_pct = _pct(competitors_hit, total_competitors)
        opportunity_score = round(
            (coverage_pct * 0.35)
            + (avg_dr * 0.35)
            + (min(avg_traffic, 10000.0) / 100.0 * 0.15)
            + ((1.0 - nofollow_rate) * 100.0 * 0.1)
            + ((1.0 - lost_rate) * 100.0 * 0.05),
            2,
        )
        opportunity_domains_rows.append(
            {
                "domain": d,
                "competitors_covered": competitors_hit,
                "competitors_covered_pct": coverage_pct,
                "links_in_dataset": int(total),
                "avg_dr": round(avg_dr, 2),
                "avg_traffic": round(avg_traffic, 2),
                "follow_pct": round((1.0 - nofollow_rate) * 100.0, 2),
                "lost_pct": round(lost_rate * 100.0, 2),
                "http_2xx_pct": round(http2xx_rate * 100.0, 2),
                "opportunity_score": opportunity_score,
            }
        )
    opportunity_domains_rows.sort(key=lambda x: (float(x.get("opportunity_score", 0.0)), float(x.get("competitors_covered", 0))), reverse=True)
    opportunity_domains_rows = opportunity_domains_rows[:500]
    ready_buy_rows: List[Dict[str, Any]] = []
    for row in opportunity_domains_rows:
        lost_pct_value = _to_float(row.get("lost_pct"))
        if lost_pct_value is None:
            lost_pct_value = 100.0
        if (
            float(row.get("avg_dr") or 0.0) >= 30.0
            and float(row.get("avg_traffic") or 0.0) >= 50.0
            and float(row.get("follow_pct") or 0.0) >= 60.0
            and lost_pct_value <= 40.0
        ):
            ready_buy_rows.append(row)
            if len(ready_buy_rows) >= 300:
                break
    for row in ready_buy_rows:
        row["popularity_score"] = round(
            (float(row.get("competitors_covered_pct") or 0.0) * 0.55)
            + (float(row.get("avg_traffic") or 0.0) / 200.0 * 0.25)
            + (float(row.get("avg_dr") or 0.0) * 0.20),
            2,
        )
    ready_buy_rows.sort(
        key=lambda x: (
            float(x.get("popularity_score") or 0.0),
            float(x.get("opportunity_score") or 0.0),
        ),
        reverse=True,
    )
    dr_bucket_rows = [{"dr_bucket": k, "links": v} for k, v in sorted(dr_bucket_counter.items(), key=lambda x: str(x[0]))]
    dr_bucket_our_rows = []
    dr_bucket_comp_rows = []
    for bucket in ["0-9", "10-29", "30-49", "50-69", "70+", "unknown"]:
        if bucket == "unknown":
            dr_bucket_our_rows.append({"dr_bucket": bucket, "links": our_dr_unknown})
            dr_bucket_comp_rows.append({"dr_bucket": bucket, "links": comp_dr_unknown})
        else:
            dr_bucket_our_rows.append({"dr_bucket": bucket, "links": sum(1 for x in our_dr_values if _dr_bucket(x) == bucket)})
            dr_bucket_comp_rows.append({"dr_bucket": bucket, "links": sum(1 for x in comp_dr_values if _dr_bucket(x) == bucket)})
    zone_rows = [{"zone": z, "links": c} for z, c in zone_counter.most_common(20)]
    redirect_301_rows = [{"domain": d, "links_with_301": c} for d, c in redirect_301_counter.most_common(200)]
    homepage_donor_rows = [{"domain": d, "links_from_homepage": c} for d, c in homepage_donor_counter.most_common(200)]
    follow_rows = [
        {"type": "dofollow", "count": follow_counter.get("dofollow", 0)},
        {"type": "nofollow", "count": follow_counter.get("nofollow", 0)},
        {"type": "unknown", "count": follow_counter.get("unknown", 0)},
    ]
    follow_mix_rows = _counter_to_pct_rows(follow_counter, value_key="type", limit=10)
    follow_domain_class_counter: Counter[str] = Counter()
    for d, fc in source_follow_profile.items():
        has_do = int(fc.get("dofollow", 0)) > 0
        has_no = int(fc.get("nofollow", 0)) > 0
        has_un = int(fc.get("unknown", 0)) > 0
        if has_do and not has_no:
            follow_domain_class_counter["follow_only_domains"] += 1
        elif has_no and not has_do:
            follow_domain_class_counter["nofollow_only_domains"] += 1
        elif has_do and has_no:
            follow_domain_class_counter["mixed_follow_nofollow_domains"] += 1
        elif has_un:
            follow_domain_class_counter["unknown_only_domains"] += 1
    follow_domain_mix_rows = _counter_to_pct_rows(follow_domain_class_counter, value_key="segment", limit=10)
    lost_status_rows = _counter_to_pct_rows(lost_status_counter, value_key="lost_status", limit=20)
    http_class_rows = _counter_to_pct_rows(http_class_counter, value_key="http_class", limit=10)
    link_type_mix_rows = _counter_to_pct_rows(link_type_counter, value_key="link_type", limit=20)
    language_mix_rows = _counter_to_pct_rows(language_counter, value_key="language", limit=20)
    anchor_mix_rows = _counter_to_pct_rows(anchor_type_counter, value_key="anchor_type", limit=20)
    follow_detail_rows = [
        {
            "segment": "our_site",
            "dofollow": follow_our_counter.get("dofollow", 0),
            "nofollow": follow_our_counter.get("nofollow", 0),
            "unknown": follow_our_counter.get("unknown", 0),
            "nofollow_share_pct": round(
                (follow_our_counter.get("nofollow", 0) / max(1, sum(follow_our_counter.values()))) * 100,
                2,
            ),
        },
        {
            "segment": "competitors",
            "dofollow": follow_comp_counter.get("dofollow", 0),
            "nofollow": follow_comp_counter.get("nofollow", 0),
            "unknown": follow_comp_counter.get("unknown", 0),
            "nofollow_share_pct": round(
                (follow_comp_counter.get("nofollow", 0) / max(1, sum(follow_comp_counter.values()))) * 100,
                2,
            ),
        },
        {
            "segment": "all",
            "dofollow": follow_counter.get("dofollow", 0),
            "nofollow": follow_counter.get("nofollow", 0),
            "unknown": follow_counter.get("unknown", 0),
            "nofollow_share_pct": round(
                (follow_counter.get("nofollow", 0) / max(1, sum(follow_counter.values()))) * 100,
                2,
            ),
        },
    ]
    dr_stats_rows = [
        {
            "segment": "our_site",
            "links_with_dr": len(our_dr_values),
            "avg_dr": round(mean(our_dr_values), 2) if our_dr_values else None,
            "median_dr": round(median(our_dr_values), 2) if our_dr_values else None,
            "min_dr": min(our_dr_values) if our_dr_values else None,
            "max_dr": max(our_dr_values) if our_dr_values else None,
        },
        {
            "segment": "competitors",
            "links_with_dr": len(comp_dr_values),
            "avg_dr": round(mean(comp_dr_values), 2) if comp_dr_values else None,
            "median_dr": round(median(comp_dr_values), 2) if comp_dr_values else None,
            "min_dr": min(comp_dr_values) if comp_dr_values else None,
            "max_dr": max(comp_dr_values) if comp_dr_values else None,
        },
        {
            "segment": "all",
            "links_with_dr": len(dr_values),
            "avg_dr": round(mean(dr_values), 2) if dr_values else None,
            "median_dr": round(median(dr_values), 2) if dr_values else None,
            "min_dr": min(dr_values) if dr_values else None,
            "max_dr": max(dr_values) if dr_values else None,
        },
    ]
    brand_rows = [{"keyword": k} for k in brand_keywords_used[:100]]

    if progress_callback:
        progress_callback(85, "Формирование итогового отчета")

    warnings: List[str] = []
    if follow_counter.get("unknown", 0) > 0:
        warnings.append("Часть ссылок без явного признака dofollow/nofollow")
    if not _parse_keywords(brand_keywords):
        warnings.append("Словарь brand keywords не передан, использованы автоматически выведенные брендовые токены")
    if batch_duplicate_domains > 0:
        warnings.append(
            f"В batch-данных обнаружены дубликаты доменов: {batch_duplicate_domains} "
            f"(лишних строк: {batch_duplicate_rows}). Для отчета выбрана наиболее заполненная запись по каждому домену."
        )
    if row_limit_hit:
        warnings.append(
            f"Обработан лимит {MAX_LINK_PROFILE_ROWS} строк для стабильной работы. "
            "Для полного охвата разделите выгрузку на части."
        )
    if (
        len(raw_homepage_links_rows) >= MAX_RAW_TABLE_ROWS
        or len(raw_redirect_links_rows) >= MAX_RAW_TABLE_ROWS
        or len(raw_competitor_links_rows) >= MAX_RAW_TABLE_ROWS
        or len(raw_duplicates_without_our_rows) >= MAX_RAW_TABLE_ROWS
    ):
        warnings.append(
            f"Сырые таблицы ограничены до {MAX_RAW_TABLE_ROWS} строк на блок для стабильной работы интерфейса."
        )

    summary = {
        "our_domain": domain,
        "backlink_files": len(backlink_files),
        "batch_file": batch_name,
        "batch_duplicate_domains": batch_duplicate_domains,
        "batch_duplicate_rows_ignored": batch_duplicate_rows,
        "rows_total": len(normalized_rows),
        "our_links": our_links,
        "unique_ref_domains": len(source_to_targets),
        "unique_competitors": len(competitor_counter),
        "avg_dr": round(mean(dr_values), 2) if dr_values else None,
        "avg_traffic": round(mean(traffic_values), 2) if traffic_values else None,
        "dofollow": follow_counter.get("dofollow", 0),
        "nofollow": follow_counter.get("nofollow", 0),
        "unknown_follow": follow_counter.get("unknown", 0),
        "duplicates_with_our_site": len(duplicates_with_our),
        "duplicates_without_our_site": len(duplicates_without_our),
        "priority_domains_scored": len(priority_score_domains),
        "our_unique_ref_domains": len(our_ref_domains),
        "donors_with_301": len(redirect_301_counter),
        "donors_homepage": len(homepage_donor_counter),
        "avg_our_dr": round(mean(our_dr_values), 2) if our_dr_values else None,
        "avg_our_traffic": round(mean(our_traffic_values), 2) if our_traffic_values else None,
        "dofollow_pct": _pct(follow_counter.get("dofollow", 0), sum(follow_counter.values())),
        "nofollow_pct": _pct(follow_counter.get("nofollow", 0), sum(follow_counter.values())),
        "lost_links_pct": _pct(sum(lost_status_counter.values()), len(normalized_rows)),
        "http_2xx_pct": _pct(http_class_counter.get("2xx", 0), sum(http_class_counter.values())),
    }

    avg_comp_follow = round(mean([float(x.get("follow_pct") or 0.0) for x in competitor_quality_rows]), 2) if competitor_quality_rows else 0.0
    avg_comp_lost = round(mean([float(x.get("lost_pct") or 0.0) for x in competitor_quality_rows]), 2) if competitor_quality_rows else 0.0
    avg_comp_quality = round(mean([float(x.get("quality_score_0_100") or 0.0) for x in competitor_quality_rows]), 2) if competitor_quality_rows else 0.0
    avg_comp_http2xx = round(mean([float(x.get("http_2xx_pct") or 0.0) for x in competitor_quality_rows]), 2) if competitor_quality_rows else 0.0

    executive_kpi_rows = [
        {
            "Показатель": "Доля dofollow ссылок, %",
            "Наш сайт": summary.get("dofollow_pct"),
            "Среднее конкурентов": avg_comp_follow,
            "Разница (п.п.)": round(float(summary.get("dofollow_pct") or 0.0) - avg_comp_follow, 2),
        },
        {
            "Показатель": "Доля потерянных ссылок, %",
            "Наш сайт": summary.get("lost_links_pct"),
            "Среднее конкурентов": avg_comp_lost,
            "Разница (п.п.)": round(float(summary.get("lost_links_pct") or 0.0) - avg_comp_lost, 2),
        },
        {
            "Показатель": "Доля 2xx доноров, %",
            "Наш сайт": summary.get("http_2xx_pct"),
            "Среднее конкурентов": avg_comp_http2xx,
            "Разница (п.п.)": round(float(summary.get("http_2xx_pct") or 0.0) - avg_comp_http2xx, 2),
        },
        {
            "Показатель": "Средний DR доноров",
            "Наш сайт": summary.get("avg_our_dr"),
            "Среднее конкурентов": round(mean([float(x) for x in comp_dr_values]), 2) if comp_dr_values else None,
            "Разница (п.п.)": round(float(summary.get("avg_our_dr") or 0.0) - (round(mean([float(x) for x in comp_dr_values]), 2) if comp_dr_values else 0.0), 2),
        },
        {
            "Показатель": "Индекс качества профиля (0-100)",
            "Наш сайт": round(
                (float(summary.get("dofollow_pct") or 0.0) * 0.35)
                + (float(summary.get("http_2xx_pct") or 0.0) * 0.25)
                + ((100.0 - float(summary.get("lost_links_pct") or 0.0)) * 0.25)
                + ((100.0 - float(our_metrics.get("homepage_pct", 0.0) or 0.0)) * 0.15),
                2,
            ),
            "Среднее конкурентов": avg_comp_quality,
            "Разница (п.п.)": round(
                round(
                    (float(summary.get("dofollow_pct") or 0.0) * 0.35)
                    + (float(summary.get("http_2xx_pct") or 0.0) * 0.25)
                    + ((100.0 - float(summary.get("lost_links_pct") or 0.0)) * 0.25)
                    + ((100.0 - float(our_metrics.get("homepage_pct", 0.0) or 0.0)) * 0.15),
                    2,
                )
                - avg_comp_quality,
                2,
            ),
        },
    ]

    anchor_total = sum(anchor_type_counter.values()) or 1
    anchor_naked = int(anchor_type_counter.get("empty", 0) + anchor_type_counter.get("naked_url", 0))
    anchor_brand = int(anchor_type_counter.get("brand", 0) + anchor_type_counter.get("brand_navigational", 0))
    anchor_commercial = int(anchor_type_counter.get("commercial", 0) + anchor_type_counter.get("brand_commercial", 0))
    anchor_navigational = int(anchor_type_counter.get("navigational", 0) + anchor_type_counter.get("generic", 0) + anchor_type_counter.get("other", 0))
    anchor_spam = int(anchor_type_counter.get("spam", 0))

    profile_structure_rows = [
        {"Метрика": "Анкоры: безанкорные/URL, %", "Значение": _pct(anchor_naked, anchor_total)},
        {"Метрика": "Анкоры: брендовые, %", "Значение": _pct(anchor_brand, anchor_total)},
        {"Метрика": "Анкоры: коммерческие, %", "Значение": _pct(anchor_commercial, anchor_total)},
        {"Метрика": "Анкоры: навигационные/генерик, %", "Значение": _pct(anchor_navigational, anchor_total)},
        {"Метрика": "Анкоры: спам, %", "Значение": _pct(anchor_spam, anchor_total)},
        {"Метрика": "Follow-ссылки на главную, %", "Значение": round(float(our_metrics.get("homepage_follow_pct", 0.0) or 0.0), 2)},
        {"Метрика": "Follow-ссылки на внутренние, %", "Значение": round(float(our_metrics.get("internal_follow_pct", 0.0) or 0.0), 2)},
    ]

    priority_dashboard_rows: List[Dict[str, Any]] = []
    action_queue_rows: List[Dict[str, Any]] = []

    spam_pct = _pct(anchor_spam, anchor_total)
    nofollow_pct = float(summary.get("nofollow_pct") or 0.0)
    lost_pct = float(summary.get("lost_links_pct") or 0.0)
    http_bad_pct = round(
        float(_pct(http_class_counter.get("4xx", 0), sum(http_class_counter.values())))
        + float(_pct(http_class_counter.get("5xx", 0), sum(http_class_counter.values()))),
        2,
    )
    donor_gap_avg = round(mean([float(x.get("donor_gap_pct") or 0.0) for x in comparison_rows]), 2) if comparison_rows else 0.0
    ready_buy_count = len(ready_buy_rows)

    def _status(metric: float, good_max: float, warn_max: float) -> str:
        if metric <= good_max:
            return "good"
        if metric <= warn_max:
            return "warning"
        return "critical"

    priority_dashboard_rows.extend(
        [
            {
                "Приоритет": "P1",
                "Блок": "Риски качества",
                "Метрика": "Потерянные ссылки, %",
                "Значение": lost_pct,
                "Порог": "<= 25",
                "Статус": _status(lost_pct, 25.0, 40.0),
                "Почему важно": "Высокая доля lost снижает стабильность ссылочного профиля.",
            },
            {
                "Приоритет": "P1",
                "Блок": "Риски качества",
                "Метрика": "Nofollow, %",
                "Значение": nofollow_pct,
                "Порог": "<= 45",
                "Статус": _status(nofollow_pct, 45.0, 60.0),
                "Почему важно": "Снижает передаваемый ссылочный вес.",
            },
            {
                "Приоритет": "P1",
                "Блок": "Риски качества",
                "Метрика": "Плохие HTTP доноров (4xx+5xx), %",
                "Значение": http_bad_pct,
                "Порог": "<= 12",
                "Статус": _status(http_bad_pct, 12.0, 20.0),
                "Почему важно": "Нестабильные/битые доноры ухудшают качество профиля.",
            },
            {
                "Приоритет": "P2",
                "Блок": "Анкор-лист",
                "Метрика": "Спам-анкоров, %",
                "Значение": spam_pct,
                "Порог": "<= 8",
                "Статус": _status(spam_pct, 8.0, 15.0),
                "Почему важно": "Риск переспама и фильтров поисковых систем.",
            },
            {
                "Приоритет": "P2",
                "Блок": "Рост",
                "Метрика": "Средний donor gap vs конкуренты, %",
                "Значение": donor_gap_avg,
                "Порог": "<= 35",
                "Статус": "warning" if donor_gap_avg > 35.0 else "good",
                "Почему важно": "Показывает незакрытые доноры конкурентов.",
            },
            {
                "Приоритет": "P2",
                "Блок": "Рост",
                "Метрика": "Готовых доноров для закупки, шт",
                "Значение": ready_buy_count,
                "Порог": ">= 50",
                "Статус": "good" if ready_buy_count >= 50 else "warning",
                "Почему важно": "Быстрый пул для GGL/Miralinks.",
            },
        ]
    )

    action_queue_rows.extend(
        [
            {
                "Приоритет": "P1",
                "Действие": "Снизить долю lost-ссылок",
                "Impact": "high",
                "Effort": "medium",
                "Что сделать": "Проверить top-доноров lost_status, восстановить/заменить ссылки.",
            },
            {
                "Приоритет": "P1",
                "Действие": "Очистить некачественные HTTP доноры",
                "Impact": "high",
                "Effort": "low",
                "Что сделать": "Исключить 4xx/5xx доноров из закупки, обновить список площадок.",
            },
            {
                "Приоритет": "P1",
                "Действие": "Поднять долю dofollow",
                "Impact": "high",
                "Effort": "medium",
                "Что сделать": "Сместить закупку в пользу площадок с follow_only/mixed с высоким follow%.",
            },
            {
                "Приоритет": "P2",
                "Действие": "Закрыть donor gap",
                "Impact": "high",
                "Effort": "high",
                "Что сделать": "Взять top из ready_buy_domains и opportunity_domains, которых нет у нас.",
            },
            {
                "Приоритет": "P2",
                "Действие": "Нормализовать анкор-лист",
                "Impact": "medium",
                "Effort": "medium",
                "Что сделать": "Держать бренд/безанкор в безопасном диапазоне, ограничить коммерцию/спам.",
            },
            {
                "Приоритет": "P3",
                "Действие": "Расширить языковый и типовой микс доноров",
                "Impact": "medium",
                "Effort": "medium",
                "Что сделать": "Добрать доноры по релевантным языкам и типам (text/image/form).",
            },
        ]
    )

    # Stage-1 data layer for structured UI/DOCX/XLSX sections.
    competitor_benchmark_rows: List[Dict[str, Any]] = []
    for d in summary_domains_order:
        m = per_target_metrics.get(d, {}) or {}
        bm = batch_metrics.get(d, {}) or {}
        if not m and not bm:
            continue
        total_links = max(1, int(m.get("total_links", 0) or 0))
        http2xx = _pct((m.get("http_class", Counter()) or Counter()).get("2xx", 0), total_links)
        quality_score = round(
            (float(m.get("follow_pct", 0.0) or 0.0) * 0.35)
            + (http2xx * 0.25)
            + ((100.0 - float(m.get("lost_pct", 0.0) or 0.0)) * 0.25)
            + ((100.0 - float(m.get("homepage_pct", 0.0) or 0.0)) * 0.15),
            2,
        )
        competitor_benchmark_rows.append(
            {
                "Домен": d,
                "Domain Rating": bm.get("dr"),
                "Total Backlinks": bm.get("backlinks_all"),
                "Referring Domains": bm.get("ref_domains_all"),
                "Follow %": m.get("follow_pct"),
                "Lost %": m.get("lost_pct"),
                "HTTP 2xx %": round(http2xx, 2),
                "Homepage %": m.get("homepage_pct"),
                "UGC %": m.get("ugc_pct"),
                "Sponsored %": m.get("sponsored_pct"),
                "Rendered %": m.get("rendered_pct"),
                "Quality Score": quality_score,
            }
        )
    if competitor_benchmark_rows:
        numeric_cols = [
            "Domain Rating",
            "Total Backlinks",
            "Referring Domains",
            "Follow %",
            "Lost %",
            "HTTP 2xx %",
            "Homepage %",
            "UGC %",
            "Sponsored %",
            "Rendered %",
            "Quality Score",
        ]
        comp_only = [r for r in competitor_benchmark_rows if not _is_our_target(str(r.get("Домен") or ""), domain)]
        if comp_only:
            avg_row: Dict[str, Any] = {"Домен": "Средние"}
            med_row: Dict[str, Any] = {"Домен": "Медиана"}
            for col in numeric_cols:
                vals = [float(x.get(col)) for x in comp_only if isinstance(x.get(col), (int, float))]
                avg_row[col] = round(mean(vals), 2) if vals else None
                med_row[col] = round(median(vals), 2) if vals else None
            competitor_benchmark_rows.extend([avg_row, med_row])

    gap_donors_priority_rows: List[Dict[str, Any]] = []
    for row in opportunity_domains_rows[:800]:
        gap_donors_priority_rows.append(
            {
                "Domain": row.get("domain"),
                "Competitors Covered": row.get("competitors_covered"),
                "Coverage %": row.get("competitors_covered_pct"),
                "Avg DR": row.get("avg_dr"),
                "Avg Traffic": row.get("avg_traffic"),
                "Follow %": row.get("follow_pct"),
                "Lost %": row.get("lost_pct"),
                "Opportunity Score": row.get("opportunity_score"),
            }
        )

    donor_overlap_matrix_rows: List[Dict[str, Any]] = []
    for src, targets in source_to_targets.items():
        if not targets:
            continue
        sorted_targets = sorted(targets)
        has_our = any(_is_our_target(t, domain) for t in sorted_targets)
        competitors = [t for t in sorted_targets if not _is_our_target(t, domain)]
        donor_overlap_matrix_rows.append(
            {
                "Domain": src,
                "Has Our Site": 1 if has_our else 0,
                "Competitors Count": len(competitors),
                "Competitors": ", ".join(competitors[:8]),
                "Targets Total": len(sorted_targets),
                "Gap Type": "gap" if (not has_our and len(competitors) >= 2) else "shared" if has_our and competitors else "single",
            }
        )
    donor_overlap_matrix_rows.sort(
        key=lambda x: (int(x.get("Has Our Site") or 0), int(x.get("Competitors Count") or 0), int(x.get("Targets Total") or 0)),
        reverse=True,
    )
    donor_overlap_matrix_rows = donor_overlap_matrix_rows[:3000]

    total_attr = max(1, len(normalized_rows))
    link_attributes_rows = [
        {"Attribute": "dofollow", "Count": follow_counter.get("dofollow", 0), "Pct": _pct(follow_counter.get("dofollow", 0), total_attr)},
        {"Attribute": "nofollow", "Count": follow_counter.get("nofollow", 0), "Pct": _pct(follow_counter.get("nofollow", 0), total_attr)},
        {"Attribute": "unknown_follow", "Count": follow_counter.get("unknown", 0), "Pct": _pct(follow_counter.get("unknown", 0), total_attr)},
        {"Attribute": "ugc_true", "Count": attr_counter.get("ugc_true", 0), "Pct": _pct(attr_counter.get("ugc_true", 0), total_attr)},
        {"Attribute": "sponsored_true", "Count": attr_counter.get("sponsored_true", 0), "Pct": _pct(attr_counter.get("sponsored_true", 0), total_attr)},
        {"Attribute": "rendered_true", "Count": attr_counter.get("rendered_true", 0), "Pct": _pct(attr_counter.get("rendered_true", 0), total_attr)},
        {"Attribute": "raw_true", "Count": attr_counter.get("raw_true", 0), "Pct": _pct(attr_counter.get("raw_true", 0), total_attr)},
    ]

    loss_recovery_rows: List[Dict[str, Any]] = []
    for k, c in lost_status_counter.most_common(50):
        loss_recovery_rows.append({"Group": "Lost status", "Value": k, "Count": c, "Pct": _pct(c, total_attr)})
    for k, c in drop_reason_counter.most_common(50):
        loss_recovery_rows.append({"Group": "Drop reason", "Value": k, "Count": c, "Pct": _pct(c, total_attr)})
    for k, c in discovered_status_counter.most_common(50):
        loss_recovery_rows.append({"Group": "Discovered status", "Value": k, "Count": c, "Pct": _pct(c, total_attr)})

    http_type_lang_platform_rows: List[Dict[str, Any]] = []
    for k, c in http_class_counter.most_common():
        http_type_lang_platform_rows.append({"Dimension": "HTTP class", "Value": k, "Count": c, "Pct": _pct(c, total_attr)})
    for k, c in link_type_counter.most_common(20):
        http_type_lang_platform_rows.append({"Dimension": "Link type", "Value": k, "Count": c, "Pct": _pct(c, total_attr)})
    for k, c in language_counter.most_common(20):
        http_type_lang_platform_rows.append({"Dimension": "Language", "Value": k, "Count": c, "Pct": _pct(c, total_attr)})
    for k, c in platform_counter.most_common(20):
        http_type_lang_platform_rows.append({"Dimension": "Platform", "Value": k, "Count": c, "Pct": _pct(c, total_attr)})

    target_structure_rows: List[Dict[str, Any]] = []
    for d in summary_domains_order:
        m = per_target_metrics.get(d, {}) or {}
        if not m:
            continue
        target_structure_rows.append(
            {
                "Domain": d,
                "Homepage %": m.get("homepage_pct"),
                "Internal %": m.get("internal_pct"),
                "Follow %": m.get("follow_pct"),
                "Nofollow %": m.get("nofollow_pct"),
                "UGC %": m.get("ugc_pct"),
                "Sponsored %": m.get("sponsored_pct"),
                "Rendered %": m.get("rendered_pct"),
                "Raw %": m.get("raw_pct"),
            }
        )

    risk_signals_rows: List[Dict[str, Any]] = []
    for src, stats in source_domain_stats.items():
        total = max(1.0, float(stats.get("count", 0.0)))
        if total < 2:
            continue
        avg_dr = float(stats.get("dr_sum", 0.0)) / max(1.0, float(stats.get("dr_n", 0.0)))
        avg_traffic = float(stats.get("traffic_sum", 0.0)) / max(1.0, float(stats.get("traffic_n", 0.0)))
        ext_avg = float(stats.get("external_links_sum", 0.0)) / max(1.0, float(stats.get("external_links_n", 0.0)))
        lost_rate = float(stats.get("lost_n", 0.0)) / total
        nofollow_rate = float(stats.get("nofollow_n", 0.0)) / total
        sponsored_rate = float(stats.get("sponsored_n", 0.0)) / total
        risk_score = round(
            (max(0.0, 35.0 - avg_dr) * 1.2)
            + (min(ext_avg, 1000.0) / 20.0)
            + (lost_rate * 45.0)
            + (nofollow_rate * 30.0)
            + (sponsored_rate * 20.0),
            2,
        )
        risk_level = "high" if risk_score >= 55 else "medium" if risk_score >= 30 else "low"
        risk_signals_rows.append(
            {
                "Domain": src,
                "Links": int(total),
                "Avg DR": round(avg_dr, 2),
                "Avg Traffic": round(avg_traffic, 2),
                "Avg External Links": round(ext_avg, 2),
                "Lost %": round(lost_rate * 100.0, 2),
                "Nofollow %": round(nofollow_rate * 100.0, 2),
                "Sponsored %": round(sponsored_rate * 100.0, 2),
                "Risk Score": risk_score,
                "Risk Level": risk_level,
            }
        )
    risk_signals_rows.sort(key=lambda x: float(x.get("Risk Score") or 0.0), reverse=True)
    risk_signals_rows = risk_signals_rows[:2000]

    action_queue_90d_rows: List[Dict[str, Any]] = [
        {"Phase": "0-30", "Priority": "P1", "Action": "Стабилизировать качество доноров", "Target KPI": "Lost % и HTTP 4xx/5xx", "Owner": "SEO"},
        {"Phase": "0-30", "Priority": "P1", "Action": "Поднять долю follow", "Target KPI": "Follow %", "Owner": "SEO/Outreach"},
        {"Phase": "31-60", "Priority": "P2", "Action": "Закрыть donor-gap из приоритетных", "Target KPI": "Competitor coverage %", "Owner": "Outreach"},
        {"Phase": "31-60", "Priority": "P2", "Action": "Скорректировать анкор-лист", "Target KPI": "Spam anchor % / Brand %", "Owner": "SEO"},
        {"Phase": "61-90", "Priority": "P2", "Action": "Расширить качественные типы/платформы", "Target KPI": "Link type/platform diversity", "Owner": "SEO"},
        {"Phase": "61-90", "Priority": "P3", "Action": "Пересчитать риск-профиль и исключить токсичные домены", "Target KPI": "Risk score median", "Owner": "SEO"},
    ]

    executive_overview_rows = [
        {"Metric": "Rows total", "Value": len(normalized_rows)},
        {"Metric": "Unique donors", "Value": len(source_to_targets)},
        {"Metric": "Unique competitors", "Value": len(competitor_counter)},
        {"Metric": "Follow %", "Value": summary.get("dofollow_pct")},
        {"Metric": "Lost %", "Value": summary.get("lost_links_pct")},
        {"Metric": "HTTP 2xx %", "Value": summary.get("http_2xx_pct")},
        {"Metric": "Ready-buy domains", "Value": len(ready_buy_rows)},
        {"Metric": "High-risk donors", "Value": sum(1 for x in risk_signals_rows if str(x.get("Risk Level")) == "high")},
    ]

    validation_checks: List[Dict[str, Any]] = []

    def _add_check(name: str, ok: bool, details: str, severity: str = "error") -> None:
        validation_checks.append(
            {
                "check": name,
                "status": "ok" if ok else ("warning" if severity == "warning" else "error"),
                "details": details,
            }
        )

    total_rows = len(normalized_rows)
    follow_total = int(follow_counter.get("dofollow", 0) + follow_counter.get("nofollow", 0) + follow_counter.get("unknown", 0))
    _add_check(
        "rows_vs_follow_total",
        total_rows == follow_total,
        f"rows_total={total_rows}, follow_split_total={follow_total}",
    )
    competitor_links_total = int(sum(int(v) for v in competitor_counter.values()))
    _add_check(
        "rows_vs_our_plus_competitors",
        total_rows == (our_links + competitor_links_total),
        f"rows_total={total_rows}, our_links={our_links}, competitor_links={competitor_links_total}",
    )
    _add_check(
        "unique_ref_domains_consistency",
        len(source_to_targets) == int(summary.get("unique_ref_domains") or 0),
        f"source_to_targets={len(source_to_targets)}, summary.unique_ref_domains={summary.get('unique_ref_domains')}",
    )
    for metric_name in ("dofollow_pct", "nofollow_pct", "lost_links_pct", "http_2xx_pct"):
        metric_value = float(summary.get(metric_name) or 0.0)
        _add_check(
            f"range_{metric_name}",
            0.0 <= metric_value <= 100.0,
            f"{metric_name}={metric_value}",
        )
    _add_check(
        "executive_tables_non_empty",
        bool(executive_overview_rows and competitor_benchmark_rows and gap_donors_priority_rows),
        f"exec={len(executive_overview_rows)}, benchmark={len(competitor_benchmark_rows)}, gap={len(gap_donors_priority_rows)}",
    )
    raw_caps_hit = (
        len(raw_homepage_links_rows) >= MAX_RAW_TABLE_ROWS
        or len(raw_redirect_links_rows) >= MAX_RAW_TABLE_ROWS
        or len(raw_competitor_links_rows) >= MAX_RAW_TABLE_ROWS
        or len(raw_duplicates_without_our_rows) >= MAX_RAW_TABLE_ROWS
    )
    _add_check(
        "raw_tables_capped",
        not raw_caps_hit,
        f"home={len(raw_homepage_links_rows)}, redirect={len(raw_redirect_links_rows)}, competitor={len(raw_competitor_links_rows)}, duplicates={len(raw_duplicates_without_our_rows)}, cap={MAX_RAW_TABLE_ROWS}",
        severity="warning",
    )

    validation_summary = {
        "ok": sum(1 for x in validation_checks if x.get("status") == "ok"),
        "warning": sum(1 for x in validation_checks if x.get("status") == "warning"),
        "error": sum(1 for x in validation_checks if x.get("status") == "error"),
    }

    def _row_get(row: Dict[str, Any], *keys: str) -> Any:
        if not isinstance(row, dict):
            return None
        for key in keys:
            if key in row and row.get(key) not in (None, ""):
                return row.get(key)
        lower = {str(k).strip().lower(): v for k, v in row.items()}
        for key in keys:
            lk = str(key).strip().lower()
            if lk in lower and lower.get(lk) not in (None, ""):
                return lower.get(lk)
        return None

    def _sample_link_rows(rows: List[Dict[str, Any]], limit: int = 3) -> List[Dict[str, Any]]:
        prepared: List[Tuple[float, float, float, Dict[str, Any]]] = []
        for row in rows or []:
            ref_url = str(_row_get(row, "Referring page URL", "source_url", "referring page url") or "").strip()
            target_url = str(_row_get(row, "Target URL", "target_url", "target url") or "").strip()
            if not ref_url or not target_url:
                continue
            dr = _to_float(_row_get(row, "Domain Rating", "Domain rating", "dr")) or 0.0
            ur = _to_float(_row_get(row, "UR", "ur", "URL Rating", "url rating")) or 0.0
            traffic = _to_float(_row_get(row, "Domain traffic", "domain traffic", "traffic")) or 0.0
            prepared.append(
                (
                    dr,
                    traffic,
                    ur,
                    {
                        "Referring page URL": ref_url,
                        "Target URL": target_url,
                        "Anchor": str(_row_get(row, "Anchor", "anchor") or "").strip(),
                        "Domain Rating": round(dr, 2) if dr else "",
                        "UR": round(ur, 2) if ur else "",
                        "Domain traffic": round(traffic, 2) if traffic else "",
                        "Nofollow": str(_row_get(row, "Nofollow", "nofollow", "follow") or "").strip(),
                        "Lost status": str(_row_get(row, "Lost status", "lost status") or "").strip(),
                    },
                )
            )
        prepared.sort(key=lambda x: (x[0], x[1], x[2]), reverse=True)
        return [x[3] for x in prepared[:limit]]

    def _format_samples_block(title: str, rows: List[Dict[str, Any]]) -> str:
        if not rows:
            return f"{title}\nНет подходящих примеров."
        lines = [title]
        for idx, row in enumerate(rows, start=1):
            lines.append(
                f"{idx}) Referring page URL: {row.get('Referring page URL')}\n"
                f"   Target URL: {row.get('Target URL')}\n"
                f"   Anchor: {row.get('Anchor') or '-'}\n"
                f"   Domain Rating: {row.get('Domain Rating') or '-'} | UR: {row.get('UR') or '-'} | Domain traffic: {row.get('Domain traffic') or '-'}\n"
                f"   Nofollow: {row.get('Nofollow') or '-'} | Lost status: {row.get('Lost status') or '-'}"
            )
        return "\n".join(lines)

    top_gap = (gap_donors_priority_rows or [{}])[0] if gap_donors_priority_rows else {}
    redirect_samples = _sample_link_rows(raw_redirect_links_rows, limit=3)
    competitor_samples = _sample_link_rows(raw_competitor_links_rows, limit=3)
    duplicates_samples = _sample_link_rows(raw_duplicates_without_our_rows, limit=3)

    anchor_brand_pct = _pct(anchor_brand, anchor_total)
    anchor_commercial_pct = _pct(anchor_commercial, anchor_total)
    anchor_naked_pct = _pct(anchor_naked, anchor_total)
    anchor_nav_pct = _pct(anchor_navigational, anchor_total)

    anchor_actions: List[str] = []
    if spam_pct > 8:
        anchor_actions.append("Снизить спам-анкоры: целевой уровень <= 8%.")
    if anchor_brand_pct < 15:
        anchor_actions.append("Увеличить долю брендовых анкоров до 15-30%.")
    if anchor_commercial_pct > 30:
        anchor_actions.append("Снизить долю коммерческих анкоров, разбавить брендом и URL-анкорами.")
    if anchor_naked_pct < 30:
        anchor_actions.append("Добавить безанкорные/URL-ссылки для естественности профиля.")
    if not anchor_actions:
        anchor_actions.append("Анкорный профиль сбалансирован, поддерживайте текущие пропорции.")

    prompts = {
        "ourSite": (
            f"Наш домен: {domain}\n"
            f"Уникальных доноров: {summary.get('our_unique_ref_domains', 0)}\n"
            f"Dofollow/Nofollow: {summary.get('dofollow_pct', 0)}% / {summary.get('nofollow_pct', 0)}%\n"
            f"Lost links: {summary.get('lost_links_pct', 0)}%\n"
            "Фокус: усиление доли качественных dofollow и снижение доли lost."
        ),
        "competitors": (
            f"Конкурентов в датасете: {summary.get('unique_competitors', 0)}\n"
            f"Топ-конкурент по quality score: {(competitor_quality_rows[0].get('competitor_domain') if competitor_quality_rows else 'н/д')}\n"
            "Приоритет анализа: конкуренты с высоким quality score и пересечением доноров с нашим доменом."
        ),
        "comparison": (
            f"Средний donor gap: {donor_gap_avg}%\n"
            f"Top gap domain: {top_gap.get('Domain', 'н/д')} (opportunity score: {top_gap.get('Opportunity Score', 'н/д')})\n"
            "Используйте comparison/gap таблицы, чтобы закрыть доноров конкурентов, которых у нас нет."
        ),
        "plan": (
            "План 30/60/90:\n"
            "1) Очистить риски качества (lost, 4xx/5xx, спам).\n"
            "2) Добрать priority/gap доноры с высоким DR и трафиком.\n"
            "3) Нормализовать анкорный микс и закрепить рост follow-ссылок."
        ),
        "anchorTemplate": (
            "Шаблон по анкорам:\n"
            f"- Безанкорные/URL: {anchor_naked_pct}%\n"
            f"- Брендовые: {anchor_brand_pct}%\n"
            f"- Коммерческие: {anchor_commercial_pct}%\n"
            f"- Навигационные/генерик: {anchor_nav_pct}%\n"
            f"- Спам: {spam_pct}%\n"
            "Рекомендации:\n- " + "\n- ".join(anchor_actions)
        ),
        "riskTemplate": (
            "Шаблон анализа риск-ссылок:\n"
            f"- Lost links: {summary.get('lost_links_pct', 0)}%\n"
            f"- Nofollow: {summary.get('nofollow_pct', 0)}%\n"
            f"- HTTP 2xx: {summary.get('http_2xx_pct', 0)}%\n"
            + "\n\n"
            + _format_samples_block("Примеры риск-ссылок (redirect/lost):", redirect_samples)
        ),
        "outreachTemplate": (
            "Шаблон для outreach и закупки:\n"
            + _format_samples_block("Примеры доноров конкурентов:", competitor_samples)
            + "\n\n"
            + _format_samples_block("Примеры дублей без нашего сайта:", duplicates_samples)
        ),
        "rowReviewTemplate": (
            "Шаблон проверки строки ссылки:\n"
            "Referring page URL\nTarget URL\nAnchor\nDomain Rating\nUR\nDomain traffic\nNofollow\nLost status\n"
            "Используйте этот шаблон для ручной валидации доноров перед закупкой."
        ),
    }
    prompt_templates_rows = [
        {"template": "ourSite", "text": prompts.get("ourSite", "")},
        {"template": "competitors", "text": prompts.get("competitors", "")},
        {"template": "comparison", "text": prompts.get("comparison", "")},
        {"template": "plan", "text": prompts.get("plan", "")},
        {"template": "anchorTemplate", "text": prompts.get("anchorTemplate", "")},
        {"template": "riskTemplate", "text": prompts.get("riskTemplate", "")},
        {"template": "outreachTemplate", "text": prompts.get("outreachTemplate", "")},
        {"template": "rowReviewTemplate", "text": prompts.get("rowReviewTemplate", "")},
    ]

    def _cap_rows(rows: List[Dict[str, Any]], limit: int = MAX_RESULT_TABLE_ROWS) -> List[Dict[str, Any]]:
        if not isinstance(rows, list):
            return []
        if len(rows) <= limit:
            return rows
        return rows[:limit]

    result = {
        "task_type": "link_profile_audit",
        "url": domain,
        "completed_at": datetime.now(timezone.utc).isoformat(),
        "results": {
            "summary": summary,
            "warnings": warnings,
            "errors": [],
            "validation": {
                "summary": validation_summary,
                "checks": _cap_rows(validation_checks, 200),
            },
            "keywords": {
                **keywords,
                "derivedBrandKeywords": derived_brand_keywords,
                "brandKeywordsUsed": brand_keywords_used,
            },
            "source_files": file_summaries,
            "outputs": {
                "competitorAnalysis": {"rows": _cap_rows(competitor_rows)},
                "anchorAnalysis": {"rows": _cap_rows(top_anchors)},
                "duplicates": {"rows": _cap_rows(duplicates_with_our)},
                "additionalMetrics": {"rows": _cap_rows(dr_bucket_rows)},
                "combinedOutput": {"rows": []},
            },
            "tables": {
                "competitor_analysis": _cap_rows(competitor_rows),
                "anchor_analysis": _cap_rows(top_anchors),
                "anchor_word_analysis": _cap_rows(anchor_word_rows),
                "duplicates_with_our_site": _cap_rows(duplicates_with_our),
                "duplicates_with_two_competitors": _cap_rows(duplicates_without_our),
                "single_competitor_domains": _cap_rows(single_competitor),
                "single_our_domains": _cap_rows(single_our),
                "priority_domains": _cap_rows(priority_domains),
                "priority_score_domains": _cap_rows(priority_score_domains),
                "our_site_overview": _cap_rows(our_site_rows),
                "comparison_overview": _cap_rows(comparison_rows),
                "benchmark_overview": _cap_rows(benchmark_rows),
                "dr_stats": _cap_rows(dr_stats_rows),
                "dr_buckets": _cap_rows(dr_bucket_rows),
                "dr_buckets_our_site": _cap_rows(dr_bucket_our_rows),
                "dr_buckets_competitors": _cap_rows(dr_bucket_comp_rows),
                "zones": _cap_rows(zone_rows),
                "follow_types": _cap_rows(follow_rows),
                "follow_mix_pct": _cap_rows(follow_mix_rows),
                "follow_domain_mix_pct": _cap_rows(follow_domain_mix_rows),
                "follow_types_detailed": _cap_rows(follow_detail_rows),
                "lost_status_mix": _cap_rows(lost_status_rows),
                "http_class_mix": _cap_rows(http_class_rows),
                "link_type_mix": _cap_rows(link_type_mix_rows),
                "language_mix": _cap_rows(language_mix_rows),
                "anchor_mix_pct": _cap_rows(anchor_mix_rows),
                "donors_with_redirect_301": _cap_rows(redirect_301_rows),
                "donors_homepage": _cap_rows(homepage_donor_rows),
                "brand_keywords_auto": _cap_rows(brand_rows),
                "source_files": _cap_rows(file_summaries),
                "link_types_by_competitor": _cap_rows(link_types_by_competitor),
                "http_codes_by_competitor": _cap_rows(http_codes_by_competitor),
                "languages_by_competitor": _cap_rows(languages_by_competitor),
                "competitor_quality": _cap_rows(competitor_quality_rows),
                "competitor_ranking": _cap_rows(competitor_rank_rows),
                "opportunity_domains": _cap_rows(opportunity_domains_rows),
                "ready_buy_domains": _cap_rows(ready_buy_rows),
                "dr_distribution_matrix": _cap_rows(dr_distribution_matrix_rows),
                "report_core_metrics": _cap_rows(report_core_metrics_rows),
                "report_follow_nofollow": _cap_rows(report_follow_nofollow_rows),
                "report_home_internal": _cap_rows(report_home_internal_rows),
                "report_follow_split": _cap_rows(report_follow_split_rows),
                "report_dr_distribution": _cap_rows(report_dr_distribution_rows),
                "report_geo_distribution": _cap_rows(report_geo_distribution_rows),
                "report_anchor_matrix": _cap_rows(report_anchor_matrix_rows),
                "report_ru_rows": _cap_rows(report_ru_rows, 5000),
                "analysis_data_sections": [{"title": s.get("title"), "rows": _cap_rows(s.get("rows") or [])} for s in imported_analysis_sections],
                "raw_homepage_links": _cap_rows(raw_homepage_links_rows),
                "raw_competitor_links": _cap_rows(raw_competitor_links_rows),
                "raw_redirect_links": _cap_rows(raw_redirect_links_rows),
                "raw_duplicates_without_our": _cap_rows(raw_duplicates_without_our_rows),
                "executive_kpi": _cap_rows(executive_kpi_rows),
                "executive_overview": _cap_rows(executive_overview_rows),
                "profile_structure": _cap_rows(profile_structure_rows),
                "priority_dashboard": _cap_rows(priority_dashboard_rows),
                "action_queue": _cap_rows(action_queue_rows),
                "action_queue_90d": _cap_rows(action_queue_90d_rows),
                "competitor_benchmark": _cap_rows(competitor_benchmark_rows),
                "gap_donors_priority": _cap_rows(gap_donors_priority_rows),
                "donor_overlap_matrix": _cap_rows(donor_overlap_matrix_rows),
                "link_attributes": _cap_rows(link_attributes_rows),
                "loss_recovery": _cap_rows(loss_recovery_rows),
                "http_type_lang_platform": _cap_rows(http_type_lang_platform_rows),
                "target_structure": _cap_rows(target_structure_rows),
                "risk_signals": _cap_rows(risk_signals_rows),
                "validation_checks": _cap_rows(validation_checks, 200),
                "prompt_templates": _cap_rows(prompt_templates_rows, 50),
                "ourSiteTables": [
                    {"title": "Приоритеты SEO (первый экран)", "rows": _cap_rows(priority_dashboard_rows)},
                    {"title": "Очередь действий (что делать первым)", "rows": _cap_rows(action_queue_rows)},
                    {"title": "План 30/60/90", "rows": _cap_rows(action_queue_90d_rows)},
                    {"title": "KPI по нашему сайту vs среднее конкурентов", "rows": _cap_rows(executive_kpi_rows)},
                    {"title": "Executive overview", "rows": _cap_rows(executive_overview_rows)},
                    {"title": "Validation checks", "rows": _cap_rows(validation_checks, 200)},
                    {"title": "Структура ссылочного профиля (наш сайт)", "rows": _cap_rows(profile_structure_rows)},
                    {"title": "Target structure (наш сайт vs конкуренты)", "rows": _cap_rows(target_structure_rows)},
                    {"title": "Наш сайт: доноры", "rows": _cap_rows(our_site_rows)},
                    {"title": "Ссылки с главных (raw)", "rows": _cap_rows(raw_homepage_links_rows)},
                    {"title": "Ссылки с конкурентов (raw)", "rows": _cap_rows(raw_competitor_links_rows)},
                    {"title": "Ссылки с редиректов (raw)", "rows": _cap_rows(raw_redirect_links_rows)},
                    {"title": "Дубликаты без нашего сайта (raw)", "rows": _cap_rows(raw_duplicates_without_our_rows)},
                ],
                "competitorTables": [
                    {"title": "Конкуренты", "rows": _cap_rows(competitor_rows)},
                    {"title": "Competitor benchmark", "rows": _cap_rows(competitor_benchmark_rows)},
                    {"title": "Рейтинг конкурентов (DR / Backlinks / Follow%)", "rows": _cap_rows(competitor_rank_rows)},
                    {"title": "Качество профиля конкурентов (0-100)", "rows": _cap_rows(competitor_quality_rows)},
                ],
                "comparisonTables": [
                    {"title": "Сравнение с конкурентами", "rows": _cap_rows(comparison_rows)},
                    {"title": "Бенчмарк avg/median по конкурентам", "rows": _cap_rows(benchmark_rows)},
                    {"title": "Бенчмарк по анкорам (avg/median)", "rows": _cap_rows(anchor_mix_benchmark_rows)},
                    {"title": "Бенчмарк по типам ссылок (avg/median)", "rows": _cap_rows(link_type_benchmark_rows)},
                    {"title": "Бенчмарк по HTTP кодам (avg/median)", "rows": _cap_rows(http_benchmark_rows)},
                    {"title": "Бенчмарк follow/home/internal (avg/median)", "rows": _cap_rows(follow_home_internal_benchmark_rows)},
                    {"title": "DR распределение доноров по доменам (%)", "rows": _cap_rows(dr_distribution_matrix_rows)},
                    {"title": "Матрица возможностей доноров", "rows": _cap_rows(opportunity_domains_rows)},
                    {"title": "Ready-to-buy доноры (GGL/Miralinks)", "rows": _cap_rows(ready_buy_rows)},
                    {"title": "Gap donors priority", "rows": _cap_rows(gap_donors_priority_rows)},
                    {"title": "Donor overlap matrix", "rows": _cap_rows(donor_overlap_matrix_rows)},
                    *[{"title": str(s.get("title") or "Анализ"), "rows": _cap_rows(s.get("rows") or [])} for s in imported_analysis_sections],
                ],
                "additionalTables": [
                    {"title": "DR статистика", "rows": _cap_rows(dr_stats_rows)},
                    {"title": "DR buckets", "rows": _cap_rows(dr_bucket_rows)},
                    {"title": "DR buckets: наш сайт", "rows": _cap_rows(dr_bucket_our_rows)},
                    {"title": "DR buckets: конкуренты", "rows": _cap_rows(dr_bucket_comp_rows)},
                    {"title": "Domain zones", "rows": _cap_rows(zone_rows)},
                    {"title": "Follow / Nofollow", "rows": _cap_rows(follow_rows)},
                    {"title": "Follow / Nofollow %", "rows": _cap_rows(follow_mix_rows)},
                    {"title": "Follow/Nofollow по доменам %", "rows": _cap_rows(follow_domain_mix_rows)},
                    {"title": "Follow / Nofollow detailed", "rows": _cap_rows(follow_detail_rows)},
                    {"title": "Anchor mix %", "rows": _cap_rows(anchor_mix_rows)},
                    {"title": "Lost status mix", "rows": _cap_rows(lost_status_rows)},
                    {"title": "HTTP class mix", "rows": _cap_rows(http_class_rows)},
                    {"title": "Link type mix", "rows": _cap_rows(link_type_mix_rows)},
                    {"title": "Language mix", "rows": _cap_rows(language_mix_rows)},
                    {"title": "Donors with redirect 301", "rows": _cap_rows(redirect_301_rows)},
                    {"title": "Donors from homepage", "rows": _cap_rows(homepage_donor_rows)},
                    {"title": "Link types by competitor", "rows": _cap_rows(link_types_by_competitor)},
                    {"title": "HTTP codes by competitor", "rows": _cap_rows(http_codes_by_competitor)},
                    {"title": "Languages by competitor", "rows": _cap_rows(languages_by_competitor)},
                    {"title": "Link attributes", "rows": _cap_rows(link_attributes_rows)},
                    {"title": "Loss & recovery", "rows": _cap_rows(loss_recovery_rows)},
                    {"title": "HTTP/Type/Lang/Platform", "rows": _cap_rows(http_type_lang_platform_rows)},
                    {"title": "Risk signals", "rows": _cap_rows(risk_signals_rows)},
                    {"title": "Prompt templates", "rows": _cap_rows(prompt_templates_rows, 50)},
                    {"title": "Auto brand keywords", "rows": _cap_rows(brand_rows)},
                    {"title": "Source files", "rows": _cap_rows(file_summaries)},
                ],
                "duplicatesTables": [
                    {"title": "Duplicates with our site", "rows": _cap_rows(duplicates_with_our)},
                    {"title": "Duplicates with two competitors", "rows": _cap_rows(duplicates_without_our)},
                    {"title": "Single competitor domains", "rows": _cap_rows(single_competitor)},
                    {"title": "Single our domains", "rows": _cap_rows(single_our)},
                    {"title": "Priority domains", "rows": _cap_rows(priority_domains)},
                    {"title": "Priority score domains", "rows": _cap_rows(priority_score_domains)},
                    {"title": "Competitors analysis", "rows": _cap_rows(competitor_rows)},
                ],
            },
            "anchor_breakdown": dict(anchor_type_counter),
            "prompts": prompts,
        },
    }

    if progress_callback:
        progress_callback(100, "Аудит ссылочного профиля завершен")
    return result
