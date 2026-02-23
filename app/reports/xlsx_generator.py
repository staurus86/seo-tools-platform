"""Excel Report Generator."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from collections import Counter, defaultdict
from typing import Dict, Any, List, Set
from datetime import datetime
import os
import re
import math
import hashlib
import tempfile
from urllib.parse import urljoin

import requests

from app.config import settings


class XLSXGenerator:
    """Генератор Excel-отчетов."""
    
    def __init__(self):
        self.reports_dir = settings.REPORTS_DIR
        os.makedirs(self.reports_dir, exist_ok=True)

    def _repair_mojibake_text(self, value: Any) -> Any:
        """Best-effort mojibake repair for text values before writing to XLSX."""
        if not isinstance(value, str):
            return value
        text = value.replace("\x00", "")
        if not text:
            return text

        bad_pattern = re.compile(r"(?:[РС][^\s]|Ð.|Ñ.|â.|Ã.|�)")

        def bad_score(s: str) -> int:
            c1 = sum(1 for ch in s if 0x80 <= ord(ch) <= 0x9F)
            return len(bad_pattern.findall(s)) + (c1 * 2)

        def cyr_score(s: str) -> int:
            return len(re.findall(r"[А-Яа-яЁё]", s))

        original_bad = bad_score(text)
        if original_bad == 0:
            return text

        def cp1251_byte_for_char(ch: str) -> int | None:
            code = ord(ch)
            if code <= 0xFF:
                return code
            if ch == "Ё":
                return 0xA8
            if ch == "ё":
                return 0xB8
            if "А" <= ch <= "я":
                return code - 0x350
            try:
                raw = ch.encode("cp1251", errors="strict")
                if len(raw) == 1:
                    return raw[0]
            except Exception:
                return None
            return None

        def recode_lossless(src: str, src_enc: str, dst_enc: str) -> str:
            # Strict mode only: skip conversion if it is lossy or invalid.
            raw = src.encode(src_enc, errors="strict")
            return raw.decode(dst_enc, errors="strict")

        def recode_mixed_cp1251_utf8(src: str) -> str | None:
            raw = bytearray()
            for ch in src:
                b = cp1251_byte_for_char(ch)
                if b is None:
                    return None
                raw.append(b)
            try:
                return raw.decode("utf-8", errors="strict")
            except Exception:
                return None

        candidates = [text]
        mixed = recode_mixed_cp1251_utf8(text)
        if mixed and mixed != text:
            candidates.append(mixed)
        pipelines = (
            ("latin1", "utf-8"),
            ("cp1251", "utf-8"),
            ("latin1", "cp1251"),
        )
        for src_enc, dst_enc in pipelines:
            try:
                fixed = recode_lossless(text, src_enc, dst_enc)
                if fixed:
                    candidates.append(fixed)
                    # Try one more pass for double-mojibake strings.
                    try:
                        fixed2 = recode_lossless(fixed, src_enc, dst_enc)
                        if fixed2:
                            candidates.append(fixed2)
                    except Exception:
                        pass
            except Exception:
                continue

        best = text
        for cand in candidates[1:]:
            if bad_score(cand) < bad_score(best):
                best = cand
            elif bad_score(cand) == bad_score(best) and cyr_score(cand) > cyr_score(best):
                best = cand

        # Accept replacement only if mojibake markers are reduced and the output is not shorter.
        if bad_score(best) <= max(0, original_bad - 1) and len(best) >= len(text):
            return best
        return text

    def _sanitize_cell_value(self, value: Any) -> Any:
        if isinstance(value, bool):
            return "✅" if value else "❌"
        if isinstance(value, str):
            return self._repair_mojibake_text(value).strip()
        return value

    def _format_engine_label(self, engine: Any) -> str:
        value = str(engine or "legacy").lower()
        if value == "legacy":
            return "базовый"
        if value == "legacy-fallback":
            return "базовый (fallback)"
        return str(engine or "legacy")
    
    def _create_header_style(self):
        """Создает стиль заголовка."""
        return {
            'font': Font(bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    
    def _create_cell_style(self):
        """Создает базовый стиль ячейки."""
        return {
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=False),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    
    def _apply_style(self, cell, style):
        """Применяет стиль к ячейке."""
        cell.value = self._sanitize_cell_value(cell.value)
        if 'font' in style:
            cell.font = style['font']
        if 'fill' in style:
            cell.fill = style['fill']
        if 'alignment' in style:
            cell.alignment = style['alignment']
        if 'border' in style:
            cell.border = style['border']

    def _sanitize_workbook_text(self, wb: Workbook) -> None:
        """Normalize all worksheet cell values before saving."""
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell, MergedCell):
                        continue
                    cell.value = self._sanitize_cell_value(cell.value)

    def _save_workbook(self, wb: Workbook, filepath: str) -> None:
        self._sanitize_workbook_text(wb)
        wb.save(filepath)

    def _severity_style(self, severity: str) -> Dict[str, Any]:
        """Return fill/font style for a severity level."""
        sev = (severity or "info").lower()
        styles = {
            "critical": {
                "fill": PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid'),
                "font": Font(color='842029', bold=True),
            },
            "warning": {
                "fill": PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid'),
                "font": Font(color='664D03', bold=True),
            },
            "info": {
                "fill": PatternFill(start_color='D1ECF1', end_color='D1ECF1', fill_type='solid'),
                "font": Font(color='0C5460'),
            },
            "ok": {
                "fill": PatternFill(start_color='D1E7DD', end_color='D1E7DD', fill_type='solid'),
                "font": Font(color='0F5132', bold=True),
            },
        }
        return styles.get(sev, styles["info"])

    def _apply_severity_cell_style(self, cell, severity: str):
        sev_style = self._severity_style(severity)
        cell.fill = sev_style["fill"]
        cell.font = sev_style["font"]
        cell.alignment = Alignment(horizontal='center', vertical='center')

    def _apply_row_severity_fill(self, ws, row_idx: int, start_col: int, end_col: int, severity: str):
        """Apply severity background to row cells only for critical severity."""
        if (severity or "").lower() != "critical":
            return
        sev_fill = self._severity_style(severity)["fill"]
        for col in range(start_col, end_col + 1):
            ws.cell(row=row_idx, column=col).fill = sev_fill

    def _status_palette(self, status: str) -> Dict[str, Any]:
        state = (status or "").lower()
        palette = {
            "good": {"fill": "D1E7DD", "font": "0F5132", "icon": "✅"},
            "warn": {"fill": "FFF3CD", "font": "664D03", "icon": "⚠️"},
            "bad": {"fill": "F8D7DA", "font": "842029", "icon": "❌"},
            "neutral": {"fill": "E2E3E5", "font": "41464B", "icon": "•"},
        }
        return palette.get(state, palette["neutral"])

    def _kpi_status_from_value(self, header: str, value: Any) -> str:
        header_l = str(header or "").lower()
        value_s = str(value or "").strip().lower()

        def as_float(v: Any) -> float | None:
            try:
                return float(v)
            except Exception:
                return None

        # Explicit string levels and states.
        if "hierarchy status" in header_l:
            if any(k in value_s for k in ("good", "ok")):
                return "good"
            if "no headers" in value_s:
                return "warn"
            if any(k in value_s for k in ("bad", "broken", "missing", "wrong start", "skip")):
                return "bad"
            return "neutral"

        if value_s in {"high", "medium", "low"}:
            if "risk level" in header_l:
                return "bad" if value_s == "high" else "warn" if value_s == "medium" else "good"
            return "good" if value_s == "high" else "warn" if value_s == "medium" else "bad"
        if value_s in {"critical", "error", "bad", "fail", "failed"}:
            return "bad"
        if value_s in {"warning", "warn", "medium"}:
            return "warn"
        if value_s in {"ok", "good", "passed", "pass", "none"}:
            return "good"

        # Boolean-ish values.
        if value_s in {"true", "✅", "yes"}:
            if any(k in header_l for k in ("risk", "alert", "overuse", "truncation", "conflict", "mismatch", "duplicate", "error")):
                return "bad"
            return "good"
        if value_s in {"false", "❌", "no"}:
            if any(k in header_l for k in ("risk", "alert", "overuse", "truncation", "conflict", "mismatch", "duplicate", "error")):
                return "good"
            return "bad"

        # HTTP/status code.
        num = as_float(value)
        if num is not None and "status" in header_l:
            code = int(num)
            if 200 <= code < 300:
                return "good"
            if 300 <= code < 400:
                return "warn"
            if code >= 400 or code == 0:
                return "bad"

        # Numeric KPI by semantic group.
        if num is not None:
            if any(k in header_l for k in ("delta",)):
                return "good" if num <= 2 else "warn" if num <= 10 else "bad"
            if any(k in header_l for k in ("toxicity", "risk", "over threshold", "waste")):
                return "good" if num <= 20 else "warn" if num <= 50 else "bad"
            if any(k in header_l for k in ("score", "health", "quality", "coverage", "confidence", "roi")):
                return "good" if num >= 80 else "warn" if num >= 60 else "bad"
        return "neutral"

    def _is_kpi_header(self, header: str) -> bool:
        h = str(header or "").lower()
        keywords = (
            "score", "risk", "status", "toxicity", "health", "quality", "coverage", "delta",
            "confidence", "alert", "severity", "level", "indexable", "indexability", "spam",
            "canonical", "orphan", "duplicate", "mismatch", "conflict", "over threshold",
        )
        return any(k in h for k in keywords)

    def _apply_kpi_cell_style(self, cell, header: str, value: Any):
        status = self._kpi_status_from_value(header, value)
        palette = self._status_palette(status)
        cell.fill = PatternFill(start_color=palette["fill"], end_color=palette["fill"], fill_type='solid')
        cell.font = Font(color=palette["font"], bold=True)
        if self._is_kpi_header(header):
            cell.alignment = Alignment(horizontal='center', vertical='center')
        return status

    def _kpi_anomaly_direction(self, header: str) -> str:
        h = str(header or "").lower()
        if any(k in h for k in ("risk", "toxicity", "delta", "waste", "duplicate", "error", "issues total", "over threshold", "ms", "ttfb")):
            return "high_worse"
        if any(k in h for k in ("score", "health", "quality", "coverage", "confidence", "roi")):
            return "low_worse"
        return "none"

    def _quantile(self, values: List[float], q: float) -> float:
        if not values:
            return 0.0
        if q <= 0:
            return min(values)
        if q >= 1:
            return max(values)
        vals = sorted(values)
        pos = (len(vals) - 1) * q
        lo = int(pos)
        hi = min(lo + 1, len(vals) - 1)
        frac = pos - lo
        return vals[lo] * (1.0 - frac) + vals[hi] * frac

    def _sitemap_issue_severity(self, item: Dict[str, Any]) -> str:
        """Infer severity for sitemap file-level row."""
        if not item.get("ok", False):
            return "critical"
        if (item.get("status_code") or 0) >= 400:
            return "critical"
        if (item.get("duplicate_count") or 0) > 0:
            return "critical"
        if item.get("errors"):
            return "critical"
        if item.get("warnings"):
            return "warning"
        return "ok"
    
    def generate_site_analyze_report(self, task_id: str, data: Dict[str, Any]) -> str:
        """Генерирует XLSX-отчет общего анализа сайта."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Анализ сайта"
        
        # Header
        ws['A1'] = 'Отчет по SEO-анализу сайта'
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        # Basic info
        ws['A3'] = 'URL:'
        ws['B3'] = data.get('url', 'н/д')
        ws['A4'] = 'Проверено страниц:'
        ws['B4'] = data.get('pages_analyzed', 0)
        ws['A5'] = 'Дата завершения:'
        ws['B5'] = data.get('completed_at', 'н/д')
        
        # Results section
        ws['A7'] = 'Результаты'
        ws['A7'].font = Font(bold=True, size=14)
        
        results = data.get('results', {})
        row = 8
        
        # Headers
        headers = ['Показатель', 'Значение', 'Статус']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            self._apply_style(cell, self._create_header_style())
        
        # Sample data (will be replaced with real data from tools)
        sample_data = [
            ['Всего страниц', results.get('total_pages', 0), 'OK'],
            ['Статус', results.get('status', 'н/д'), 'OK'],
            ['Сводка', results.get('summary', 'н/д'), 'OK']
        ]
        
        for data_row in sample_data:
            row += 1
            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=col, value=value)
                self._apply_style(cell, self._create_cell_style())
        
        # Auto-adjust column widths
        for col in range(1, 4):
            ws.column_dimensions[get_column_letter(col)].width = 25
        
        # Save
        filepath = os.path.join(self.reports_dir, f"{task_id}.xlsx")
        self._save_workbook(wb, filepath)
        wb.close()
        return filepath
    
    def generate_robots_report(self, task_id: str, data: Dict[str, Any]) -> str:
        """Генерирует краткий XLSX-отчет по robots.txt."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Проверка Robots"
        
        ws['A1'] = 'Отчет по robots.txt'
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        ws['A3'] = 'URL:'
        ws['B3'] = data.get('url', 'н/д')
        
        results = data.get('results', {})
        ws['A5'] = 'Файл robots.txt найден:'
        ws['B5'] = 'Да' if results.get('robots_txt_found') else 'Нет'
        
        filepath = os.path.join(self.reports_dir, f"{task_id}.xlsx")
        self._save_workbook(wb, filepath)
        wb.close()
        return filepath
    
    def generate_sitemap_report(self, task_id: str, data: Dict[str, Any]) -> str:
        """Generate a complete sitemap validation XLSX report."""
        wb = Workbook()
        header_style = self._create_header_style()
        cell_style = self._create_cell_style()
        results = data.get('results', {}) or {}
        report_url = results.get("resolved_sitemap_url") or data.get('url', 'н/д')

        ws = wb.active
        ws.title = "Сводка"
        ws['A1'] = 'Отчет по валидации Sitemap'
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:E1')

        summary_rows = [
            ("URL", report_url),
            ("Валиден", "Да" if results.get("valid") else "Нет"),
            ("HTTP статус", results.get("status_code", "н/д")),
            ("Sitemap-файлов просканировано", results.get("sitemaps_scanned", 0)),
            ("Валидных sitemap-файлов", results.get("sitemaps_valid", 0)),
            ("Всего URL обнаружено", results.get("urls_count", 0)),
            ("Уникальных URL", results.get("unique_urls_count", 0)),
            ("Дубли URL", results.get("duplicate_urls_count", 0)),
            ("Некорректные URL", results.get("invalid_urls_count", 0)),
            ("Некорректный lastmod", results.get("invalid_lastmod_count", 0)),
            ("Некорректный changefreq", results.get("invalid_changefreq_count", 0)),
            ("Некорректный priority", results.get("invalid_priority_count", 0)),
            ("Общий размер (байт)", results.get("size", 0)),
            ("Экспорт URL ограничен", "Да" if results.get("urls_export_truncated") else "Нет"),
            ("Лимит URL для экспорта", results.get("max_export_urls", 0)),
            ("Размер части экспорта", results.get("export_chunk_size", 0)),
            ("Число частей экспорта", results.get("export_parts_count", 0)),
        ]
        row = 3
        for key, value in summary_rows:
            ws.cell(row=row, column=1, value=key).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
        ws.column_dimensions['A'].width = 34
        ws.column_dimensions['B'].width = 80

        files_ws = wb.create_sheet("Sitemap-файлы")
        files_headers = [
            "URL sitemap", "Тип", "HTTP", "OK", "URL",
            "Дубли", "Размер, байт", "Ошибки", "Предупреждения", "Критичность"
        ]
        for col, header in enumerate(files_headers, 1):
            cell = files_ws.cell(row=1, column=col, value=header)
            self._apply_style(cell, header_style)

        for row_idx, item in enumerate((results.get("sitemap_files", []) or []), start=2):
            severity = self._sitemap_issue_severity(item)
            values = [
                item.get("sitemap_url", ""),
                item.get("type", ""),
                item.get("status_code", ""),
                "Да" if item.get("ok") else "Нет",
                item.get("urls_count", 0),
                item.get("duplicate_count", 0),
                item.get("size_bytes", 0),
                " | ".join(item.get("errors", [])[:5]),
                " | ".join(item.get("warnings", [])[:5]),
                severity.capitalize(),
            ]
            for col, value in enumerate(values, 1):
                cell = files_ws.cell(row=row_idx, column=col, value=value)
                self._apply_style(cell, cell_style)
            self._apply_row_severity_fill(files_ws, row_idx, 1, len(files_headers), severity)
            self._apply_severity_cell_style(files_ws.cell(row=row_idx, column=len(files_headers)), severity)
        files_ws.freeze_panes = "A2"
        files_ws.auto_filter.ref = f"A1:{get_column_letter(len(files_headers))}1"
        for col, width in enumerate([72, 14, 10, 8, 12, 12, 14, 60, 60, 14], 1):
            files_ws.column_dimensions[get_column_letter(col)].width = width

        errors_ws = wb.create_sheet("Ошибки")
        errors_ws.cell(row=1, column=1, value="Ошибка")
        errors_ws.cell(row=1, column=2, value="Критичность")
        self._apply_style(errors_ws.cell(row=1, column=1), header_style)
        self._apply_style(errors_ws.cell(row=1, column=2), header_style)
        for idx, err in enumerate((results.get("errors", []) or []), start=2):
            err_cell = errors_ws.cell(row=idx, column=1, value=err)
            sev_cell = errors_ws.cell(row=idx, column=2, value="Критично")
            self._apply_style(err_cell, cell_style)
            self._apply_style(sev_cell, cell_style)
            self._apply_row_severity_fill(errors_ws, idx, 1, 2, "critical")
            self._apply_severity_cell_style(sev_cell, "critical")
        errors_ws.column_dimensions['A'].width = 140
        errors_ws.column_dimensions['B'].width = 14
        errors_ws.freeze_panes = "A2"
        errors_ws.auto_filter.ref = "A1:B1"

        warnings_ws = wb.create_sheet("Предупреждения")
        warnings_ws.cell(row=1, column=1, value="Предупреждение")
        warnings_ws.cell(row=1, column=2, value="Критичность")
        self._apply_style(warnings_ws.cell(row=1, column=1), header_style)
        self._apply_style(warnings_ws.cell(row=1, column=2), header_style)
        for idx, warn in enumerate((results.get("warnings", []) or []), start=2):
            warn_cell = warnings_ws.cell(row=idx, column=1, value=warn)
            sev_cell = warnings_ws.cell(row=idx, column=2, value="Предупреждение")
            self._apply_style(warn_cell, cell_style)
            self._apply_style(sev_cell, cell_style)
            self._apply_row_severity_fill(warnings_ws, idx, 1, 2, "warning")
            self._apply_severity_cell_style(sev_cell, "warning")
        warnings_ws.column_dimensions['A'].width = 140
        warnings_ws.column_dimensions['B'].width = 14
        warnings_ws.freeze_panes = "A2"
        warnings_ws.auto_filter.ref = "A1:B1"

        dup_ws = wb.create_sheet("Дубли")
        dup_headers = ["URL", "Первый sitemap", "Сitemap-дубль", "Критичность"]
        for col, header in enumerate(dup_headers, 1):
            cell = dup_ws.cell(row=1, column=col, value=header)
            self._apply_style(cell, header_style)
        for row_idx, item in enumerate((results.get("duplicate_details", []) or []), start=2):
            dup_ws.cell(row=row_idx, column=1, value=item.get("url", ""))
            dup_ws.cell(row=row_idx, column=2, value=item.get("first_sitemap", ""))
            dup_ws.cell(row=row_idx, column=3, value=item.get("duplicate_sitemap", ""))
            dup_ws.cell(row=row_idx, column=4, value="Критично")
            for col in range(1, 5):
                self._apply_style(dup_ws.cell(row=row_idx, column=col), cell_style)
            self._apply_row_severity_fill(dup_ws, row_idx, 1, 4, "critical")
            self._apply_severity_cell_style(dup_ws.cell(row=row_idx, column=4), "critical")
        dup_ws.freeze_panes = "A2"
        dup_ws.auto_filter.ref = "A1:D1"
        dup_ws.column_dimensions['A'].width = 80
        dup_ws.column_dimensions['B'].width = 60
        dup_ws.column_dimensions['C'].width = 60
        dup_ws.column_dimensions['D'].width = 14

        rec_ws = wb.create_sheet("Рекомендации")
        rec_ws.cell(row=1, column=1, value="Рекомендация")
        rec_ws.cell(row=1, column=2, value="Критичность")
        self._apply_style(rec_ws.cell(row=1, column=1), header_style)
        self._apply_style(rec_ws.cell(row=1, column=2), header_style)
        for idx, rec in enumerate((results.get("recommendations", []) or []), start=2):
            rec_cell = rec_ws.cell(row=idx, column=1, value=rec)
            sev_cell = rec_ws.cell(row=idx, column=2, value="Инфо")
            self._apply_style(rec_cell, cell_style)
            self._apply_style(sev_cell, cell_style)
            self._apply_row_severity_fill(rec_ws, idx, 1, 2, "info")
            self._apply_severity_cell_style(sev_cell, "info")
        rec_ws.column_dimensions['A'].width = 140
        rec_ws.column_dimensions['B'].width = 14
        rec_ws.freeze_panes = "A2"
        rec_ws.auto_filter.ref = "A1:B1"

        tool_notes_ws = wb.create_sheet("Служебные заметки")
        tool_notes_ws.cell(row=1, column=1, value="Заметка")
        self._apply_style(tool_notes_ws.cell(row=1, column=1), header_style)
        tool_notes = results.get("tool_notes", []) or []
        if tool_notes:
            for idx, note in enumerate(tool_notes, start=2):
                c = tool_notes_ws.cell(row=idx, column=1, value=str(note))
                self._apply_style(c, cell_style)
        else:
            c = tool_notes_ws.cell(row=2, column=1, value="Служебные заметки отсутствуют")
            self._apply_style(c, cell_style)
        tool_notes_ws.column_dimensions['A'].width = 140
        tool_notes_ws.freeze_panes = "A2"

        insights_ws = wb.create_sheet("Инсайты")
        insights_ws.cell(row=1, column=1, value="Метрика")
        insights_ws.cell(row=1, column=2, value="Значение")
        insights_ws.cell(row=1, column=3, value="Интерпретация")
        for col in range(1, 4):
            self._apply_style(insights_ws.cell(row=1, column=col), header_style)

        total_urls = int(results.get("urls_count", 0) or 0)
        unique_urls = int(results.get("unique_urls_count", 0) or 0)
        scanned = int(results.get("sitemaps_scanned", 0) or 0)
        valid_scanned = int(results.get("sitemaps_valid", 0) or 0)
        exported_urls = len(results.get("export_urls", []) or [])
        warnings_count = len(results.get("warnings", []) or [])
        errors_count = len(results.get("errors", []) or [])
        unique_ratio = round((unique_urls / total_urls) * 100, 2) if total_urls > 0 else 0.0
        valid_ratio = round((valid_scanned / scanned) * 100, 2) if scanned > 0 else 0.0
        export_ratio = round((exported_urls / total_urls) * 100, 2) if total_urls > 0 else 0.0

        insight_rows = [
            ("Оценка качества", results.get("quality_score", "н/д"), f"Оценка {results.get('quality_grade', 'н/д')}"),
            ("Ошибки / Предупреждения", f"{errors_count} / {warnings_count}", "Чем меньше, тем лучше"),
            ("Доля уникальных URL", f"{unique_ratio}%", "Уникальные URL от общего числа обнаруженных URL"),
            ("Доля валидных sitemap-файлов", f"{valid_ratio}%", "Валидные sitemap-файлы от числа просканированных"),
            ("Покрытие экспорта", f"{export_ratio}%", "Экспортированные URL от общего числа обнаруженных URL"),
            ("Ограничение обхода достигнуто", "Да" if any("лимит обхода" in str(w).lower() for w in (results.get("tool_notes", []) or [])) else "Нет", "Если Да, просканированы не все sitemap-файлы"),
            ("Экспорт ограничен", "Да" if results.get("urls_export_truncated") else "Нет", "Если Да, используйте экспорт частями для полного списка URL"),
        ]
        for idx, (k, v, i_txt) in enumerate(insight_rows, start=2):
            self._apply_style(insights_ws.cell(row=idx, column=1, value=k), cell_style)
            self._apply_style(insights_ws.cell(row=idx, column=2, value=v), cell_style)
            self._apply_style(insights_ws.cell(row=idx, column=3, value=i_txt), cell_style)

        highlights = results.get("highlights", []) or []
        start_row = len(insight_rows) + 4
        self._apply_style(insights_ws.cell(row=start_row, column=1, value="Ключевые выводы"), header_style)
        if highlights:
            for offset, item in enumerate(highlights, start=1):
                self._apply_style(insights_ws.cell(row=start_row + offset, column=1, value=str(item)), cell_style)
        else:
            self._apply_style(insights_ws.cell(row=start_row + 1, column=1, value="Ключевые выводы отсутствуют"), cell_style)

        insights_ws.column_dimensions['A'].width = 34
        insights_ws.column_dimensions['B'].width = 24
        insights_ws.column_dimensions['C'].width = 80
        insights_ws.freeze_panes = "A2"

        urls_ws = wb.create_sheet("Экспорт URL")
        urls_ws.cell(row=1, column=1, value="URL")
        self._apply_style(urls_ws.cell(row=1, column=1), header_style)
        for idx, u in enumerate((results.get("export_urls", []) or []), start=2):
            c = urls_ws.cell(row=idx, column=1, value=u)
            self._apply_style(c, cell_style)
        urls_ws.column_dimensions['A'].width = 120
        urls_ws.freeze_panes = "A2"
        urls_ws.auto_filter.ref = "A1:A1"

        filepath = os.path.join(self.reports_dir, f"{task_id}.xlsx")
        self._save_workbook(wb, filepath)
        wb.close()
        return filepath

    def generate_render_report(self, task_id: str, data: Dict[str, Any]) -> str:
        """Генерирует детальный XLSX-отчет по аудиту рендеринга (С„окус на проблемах)."""
        wb = Workbook()
        header_style = self._create_header_style()
        cell_style = self._create_cell_style()

        results = data.get('results', {}) or {}
        summary = results.get('summary', {}) or {}
        variants = results.get('variants', []) or []
        issues = results.get('issues', []) or []
        recommendations = results.get('recommendations', []) or []

        ws = wb.active
        ws.title = 'Сводка'
        ws['A1'] = 'Отчет аудита рендеринга (JS и без JS)'
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:E1')

        rows = [
            ('Адрес URL', data.get('url', 'н/д')),
            ('Движок', self._format_engine_label(results.get('engine', 'legacy'))),
            ('Профилей', summary.get('variants_total', len(variants))),
            ('Оценка', summary.get('score', 'н/д')),
            ('Критичные', summary.get('critical_issues', 0)),
            ('Предупреждения', summary.get('warning_issues', 0)),
            ('Потерянных элементов всего', summary.get('missing_total', 0)),
            ('Потери средний %', summary.get('avg_missing_pct', 0)),
            ('Ср. загрузка без JS (мс)', summary.get('avg_raw_load_ms', 0)),
            ('Ср. загрузка JS (мс)', summary.get('avg_js_load_ms', 0)),
        ]
        row_num = 3
        for key, value in rows:
            ws.cell(row=row_num, column=1, value=key).font = Font(bold=True)
            ws.cell(row=row_num, column=2, value=value)
            row_num += 1
        ws.column_dimensions['A'].width = 32
        ws.column_dimensions['B'].width = 80

        variant_ws = wb.create_sheet('Профили')
        headers = ['Профиль', 'Оценка', 'Потери', 'Потери %', 'H1 без JS', 'H1 с JS', 'Ссылки без JS', 'Ссылки с JS', 'Структурированные данные без JS', 'Структурированные данные с JS']
        for col, header in enumerate(headers, 1):
            self._apply_style(variant_ws.cell(row=1, column=col, value=header), header_style)

        for row_idx, variant in enumerate(variants, start=2):
            metrics = variant.get('metrics', {}) or {}
            raw = variant.get('raw', {}) or {}
            rendered = variant.get('rendered', {}) or {}
            values = [
                variant.get('variant_label') or variant.get('variant_id', ''),
                float(metrics.get('score', 0.0) or 0.0),
                int(metrics.get('total_missing', 0) or 0),
                float(metrics.get('missing_pct', 0.0) or 0.0),
                int(raw.get('h1_count', 0) or 0),
                int(rendered.get('h1_count', 0) or 0),
                int(raw.get('links_count', 0) or 0),
                int(rendered.get('links_count', 0) or 0),
                int(raw.get('structured_data_count', 0) or 0),
                int(rendered.get('structured_data_count', 0) or 0),
            ]
            for col, value in enumerate(values, 1):
                self._apply_style(variant_ws.cell(row=row_idx, column=col, value=value), cell_style)

            score = float(metrics.get('score', 0.0) or 0.0)
            severity = 'ok' if score >= 80 else ('warning' if score >= 60 else 'critical')
            self._apply_row_severity_fill(variant_ws, row_idx, 1, len(headers), severity)
            self._apply_severity_cell_style(variant_ws.cell(row=row_idx, column=2), severity)

        variant_ws.freeze_panes = 'A2'
        variant_ws.auto_filter.ref = f"A1:J{max(2, len(variants)+1)}"
        for col, width in enumerate([28, 12, 12, 12, 12, 10, 14, 10, 14, 10], 1):
            variant_ws.column_dimensions[get_column_letter(col)].width = width

        issues_ws = wb.create_sheet('Проблемы')
        issue_headers = ['Критичность', 'Профиль', 'Код', 'Заголовок', 'Детали']
        for col, header in enumerate(issue_headers, 1):
            self._apply_style(issues_ws.cell(row=1, column=col, value=header), header_style)

        for row_idx, issue in enumerate(issues, start=2):
            severity = (issue.get('severity') or 'info').lower()
            values = [
                severity.upper(),
                issue.get('variant', ''),
                issue.get('code', ''),
                issue.get('title', ''),
                issue.get('details', ''),
            ]
            for col, value in enumerate(values, 1):
                self._apply_style(issues_ws.cell(row=row_idx, column=col, value=value), cell_style)
            self._apply_row_severity_fill(issues_ws, row_idx, 1, len(issue_headers), severity)
            self._apply_severity_cell_style(issues_ws.cell(row=row_idx, column=1), severity)

        issues_ws.freeze_panes = 'A2'
        issues_ws.auto_filter.ref = f"A1:E{max(2, len(issues)+1)}"
        for col, width in enumerate([12, 24, 24, 34, 96], 1):
            issues_ws.column_dimensions[get_column_letter(col)].width = width

        rec_ws = wb.create_sheet('Рекомендации')
        self._apply_style(rec_ws.cell(row=1, column=1, value='Рекомендация'), header_style)
        for idx, text in enumerate(recommendations, start=2):
            self._apply_style(rec_ws.cell(row=idx, column=1, value=text), cell_style)
        rec_ws.column_dimensions['A'].width = 160
        rec_ws.freeze_panes = 'A2'
        rec_ws.auto_filter.ref = 'A1:A1'

        missing_ws = wb.create_sheet('Потерянные элементы')
        missing_headers = ['Профиль', 'Категория', 'Элемент']
        for col, header in enumerate(missing_headers, 1):
            self._apply_style(missing_ws.cell(row=1, column=col, value=header), header_style)
        row_idx = 2
        for variant in variants:
            profile = variant.get('variant_label') or variant.get('variant_id', '')
            missing = variant.get('missing', {}) or {}
            for key, label in [
                ('visible_text', 'Текст только в JS'),
                ('headings', 'Заголовки только в JS'),
                ('links', 'Ссылки только в JS'),
                ('structured_data', 'Структурированные данные только в JS'),
            ]:
                values = missing.get(key, []) or []
                for value in values:
                    self._apply_style(missing_ws.cell(row=row_idx, column=1, value=profile), cell_style)
                    self._apply_style(missing_ws.cell(row=row_idx, column=2, value=label), cell_style)
                    self._apply_style(missing_ws.cell(row=row_idx, column=3, value=str(value)), cell_style)
                    row_idx += 1
        missing_ws.freeze_panes = 'A2'
        missing_ws.auto_filter.ref = f"A1:C{max(2, row_idx-1)}"
        missing_ws.column_dimensions['A'].width = 24
        missing_ws.column_dimensions['B'].width = 28
        missing_ws.column_dimensions['C'].width = 140

        meta_ws = wb.create_sheet('Мета (не SEO)')
        meta_headers = ['Профиль', 'Ключ', 'Без JS', 'С JS', 'Статус']
        for col, header in enumerate(meta_headers, 1):
            self._apply_style(meta_ws.cell(row=1, column=col, value=header), header_style)
        row_idx = 2
        status_map = {
            'same': 'Совпадает',
            'changed': 'Изменено',
            'only_rendered': 'Только в JS',
            'only_raw': 'Только без JS',
        }
        for variant in variants:
            profile = variant.get('variant_label') or variant.get('variant_id', '')
            comparison = ((variant.get('meta_non_seo') or {}).get('comparison') or {})
            for item in (comparison.get('items') or []):
                self._apply_style(meta_ws.cell(row=row_idx, column=1, value=profile), cell_style)
                self._apply_style(meta_ws.cell(row=row_idx, column=2, value=item.get('key', '')), cell_style)
                self._apply_style(meta_ws.cell(row=row_idx, column=3, value=item.get('raw', '')), cell_style)
                self._apply_style(meta_ws.cell(row=row_idx, column=4, value=item.get('rendered', '')), cell_style)
                self._apply_style(
                    meta_ws.cell(row=row_idx, column=5, value=status_map.get(item.get('status', ''), item.get('status', ''))),
                    cell_style,
                )
                row_idx += 1
        meta_ws.freeze_panes = 'A2'
        meta_ws.auto_filter.ref = f"A1:E{max(2, row_idx-1)}"
        meta_ws.column_dimensions['A'].width = 24
        meta_ws.column_dimensions['B'].width = 36
        meta_ws.column_dimensions['C'].width = 80
        meta_ws.column_dimensions['D'].width = 80
        meta_ws.column_dimensions['E'].width = 18

        seo_ws = wb.create_sheet('SEO обязательно')
        seo_headers = ['Профиль', 'Элемент', 'Без JS', 'С JS', 'Статус', 'Исправление']
        for col, header in enumerate(seo_headers, 1):
            self._apply_style(seo_ws.cell(row=1, column=col, value=header), header_style)
        row_idx = 2
        status_map = {'pass': 'Пройдено', 'warn': 'Предупреждение', 'fail': 'Критично'}
        for variant in variants:
            profile = variant.get('variant_label') or variant.get('variant_id', '')
            for item in ((variant.get('seo_required') or {}).get('items') or []):
                self._apply_style(seo_ws.cell(row=row_idx, column=1, value=profile), cell_style)
                self._apply_style(seo_ws.cell(row=row_idx, column=2, value=item.get('label', '')), cell_style)
                self._apply_style(seo_ws.cell(row=row_idx, column=3, value=str(item.get('raw', ''))), cell_style)
                self._apply_style(seo_ws.cell(row=row_idx, column=4, value=str(item.get('rendered', ''))), cell_style)
                status = status_map.get(item.get('status', ''), item.get('status', ''))
                self._apply_style(seo_ws.cell(row=row_idx, column=5, value=status), cell_style)
                self._apply_style(seo_ws.cell(row=row_idx, column=6, value=item.get('fix', '')), cell_style)
                sev = 'critical' if item.get('status') == 'fail' else ('warning' if item.get('status') == 'warn' else 'ok')
                self._apply_row_severity_fill(seo_ws, row_idx, 1, len(seo_headers), sev)
                row_idx += 1
        seo_ws.freeze_panes = 'A2'
        seo_ws.auto_filter.ref = f"A1:F{max(2, row_idx-1)}"
        seo_ws.column_dimensions['A'].width = 24
        seo_ws.column_dimensions['B'].width = 44
        seo_ws.column_dimensions['C'].width = 40
        seo_ws.column_dimensions['D'].width = 40
        seo_ws.column_dimensions['E'].width = 12
        seo_ws.column_dimensions['F'].width = 70

        filepath = os.path.join(self.reports_dir, f"{task_id}.xlsx")
        self._save_workbook(wb, filepath)
        return filepath
    def generate_mobile_report(self, task_id: str, data: Dict[str, Any]) -> str:
        """Generate detailed mobile XLSX report."""
        wb = Workbook()
        header_style = self._create_header_style()
        cell_style = self._create_cell_style()

        results = data.get("results", {}) or {}
        summary = results.get("summary", {}) or {}
        devices = results.get("device_results", []) or []
        issues = results.get("issues", []) or []
        recommendations = results.get("recommendations", []) or []
        artifacts = results.get("artifacts", {}) or {}
        screenshot_dir = str(artifacts.get("screenshot_dir") or "").strip()
        server_base_url = str(data.get("server_base_url") or "").strip()
        temp_screenshots: List[str] = []

        def _resolve_mobile_screenshot_path(device: Dict[str, Any]) -> str:
            candidates: List[str] = []
            raw_path = str(device.get("screenshot_path") or "").strip()
            shot_name = str(device.get("screenshot_name") or "").strip()
            shot_url = str(device.get("screenshot_url") or "").strip()

            if raw_path:
                candidates.append(raw_path)
            if shot_name and screenshot_dir:
                candidates.append(os.path.join(screenshot_dir, shot_name))
            if shot_name:
                candidates.append(os.path.join(self.reports_dir, "mobile", task_id, "screenshots", shot_name))
            if raw_path:
                candidates.append(
                    os.path.join(self.reports_dir, "mobile", task_id, "screenshots", os.path.basename(raw_path))
                )

            for candidate in candidates:
                if candidate and os.path.exists(candidate):
                    return candidate

            if shot_url:
                if shot_url.startswith("/"):
                    if server_base_url:
                        shot_url = urljoin(server_base_url, shot_url)
                    else:
                        shot_url = ""
                if shot_url:
                    try:
                        response = requests.get(shot_url, timeout=25)
                        if response.status_code == 200 and response.content:
                            fd, temp_path = tempfile.mkstemp(prefix=f"mobile_xlsx_{task_id}_", suffix=".png")
                            os.close(fd)
                            with open(temp_path, "wb") as f:
                                f.write(response.content)
                            temp_screenshots.append(temp_path)
                            return temp_path
                    except Exception:
                        pass
            return ""

        ws = wb.active
        ws.title = "Сводка"
        ws["A1"] = "Отчет по мобильной совместимости"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:E1")

        rows = [
            ("URL", data.get("url", "н/д")),
            ("Движок", self._format_engine_label(results.get("engine", "legacy"))),
            ("Режим", results.get("mode", "full")),
            ("Оценка", results.get("score", 0)),
            ("Мобильно-дружелюбно", "Да" if results.get("mobile_friendly") else "Нет"),
            ("Всего устройств", summary.get("total_devices", len(results.get("devices_tested", [])))),
            ("Дружелюбных устройств", summary.get("mobile_friendly_devices", 0)),
            ("Недружелюбных устройств", summary.get("non_friendly_devices", 0)),
            ("Среднее время загрузки (мс)", summary.get("avg_load_time_ms", 0)),
            ("Всего проблем", results.get("issues_count", 0)),
            ("Критично", summary.get("critical_issues", 0)),
            ("Предупреждение", summary.get("warning_issues", 0)),
            ("Инфо", summary.get("info_issues", 0)),
        ]
        r = 3
        for key, val in rows:
            ws.cell(row=r, column=1, value=key).font = Font(bold=True)
            ws.cell(row=r, column=2, value=val)
            r += 1
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 80

        dws = wb.create_sheet("Устройства")
        headers = ["Устройство", "Категория", "HTTP", "Мобильно-дружелюбно", "Проблемы", "Время загрузки (мс)", "Скриншот", "Критичность"]
        for col, header in enumerate(headers, 1):
            self._apply_style(dws.cell(row=1, column=col, value=header), header_style)

        for row_idx, d in enumerate(devices, start=2):
            if d.get("issues_count", 0) > 0 and not d.get("mobile_friendly"):
                severity = "warning"
            elif d.get("issues_count", 0) > 0:
                severity = "info"
            else:
                severity = "ok"

            values = [
                d.get("device_name", ""),
                d.get("category", ""),
                d.get("status_code", "н/д"),
                "Да" if d.get("mobile_friendly") else "Нет",
                d.get("issues_count", 0),
                d.get("load_time_ms", 0),
                d.get("screenshot_name", ""),
                severity.capitalize(),
            ]
            for col, value in enumerate(values, 1):
                self._apply_style(dws.cell(row=row_idx, column=col, value=value), cell_style)
            self._apply_row_severity_fill(dws, row_idx, 1, len(headers), severity)
            self._apply_severity_cell_style(dws.cell(row=row_idx, column=len(headers)), severity)
        dws.freeze_panes = "A2"
        dws.auto_filter.ref = "A1:H1"
        for col, width in enumerate([28, 12, 10, 14, 10, 12, 40, 12], 1):
            dws.column_dimensions[get_column_letter(col)].width = width

        iws = wb.create_sheet("Проблемы")
        issue_headers = ["Критичность", "Устройство", "Код", "Проблема", "Детали"]
        for col, header in enumerate(issue_headers, 1):
            self._apply_style(iws.cell(row=1, column=col, value=header), header_style)
        for row_idx, issue in enumerate(issues, start=2):
            severity = (issue.get("severity") or "info").lower()
            values = [
                severity.capitalize(),
                issue.get("device", ""),
                issue.get("code", ""),
                issue.get("title", ""),
                issue.get("details", ""),
            ]
            for col, value in enumerate(values, 1):
                self._apply_style(iws.cell(row=row_idx, column=col, value=value), cell_style)
            self._apply_row_severity_fill(iws, row_idx, 1, len(issue_headers), severity)
            self._apply_severity_cell_style(iws.cell(row=row_idx, column=1), severity)
        iws.freeze_panes = "A2"
        iws.auto_filter.ref = "A1:E1"
        for col, width in enumerate([12, 24, 20, 30, 80], 1):
            iws.column_dimensions[get_column_letter(col)].width = width

        rws = wb.create_sheet("Рекомендации")
        self._apply_style(rws.cell(row=1, column=1, value="Рекомендация"), header_style)
        for idx, rec in enumerate(recommendations, start=2):
            self._apply_style(rws.cell(row=idx, column=1, value=rec), cell_style)
        rws.column_dimensions["A"].width = 160
        rws.freeze_panes = "A2"

        sws = wb.create_sheet("Скриншоты")
        shot_headers = ["Устройство", "Имя скриншота", "Путь", "URL", "Превью"]
        for col, header in enumerate(shot_headers, 1):
            self._apply_style(sws.cell(row=1, column=col, value=header), header_style)
        for row_idx, d in enumerate(devices, start=2):
            vals = [d.get("device_name", ""), d.get("screenshot_name", ""), d.get("screenshot_path", ""), d.get("screenshot_url", "")]
            for col, value in enumerate(vals, 1):
                self._apply_style(sws.cell(row=row_idx, column=col, value=value), cell_style)
            resolved_shot = _resolve_mobile_screenshot_path(d)
            if resolved_shot and os.path.exists(resolved_shot):
                try:
                    img = XLImage(resolved_shot)
                    img.width = 360
                    img.height = 200
                    sws.row_dimensions[row_idx].height = 155
                    sws.add_image(img, f"E{row_idx}")
                except Exception:
                    pass
        sws.freeze_panes = "A2"
        sws.auto_filter.ref = "A1:E1"
        for col, width in enumerate([26, 40, 80, 48, 52], 1):
            sws.column_dimensions[get_column_letter(col)].width = width

        filepath = os.path.join(self.reports_dir, f"{task_id}.xlsx")
        try:
            self._save_workbook(wb, filepath)
            return filepath
        finally:
            for temp_path in temp_screenshots:
                try:
                    if temp_path and os.path.exists(temp_path):
                        os.remove(temp_path)
                except Exception:
                    pass

    def generate_bot_report(self, task_id: str, data: Dict[str, Any]) -> str:
        """Generate detailed bot accessibility report with severity styling."""
        wb = Workbook()
        header_style = self._create_header_style()
        cell_style = self._create_cell_style()
        results = data.get("results", {}) or {}
        report_url = data.get("url", "н/д")
        summary = results.get("summary", {}) or {}
        bot_rows = results.get("bot_rows", []) or []
        bot_results = results.get("bot_results", {}) or {}
        category_stats = results.get("category_stats", []) or []
        priority_blockers = results.get("priority_blockers", []) or []
        playbooks = results.get("playbooks", []) or []
        baseline_diff = results.get("baseline_diff", {}) or {}
        trend = results.get("trend", {}) or {}
        alerts = results.get("alerts", []) or []
        robots_linter = (results.get("robots_linter", {}) or {}).get("findings", []) or []
        allowlist_sim = (results.get("allowlist_simulator", {}) or {}).get("scenarios", []) or []
        action_center = (results.get("action_center", {}) or {}).get("by_owner", {}) or {}
        evidence_pack = (results.get("evidence_pack", {}) or {}).get("rows", []) or []
        batch_runs = results.get("batch_runs", []) or []
        issues = results.get("issues", []) or []
        recommendations = results.get("recommendations", []) or []

        ws = wb.active
        ws.title = "Сводка"
        ws["A1"] = "Отчет по доступности ботов"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:E1")

        summary_rows = [
            ("URL", report_url),
            ("Движок", self._format_engine_label(results.get("engine", "legacy"))),
            ("Домен", results.get("domain", "")),
            ("Ботов проверено", len(results.get("bots_checked", []) or list(bot_results.keys()))),
            ("Доступно", summary.get("accessible", 0)),
            ("Недоступно", summary.get("unavailable", 0)),
            ("С контентом", summary.get("with_content", 0)),
            ("Без контента", summary.get("without_content", 0)),
            ("Заблокировано robots", summary.get("robots_disallowed", 0)),
            ("Заблокировано X-Robots", summary.get("x_robots_forbidden", 0)),
            ("Заблокировано Meta Robots", summary.get("meta_forbidden", 0)),
            ("Режим AI-политики", "намеренные блокировки AI" if results.get("ai_block_expected") else "строгая доступность"),
            ("Ожидаемо заблокировано AI", summary.get("expected_ai_policy_blocked", 0)),
            ("Среднее время ответа (мс)", summary.get("avg_response_time_ms", "")),
        ]
        row = 3
        for key, value in summary_rows:
            ws.cell(row=row, column=1, value=key).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 90

        exec_ws = wb.create_sheet("Краткая сводка", 1)
        exec_ws["A1"] = "Проверка доступности ботов - краткая сводка"
        exec_ws["A1"].font = Font(bold=True, size=16)
        exec_ws.merge_cells("A1:G1")

        total_bots = int(summary.get("total", 0) or len(bot_rows) or 0)
        indexable_bots = int(summary.get("indexable", 0) or 0)
        indexable_pct = round((indexable_bots / total_bots) * 100, 1) if total_bots > 0 else 0.0
        sla_breaches = sum(1 for c in category_stats if not bool(c.get("sla_met")))
        waf_detected = int(summary.get("waf_cdn_detected", 0) or 0)

        executive_kpis = [
            ("URL", report_url, "Область"),
            ("Ботов проверено", total_bots, "Всего протестированных user-agent"),
            ("Индексируемые боты", indexable_bots, "Боты, которые могут индексировать контент"),
            ("Доля индексируемых", f"{indexable_pct}%", "Цель зависит от SLA-профиля"),
            ("Категорий ниже SLA", sla_breaches, "Категории ниже SLA-цели"),
            ("Обнаружений WAF/CDN", waf_detected, "Потенциальные anti-bot ограничения"),
        ]
        for row_idx, (metric, value, note) in enumerate(executive_kpis, start=3):
            self._apply_style(exec_ws.cell(row=row_idx, column=1, value=metric), header_style)
            self._apply_style(exec_ws.cell(row=row_idx, column=2, value=value), cell_style)
            self._apply_style(exec_ws.cell(row=row_idx, column=3, value=note), cell_style)

        playbook_by_code = {}
        for item in playbooks:
            code = str(item.get("blocker_code", "") or "").strip()
            if code and code not in playbook_by_code:
                playbook_by_code[code] = item

        top_blockers = sorted(
            (priority_blockers or []),
            key=lambda x: float(x.get("priority_score", 0) or 0),
            reverse=True,
        )[:5]
        blocker_header_row = 11
        blocker_headers = ["Ранг", "Код", "Заголовок", "Приоритет", "Затронуто ботов", "Владелец", "Топ действий"]
        for col, header in enumerate(blocker_headers, 1):
            self._apply_style(exec_ws.cell(row=blocker_header_row, column=col, value=header), header_style)

        for idx, blocker in enumerate(top_blockers, start=1):
            code = str(blocker.get("code", "") or "")
            pb = playbook_by_code.get(code, {})
            actions = pb.get("actions", []) or []
            action_preview = " | ".join(actions[:2]) if actions else "-"
            priority = float(blocker.get("priority_score", 0) or 0)
            values = [
                idx,
                code,
                blocker.get("title", ""),
                priority,
                int(blocker.get("affected_bots", 0) or 0),
                pb.get("owner", ""),
                action_preview,
            ]
            row_idx = blocker_header_row + idx
            for col, value in enumerate(values, 1):
                self._apply_style(exec_ws.cell(row=row_idx, column=col, value=value), cell_style)

            severity = "critical" if priority >= 20 else ("warning" if priority >= 10 else "info")
            self._apply_row_severity_fill(exec_ws, row_idx, 1, len(blocker_headers), severity)
            self._apply_severity_cell_style(exec_ws.cell(row=row_idx, column=4), severity)

        bucket_headers = ["Этап спринта", "Владелец", "Заголовок действия", "Приоритет", "Действия"]
        bucket_start = blocker_header_row + max(2, len(top_blockers) + 3)
        for col, header in enumerate(bucket_headers, 1):
            self._apply_style(exec_ws.cell(row=bucket_start, column=col, value=header), header_style)

        def _bucket_name(priority_score: float) -> str:
            if priority_score >= 20:
                return "Сейчас"
            if priority_score >= 12:
                return "Далее"
            return "Позже"

        for idx, pb in enumerate(playbooks[:15], start=1):
            score = float(pb.get("priority_score", 0) or 0)
            row_idx = bucket_start + idx
            values = [
                _bucket_name(score),
                pb.get("owner", ""),
                pb.get("title", ""),
                score,
                " | ".join(pb.get("actions", []) or []),
            ]
            for col, value in enumerate(values, 1):
                self._apply_style(exec_ws.cell(row=row_idx, column=col, value=value), cell_style)

        for col, width in enumerate([16, 20, 40, 12, 16, 18, 72], 1):
            exec_ws.column_dimensions[get_column_letter(col)].width = width
        exec_ws.freeze_panes = "A12"
        exec_ws.auto_filter.ref = "A11:G11"

        results_ws = wb.create_sheet("Результаты ботов")
        result_headers = [
            "Бот",
            "Категория",
            "HTTP",
            "Доступен",
            "Есть контент",
            "Разрешен robots",
            "X-Robots-Tag",
            "Заблокирован X-Robots",
            "Meta Robots",
            "Заблокирован Meta",
            "Время ответа (мс)",
            "Финальный URL",
            "Ошибка",
            "Критичность",
        ]
        for col, header in enumerate(result_headers, 1):
            self._apply_style(results_ws.cell(row=1, column=col, value=header), header_style)

        if not bot_rows and bot_results:
            for bot, item in bot_results.items():
                bot_rows.append({
                    "bot_name": bot,
                    "category": item.get("category", ""),
                    "status": item.get("status"),
                    "accessible": item.get("accessible"),
                    "has_content": item.get("has_content"),
                    "robots_allowed": item.get("robots_allowed"),
                    "x_robots_tag": item.get("x_robots_tag"),
                    "x_robots_forbidden": item.get("x_robots_forbidden"),
                    "meta_robots": item.get("meta_robots"),
                    "meta_forbidden": item.get("meta_forbidden"),
                    "response_time_ms": item.get("response_time_ms"),
                    "final_url": item.get("final_url"),
                    "error": item.get("error"),
                })

        for row_idx, item in enumerate(bot_rows, start=2):
            if item.get("error") or not item.get("accessible"):
                severity = "critical"
            elif not item.get("has_content"):
                severity = "warning"
            elif item.get("x_robots_forbidden") or item.get("meta_forbidden"):
                severity = "info"
            else:
                severity = "ok"

            values = [
                item.get("bot_name", ""),
                item.get("category", ""),
                item.get("status", ""),
                "Да" if item.get("accessible") else "Нет",
                "Да" if item.get("has_content") else "Нет",
                "Да" if item.get("robots_allowed") is True else ("Нет" if item.get("robots_allowed") is False else "н/д"),
                item.get("x_robots_tag", ""),
                "Да" if item.get("x_robots_forbidden") else "Нет",
                item.get("meta_robots", ""),
                "Да" if item.get("meta_forbidden") else "Нет",
                item.get("response_time_ms", ""),
                item.get("final_url", ""),
                item.get("error", ""),
                severity.capitalize(),
            ]
            for col, value in enumerate(values, 1):
                self._apply_style(results_ws.cell(row=row_idx, column=col, value=value), cell_style)
            self._apply_row_severity_fill(results_ws, row_idx, 1, len(result_headers), severity)
            self._apply_severity_cell_style(results_ws.cell(row=row_idx, column=len(result_headers)), severity)

        results_ws.freeze_panes = "A2"
        results_ws.auto_filter.ref = f"A1:{get_column_letter(len(result_headers))}1"
        for col, width in enumerate([26, 16, 10, 10, 12, 14, 28, 16, 28, 14, 14, 40, 38, 12], 1):
            results_ws.column_dimensions[get_column_letter(col)].width = width

        categories_ws = wb.create_sheet("Категории")
        category_headers = ["Категория", "Всего", "Доступно", "С контентом", "Ограничивающие директивы", "Критичность"]
        for col, header in enumerate(category_headers, 1):
            self._apply_style(categories_ws.cell(row=1, column=col, value=header), header_style)
        for row_idx, item in enumerate(category_stats, start=2):
            total = item.get("total", 0) or 0
            accessible = item.get("accessible", 0) or 0
            ratio = (accessible / total) if total else 0
            severity = "ok" if ratio >= 0.9 else ("warning" if ratio >= 0.6 else "critical")
            values = [
                item.get("category", ""),
                total,
                accessible,
                item.get("with_content", 0),
                item.get("restrictive_directives", 0),
                severity.capitalize(),
            ]
            for col, value in enumerate(values, 1):
                self._apply_style(categories_ws.cell(row=row_idx, column=col, value=value), cell_style)
            self._apply_row_severity_fill(categories_ws, row_idx, 1, len(category_headers), severity)
            self._apply_severity_cell_style(categories_ws.cell(row=row_idx, column=len(category_headers)), severity)
        categories_ws.freeze_panes = "A2"
        categories_ws.auto_filter.ref = "A1:F1"
        for col, width in enumerate([24, 10, 12, 14, 22, 12], 1):
            categories_ws.column_dimensions[get_column_letter(col)].width = width

        matrix_ws = wb.create_sheet("Матрица")
        matrix_headers = ["Категория", "Всего", "Сканируемо", "Рендерится", "Индексируемо", "Не индексируемо", "Индексируемо %", "SLA цель %", "SLA выполнен", "Оценка риска"]
        for col, header in enumerate(matrix_headers, 1):
            self._apply_style(matrix_ws.cell(row=1, column=col, value=header), header_style)
        for row_idx, item in enumerate(category_stats, start=2):
            values = [
                item.get("category", ""),
                item.get("total", 0),
                item.get("crawlable", 0),
                item.get("renderable", 0),
                item.get("indexable", 0),
                item.get("non_indexable", 0),
                item.get("indexable_pct", 0),
                item.get("sla_target_pct", 0),
                "Да" if item.get("sla_met") else "Нет",
                item.get("priority_risk_score", 0),
            ]
            for col, value in enumerate(values, 1):
                self._apply_style(matrix_ws.cell(row=row_idx, column=col, value=value), cell_style)
        matrix_ws.freeze_panes = "A2"
        matrix_ws.auto_filter.ref = "A1:J1"
        for col, width in enumerate([20, 10, 12, 12, 12, 14, 12, 12, 10, 12], 1):
            matrix_ws.column_dimensions[get_column_letter(col)].width = width

        blockers_ws = wb.create_sheet("Приоритетные блокеры")
        blocker_headers = ["Код", "Заголовок", "Затронуто ботов", "Взвешенное влияние", "Оценка приоритета", "Примеры ботов", "Детали"]
        for col, header in enumerate(blocker_headers, 1):
            self._apply_style(blockers_ws.cell(row=1, column=col, value=header), header_style)
        for row_idx, item in enumerate(priority_blockers, start=2):
            values = [
                item.get("code", ""),
                item.get("title", ""),
                item.get("affected_bots", 0),
                item.get("weighted_impact", 0),
                item.get("priority_score", 0),
                ", ".join(item.get("sample_bots", []) or []),
                item.get("details", ""),
            ]
            for col, value in enumerate(values, 1):
                self._apply_style(blockers_ws.cell(row=row_idx, column=col, value=value), cell_style)
        blockers_ws.freeze_panes = "A2"
        blockers_ws.auto_filter.ref = "A1:G1"
        for col, width in enumerate([20, 30, 14, 14, 14, 48, 62], 1):
            blockers_ws.column_dimensions[get_column_letter(col)].width = width

        playbooks_ws = wb.create_sheet("Плейбуки")
        playbook_headers = ["Код блокера", "Владелец", "Заголовок", "Оценка приоритета", "Действия"]
        for col, header in enumerate(playbook_headers, 1):
            self._apply_style(playbooks_ws.cell(row=1, column=col, value=header), header_style)
        for row_idx, item in enumerate(playbooks, start=2):
            values = [
                item.get("blocker_code", ""),
                item.get("owner", ""),
                item.get("title", ""),
                item.get("priority_score", 0),
                " | ".join(item.get("actions", []) or []),
            ]
            for col, value in enumerate(values, 1):
                self._apply_style(playbooks_ws.cell(row=row_idx, column=col, value=value), cell_style)
        playbooks_ws.freeze_panes = "A2"
        playbooks_ws.auto_filter.ref = "A1:E1"
        for col, width in enumerate([20, 16, 32, 14, 120], 1):
            playbooks_ws.column_dimensions[get_column_letter(col)].width = width

        raw_ws = wb.create_sheet("Сырые строки ботов")
        raw_headers = [
            "Бот", "Категория", "HTTP", "Сканируемо", "Рендерится", "Индексируемо", "Разрешен robots",
            "Совпавший UA", "Совпавшее правило", "Совпавший шаблон", "WAF обнаружен", "Провайдер WAF", "Причина WAF",
            "Время ответа мс", "Финальный URL", "Причины блокировки", "Ошибка",
        ]
        for col, header in enumerate(raw_headers, 1):
            self._apply_style(raw_ws.cell(row=1, column=col, value=header), header_style)
        for row_idx, item in enumerate(bot_rows, start=2):
            robots_eval = item.get("robots_evaluation", {}) or {}
            waf = item.get("waf_cdn_signal", {}) or {}
            values = [
                item.get("bot_name", ""),
                item.get("category", ""),
                item.get("status", ""),
                "Да" if item.get("crawlable") else "Нет",
                "Да" if item.get("renderable") else "Нет",
                "Да" if item.get("indexable") else "Нет",
                "Да" if item.get("robots_allowed") is True else ("Нет" if item.get("robots_allowed") is False else "н/д"),
                robots_eval.get("matched_user_agent", ""),
                robots_eval.get("matched_rule", ""),
                robots_eval.get("matched_pattern", ""),
                "Да" if waf.get("detected") else "Нет",
                waf.get("provider", ""),
                waf.get("reason", ""),
                item.get("response_time_ms", ""),
                item.get("final_url", ""),
                ", ".join(item.get("blocked_reasons", []) or []),
                item.get("error", ""),
            ]
            for col, value in enumerate(values, 1):
                self._apply_style(raw_ws.cell(row=row_idx, column=col, value=value), cell_style)
        raw_ws.freeze_panes = "A2"
        raw_ws.auto_filter.ref = "A1:Q1"
        for col, width in enumerate([24, 14, 10, 10, 10, 10, 12, 16, 12, 24, 12, 14, 28, 14, 42, 34, 46], 1):
            raw_ws.column_dimensions[get_column_letter(col)].width = width

        diff_ws = wb.create_sheet("Сравнение с baseline")
        diff_headers = ["Метрика", "Текущее", "Базовое", "Дельта"]
        for col, header in enumerate(diff_headers, 1):
            self._apply_style(diff_ws.cell(row=1, column=col, value=header), header_style)
        diff_rows = baseline_diff.get("metrics", []) or []
        if diff_rows:
            for row_idx, item in enumerate(diff_rows, start=2):
                values = [item.get("metric", ""), item.get("current", ""), item.get("baseline", ""), item.get("delta", "")]
                for col, value in enumerate(values, 1):
                    self._apply_style(diff_ws.cell(row=row_idx, column=col, value=value), cell_style)
        else:
            self._apply_style(diff_ws.cell(row=2, column=1, value=baseline_diff.get("message", "Нет данных baseline")), cell_style)
        diff_ws.column_dimensions["A"].width = 28
        diff_ws.column_dimensions["B"].width = 14
        diff_ws.column_dimensions["C"].width = 14
        diff_ws.column_dimensions["D"].width = 14

        trend_ws = wb.create_sheet("История тренда")
        trend_headers = [
            "Время запуска",
            "Индексируемо",
            "Сканируемо",
            "Рендерится",
            "Доступно",
            "Средний ответ, мс",
            "Критично",
            "Предупреждения",
            "WAF/CDN",
            "Retry profile",
            "Критичность",
            "Профиль SLA",
        ]
        for col, header in enumerate(trend_headers, 1):
            self._apply_style(trend_ws.cell(row=1, column=col, value=header), header_style)
        trend_rows = trend.get("history", []) or []
        if trend_rows:
            for row_idx, item in enumerate(trend_rows, start=2):
                values = [
                    item.get("timestamp", ""),
                    item.get("indexable", 0),
                    item.get("crawlable", 0),
                    item.get("renderable", 0),
                    item.get("accessible", 0),
                    item.get("avg_response_time_ms", 0),
                    item.get("critical_issues", 0),
                    item.get("warning_issues", 0),
                    item.get("waf_cdn_detected", 0),
                    item.get("retry_profile", ""),
                    item.get("criticality_profile", ""),
                    item.get("sla_profile", ""),
                ]
                for col, value in enumerate(values, 1):
                    self._apply_style(trend_ws.cell(row=row_idx, column=col, value=value), cell_style)
        else:
            self._apply_style(trend_ws.cell(row=2, column=1, value="История тренда недоступна"), cell_style)
        trend_ws.freeze_panes = "A2"
        trend_ws.auto_filter.ref = "A1:L1"
        for col, width in enumerate([24, 10, 10, 10, 10, 14, 10, 10, 10, 14, 14, 12], 1):
            trend_ws.column_dimensions[get_column_letter(col)].width = width

        alerts_ws = wb.create_sheet("Алерты")
        for col, header in enumerate(["Критичность", "Код", "Сообщение"], 1):
            self._apply_style(alerts_ws.cell(row=1, column=col, value=header), header_style)
        if alerts:
            for row_idx, a in enumerate(alerts, start=2):
                sev = str(a.get("severity", "info")).lower()
                values = [sev.upper(), a.get("code", ""), a.get("message", "")]
                for col, value in enumerate(values, 1):
                    self._apply_style(alerts_ws.cell(row=row_idx, column=col, value=value), cell_style)
                self._apply_row_severity_fill(alerts_ws, row_idx, 1, 3, sev)
                self._apply_severity_cell_style(alerts_ws.cell(row=row_idx, column=1), sev)
        else:
            self._apply_style(alerts_ws.cell(row=2, column=1, value="Нет алертов"), cell_style)
        alerts_ws.freeze_panes = "A2"
        alerts_ws.auto_filter.ref = "A1:C1"
        for col, width in enumerate([12, 24, 110], 1):
            alerts_ws.column_dimensions[get_column_letter(col)].width = width

        linter_ws = wb.create_sheet("Линтер Robots")
        for col, header in enumerate(["Критичность", "Код", "Сообщение"], 1):
            self._apply_style(linter_ws.cell(row=1, column=col, value=header), header_style)
        if robots_linter:
            for row_idx, a in enumerate(robots_linter, start=2):
                sev = str(a.get("severity", "info")).lower()
                values = [sev.upper(), a.get("code", ""), a.get("message", "")]
                for col, value in enumerate(values, 1):
                    self._apply_style(linter_ws.cell(row=row_idx, column=col, value=value), cell_style)
                self._apply_row_severity_fill(linter_ws, row_idx, 1, 3, sev)
                self._apply_severity_cell_style(linter_ws.cell(row=row_idx, column=1), sev)
        else:
            self._apply_style(linter_ws.cell(row=2, column=1, value="Линтер не выявил замечаний"), cell_style)
        linter_ws.freeze_panes = "A2"
        linter_ws.auto_filter.ref = "A1:C1"
        for col, width in enumerate([12, 24, 110], 1):
            linter_ws.column_dimensions[get_column_letter(col)].width = width

        sim_ws = wb.create_sheet("Симулятор allowlist")
        sim_headers = ["Сценарий", "Категория", "Затронуто", "Дельта рендеринга", "Дельта индексации", "Прогноз индексируемых", "Прогноз %"]
        for col, header in enumerate(sim_headers, 1):
            self._apply_style(sim_ws.cell(row=1, column=col, value=header), header_style)
        if allowlist_sim:
            for row_idx, s in enumerate(allowlist_sim, start=2):
                values = [
                    s.get("scenario", ""),
                    s.get("category", ""),
                    s.get("affected_bots", 0),
                    s.get("delta_renderable", 0),
                    s.get("delta_indexable", 0),
                    s.get("projected_indexable", 0),
                    s.get("projected_indexable_pct", 0),
                ]
                for col, value in enumerate(values, 1):
                    self._apply_style(sim_ws.cell(row=row_idx, column=col, value=value), cell_style)
        else:
            self._apply_style(sim_ws.cell(row=2, column=1, value="Нет данных симуляции"), cell_style)
        sim_ws.freeze_panes = "A2"
        sim_ws.auto_filter.ref = "A1:G1"
        for col, width in enumerate([22, 14, 10, 14, 14, 16, 12], 1):
            sim_ws.column_dimensions[get_column_letter(col)].width = width

        action_ws = wb.create_sheet("Центр действий")
        action_headers = ["Владелец", "Заголовок", "Приоритет", "Код блокера", "Действия"]
        for col, header in enumerate(action_headers, 1):
            self._apply_style(action_ws.cell(row=1, column=col, value=header), header_style)
        row_idx = 2
        for owner, rows in action_center.items():
            for item in (rows or []):
                values = [owner, item.get("title", ""), item.get("priority_score", 0), item.get("blocker_code", ""), " | ".join(item.get("actions", []) or [])]
                for col, value in enumerate(values, 1):
                    self._apply_style(action_ws.cell(row=row_idx, column=col, value=value), cell_style)
                row_idx += 1
        if row_idx == 2:
            self._apply_style(action_ws.cell(row=2, column=1, value="Нет действий по владельцам"), cell_style)
        action_ws.freeze_panes = "A2"
        action_ws.auto_filter.ref = "A1:E1"
        for col, width in enumerate([20, 40, 10, 16, 110], 1):
            action_ws.column_dimensions[get_column_letter(col)].width = width

        evidence_ws = wb.create_sheet("Пакет доказательств")
        evidence_headers = ["Бот", "Категория", "HTTP", "Причина индексируемости", "WAF обнаружен", "Уверенность WAF", "Причина WAF", "Пояснение robots", "Пример ответа"]
        for col, header in enumerate(evidence_headers, 1):
            self._apply_style(evidence_ws.cell(row=1, column=col, value=header), header_style)
        if evidence_pack:
            for row_idx, e in enumerate(evidence_pack, start=2):
                values = [
                    e.get("bot", ""),
                    e.get("category", ""),
                    e.get("status", ""),
                    e.get("indexability_reason", ""),
                    "Да" if e.get("waf_detected") else "Нет",
                    e.get("waf_confidence", 0),
                    e.get("waf_reason", ""),
                    e.get("robots_explain", ""),
                    e.get("response_sample", ""),
                ]
                for col, value in enumerate(values, 1):
                    self._apply_style(evidence_ws.cell(row=row_idx, column=col, value=value), cell_style)
        else:
            self._apply_style(evidence_ws.cell(row=2, column=1, value="Нет строк доказательств"), cell_style)
        evidence_ws.freeze_panes = "A2"
        evidence_ws.auto_filter.ref = "A1:I1"
        for col, width in enumerate([22, 14, 10, 34, 12, 12, 24, 40, 70], 1):
            evidence_ws.column_dimensions[get_column_letter(col)].width = width

        if batch_runs:
            batch_ws = wb.create_sheet("Пакетные прогоны")
            batch_headers = ["URL", "Индексируемо", "Всего", "Сканируемо", "Рендерится", "Критично", "Предупреждения", "Средний ответ, мс"]
            for col, header in enumerate(batch_headers, 1):
                self._apply_style(batch_ws.cell(row=1, column=col, value=header), header_style)
            for row_idx, b in enumerate(batch_runs, start=2):
                values = [
                    b.get("url", ""),
                    b.get("indexable", 0),
                    b.get("total", 0),
                    b.get("crawlable", 0),
                    b.get("renderable", 0),
                    b.get("critical_issues", 0),
                    b.get("warning_issues", 0),
                    b.get("avg_response_time_ms", 0),
                ]
                for col, value in enumerate(values, 1):
                    self._apply_style(batch_ws.cell(row=row_idx, column=col, value=value), cell_style)
            batch_ws.freeze_panes = "A2"
            batch_ws.auto_filter.ref = "A1:H1"
            for col, width in enumerate([70, 10, 8, 10, 10, 10, 10, 14], 1):
                batch_ws.column_dimensions[get_column_letter(col)].width = width

        issues_ws = wb.create_sheet("Проблемы")
        issue_headers = ["Критичность", "Бот", "Категория", "Заголовок", "Детали"]
        for col, header in enumerate(issue_headers, 1):
            self._apply_style(issues_ws.cell(row=1, column=col, value=header), header_style)
        for row_idx, item in enumerate(issues, start=2):
            severity = (item.get("severity") or "info").lower()
            values = [
                severity.capitalize(),
                item.get("bot", ""),
                item.get("category", ""),
                item.get("title", ""),
                item.get("details", ""),
            ]
            for col, value in enumerate(values, 1):
                self._apply_style(issues_ws.cell(row=row_idx, column=col, value=value), cell_style)
            self._apply_row_severity_fill(issues_ws, row_idx, 1, len(issue_headers), severity)
            self._apply_severity_cell_style(issues_ws.cell(row=row_idx, column=1), severity)
        issues_ws.freeze_panes = "A2"
        issues_ws.auto_filter.ref = "A1:E1"
        for col, width in enumerate([12, 24, 16, 28, 80], 1):
            issues_ws.column_dimensions[get_column_letter(col)].width = width

        rec_ws = wb.create_sheet("Рекомендации")
        self._apply_style(rec_ws.cell(row=1, column=1, value="Рекомендация"), header_style)
        for idx, text in enumerate(recommendations, start=2):
            cell = rec_ws.cell(row=idx, column=1, value=text)
            self._apply_style(cell, cell_style)
        rec_ws.column_dimensions["A"].width = 160
        rec_ws.freeze_panes = "A2"
        rec_ws.auto_filter.ref = "A1:A1"

        filepath = os.path.join(self.reports_dir, f"{task_id}.xlsx")
        self._save_workbook(wb, filepath)
        return filepath

    def generate_site_audit_pro_report(self, task_id: str, data: Dict[str, Any]) -> str:
        """Generate 8-sheet Site Audit Pro XLSX with deduplicated metric ownership."""
        wb = Workbook()
        header_style = self._create_header_style()
        cell_style = self._create_cell_style()

        results = data.get("results", {}) or {}
        summary = results.get("summary", {}) or {}
        pages = results.get("pages", []) or []
        issues = results.get("issues", []) or []
        pipeline = results.get("pipeline", {}) or {}
        pipeline_metrics = pipeline.get("metrics", {}) or {}
        report_url = data.get("url", "н/д")
        mode = results.get("mode", "quick")

        tfidf_by_url = {row.get("url"): row.get("top_terms", []) for row in (pipeline.get("tf_idf") or [])}
        issues_by_url: Dict[str, List[Dict[str, Any]]] = {}
        for issue in issues:
            issues_by_url.setdefault(issue.get("url", ""), []).append(issue)

        def bool_icon(value: Any, positive: str = "YES", negative: str = "NO") -> str:
            return positive if bool(value) else negative

        def icon_score(value: Any, low: float = 60.0, high: float = 80.0) -> str:
            v = to_float(value, 0.0)
            if v >= high:
                return f"GOOD {v:.0f}"
            if v >= low:
                return f"WARN {v:.0f}"
            return f"BAD {v:.0f}"

        def icon_count(count: int, low: float = 20.0, high: float = 50.0) -> str:
            # Smaller count is better; use inverted score scale.
            score = 100.0 - min(100.0, float(count) * 5.0)
            if score >= high:
                return f"GOOD {count}"
            if score >= low:
                return f"WARN {count}"
            return f"BAD {count}"

        def as_percent(value: Any) -> str:
            try:
                return f"{float(value):.1f}%"
            except Exception:
                return "0.0%"

        def to_float(value: Any, default: float = 0.0) -> float:
            try:
                return float(value)
            except Exception:
                return default

        def ok_if_empty(recs: List[str]) -> str:
            return "OK" if not recs else "; ".join(recs[:2])

        def page_solution(tab: str, page: Dict[str, Any], page_issues: List[Dict[str, Any]] | None = None) -> str:
            recs: List[str] = []
            page_issues = page_issues or []
            if tab == "main":
                if not page.get("indexable"):
                    recs.append("Исправить noindex/status/robots")
                if to_float(page.get("health_score"), 100.0) < 60:
                    recs.append("Повысить техническое и контентное качество до 60+")
                if not page_issues and not recs:
                    return "OK"
                if recs:
                    return ok_if_empty(recs)
                issue_titles = [str(i.get("title") or i.get("code") or "") for i in page_issues if (i.get("title") or i.get("code"))]
                return ok_if_empty(issue_titles)

            if tab == "hierarchy":
                status = str(page.get("h_hierarchy") or "").lower()
                if "wrong start" in status:
                    return "Добавить H1 в начале основного контента"
                if "level skip" in status:
                    return "Исправить пропуски уровней заголовков (H1->H2->H3)"
                if "multiple h1" in status:
                    return "Оставить один H1, остальные перевести в H2"
                if "missing h1" in status:
                    return "Добавить один описательный H1"
                return "OK"

            if tab == "onpage":
                title_len = int(page.get("title_len") or len(str(page.get("title") or "")))
                desc_len = int(page.get("description_len") or len(str(page.get("meta_description") or "")))
                if title_len < 30 or title_len > 60:
                    recs.append("Держите длину title в диапазоне 30-60 символов и обеспечьте уникальность")
                if desc_len < 50 or desc_len > 160:
                    recs.append("Держите meta description в диапазоне 100-160 символов")
                if int(page.get("h1_count") or 0) != 1:
                    recs.append("Оставьте ровно один H1")
                if (page.get("canonical_status") or "") in ("missing", "external", "invalid"):
                    recs.append("Укажите валидный canonical URL")
                if int(page.get("duplicate_title_count") or 0) > 1:
                    recs.append("Уберите дубли title")
                if int(page.get("duplicate_description_count") or 0) > 1:
                    recs.append("Уберите дубли meta description")
                if not bool(page.get("charset_declared")):
                    recs.append("Добавьте <meta charset=\"utf-8\">")
                if not bool(page.get("viewport_declared")):
                    recs.append("Добавьте meta viewport для мобильных устройств")
                if bool(page.get("multiple_meta_robots")):
                    recs.append("Оставьте только один meta robots тег")
                return ok_if_empty(recs)

            if tab == "content":
                unique_percent = to_float(page.get("unique_percent"), 0.0)
                words = int(page.get("word_count") or 0)
                if unique_percent < 30:
                    recs.append("Добавьте уникальные блоки контента (уникальность <30%)")
                elif unique_percent < 50:
                    recs.append("Повышайте уникальность выше 50%")
                if words < 300:
                    recs.append("Увеличьте объем контента до 300+ слов")
                if to_float(page.get("toxicity_score"), 0.0) > 40:
                    recs.append("Снизьте переспам и злоупотребление ключевыми словами")
                if int(page.get("near_duplicate_count") or 0) > 0:
                    recs.append("Перепишите близкие дубли и усилите соответствие интенту страницы")
                if bool(page.get("hidden_content")):
                    recs.append("Удалите скрытые блоки контента (display:none/offscreen/small font)")
                if bool(page.get("cloaking_detected")):
                    recs.append("Устраните расхождения скрытого/видимого контента (признаки cloaking)")
                if int(page.get("cta_count") or 0) <= 0 and str(page.get("page_type") or "") in {"home", "service", "product", "category"}:
                    recs.append("Добавьте конверсионные CTA-блоки (форма/кнопка/обратный звонок)")
                if words >= 600 and int(page.get("lists_count") or 0) == 0 and int(page.get("tables_count") or 0) == 0:
                    recs.append("Структурируйте длинный контент списками и таблицами")
                return ok_if_empty(recs)

            if tab == "technical":
                if not page.get("indexable"):
                    recs.append("Сделайте страницу индексируемой")
                if not page.get("is_https"):
                    recs.append("Включите HTTPS")
                if not page.get("compression_enabled"):
                    recs.append("Включите gzip или brotli")
                if not page.get("cache_enabled"):
                    recs.append("Настройте Cache-Control")
                if (page.get("canonical_status") or "") in ("missing", "external", "invalid"):
                    recs.append("Проверьте и исправьте canonical")
                if len(page.get("deprecated_tags") or []) > 0:
                    recs.append("Удалите устаревшие HTML-теги")
                rt = page.get("response_time_ms")
                if rt is not None and int(rt) > 2000:
                    recs.append("Сократите время ответа сервера")
                if to_float(page.get("perf_light_score"), 100.0) < 60:
                    recs.append("Снизьте render-blocking JS и размер DOM/HTML")
                return ok_if_empty(recs)

            if tab == "eeat":
                eeat = to_float(page.get("eeat_score"), 0.0)
                if eeat < 50:
                    recs.append("Добавьте профиль автора, био, источники и кейсы")
                elif eeat < 70:
                    recs.append("Усильте сигналы экспертности и доверия")
                if not page.get("has_author_info"):
                    recs.append("Добавьте блок автора с контактами и соцсетями")
                if not page.get("has_reviews"):
                    recs.append("Добавьте отзывы и кейсы")
                return ok_if_empty(recs)

            if tab == "trust":
                if not page.get("has_contact_info"):
                    recs.append("Добавьте контакты, адрес и телефон")
                if not page.get("has_legal_docs"):
                    recs.append("Добавить юридические страницы и политики")
                if not page.get("has_reviews"):
                    recs.append("Добавить отзывы")
                if not page.get("trust_badges"):
                    recs.append("Добавить бейджи доверия/сертификаты")
                return ok_if_empty(recs)

            if tab == "health":
                if to_float(page.get("health_score"), 100.0) < 60:
                    recs.append("Повысить общий health score до 60+")
                if not page.get("indexable"):
                    recs.append("Исправить индексируемость")
                if int(page.get("duplicate_title_count") or 0) > 1:
                    recs.append("Убрать дубли title")
                if int(page.get("duplicate_description_count") or 0) > 1:
                    recs.append("Убрать дубли meta description")
                return ok_if_empty(recs)

            if tab == "links":
                if page.get("orphan_page"):
                    recs.append("Добавить внутренние ссылки на эту страницу")
                if int(page.get("outgoing_internal_links") or 0) == 0:
                    recs.append("Добавьте исходящие внутренние ссылки на релевантные страницы")
                return ok_if_empty(recs)

            if tab == "images":
                img_opt = page.get("images_optimization") or {}
                if int(img_opt.get("no_alt") or page.get("images_without_alt") or 0) > 0:
                    recs.append("Добавьте ALT-тексты для изображений")
                if int(img_opt.get("no_width_height") or 0) > 0:
                    recs.append("Добавьте атрибуты width/height")
                if int(img_opt.get("no_lazy_load") or 0) > 0:
                    recs.append("Включите lazy loading")
                if int(page.get("images_modern_format_count") or 0) <= 0 and int(page.get("images_count") or 0) > 0:
                    recs.append("Используйте WebP/AVIF для ключевых изображений")
                if int(page.get("generic_alt_count") or 0) > 0:
                    recs.append("Замените общие ALT-тексты на описательные")
                if int(page.get("decorative_non_empty_alt_count") or 0) > 0:
                    recs.append("Используйте пустой ALT для декоративных изображений")
                return ok_if_empty(recs)

            if tab == "external":
                total = int(page.get("outgoing_external_links") or 0)
                if total == 0:
                    recs.append("Добавьте релевантные внешние источники")
                return ok_if_empty(recs)

            if tab == "structured":
                if int(page.get("structured_data") or 0) == 0:
                    recs.append("Внедрите schema.org (JSON-LD)")
                if int(page.get("hreflang_count") or 0) == 0:
                    recs.append("Добавьте hreflang для языковых версий")
                return ok_if_empty(recs)

            if tab == "keywords":
                if not (page.get("top_keywords") or page.get("top_terms") or page.get("tf_idf_keywords")):
                    recs.append("Уточните семантическое ядро и целевые ключи")
                if to_float(page.get("toxicity_score"), 0.0) > 40:
                    recs.append("Снизьте переоптимизацию ключевых слов")
                return ok_if_empty(recs)

            if tab == "topics":
                if not page.get("topic_hub"):
                    recs.append("Свяжите страницу с релевантными hub-страницами")
                if not page.get("topic_label"):
                    recs.append("Определите четкий тематический кластер")
                return ok_if_empty(recs)

            if tab == "advanced":
                freshness = page.get("content_freshness_days")
                if freshness is not None and int(freshness) > 365:
                    recs.append("Обновите устаревший контент")
                if bool(page.get("hidden_content")):
                    recs.append("Удалите скрытый контент")
                if bool(page.get("cloaking_detected")):
                    recs.append("Устраните признаки cloaking")
                return ok_if_empty(recs)

            if tab == "link_quality":
                if to_float(page.get("link_quality_score"), 0.0) < 60:
                    recs.append("Улучшите внутреннюю перелинковку")
                if page.get("orphan_page"):
                    recs.append("Добавьте входящие внутренние ссылки")
                return ok_if_empty(recs)

            return "ОК"

        def hierarchy_problem(page: Dict[str, Any]) -> str:
            errors = page.get("h_errors") or []
            if errors:
                mapping = {
                    "wrong_start": "Иерархия начинается с H2+ вместо H1",
                    "missing_h1": "Отсутствует H1",
                    "multiple_h1": "Несколько H1",
                    "heading_level_skip": "Обнаружен пропуск уровня заголовков",
                }
                return "; ".join(mapping.get(str(e), str(e)) for e in errors)
            return "No hierarchy issues"

        def ai_found_list(page: Dict[str, Any]) -> str:
            markers = page.get("ai_markers_list") or []
            return ", ".join(markers[:10]) if markers else "No AI markers detected"

        def infer_page_severity(page: Dict[str, Any]) -> str:
            severity = "ok"
            for issue in (page.get("issues") or []):
                sev = (issue.get("severity") or "info").lower()
                if sev == "critical":
                    return "critical"
                if sev == "warning":
                    severity = "warning"
                elif sev == "info" and severity == "ok":
                    severity = "info"
            return severity

        def derive_page_metrics(page: Dict[str, Any]) -> Dict[str, Any]:
            # Local import keeps entropy math robust even if runtime loads stale globals.
            import math as _math
            title = str(page.get("title") or "")
            meta_description = str(page.get("meta_description") or "")
            title_len = int(page.get("title_len") or len(title))
            description_len = int(page.get("description_len") or len(meta_description))

            status_code = int(page.get("status_code") or 0)
            response_ms = int(page.get("response_time_ms") or 0)
            html_bytes = int(page.get("html_size_bytes") or 0)
            content_kb = round(to_float(page.get("content_kb"), html_bytes / 1024.0), 1)

            sdetail = page.get("structured_data_detail") or {}
            img_opt = page.get("images_optimization") or {}
            ext_follow = int(page.get("external_follow_links", 0) or 0)
            ext_nofollow = int(page.get("external_nofollow_links", 0) or 0)
            ext_total = int(page.get("outgoing_external_links", ext_follow + ext_nofollow) or 0)
            follow_ratio = round((ext_follow / ext_total * 100.0), 1) if ext_total > 0 else 0.0

            no_alt = int(page.get("images_without_alt", img_opt.get("no_alt", 0)) or 0)
            no_wh = int(img_opt.get("no_width_height", 0) or 0)
            no_lazy = int(img_opt.get("no_lazy_load", 0) or 0)
            image_issues_total = no_alt + no_wh + no_lazy

            ai_markers_count = int(page.get("ai_markers_count", 0) or 0)
            toxicity = to_float(page.get("toxicity_score"), 0.0)
            filler_ratio = to_float(page.get("filler_ratio"), 0.0)
            ai_density = to_float(page.get("ai_markers_density_1k"), 0.0)
            ai_risk = to_float(page.get("ai_risk_score"), round(min(100.0, ai_markers_count * 6.0 + toxicity * 0.7 + filler_ratio * 0.5), 1))
            ai_risk_level = str(page.get("ai_risk_level") or ("high" if ai_risk >= 70 else "medium" if ai_risk >= 40 else "low"))

            onpage_score = 100
            if title_len < 30 or title_len > 60:
                onpage_score -= 15
            if description_len < 50 or description_len > 160:
                onpage_score -= 10
            if int(page.get("h1_count") or 0) != 1:
                onpage_score -= 20
            if (page.get("canonical_status") or "") in ("missing", "external", "invalid"):
                onpage_score -= 20
            onpage_score = max(0, onpage_score)

            technical_score = 100
            if status_code >= 400:
                technical_score -= 30
            if response_ms > 2000:
                technical_score -= 15
            if not bool(page.get("is_https")):
                technical_score -= 15
            if not bool(page.get("compression_enabled")):
                technical_score -= 10
            if not bool(page.get("cache_enabled")):
                technical_score -= 10
            if to_float(page.get("perf_light_score"), 100.0) < 60:
                technical_score -= 10
            technical_score = max(0, technical_score)

            content_score = 100
            if int(page.get("word_count") or 0) < 300:
                content_score -= 25
            if to_float(page.get("unique_percent"), 0.0) < 50:
                content_score -= 20
            if toxicity > 40:
                content_score -= 20
            if ai_risk >= 70:
                content_score -= 20
            content_score = max(0, content_score)

            link_score = int(round(max(0.0, min(100.0, to_float(page.get("link_quality_score"), 0.0)))))
            media_score = max(0, 100 - min(80, image_issues_total * 10))
            if int(page.get("images_count") or 0) > 0 and int(page.get("images_modern_format_count") or 0) == 0:
                media_score = max(0, media_score - 10)
            if int(page.get("generic_alt_count") or 0) > 0:
                media_score = max(0, media_score - 8)
            if int(page.get("decorative_non_empty_alt_count") or 0) > 0:
                media_score = max(0, media_score - 6)
            hierarchy_score = 100 if not (page.get("h_errors") or []) else max(0, 100 - len(page.get("h_errors") or []) * 25)
            keyword_stuffing = to_float(page.get("keyword_stuffing_score"), 0.0)
            lexical_div = to_float(page.get("lexical_diversity"), 0.0)
            tfidf_map = page.get("tf_idf_keywords") or {}
            tfidf_terms_count = len(tfidf_map)
            density_map = page.get("keyword_density_profile") or {}
            density_values = [to_float(v, 0.0) for v in density_map.values()]
            density_sum = max(0.0001, sum(density_values))
            keyword_entropy = 0.0
            for v in density_values:
                p = v / density_sum if density_sum > 0 else 0.0
                if p > 0:
                    try:
                        keyword_entropy -= p * _math.log(p, 2)
                    except Exception:
                        # Safe fallback: skip malformed contribution instead of failing XLSX export.
                        continue
            keyword_entropy = round(keyword_entropy, 3)
            keyword_score = 100.0
            keyword_score -= min(55.0, keyword_stuffing)
            if lexical_div < 0.35:
                keyword_score -= 15.0
            elif lexical_div < 0.45:
                keyword_score -= 8.0
            if tfidf_terms_count < 5:
                keyword_score -= 10.0
            if len(page.get("top_keywords") or []) < 3:
                keyword_score -= 5.0
            if int(page.get("word_count") or 0) < 200:
                keyword_score -= 8.0
            if toxicity > 40:
                keyword_score -= 8.0
            keyword_score = max(0, int(round(keyword_score)))

            return {
                "title": title,
                "meta_description": meta_description,
                "title_len": title_len,
                "description_len": description_len,
                "status_code": status_code,
                "response_ms": response_ms,
                "html_bytes": html_bytes,
                "content_kb": content_kb,
                "structured_detail": sdetail,
                "ext_follow": ext_follow,
                "ext_nofollow": ext_nofollow,
                "ext_total": ext_total,
                "follow_ratio": follow_ratio,
                "no_alt": no_alt,
                "no_wh": no_wh,
                "no_lazy": no_lazy,
                "image_issues_total": image_issues_total,
                "ai_risk": ai_risk,
                "ai_risk_level": ai_risk_level,
                "ai_density": ai_density,
                "onpage_score": onpage_score,
                "technical_score": technical_score,
                "content_score": content_score,
                "link_score": link_score,
                "media_score": media_score,
                "hierarchy_score": hierarchy_score,
                "keyword_score": keyword_score,
                "tfidf_terms_count": tfidf_terms_count,
                "keyword_entropy": keyword_entropy,
            }

        derived_by_url = {str(page.get("url", "")): derive_page_metrics(page) for page in pages}
        page_by_url = {str(page.get("url", "")): page for page in pages}
        sheet_stats: List[Dict[str, Any]] = []

        def issue_recommendation(issue: Dict[str, Any]) -> str:
            code = str(issue.get("code") or "").lower()
            title = str(issue.get("title") or "").lower()
            details = str(issue.get("details") or "")
            severity = str(issue.get("severity") or "info").lower()
            critical_codes = {
                "http_status_error",
                "non_https_url",
                "canonical_target_error_status",
                "canonical_target_redirect",
                "noindex_canonical_conflict",
                "security_mixed_content_homepage",
                "ai_risk_high",
            }
            warning_codes = {
                "missing_title",
                "missing_meta_description",
                "missing_canonical",
                "compression_disabled",
                "cache_disabled",
                "h1_hierarchy_issue",
                "structured_data_common_errors",
                "multiple_title_tags",
                "multiple_meta_descriptions",
                "multiple_meta_robots",
                "missing_charset_meta",
                "missing_viewport_meta",
                "generic_alt_texts",
                "decorative_images_with_alt",
                "duplicate_image_sources",
                "low_modern_image_formats",
                "crawl_budget_risk_high",
                "crawl_budget_risk_medium",
                "hreflang_extended_check",
            }
            mapping = [
                ("duplicate_title", "Сделать уникальный title для каждой дублирующейся страницы."),
                ("duplicate_meta_description", "Сделать уникальный meta description для каждой дублирующейся страницы."),
                ("missing_title", "Добавить описательный title (30-60 символов)."),
                ("missing_meta_description", "Добавить meta description (100-160 символов)."),
                ("missing_canonical", "Добавить canonical URL с указанием на финальную страницу со статусом 200."),
                ("canonical_target_noindex", "Направить canonical на индексируемый URL."),
                ("canonical_target_redirect", "Направить canonical сразу на финальный URL со статусом 200."),
                ("canonical_target_error_status", "Исправить HTTP-статус canonical-цели."),
                ("noindex_canonical_conflict", "Выровнять сигналы индексируемости: canonical и robots noindex."),
                ("http_status_error", "Исправить HTTP-статус страницы и внутренние ссылки на этот URL."),
                ("non_https_url", "Включить HTTPS и настроить редирект HTTP -> HTTPS."),
                ("compression_disabled", "Включить сжатие gzip/brotli."),
                ("cache_disabled", "Добавить заголовки Cache-Control/ETag."),
                ("light_perf_low_score", "Снизить render-blocking JS и тяжелый DOM."),
                ("thin_content", "Расширить контент и убрать шаблонную «воду»."),
                ("near_duplicate_content", "Переписать близко-дублирующиеся блоки текста."),
                ("deep_click_depth", "Добавить внутренние ссылки со страниц верхнего уровня."),
                ("h1_hierarchy_issue", "Исправить структуру H1 и иерархию заголовков."),
                ("structured_data_common_errors", "Исправить обязательные поля в structured data (цена, наличие, рейтинг и т.д.)."),
                ("ai_risk_high", "Очеловечить текст и убрать повторяющиеся AI-паттерны."),
                ("security_missing_csp", "Добавить заголовок Content-Security-Policy."),
                ("security_missing_hsts", "Добавить заголовок Strict-Transport-Security."),
                ("security_missing_xfo", "Добавить заголовок X-Frame-Options."),
                ("security_missing_referrer_policy", "Добавить заголовок Referrer-Policy."),
                ("security_missing_permissions_policy", "Добавить заголовок Permissions-Policy."),
                ("security_mixed_content_homepage", "Заменить все ресурсы http:// на https://."),
                ("crawl_budget_risk_high", "Снизить число URL-параметров и уменьшить глубину URL."),
                ("crawl_budget_risk_medium", "Нормализовать URL-параметры для контроля обхода."),
                ("multiple_title_tags", "Оставить один title-тег на страницу."),
                ("multiple_meta_descriptions", "Оставить один meta description."),
                ("multiple_meta_robots", "Оставить один meta robots."),
                ("missing_charset_meta", "Добавить charset в <head>."),
                ("missing_viewport_meta", "Добавить meta viewport."),
                ("generic_alt_texts", "Заменить общие ALT-тексты на конкретные описания."),
                ("decorative_images_with_alt", "Использовать пустой ALT для декоративных изображений."),
                ("duplicate_image_sources", "Убрать дубли источников изображений."),
                ("low_modern_image_formats", "Конвертировать ключевые изображения в WebP/AVIF."),
                ("hreflang_extended_check", "Исправить взаимность hreflang/x-default/коды языков."),
                ("hidden_content_css", "Убрать скрытый SEO-текст (display:none/offscreen/opacity/font-size<5px)."),
                ("cloaking_detected", "Выровнять видимый и скрытый контент, убрать признаки клоакинга."),
                ("cta_missing", "Добавить четкие CTA-блоки (форма/кнопка/обратный звонок/чекаут)."),
                ("no_lists_tables_on_long_content", "Структурировать длинный текст списками и таблицами."),
            ]
            prefix = "Инфо:"
            if severity == "critical" or code in critical_codes:
                prefix = "Критично:"
            elif severity == "warning" or code in warning_codes:
                prefix = "Предупреждение:"
            for key, rec in mapping:
                if key in code:
                    return f"{prefix} {rec}"
            if "canonical" in title:
                return f"{prefix} Проверить canonical URL и его консистентность."
            if "index" in title:
                return f"{prefix} Исправить индексируемость и директивы robots."
            if "schema" in title or "structured" in title:
                return f"{prefix} Исправить обязательные поля structured data."
            if details:
                return f"{prefix} Исследовать и исправить: {details[:120]}"
            return f"{prefix} Исправить проблему с учетом контекста страницы и SEO best practices."

        def fill_sheet(
            sheet_name: str,
            headers: List[str],
            rows: List[List[Any]],
            severity_idx: int = -1,
            widths: List[int] = None,
            score_idx: int = -1,
        ):
            wsx = wb.create_sheet(sheet_name)
            for col, header in enumerate(headers, 1):
                self._apply_style(wsx.cell(row=1, column=col, value=self._sanitize_cell_value(header)), header_style)
            ordered_rows = list(rows)
            if severity_idx >= 0:
                severity_rank = {"critical": 0, "warning": 1, "info": 2, "ok": 3}
                ordered_rows.sort(
                    key=lambda row: severity_rank.get(str(row[severity_idx]).lower(), 4) if severity_idx < len(row) else 4
                )
            sev_counts = {"critical": 0, "warning": 0, "info": 0}
            score_values: List[float] = []
            numeric_kpi_by_col: Dict[int, List[Dict[str, Any]]] = defaultdict(list)
            row_severity: Dict[int, str] = {}
            for row_idx, row_data in enumerate(ordered_rows, start=2):
                for col, value in enumerate(row_data, 1):
                    cell = wsx.cell(row=row_idx, column=col, value=self._sanitize_cell_value(value))
                    self._apply_style(cell, cell_style)
                    header = headers[col - 1] if (col - 1) < len(headers) else ""
                    if self._is_kpi_header(header):
                        base_status = self._apply_kpi_cell_style(cell, header, value)
                        direction = self._kpi_anomaly_direction(header)
                        try:
                            num_val = float(value)
                            if direction != "none":
                                numeric_kpi_by_col[col].append({"row": row_idx, "value": num_val, "status": base_status, "header": header})
                        except Exception:
                            pass
                if severity_idx >= 0:
                    sev_value = str(row_data[severity_idx]).lower()
                    row_severity[row_idx] = sev_value
                    if sev_value in sev_counts:
                        sev_counts[sev_value] += 1
                    self._apply_row_severity_fill(wsx, row_idx, 1, len(headers), sev_value)
                    sev_cell = wsx.cell(row=row_idx, column=severity_idx + 1)
                    if sev_value == "critical":
                        self._apply_severity_cell_style(sev_cell, sev_value)
                    else:
                        sev_cell.alignment = Alignment(horizontal='center', vertical='center')
                if score_idx >= 0 and score_idx < len(row_data):
                    try:
                        score_values.append(float(row_data[score_idx]))
                    except Exception:
                        pass

            # Column-relative anomaly highlight for numeric KPI cells.
            for col_idx, items in numeric_kpi_by_col.items():
                if len(items) < 6:
                    continue
                direction = self._kpi_anomaly_direction(items[0]["header"])
                values = [float(item["value"]) for item in items]
                hi_warn = self._quantile(values, 0.90)
                hi_bad = self._quantile(values, 0.97)
                lo_warn = self._quantile(values, 0.10)
                lo_bad = self._quantile(values, 0.03)
                for item in items:
                    row_idx = int(item["row"])
                    if row_severity.get(row_idx, "") == "critical":
                        continue
                    if str(item.get("status", "")).lower() in {"bad", "warn"}:
                        continue
                    value = float(item["value"])
                    anomaly_state = ""
                    if direction == "high_worse":
                        if value >= hi_bad:
                            anomaly_state = "bad"
                        elif value >= hi_warn:
                            anomaly_state = "warn"
                    elif direction == "low_worse":
                        if value <= lo_bad:
                            anomaly_state = "bad"
                        elif value <= lo_warn:
                            anomaly_state = "warn"
                    if not anomaly_state:
                        continue
                    cell = wsx.cell(row=row_idx, column=col_idx)
                    palette = self._status_palette(anomaly_state)
                    cell.fill = PatternFill(start_color=palette["fill"], end_color=palette["fill"], fill_type='solid')
                    cell.font = Font(color=palette["font"], bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            wsx.freeze_panes = "A2"
            wsx.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
            if widths:
                for col, width in enumerate(widths, 1):
                    wsx.column_dimensions[get_column_letter(col)].width = width
            sheet_stats.append(
                {
                    "sheet": sheet_name,
                    "critical": sev_counts["critical"],
                    "warning": sev_counts["warning"],
                    "info": sev_counts["info"],
                    "avg_score": round(sum(score_values) / len(score_values), 1) if score_values else "",
                }
            )

        def sort_rows(rows: List[List[Any]], idx: int, reverse: bool = False):
            rows.sort(key=lambda row: to_float(row[idx], 0.0) if idx < len(row) else 0.0, reverse=reverse)

        def delta_to_target(score: Any, target: float) -> float:
            value = to_float(score, 0.0)
            return round(max(0.0, float(target) - value), 1)

        def score_level(value: Any, high: float = 80.0, medium: float = 60.0) -> str:
            v = to_float(value, 0.0)
            if v >= high:
                return "high"
            if v >= medium:
                return "medium"
            return "low"

        def approximate_pixel_width(text: Any) -> int:
            s = str(text or "")
            width = 0
            for ch in s:
                if ch.isspace():
                    width += 3
                elif re.match(r"[A-Za-z0-9]", ch):
                    width += 7
                elif re.match(r"[А-Яа-яЁё]", ch):
                    width += 8
                else:
                    width += 6
            return width

        def normalize_url_value(url: Any) -> str:
            value = str(url or "").strip().rstrip("/")
            return value.lower()

        def detect_issue_tab(code: str) -> str:
            code_l = str(code or "").lower()
            if any(x in code_l for x in ("title", "meta", "h1", "canonical", "robots", "viewport", "charset", "schema", "structured")):
                return "2_OnPage+Structured"
            if any(x in code_l for x in ("http_status", "https", "cache", "compression", "security", "mixed_content", "crawl_budget", "redirect")):
                return "3_Technical"
            if any(x in code_l for x in ("thin_content", "duplicate_content", "ai_", "hidden_content", "cloaking", "cta", "list", "table")):
                return "4_Content+AI"
            if any(x in code_l for x in ("anchor", "link", "orphan", "pagerank")):
                return "5_LinkGraph"
            if any(x in code_l for x in ("image", "alt", "webp", "avif", "external")):
                return "6_Images+External"
            if any(x in code_l for x in ("hierarchy", "heading", "h1_hierarchy")):
                return "7_HierarchyErrors"
            if any(x in code_l for x in ("keyword", "tf_idf", "intent", "cannibal")):
                return "8_Keywords"
            return "14_Issues_Raw"

        def owner_hint_by_code(code: str) -> str:
            code_l = str(code).lower()
            if any(x in code_l for x in ("title", "meta", "h1", "keyword", "content", "ai_", "duplicate_")):
                return "Content+SEO"
            if any(x in code_l for x in ("schema", "structured", "hreflang", "canonical", "index", "http_status")):
                return "SEO+Dev"
            if any(x in code_l for x in ("security", "cache", "compression", "https", "crawl_budget")):
                return "Dev+Infra"
            return "SEO"

        def root_cause_cluster(code: str) -> str:
            code_l = str(code).lower()
            if any(x in code_l for x in ("title", "meta", "h1", "keyword", "content", "duplicate", "ai_")):
                return "Content model"
            if any(x in code_l for x in ("canonical", "index", "robots", "http_status", "redirect", "crawl_budget")):
                return "Indexing flow"
            if any(x in code_l for x in ("schema", "structured", "hreflang")):
                return "Structured data"
            if any(x in code_l for x in ("security", "https", "cache", "compression", "mixed_content")):
                return "Platform infra"
            if any(x in code_l for x in ("image", "alt", "webp", "avif")):
                return "Media pipeline"
            return "General SEO"

        def dependency_codes_for(code: str) -> str:
            code_l = str(code).lower()
            deps: List[str] = []
            if any(x in code_l for x in ("meta", "title", "h1", "canonical", "robots")):
                deps.append("template_render")
            if any(x in code_l for x in ("schema", "structured")):
                deps.append("structured_template")
            if any(x in code_l for x in ("http_status", "redirect", "https", "security", "cache", "compression")):
                deps.append("server_config")
            if any(x in code_l for x in ("image", "alt", "webp", "avif")):
                deps.append("media_cdn")
            if any(x in code_l for x in ("crawl_budget", "index", "canonical")):
                deps.append("crawl_policy")
            return ", ".join(deps[:4]) if deps else "none"

        # Sheet 1: Executive summary + formula-based issue rates
        ws = wb.active
        ws.title = "1_Сводка"
        ws["A1"] = "Отчет Site Audit Pro"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:D1")

        summary_rows = [
            ("URL", report_url),
            ("Режим", mode),
            ("Всего страниц", summary.get("total_pages", len(pages))),
            ("Внутренние страницы", summary.get("internal_pages", "")),
            ("Всего проблем", summary.get("issues_total", len(issues))),
            ("Критично", summary.get("critical_issues", 0)),
            ("Предупреждение", summary.get("warning_issues", 0)),
            ("Инфо", summary.get("info_issues", 0)),
            ("Оценка", summary.get("score", "")),
        ]
        row = 3
        for key, value in summary_rows:
            ws.cell(row=row, column=1, value=self._sanitize_cell_value(key)).font = Font(bold=True)
            ws.cell(row=row, column=2, value=self._sanitize_cell_value(value))
            row += 1
        ws["A13"] = "Доля критичных"
        ws["B13"] = "=IF(B5=0,0,B8/B5)"
        ws["A14"] = "Доля предупреждений"
        ws["B14"] = "=IF(B5=0,0,B9/B5)"
        ws["A15"] = "Доля инфо"
        ws["B15"] = "=IF(B5=0,0,B10/B5)"
        ws["B13"].number_format = "0.00%"
        ws["B14"].number_format = "0.00%"
        ws["B15"].number_format = "0.00%"

        ws["A17"] = "Ср. ответ (мс)"
        ws["B17"] = pipeline_metrics.get("avg_response_time_ms", "")
        ws["A18"] = "Ср. читабельность"
        ws["B18"] = pipeline_metrics.get("avg_readability_score", "")
        ws["A19"] = "Ср. качество ссылок"
        ws["B19"] = pipeline_metrics.get("avg_link_quality_score", "")
        ws["A20"] = "Страницы-сироты"
        ws["B20"] = pipeline_metrics.get("orphan_pages", "")
        ws["A21"] = "Тематические хабы"
        ws["B21"] = pipeline_metrics.get("topic_hubs", "")
        critical_pages = sum(
            1
            for page in pages
            if any(str(i.get("severity", "")).lower() == "critical" for i in issues_by_url.get(str(page.get("url", "")), []))
        )
        ws["D17"] = "Критичные страницы"
        ws["E17"] = critical_pages
        ws["D18"] = "Критичные страницы %"
        ws["E18"] = round((critical_pages / float(max(1, len(pages)))) * 100.0, 1)

        ws["A23"] = "Матрица качества вкладок"
        ws["A23"].font = Font(bold=True)
        ws["A24"] = "Вкладка"
        ws["B24"] = "Критично"
        ws["C24"] = "Предупреждение"
        ws["D24"] = "Инфо"
        ws["E24"] = "Нагрузка проблем"
        ws["F24"] = "Ср. оценка"
        ws["G24"] = "Индекс здоровья"
        ws["H24"] = "Вес влияния"
        for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
            ws[f"{col}24"].font = Font(bold=True)

        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 80
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 14
        ws.column_dimensions["H"].width = 14
        ws.column_dimensions["I"].width = 56
        ws.column_dimensions["J"].width = 14
        ws.column_dimensions["K"].width = 14
        ws.column_dimensions["M"].width = 28
        ws.column_dimensions["N"].width = 14
        ws.column_dimensions["O"].width = 14
        ws.column_dimensions["P"].width = 52

        ws["I3"] = "Топ критичных страниц"
        ws["I3"].font = Font(bold=True)
        ws["I4"] = "URL"
        ws["J4"] = "Критично"
        ws["K4"] = "Предупреждение"
        for col in ["I", "J", "K"]:
            ws[f"{col}4"].font = Font(bold=True)
        page_issue_stats = []
        for page in pages:
            url = str(page.get("url", ""))
            p_issues = issues_by_url.get(url, [])
            c_count = sum(1 for i in p_issues if str(i.get("severity", "")).lower() == "critical")
            w_count = sum(1 for i in p_issues if str(i.get("severity", "")).lower() == "warning")
            if c_count > 0 or w_count > 0:
                page_issue_stats.append((url, c_count, w_count))
        page_issue_stats.sort(key=lambda x: (x[1], x[2]), reverse=True)
        for offset, (p_url, c_count, w_count) in enumerate(page_issue_stats[:10], start=5):
            ws.cell(row=offset, column=9, value=self._sanitize_cell_value(p_url))
            ws.cell(row=offset, column=10, value=c_count)
            ws.cell(row=offset, column=11, value=w_count)
            for col in range(9, 12):
                self._apply_style(ws.cell(row=offset, column=col), cell_style)

        ws["M3"] = "Топ кодов проблем"
        ws["M3"].font = Font(bold=True)
        ws["M4"] = "Код"
        ws["N4"] = "Количество"
        ws["O4"] = "Доля %"
        ws["P4"] = "Владелец"
        for col in ["M", "N", "O", "P"]:
            ws[f"{col}4"].font = Font(bold=True)
        issue_code_stats = Counter(str(i.get("code") or "unknown") for i in issues)
        for offset, (code, count) in enumerate(issue_code_stats.most_common(10), start=5):
            ws.cell(row=offset, column=13, value=self._sanitize_cell_value(code))
            ws.cell(row=offset, column=14, value=count)
            ws.cell(row=offset, column=15, value=round((count / float(max(1, len(issues)))) * 100.0, 1))
            ws.cell(row=offset, column=16, value=self._sanitize_cell_value(owner_hint_by_code(code)))
            for col in range(13, 17):
                self._apply_style(ws.cell(row=offset, column=col), cell_style)

        # Sheet 2: OnPage + Structured
        onpage_headers = [
            "URL", "Title", "Длина title", "Ширина title, px", "Meta description", "Длина description", "Ширина description, px", "Риск обрезки в SERP", "Кол-во H1", "Текст H1",
            "Canonical URL", "Статус canonical", "Meta robots", "X-Robots", "Кол-во schema",
            "JSON-LD", "Microdata", "RDFa", "Типы structured", "Кол-во hreflang",
            "Хлебные крошки", "Mobile hint", "Charset", "Viewport", "Дубликаты meta robots",
            "Кол-во title-тегов", "Кол-во description-тегов", "Дубли title", "Дубли description", "Self-match canonical", "OnPage-скор", "OnPage-дельта до цели", "OnPage-решение", "Критичность",
        ]
        onpage_rows = []
        for page in pages:
            sev = infer_page_severity(page)
            d = derived_by_url.get(str(page.get("url", "")), {})
            sdetail = d.get("structured_detail", {})
            title_px = approximate_pixel_width(d.get("title", page.get("title", "")))
            desc_px = approximate_pixel_width(d.get("meta_description", page.get("meta_description", "")))
            serp_risk = "✅" if (title_px > 580 or desc_px > 920) else "❌"
            canonical_self_match = "✅" if normalize_url_value(page.get("canonical", "")) == normalize_url_value(page.get("final_url", page.get("url", ""))) else "❌"
            onpage_rows.append([
                page.get("url", ""),
                d.get("title", page.get("title", "")),
                d.get("title_len", 0),
                title_px,
                d.get("meta_description", page.get("meta_description", "")),
                d.get("description_len", 0),
                desc_px,
                serp_risk,
                page.get("h1_count", ""),
                page.get("h1_text", ""),
                page.get("canonical", ""),
                page.get("canonical_status", ""),
                page.get("meta_robots", ""),
                page.get("x_robots_tag", ""),
                page.get("structured_data", 0),
                sdetail.get("json_ld", 0),
                sdetail.get("microdata", 0),
                sdetail.get("rdfa", 0),
                ", ".join((page.get("structured_types") or [])[:6]),
                page.get("hreflang_count", 0),
                page.get("breadcrumbs", ""),
                page.get("mobile_friendly_hint", ""),
                page.get("charset_declared", ""),
                page.get("viewport_declared", ""),
                page.get("multiple_meta_robots", ""),
                page.get("title_tags_count", 0),
                page.get("meta_description_tags_count", 0),
                page.get("duplicate_title_count", 0),
                page.get("duplicate_description_count", 0),
                canonical_self_match,
                d.get("onpage_score", ""),
                delta_to_target(d.get("onpage_score", ""), 85.0),
                page_solution("onpage", page),
                sev,
            ])
        fill_sheet(
            "2_OnPage+Structured",
            onpage_headers,
            onpage_rows,
            severity_idx=33,
            widths=[56, 28, 10, 10, 32, 10, 10, 14, 10, 22, 32, 16, 20, 20, 12, 8, 10, 8, 36, 10, 10, 10, 10, 12, 10, 12, 10, 10, 10, 10, 12, 12, 48, 10],
            score_idx=30,
        )

        # Sheet 3: Technical
        tech_headers = [
            "URL", "Итоговый URL", "Статус", "Строка статуса", "Ответ, мс", "Размер, KB", "HTML байт", "DOM-узлы", "Редиректы",
            "HTTPS", "Compression", "Compression algo", "Cache enabled", "Cache-Control",
            "Last-Modified", "Freshness days", "JS assets", "CSS assets", "Render-blocking JS", "Preload hints",
            "Perf light score", "Path depth", "URL params", "Crawl budget risk",
            "Security headers score", "CSP", "HSTS", "X-Frame-Options", "Referrer-Policy", "Permissions-Policy", "Mixed content refs",
            "Оценка качества HTML", "Кол-во устаревших тегов", "Причина неиндексируемости", "TTFB, мс", "Соотношение HTML/JS", "Риск цепочки редиректов", "Транспортный риск", "Уровень транспорта", "Технический скор", "Техническая дельта до цели", "Техническое решение", "Критичность",
        ]
        tech_rows = []
        for page in pages:
            sev = infer_page_severity(page)
            d = derived_by_url.get(str(page.get("url", "")), {})
            ttfb_ms = int(page.get("ttfb_ms") or d.get("response_ms", 0) or 0)
            js_assets_count = int(page.get("js_assets_count", 0) or 0)
            html_js_ratio = round((to_float(d.get("content_kb", 0.0), 0.0) / max(1.0, float(js_assets_count))), 2)
            redirect_chain_risk = "✅" if int(page.get("redirect_count", 0) or 0) > 1 else "❌"
            transport_risk = 100
            if not bool(page.get("is_https")):
                transport_risk -= 40
            if not bool(page.get("hsts_present")):
                transport_risk -= 25
            if int(page.get("mixed_content_count", 0) or 0) > 0:
                transport_risk -= 25
            if not bool(page.get("csp_present")):
                transport_risk -= 10
            transport_risk = max(0, transport_risk)
            transport_level = score_level(transport_risk, high=85.0, medium=65.0)
            tech_rows.append([
                page.get("url", ""),
                page.get("final_url", ""),
                d.get("status_code", ""),
                page.get("status_line", "") or "",
                d.get("response_ms", ""),
                d.get("content_kb", ""),
                d.get("html_bytes", ""),
                page.get("dom_nodes_count", ""),
                page.get("redirect_count", 0),
                page.get("is_https", ""),
                page.get("compression_enabled", ""),
                page.get("compression_algorithm", ""),
                page.get("cache_enabled", ""),
                page.get("cache_control", ""),
                page.get("last_modified", ""),
                page.get("content_freshness_days", ""),
                page.get("js_assets_count", 0),
                page.get("css_assets_count", 0),
                page.get("render_blocking_js_count", 0),
                page.get("preload_hints_count", 0),
                page.get("perf_light_score", ""),
                page.get("path_depth", 0),
                page.get("url_params_count", 0),
                page.get("crawl_budget_risk", ""),
                page.get("security_headers_score", ""),
                page.get("csp_present", ""),
                page.get("hsts_present", ""),
                page.get("x_frame_options_present", ""),
                page.get("referrer_policy_present", ""),
                page.get("permissions_policy_present", ""),
                page.get("mixed_content_count", 0),
                page.get("html_quality_score", ""),
                len(page.get("deprecated_tags") or []),
                page.get("indexability_reason", ""),
                ttfb_ms,
                html_js_ratio,
                redirect_chain_risk,
                transport_risk,
                transport_level,
                d.get("technical_score", ""),
                delta_to_target(d.get("technical_score", ""), 85.0),
                page_solution("technical", page),
                sev,
            ])
        fill_sheet(
            "3_Technical",
            tech_headers,
            tech_rows,
            severity_idx=42,
            widths=[50, 50, 10, 22, 12, 10, 12, 10, 10, 8, 10, 14, 10, 22, 24, 14, 10, 10, 14, 10, 12, 10, 10, 14, 14, 8, 8, 12, 12, 14, 14, 10, 12, 12, 10, 10, 12, 12, 12, 12, 46, 10],
            score_idx=39,
        )

        # Sheet 4: Content + AI
        content_headers = [
            "URL", "Кол-во слов", "Уникальных слов", "Уникальность, %", "Лексическое разнообразие", "Скор читабельности",
            "Ср. длина предложения", "Ср. длина слова", "Сложные слова, %", "Скор keyword stuffing",
            "Плотность контента, %", "Boilerplate, %", "Toxicity-скор", "Доля «воды»",
            "Фразы-наполнители", "Кол-во AI-маркеров", "Список AI-маркеров", "Пример AI-маркеров",
            "Плотность AI /1k", "AI-риск", "Уровень AI-риска", "Тип страницы",
            "Скрытый контент", "Скрытые узлы", "Символов скрытого текста", "Клоакинг",
            "Соотношение контент/шаблон", "Кол-во абзацев", "Ср. длина абзаца", "Критичность скрытого",
            "Кол-во CTA", "Качество CTA", "Микс типов CTA", "Кол-во списков", "Кол-во таблиц",
            "Близкие дубли", "URL близких дублей",
            "Контент-скор", "Контент-дельта до цели", "Контент-решение", "Критичность",
        ]
        content_rows = []
        for page in pages:
            sev = infer_page_severity(page)
            d = derived_by_url.get(str(page.get("url", "")), {})
            content_density = to_float(page.get("content_density", 0), 0.0)
            boilerplate_percent = to_float(page.get("boilerplate_percent", 0), 0.0)
            content_template_ratio = round(content_density / max(1.0, boilerplate_percent), 2)
            paragraph_count = int(page.get("paragraph_count", 0) or max(1, int(page.get("word_count", 0) or 0) // 80))
            avg_paragraph_len = round((to_float(page.get("word_count", 0), 0.0) / max(1, paragraph_count)), 1)
            hidden_nodes = int(page.get("hidden_nodes_count", 0) or 0)
            hidden_chars = int(page.get("hidden_text_chars", 0) or 0)
            if not bool(page.get("hidden_content", False)):
                hidden_severity = "none"
            elif hidden_nodes >= 20 or hidden_chars >= 1200:
                hidden_severity = "high"
            elif hidden_nodes >= 8 or hidden_chars >= 400:
                hidden_severity = "medium"
            else:
                hidden_severity = "low"
            cta_type_mix = page.get("cta_type_mix", "")
            if not cta_type_mix:
                cta_parts = []
                for key, label in (("cta_form_count", "form"), ("cta_button_count", "button"), ("cta_link_count", "link")):
                    val = int(page.get(key, 0) or 0)
                    if val > 0:
                        cta_parts.append(f"{label}:{val}")
                cta_type_mix = ", ".join(cta_parts) if cta_parts else "н/д"
            content_rows.append([
                page.get("url", ""),
                page.get("word_count", 0),
                page.get("unique_word_count", 0),
                page.get("unique_percent", 0),
                page.get("lexical_diversity", 0),
                page.get("readability_score", 0),
                page.get("avg_sentence_length", 0),
                page.get("avg_word_length", 0),
                page.get("complex_words_percent", 0),
                page.get("keyword_stuffing_score", 0),
                page.get("content_density", 0),
                page.get("boilerplate_percent", 0),
                page.get("toxicity_score", 0),
                page.get("filler_ratio", 0),
                len(page.get("filler_phrases") or []),
                page.get("ai_markers_count", 0),
                ", ".join((page.get("ai_markers_list") or [])[:10]),
                page.get("ai_marker_sample", "") or "No sample",
                d.get("ai_density", page.get("ai_markers_density_1k", 0)),
                d.get("ai_risk", page.get("ai_risk_score", 0)),
                d.get("ai_risk_level", page.get("ai_risk_level", "")),
                page.get("page_type", ""),
                page.get("hidden_content", False),
                page.get("hidden_nodes_count", 0),
                page.get("hidden_text_chars", 0),
                page.get("cloaking_detected", False),
                content_template_ratio,
                paragraph_count,
                avg_paragraph_len,
                hidden_severity,
                page.get("cta_count", 0),
                page.get("cta_text_quality", 0),
                cta_type_mix,
                page.get("lists_count", 0),
                page.get("tables_count", 0),
                page.get("near_duplicate_count", 0),
                ", ".join((page.get("near_duplicate_urls") or [])[:5]),
                d.get("content_score", ""),
                delta_to_target(d.get("content_score", ""), 80.0),
                page_solution("content", page),
                sev,
            ])
        sort_rows(content_rows, 15, reverse=True)
        fill_sheet(
            "4_Content+AI",
            content_headers,
            content_rows,
            severity_idx=40,
            widths=[52, 10, 12, 10, 12, 12, 12, 10, 12, 12, 12, 12, 10, 10, 12, 12, 50, 62, 12, 10, 12, 12, 10, 10, 12, 10, 12, 12, 12, 12, 10, 10, 18, 10, 10, 12, 12, 36, 12, 12, 46, 10],
            score_idx=37,
        )

        # Sheet 5: Link Graph
        link_headers = [
            "URL", "Входящие внутренние", "Исходящие внутренние", "Исходящие внешние", "Страница-сирота",
            "Тематический хаб", "Глубина клика", "PageRank", "Доля слабых анкоров", "Качество анкоров", "Качество ссылок",
            "Всего follow-ссылок", "Всего nofollow-ссылок", "Кол-во семантических ссылок", "Возможности внутренней перелинковки", "Битые внутренние цели", "Алерт переиспользования анкоров",
            "Link-скор", "Link-дельта до цели", "Решение по перелинковке", "Критичность",
        ]
        link_rows = []
        for page in pages:
            sev = infer_page_severity(page)
            d = derived_by_url.get(str(page.get("url", "")), {})
            incoming_internal = int(page.get("incoming_internal_links", 0) or 0)
            outgoing_internal = int(page.get("outgoing_internal_links", 0) or 0)
            semantic_links_count = len(page.get("semantic_links") or [])
            internal_link_opportunities = int(page.get("internal_link_opportunities", 0) or max(0, 8 - incoming_internal))
            broken_internal_targets = int(page.get("broken_internal_links_count", page.get("broken_internal_targets", 0)) or 0)
            anchor_overuse_alert = "✅" if (to_float(page.get("weak_anchor_ratio"), 0.0) >= 0.35 or to_float(page.get("anchor_text_quality_score"), 100.0) < 40.0) else "❌"
            link_rows.append([
                page.get("url", ""),
                incoming_internal,
                outgoing_internal,
                page.get("outgoing_external_links", 0),
                page.get("orphan_page", ""),
                page.get("topic_hub", ""),
                page.get("click_depth", ""),
                page.get("pagerank", 0),
                page.get("weak_anchor_ratio", 0),
                page.get("anchor_text_quality_score", 0),
                page.get("link_quality_score", 0),
                page.get("follow_links_total", 0),
                page.get("nofollow_links_total", 0),
                semantic_links_count,
                internal_link_opportunities,
                broken_internal_targets,
                anchor_overuse_alert,
                d.get("link_score", ""),
                delta_to_target(d.get("link_score", ""), 80.0),
                page_solution("link_quality", page),
                sev,
            ])
        sort_rows(link_rows, 17, reverse=False)
        fill_sheet(
            "5_LinkGraph",
            link_headers,
            link_rows,
            severity_idx=20,
            widths=[50, 12, 12, 12, 10, 10, 10, 10, 14, 12, 12, 14, 16, 14, 12, 12, 12, 10, 12, 46, 10],
            score_idx=17,
        )

        # Sheet 6: Images + External
        img_headers = [
            "URL", "Всего изображений", "Без alt", "Без width/height", "Без lazy-load", "Всего image-проблем",
            "Современные форматы", "Дубли src", "Внешние изображения", "Generic ALT", "Декоративные с ALT",
            "Всего внешних", "Внешние follow", "Внешние nofollow", "Доля follow, %",
            "Макс. размер изображения, KB", "Без srcset", "Домены внешних изображений", "Релевантность ALT",
            "Медиа-скор", "Медиа-дельта до цели", "Решение по изображениям+внешним", "Критичность",
        ]
        img_rows = []
        for page in pages:
            sev = infer_page_severity(page)
            d = derived_by_url.get(str(page.get("url", "")), {})
            img_opt = page.get("images_optimization") or {}
            largest_image_kb = round(to_float(page.get("largest_image_kb", img_opt.get("largest_kb", 0)), 0.0), 1)
            no_srcset_count = int(page.get("images_without_srcset", img_opt.get("no_srcset", 0)) or 0)
            external_domains_count = int(page.get("external_image_domains_count", 0) or len(set(page.get("external_image_domains") or [])))
            generic_alt = int(page.get("generic_alt_count", 0) or 0)
            missing_alt = int(d.get("no_alt", 0) or 0)
            alt_relevance = "good"
            if generic_alt > 0 and missing_alt > 0:
                alt_relevance = "weak"
            elif generic_alt > 0 or missing_alt > 0:
                alt_relevance = "medium"
            img_rows.append([
                page.get("url", ""),
                page.get("images_count", img_opt.get("total", 0)),
                d.get("no_alt", 0),
                d.get("no_wh", 0),
                d.get("no_lazy", 0),
                d.get("image_issues_total", 0),
                page.get("images_modern_format_count", 0),
                page.get("image_duplicate_src_count", 0),
                page.get("images_external_count", 0),
                page.get("generic_alt_count", 0),
                page.get("decorative_non_empty_alt_count", 0),
                d.get("ext_total", 0),
                d.get("ext_follow", 0),
                d.get("ext_nofollow", 0),
                d.get("follow_ratio", 0.0),
                largest_image_kb,
                no_srcset_count,
                external_domains_count,
                alt_relevance,
                d.get("media_score", ""),
                delta_to_target(d.get("media_score", ""), 85.0),
                f"{page_solution('images', page)}; {page_solution('external', page)}",
                sev,
            ])
        sort_rows(img_rows, 5, reverse=True)
        fill_sheet(
            "6_Images+External",
            img_headers,
            img_rows,
            severity_idx=22,
            widths=[52, 10, 10, 12, 12, 12, 12, 10, 10, 10, 12, 12, 12, 14, 12, 12, 10, 12, 12, 10, 12, 58, 10],
            score_idx=19,
        )

        # Sheet 7: Hierarchy + Errors
        issue_headers = [
            "URL", "Статус иерархии", "Проблемы иерархии", "Всего заголовков", "Кол-во H1",
            "Outline заголовков", "Код", "Заголовок проблемы", "Детали проблемы", "H2 до H1", "Скор глубины outline", "Дубли текстов заголовков", "TOC-ready",
            "Hierarchy-скор", "Hierarchy-дельта до цели", "Решение по иерархии", "Критичность",
        ]
        issue_rows = []
        for page in pages:
            sev = infer_page_severity(page)
            d = derived_by_url.get(str(page.get("url", "")), {})
            h_details = page.get("h_details") or {}
            h_outline = h_details.get("heading_outline") or []
            outline_text = " | ".join(f"H{item.get('level')}:{item.get('text')}" for item in h_outline[:8])
            first_level = int((h_outline[0] or {}).get("level", 0) or 0) if h_outline else 0
            h2_before_h1 = "✅" if (first_level > 1 or (page.get("h1_count", 0) or 0) == 0) else "❌"
            levels = [int(item.get("level", 0) or 0) for item in h_outline if int(item.get("level", 0) or 0) > 0]
            max_level = max(levels) if levels else 1
            level_skips = sum(1 for prev, cur in zip(levels, levels[1:]) if cur - prev > 1)
            outline_depth_score = max(0, 100 - ((max_level - 3) * 10) - (level_skips * 20))
            header_texts = [str(item.get("text", "")).strip().lower() for item in h_outline if str(item.get("text", "")).strip()]
            heading_dup_count = max(0, len(header_texts) - len(set(header_texts)))
            toc_ready = "✅" if (page.get("h1_count", 0) == 1 and level_skips == 0 and len(h_outline) >= 3) else "❌"
            page_issues = issues_by_url.get(page.get("url", ""), [])
            hierarchy_issue = next(
                (i for i in page_issues if "h1" in str(i.get("code", "")).lower() or "hierarchy" in str(i.get("code", "")).lower()),
                page_issues[0] if page_issues else {},
            )
            issue_rows.append([
                page.get("url", ""),
                page.get("h_hierarchy", ""),
                hierarchy_problem(page),
                h_details.get("total_headers", sum((page.get("heading_distribution") or {}).values())),
                page.get("h1_count", 0),
                outline_text,
                hierarchy_issue.get("code", ""),
                hierarchy_issue.get("title", ""),
                hierarchy_issue.get("details", ""),
                h2_before_h1,
                outline_depth_score,
                heading_dup_count,
                toc_ready,
                d.get("hierarchy_score", ""),
                delta_to_target(d.get("hierarchy_score", ""), 90.0),
                page_solution("hierarchy", page, page_issues),
                sev,
            ])
        sort_rows(issue_rows, 13, reverse=False)
        fill_sheet(
            "7_HierarchyErrors",
            issue_headers,
            issue_rows,
            severity_idx=16,
            widths=[50, 18, 32, 12, 10, 72, 20, 28, 40, 12, 12, 12, 10, 12, 12, 48, 10],
            score_idx=13,
        )

        # Sheet 8: Keywords
        keyword_headers = [
            "URL", "Тема", "Топ терминов (TF-IDF)", "Топ ключей", "TF-IDF #1", "TF-IDF #2", "TF-IDF #3",
            "Профиль плотности ключей", "TF-IDF термины", "Энтропия ключей", "Доля топ-ключа, %",
            "SPAM-алерт", "Вода, %", "BM25-подобная релевантность", "Точное в Title", "Точное в H1", "Точное в URL", "Уверенность интента",
            "Уровень интента", "Keyword-скор", "Keyword-дельта до цели", "Решение по ключам", "Критичность",
        ]
        keyword_rows = []
        total_pages = max(1, int(summary.get("total_pages", len(pages)) or len(pages) or 1))
        for page in pages:
            sev = infer_page_severity(page)
            d = derived_by_url.get(str(page.get("url", "")), {})
            url = page.get("url", "")
            top_terms = tfidf_by_url.get(url, page.get("top_terms", [])) or list((page.get("tf_idf_keywords") or {}).keys())
            kw_profile = page.get("keyword_density_profile") or {}
            kw_profile_text = ", ".join(f"{k}:{v}%" for k, v in list(kw_profile.items())[:6])
            density_vals = [to_float(v, 0.0) for v in kw_profile.values()]
            top_keyword_share = round(max(density_vals), 2) if density_vals else 0.0
            water_words_pct = round(to_float(page.get("filler_ratio"), 0.0) * 100.0, 2)
            primary_term = str(top_terms[0] if len(top_terms) > 0 else "").lower().strip()
            title_l = str(page.get("title", "")).lower()
            h1_l = str(page.get("h1_text", page.get("h1", ""))).lower()
            url_l = str(url).lower()
            exact_title = "✅" if primary_term and primary_term in title_l else "❌"
            exact_h1 = "✅" if primary_term and primary_term in h1_l else "❌"
            exact_url = "✅" if primary_term and primary_term in url_l else "❌"
            bm25_like = round((to_float(d.get("tfidf_terms_count", 0), 0.0) * 4.0) + max(0.0, 6.0 - top_keyword_share) + (to_float(d.get("keyword_entropy", 0), 0.0) * 5.0), 2)
            intent_confidence = min(100, int(30 + (20 if exact_title == "✅" else 0) + (25 if exact_h1 == "✅" else 0) + (25 if exact_url == "✅" else 0)))
            intent_level = score_level(intent_confidence, high=75.0, medium=50.0)
            spam_alert = "✅" if (
                to_float(page.get("keyword_stuffing_score"), 0.0) >= 35.0
                or top_keyword_share >= 12.0
                or to_float(page.get("toxicity_score"), 0.0) >= 40.0
            ) else "❌"
            keyword_rows.append([
                url,
                page.get("topic_label", ""),
                ", ".join(top_terms[:10]),
                ", ".join((page.get("top_keywords") or [])[:8]),
                top_terms[0] if len(top_terms) > 0 else "",
                top_terms[1] if len(top_terms) > 1 else "",
                top_terms[2] if len(top_terms) > 2 else "",
                kw_profile_text,
                d.get("tfidf_terms_count", len(page.get("tf_idf_keywords") or {})),
                d.get("keyword_entropy", 0.0),
                top_keyword_share,
                spam_alert,
                water_words_pct,
                bm25_like,
                exact_title,
                exact_h1,
                exact_url,
                intent_confidence,
                intent_level,
                d.get("keyword_score", ""),
                delta_to_target(d.get("keyword_score", ""), 80.0),
                page_solution("keywords", page),
                sev,
            ])
        sort_rows(keyword_rows, 19, reverse=False)
        fill_sheet(
            "8_Keywords",
            keyword_headers,
            keyword_rows,
            severity_idx=22,
            widths=[48, 16, 42, 36, 14, 14, 14, 42, 10, 12, 12, 10, 12, 12, 10, 10, 10, 12, 12, 10, 12, 46, 10],
            score_idx=19,
        )

        token_re = re.compile(r"[a-zA-Z\u0400-\u04FF0-9]{3,}")
        stop_words = {
            "the", "and", "for", "with", "that", "this", "from", "your", "you", "are",
            "http", "https", "www", "com", "page", "site", "about",
            "это", "как", "для", "что", "или", "при", "также", "если", "чтобы", "когда",
        }
        water_words = {
            "очень", "просто", "лучший", "максимально", "идеально", "качественно", "эффективно",
            "really", "very", "super", "best", "amazing", "quality", "effective", "efficient",
        }
        unigram_counter: Counter = Counter()
        bigram_counter: Counter = Counter()
        trigram_counter: Counter = Counter()
        term_page_presence: Dict[str, Set[str]] = defaultdict(set)
        tfidf_sum: Dict[str, float] = defaultdict(float)
        tfidf_count: Dict[str, int] = defaultdict(int)
        top_density_by_term: Dict[str, float] = defaultdict(float)
        total_unigram_tokens = 0

        for page in pages:
            url = str(page.get("url") or "")
            src_parts = [
                str(page.get("title") or ""),
                str(page.get("meta_description") or ""),
                " ".join(page.get("top_terms") or []),
                " ".join(page.get("top_keywords") or []),
                " ".join(list((page.get("tf_idf_keywords") or {}).keys())[:20]),
            ]
            source = " ".join(src_parts).lower()
            tokens = [t for t in token_re.findall(source) if t not in stop_words]
            if not tokens:
                continue
            total_unigram_tokens += len(tokens)
            uniq_tokens = set(tokens)
            for t in tokens:
                unigram_counter[t] += 1
            for term in uniq_tokens:
                term_page_presence[term].add(url)
            for a, b in zip(tokens, tokens[1:]):
                bigram_counter[f"{a} {b}"] += 1
            for a, b, c in zip(tokens, tokens[1:], tokens[2:]):
                trigram_counter[f"{a} {b} {c}"] += 1
            for term, val in (page.get("tf_idf_keywords") or {}).items():
                term_l = str(term).lower().strip()
                if not term_l:
                    continue
                tfidf_sum[term_l] += to_float(val, 0.0)
                tfidf_count[term_l] += 1
            for term, dens in (page.get("keyword_density_profile") or {}).items():
                term_l = str(term).lower().strip()
                if not term_l:
                    continue
                top_density_by_term[term_l] = max(top_density_by_term.get(term_l, 0.0), to_float(dens, 0.0))

        summary_headers = [
            "N-gram", "Ключ", "Общая частота", "Страниц с термином", "Страниц %",
            "Доля токена %", "Средний TF-IDF", "Пиковая плотность %", "SPAM-сигнал", "Вода/шум", "Межстраничный повтор", "Риск-скор", "Брендовый термин", "Интент термина", "Сводная заметка",
        ]
        summary_rows: List[List[Any]] = []
        brand_host = re.sub(r"^https?://", "", str(report_url or "").lower()).split("/")[0]
        brand_tokens = {t for t in re.split(r"[^a-z0-9а-яё]+", brand_host) if len(t) >= 3 and t not in {"www", "com", "net", "org", "ru"}}

        def append_ngram_rows(counter: Counter, ngram_size: int, limit: int) -> None:
            for term, freq in counter.most_common(limit):
                if freq <= 0:
                    continue
                pages_with_term = len(term_page_presence.get(term, set())) if ngram_size == 1 else 0
                pages_pct = round((pages_with_term / float(total_pages)) * 100.0, 1) if total_pages else 0.0
                token_share = round((freq / max(1, total_unigram_tokens)) * 100.0, 2) if ngram_size == 1 else ""
                avg_tfidf = round(tfidf_sum.get(term, 0.0) / max(1, tfidf_count.get(term, 0)), 4) if ngram_size == 1 else ""
                top_density = round(top_density_by_term.get(term, 0.0), 2) if ngram_size == 1 else ""
                spam_alert = "✅" if (ngram_size == 1 and (top_density_by_term.get(term, 0.0) >= 12.0 or pages_pct >= 80.0)) else "❌"
                water_noise = "✅" if term in stop_words or term in water_words else "❌"
                cross_page_repeat = "✅" if (ngram_size == 1 and pages_with_term >= max(3, int(total_pages * 0.4))) else "❌"
                risk_score = 0
                if spam_alert == "✅":
                    risk_score += 45
                if cross_page_repeat == "✅":
                    risk_score += 25
                if water_noise == "✅":
                    risk_score += 15
                if ngram_size == 1:
                    risk_score += min(15, int(round(to_float(top_density_by_term.get(term, 0.0), 0.0))))
                brand_term = "✅" if any(bt and bt in term for bt in brand_tokens) else "❌"
                term_intent = "informational"
                if any(x in term for x in ("buy", "price", "order", "куп", "цен", "заказ")):
                    term_intent = "transactional"
                elif any(x in term for x in ("service", "catalog", "product", "обзор", "каталог", "услуг")):
                    term_intent = "commercial"
                note = ""
                if spam_alert == "✅":
                    note = "Potential over-optimization across pages"
                elif cross_page_repeat == "✅":
                    note = "Repeated across many pages; validate intent relevance"
                elif ngram_size == 1 and tfidf_count.get(term, 0) <= 1 and freq >= 3:
                    note = "Frequent term with weak TF-IDF support"
                summary_rows.append(
                    [
                        f"{ngram_size}-gram",
                        term,
                        freq,
                        pages_with_term if ngram_size == 1 else "",
                        pages_pct if ngram_size == 1 else "",
                        token_share,
                        avg_tfidf,
                        top_density,
                        spam_alert,
                        water_noise,
                        cross_page_repeat,
                        risk_score,
                        brand_term,
                        term_intent,
                        note,
                    ]
                )

        append_ngram_rows(unigram_counter, 1, 120)
        append_ngram_rows(bigram_counter, 2, 80)
        append_ngram_rows(trigram_counter, 3, 60)
        summary_rows.sort(key=lambda row: (to_float(row[11], 0.0), to_float(row[2], 0.0)), reverse=True)
        fill_sheet(
            "8b_Keywords_Summary",
            summary_headers,
            summary_rows,
            widths=[10, 30, 10, 14, 10, 12, 12, 12, 10, 10, 12, 10, 10, 14, 56],
        )

        # Sheet 8c: Keywords insights (intent, cannibalization, gaps, section anomalies)
        def normalize_terms(values: List[Any]) -> List[str]:
            terms: List[str] = []
            for val in values:
                txt = str(val or "").strip().lower()
                if txt:
                    terms.append(txt)
            return terms

        def extract_path(url: str) -> str:
            m = re.match(r"^https?://[^/]+(?P<path>/[^?#]*)?", str(url or "").strip().lower())
            path = (m.group("path") if m else "") or "/"
            return path if path else "/"

        def extract_section(url: str) -> str:
            path = extract_path(url).strip("/")
            if not path:
                return "homepage"
            return path.split("/", 1)[0]

        def detect_intent(url: str, text: str) -> tuple:
            url_l = str(url or "").lower()
            text_l = str(text or "").lower()
            tokens = set(token_re.findall(text_l))

            intent_lex = {
                "transactional": {
                    "buy", "order", "checkout", "cart", "price", "pricing", "quote", "purchase",
                    "купить", "заказать", "цена", "стоимость", "оформить", "заказ",
                },
                "commercial": {
                    "best", "compare", "review", "service", "solutions", "product", "catalog",
                    "лучший", "сравнение", "обзор", "услуга", "товар", "каталог", "решения",
                },
                "informational": {
                    "how", "what", "why", "guide", "faq", "blog", "news", "article",
                    "как", "что", "почему", "инструкция", "faq", "блог", "новости", "статья",
                },
                "navigational": {
                    "about", "contact", "company", "login", "signin", "profile",
                    "о", "контакты", "компания", "вход", "профиль",
                },
            }
            url_boost = {
                "transactional": ("buy", "order", "checkout", "cart", "price", "pricing", "shop", "купить", "заказ", "цена"),
                "commercial": ("product", "catalog", "service", "solutions", "товар", "каталог", "услуги"),
                "informational": ("blog", "news", "guide", "faq", "help", "блог", "новости", "статья"),
                "navigational": ("about", "company", "contact", "login", "about-us", "contacts", "о-компании"),
            }

            scores = {}
            for intent, lex in intent_lex.items():
                token_hits = sum(1 for t in tokens if t in lex)
                url_hits = sum(1 for cue in url_boost[intent] if cue in url_l)
                scores[intent] = token_hits + (url_hits * 2)

            best_intent = max(scores, key=scores.get) if scores else "informational"
            ordered = sorted(scores.items(), key=lambda x: x[1], reverse=True)
            top_score = ordered[0][1] if ordered else 0
            second_score = ordered[1][1] if len(ordered) > 1 else 0
            confidence = min(100, max(15, 55 + (top_score - second_score) * 12 + top_score * 4))
            return best_intent, confidence

        insights_headers = [
            "Метод", "Область", "Интент", "Термин/паттерн", "Метрика", "Критичность", "Действие", "Примеры", "Приоритет",
        ]
        insights_rows: List[List[Any]] = []
        intent_counter: Counter = Counter()

        density_term_urls: Dict[str, set] = defaultdict(set)
        density_term_sum: Dict[str, float] = defaultdict(float)
        density_term_count: Dict[str, int] = defaultdict(int)
        density_term_max: Dict[str, float] = defaultdict(float)
        section_term_counts: Dict[str, Counter] = defaultdict(Counter)
        section_term_pages: Dict[str, Dict[str, set]] = defaultdict(lambda: defaultdict(set))
        section_token_total: Dict[str, int] = defaultdict(int)

        def choose_primary_url(urls_set: set, term: str) -> str:
            best_url = ""
            best_score = -10**9
            for candidate in urls_set:
                page = page_by_url.get(candidate, {})
                derived = derived_by_url.get(candidate, {})
                incoming = int(page.get("incoming_internal_links", 0) or 0)
                path_depth = int(page.get("path_depth", 0) or 0)
                keyword_score = to_float(derived.get("keyword_score", 0.0), 0.0)
                indexable_bonus = 20 if bool(page.get("indexable", False)) else -20
                exact_title = 12 if str(term or "").lower() in str(page.get("title", "")).lower() else 0
                score = keyword_score + (incoming * 2.0) - (path_depth * 5.0) + indexable_bonus + exact_title
                if score > best_score:
                    best_score = score
                    best_url = candidate
            return best_url

        for page in pages:
            url = str(page.get("url") or "")
            title = str(page.get("title") or "")
            meta = str(page.get("meta_description") or "")
            h1 = page.get("h1") or page.get("h1_text") or ""
            if isinstance(h1, list):
                h1 = " ".join(str(x or "") for x in h1[:3])
            top_terms_page = normalize_terms(page.get("top_terms") or [])
            top_keywords_page = normalize_terms(page.get("top_keywords") or [])
            tfidf_pairs = sorted((page.get("tf_idf_keywords") or {}).items(), key=lambda kv: to_float(kv[1], 0.0), reverse=True)
            tfidf_terms_page = normalize_terms([k for k, _ in tfidf_pairs[:8]])
            merged_text = " ".join([title, meta, str(h1), " ".join(top_terms_page), " ".join(top_keywords_page), " ".join(tfidf_terms_page)]).strip()

            # 1) Intent clustering (per page)
            intent, confidence = detect_intent(url, merged_text)
            intent_counter[intent] += 1
            intent_sev = "info"
            if confidence < 45:
                intent_sev = "warning"
            if confidence < 30:
                intent_sev = "critical"
            url_intent_hint = "unknown"
            path_l = extract_path(url)
            if any(k in path_l for k in ("buy", "order", "shop", "checkout", "купить", "заказ")):
                url_intent_hint = "transactional"
            elif any(k in path_l for k in ("catalog", "product", "service", "каталог", "товар", "услуг")):
                url_intent_hint = "commercial"
            elif any(k in path_l for k in ("blog", "news", "guide", "faq", "блог", "новост")):
                url_intent_hint = "informational"
            elif any(k in path_l for k in ("about", "contact", "company", "о-компании", "контакт")):
                url_intent_hint = "navigational"
            if url_intent_hint != "unknown" and url_intent_hint != intent:
                insights_rows.append([
                    "Конфликт интента",
                    url,
                    intent,
                    f"url_hint={url_intent_hint}",
                    f"confidence={confidence}%",
                    "warning" if confidence >= 45 else "critical",
                    "Привести контент и метаданные в соответствие с интентом URL или обновить slug.",
                    extract_section(url),
                ])
            insights_rows.append([
                "Кластер интента",
                url,
                intent,
                ", ".join((top_terms_page or top_keywords_page or tfidf_terms_page)[:4]),
                f"confidence={confidence}%",
                intent_sev,
                "Align title/H1/meta with one dominant intent.",
                extract_section(url),
            ])

            # 2) Gap analysis (per page, target terms vs title/h1/meta)
            title_terms = set(token_re.findall(title.lower()))
            h1_terms = set(token_re.findall(str(h1).lower()))
            meta_terms = set(token_re.findall(meta.lower()))
            targets = []
            seen = set()
            for t in top_terms_page + top_keywords_page + tfidf_terms_page:
                if t in seen:
                    continue
                seen.add(t)
                targets.append(t)
                if len(targets) >= 10:
                    break
            missing_title = [t for t in targets if t not in title_terms]
            missing_h1 = [t for t in targets if t not in h1_terms]
            missing_meta = [t for t in targets if t not in meta_terms]
            gap_risk = len(missing_title) + len(missing_h1) + len(missing_meta)
            gap_sev = "info"
            if gap_risk >= 10:
                gap_sev = "critical"
            elif gap_risk >= 6:
                gap_sev = "warning"
            insights_rows.append([
                "Gap-анализ",
                url,
                intent,
                ", ".join(targets[:5]),
                f"пропуски title/h1/meta = {len(missing_title)}/{len(missing_h1)}/{len(missing_meta)}",
                gap_sev,
                "Добавьте ключевые термины в title/H1/meta естественно, без переспама.",
                "; ".join([
                    f"title:{', '.join(missing_title[:3])}" if missing_title else "",
                    f"h1:{', '.join(missing_h1[:3])}" if missing_h1 else "",
                    f"meta:{', '.join(missing_meta[:3])}" if missing_meta else "",
                ]).strip("; ").strip(),
            ])

            # 3) Data for cannibalization
            for term, dens in (page.get("keyword_density_profile") or {}).items():
                term_l = str(term or "").strip().lower()
                dens_v = to_float(dens, 0.0)
                if not term_l or dens_v <= 0:
                    continue
                density_term_urls[term_l].add(url)
                density_term_sum[term_l] += dens_v
                density_term_count[term_l] += 1
                density_term_max[term_l] = max(density_term_max.get(term_l, 0.0), dens_v)

            # 4) Data for n-gram anomalies by section
            section = extract_section(url)
            tokens = [t for t in token_re.findall(merged_text.lower()) if t not in stop_words]
            if tokens:
                section_token_total[section] += len(tokens)
                term_counter = Counter(tokens)
                section_term_counts[section].update(term_counter)
                for term in set(tokens):
                    section_term_pages[section][term].add(url)

        # Cannibalization insights
        for term, urls_set in density_term_urls.items():
            urls_count = len(urls_set)
            if urls_count < 2:
                continue
            avg_density = density_term_sum.get(term, 0.0) / max(1, density_term_count.get(term, 0))
            max_density = density_term_max.get(term, 0.0)
            primary_url = choose_primary_url(urls_set, term)
            sev = "warning"
            if urls_count >= 4 and (max_density >= 6.0 or avg_density >= 3.5):
                sev = "critical"
            elif urls_count <= 2 and avg_density < 2.0:
                sev = "info"
            insights_rows.append([
                "Cannibalization",
                f"{urls_count} URLs",
                "",
                term,
                f"avg_density={round(avg_density, 2)}%, max_density={round(max_density, 2)}%, primary={primary_url}",
                sev,
                f"Differentiate intent; keep '{term}' primary on selected URL and de-optimize siblings.",
                ", ".join(sorted(list(urls_set))[:6]) + (f" | primary={primary_url}" if primary_url else ""),
            ])

        # Section n-gram anomalies
        for section, counts in section_term_counts.items():
            sec_total = max(1, section_token_total.get(section, 0))
            for term, freq in counts.items():
                if freq < 3:
                    continue
                global_freq = unigram_counter.get(term, 0)
                if global_freq <= 0:
                    continue
                global_share = global_freq / max(1, total_unigram_tokens)
                section_share = freq / sec_total
                lift = section_share / max(0.0001, global_share)
                pages_with_term = len(section_term_pages.get(section, {}).get(term, set()))
                if lift < 2.5 or pages_with_term < 2:
                    continue
                sev = "warning" if lift >= 3.0 else "info"
                if lift >= 4.5:
                    sev = "critical"
                insights_rows.append([
                    "Section anomaly",
                    section,
                    "",
                    term,
                    f"share={round(section_share*100,2)}% vs global={round(global_share*100,2)}%, lift={round(lift,2)}",
                    sev,
                    "Проверить, намеренный ли повтор термина для этого кластера разделов.",
                    f"pages={pages_with_term}",
                ])

        for intent, count in sorted(intent_counter.items(), key=lambda x: x[1], reverse=True):
            share = round((count / max(1, len(pages))) * 100.0, 1)
            sev = "warning" if share >= 70.0 else "info"
            insights_rows.append([
                "Обзор интентов",
                "весь сайт",
                intent,
                "",
                f"pages={count}, share={share}%",
                sev,
                "Проверить распределение относительно бизнес-целей и структуры воронки.",
                "",
            ])

        for row in insights_rows:
            if len(row) >= 9:
                continue
            sev_val = str(row[5]).lower() if len(row) > 5 else "info"
            priority = "P3"
            if sev_val == "critical":
                priority = "P1"
            elif sev_val == "warning":
                priority = "P2"
            row.append(priority)

        severity_order = {"critical": 0, "warning": 1, "info": 2, "ok": 3}
        insights_rows.sort(key=lambda row: (severity_order.get(str(row[5]).lower(), 9), str(row[0]), str(row[1])))
        fill_sheet(
            "8c_Keywords_Insights",
            insights_headers,
            insights_rows,
            widths=[18, 46, 16, 28, 34, 10, 48, 56, 10],
        )

        # Optional full-mode optimized deep sheets (no compatibility duplication).
        if str(mode).lower() == "full":
            indexability_headers = [
                "URL", "Статус", "Индексируемо", "Noindex", "Заблокировано robots",
                "Причина индексируемости", "Canonical URL", "Статус canonical",
                "Meta robots", "X-Robots-Tag", "Тип конфликта", "В sitemap", "Риск обнаружения",
                "Скор индексируемости", "Уровень индексируемости", "Дельта индексируемости до цели", "Решение по индексируемости", "Критичность",
            ]
            indexability_rows = []
            for page in pages:
                sev = infer_page_severity(page)
                status_code = int(page.get("status_code") or 0)
                is_indexable = bool(page.get("indexable"))
                noindex_flag = bool(page.get("noindex"))
                blocked_robots = bool(page.get("blocked_by_robots"))
                canonical_status = str(page.get("canonical_status") or "").lower()
                conflict_type = "none"
                if noindex_flag and canonical_status in ("ok", "self"):
                    conflict_type = "noindex_vs_canonical"
                elif blocked_robots and is_indexable:
                    conflict_type = "robots_vs_indexable"
                elif status_code >= 400 and is_indexable:
                    conflict_type = "status_vs_indexable"
                in_sitemap = bool(page.get("in_sitemap", False))
                discovery_risk = "✅" if (is_indexable and int(page.get("incoming_internal_links", 0) or 0) < 2) else "❌"
                score = 100
                if status_code >= 400:
                    score -= 40
                if not is_indexable:
                    score -= 30
                if noindex_flag:
                    score -= 20
                if blocked_robots:
                    score -= 20
                if canonical_status in ("missing", "external", "invalid"):
                    score -= 10
                score = max(0, score)
                indexability_level = score_level(score, high=90.0, medium=75.0)
                indexability_rows.append([
                    page.get("url", ""),
                    status_code,
                    is_indexable,
                    noindex_flag,
                    blocked_robots,
                    page.get("indexability_reason", ""),
                    page.get("canonical", ""),
                    page.get("canonical_status", ""),
                    page.get("meta_robots", ""),
                    page.get("x_robots_tag", ""),
                    conflict_type,
                    in_sitemap,
                    discovery_risk,
                    score,
                    indexability_level,
                    delta_to_target(score, 95.0),
                    page_solution("technical", page),
                    sev,
                ])
            sort_rows(indexability_rows, 13, reverse=False)
            fill_sheet(
                "9_Indexability",
                indexability_headers,
                indexability_rows,
                severity_idx=17,
                widths=[52, 10, 10, 10, 14, 20, 30, 16, 20, 20, 20, 10, 12, 14, 12, 14, 46, 10],
                score_idx=13,
            )

            structured_headers = [
                "URL", "Всего structured", "JSON-LD", "Microdata", "RDFa",
                "Типы structured", "Hreflang", "Хлебные крошки",
                "FAQ schema", "Product schema", "Article schema",
                "Конфликт schema", "Подходит для rich result", "Критичные schema-ошибки", "Кол-во типовых ошибок", "Коды типовых ошибок",
                "Покрытие structured, %", "Structured-дельта до цели", "Решение по structured", "Критичность",
            ]
            structured_rows = []
            for page in pages:
                sev = infer_page_severity(page)
                detail = page.get("structured_data_detail") or {}
                types = [str(t).lower() for t in (page.get("structured_types") or [])]
                has_faq = any("faq" in t for t in types)
                has_product = any("product" in t for t in types)
                has_article = any("article" in t for t in types)
                page_type_l = str(page.get("page_type", "")).lower()
                schema_mismatch = "❌"
                if "product" in page_type_l and not has_product:
                    schema_mismatch = "✅"
                if "article" in page_type_l and not has_article:
                    schema_mismatch = "✅"
                if "faq" in page_type_l and not has_faq:
                    schema_mismatch = "✅"
                error_codes = [str(c).lower() for c in (page.get("structured_error_codes") or [])]
                critical_schema_errors = [c for c in error_codes if any(k in c for k in ("missing_price", "missing_availability", "missing_name", "missing_offers", "invalid_type"))]
                missing_fields = [str(f).lower() for f in (page.get("structured_missing_fields") or [])]
                expected_by_type = []
                if "product" in page_type_l:
                    expected_by_type = ["product"]
                elif "article" in page_type_l or "blog" in page_type_l or "news" in page_type_l:
                    expected_by_type = ["article", "blogposting", "newsarticle"]
                elif "faq" in page_type_l:
                    expected_by_type = ["faq"]
                if expected_by_type and not any(any(exp in t for exp in expected_by_type) for t in types):
                    critical_schema_errors.append(f"missing_schema_for_page_type:{page_type_l}")
                if "product" in page_type_l:
                    if any(x in missing_fields for x in ("price", "offers.price")):
                        critical_schema_errors.append("missing_price")
                    if any(x in missing_fields for x in ("availability", "offers.availability")):
                        critical_schema_errors.append("missing_availability")
                    if any(x in missing_fields for x in ("name",)):
                        critical_schema_errors.append("missing_name")
                if "article" in page_type_l or "blog" in page_type_l or "news" in page_type_l:
                    if any(x in missing_fields for x in ("headline",)):
                        critical_schema_errors.append("missing_headline")
                    if any(x in missing_fields for x in ("datepublished", "datemodified")):
                        critical_schema_errors.append("missing_article_date")
                if "faq" in page_type_l and any(x in missing_fields for x in ("mainentity", "acceptedanswer")):
                    critical_schema_errors.append("missing_faq_entities")
                critical_schema_errors = sorted(list(dict.fromkeys(critical_schema_errors)))
                rich_result_eligible = "✅" if (len(types) > 0 and len(critical_schema_errors) == 0 and int(page.get("structured_errors_count", 0) or 0) <= 2) else "❌"
                coverage = 0.0
                coverage += min(70.0, float(page.get("structured_data", 0) or 0) * 20.0)
                if int(page.get("hreflang_count", 0) or 0) > 0:
                    coverage += 15.0
                if bool(page.get("breadcrumbs")):
                    coverage += 15.0
                structured_rows.append([
                    page.get("url", ""),
                    page.get("structured_data", 0),
                    detail.get("json_ld", 0),
                    detail.get("microdata", 0),
                    detail.get("rdfa", 0),
                    ", ".join((page.get("structured_types") or [])[:8]),
                    page.get("hreflang_count", 0),
                    page.get("breadcrumbs", ""),
                    has_faq,
                    has_product,
                    has_article,
                    schema_mismatch,
                    rich_result_eligible,
                    ", ".join(critical_schema_errors[:5]),
                    page.get("structured_errors_count", 0),
                    ", ".join((page.get("structured_error_codes") or [])[:8]),
                    round(min(100.0, coverage), 1),
                    delta_to_target(round(min(100.0, coverage), 1), 70.0),
                    page_solution("structured", page),
                    sev,
                ])
            sort_rows(structured_rows, 1, reverse=True)
            fill_sheet(
                "10_StructuredData",
                structured_headers,
                structured_rows,
                severity_idx=19,
                widths=[50, 12, 10, 10, 8, 40, 10, 12, 10, 12, 12, 12, 12, 36, 12, 12, 36, 18, 14, 48, 10],
                score_idx=16,
            )

            trust_eeat_headers = [
                "URL", "Trust-скор", "EEAT-скор", "Экспертиза", "Авторитет",
                "Надежность", "Опыт", "Инфо об авторе", "Контакты", "Юридическая инфо", "Отзывы", "Бейджи",
                "Редакционная политика", "Указанные источники", "EEAT-матрица", "Trust-gap", "Кол-во trust-доказательств", "Trust-дельта до цели", "Решение Trust+EEAT", "Критичность",
            ]
            trust_eeat_rows = []
            for page in pages:
                sev = infer_page_severity(page)
                comp = page.get("eeat_components") or {}
                trust_score = to_float(page.get("trust_score"), 0.0)
                eeat_score = to_float(page.get("eeat_score"), 0.0)
                trust_gap = round(max(0.0, 70.0 - ((trust_score + eeat_score) / 2.0)), 1)
                trust_evidence_count = sum(
                    1 for flag in (
                        page.get("has_author_info", False),
                        page.get("has_contact_info", False),
                        page.get("has_legal_docs", False),
                        page.get("has_reviews", False),
                        page.get("trust_badges", False),
                    )
                    if bool(flag)
                )
                editorial_policy = bool(page.get("editorial_policy_present", False))
                sources_cited = bool(page.get("sources_cited", page.get("citations_present", False)))
                eeat_matrix = f"E:{comp.get('expertise', 0)}|A:{comp.get('authoritativeness', 0)}|T:{comp.get('trustworthiness', 0)}|Ex:{comp.get('experience', 0)}"
                trust_eeat_rows.append([
                    page.get("url", ""),
                    trust_score,
                    eeat_score,
                    comp.get("expertise", ""),
                    comp.get("authoritativeness", ""),
                    comp.get("trustworthiness", ""),
                    comp.get("experience", ""),
                    page.get("has_author_info", ""),
                    page.get("has_contact_info", ""),
                    page.get("has_legal_docs", ""),
                    page.get("has_reviews", ""),
                    page.get("trust_badges", ""),
                    editorial_policy,
                    sources_cited,
                    eeat_matrix,
                    trust_gap,
                    trust_evidence_count,
                    delta_to_target(trust_score, 75.0),
                    f"{page_solution('eeat', page)}; {page_solution('trust', page)}",
                    sev,
                ])
            sort_rows(trust_eeat_rows, 1, reverse=False)
            fill_sheet(
                "11_Trust_EEAT",
                trust_eeat_headers,
                trust_eeat_rows,
                severity_idx=19,
                widths=[46, 10, 10, 10, 10, 12, 10, 10, 10, 10, 10, 10, 12, 10, 22, 12, 12, 14, 54, 10],
                score_idx=1,
            )

            topics_headers = [
                "URL", "Тема", "Хаб", "Входящие ссылки", "Исходящие внутренние ссылки", "Кол-во семантических ссылок",
                "Рекомендуемые ссылки", "Детали семантических ссылок", "Топ терминов", "Топ ключей", "Скор тематической глубины", "Тематическая дельта до цели",
                "Полнота кластера", "Сиротский узел темы", "Перегруз хаба", "Консистентность сущностей",
                "Решение по темам", "Критичность",
            ]
            topics_rows = []
            semantic_by_source: Dict[str, List[Dict[str, Any]]] = {}
            for row in (pipeline.get("semantic_linking_map") or []):
                semantic_by_source.setdefault(row.get("source_url", ""), []).append(row)
            for page in pages:
                sev = infer_page_severity(page)
                src = page.get("url", "")
                semantic_rows = semantic_by_source.get(src, [])
                if not semantic_rows:
                    semantic_rows = page.get("semantic_links") or []
                semantic_summary = []
                for item in semantic_rows[:4]:
                    target = item.get("target_url", "")
                    topic = item.get("topic") or item.get("suggested_anchor") or ""
                    reason = item.get("reason", "")
                    semantic_summary.append(f"[{topic}] {target} {reason}".strip())
                semantic_count = len(semantic_rows)
                outgoing_internal = int(page.get("outgoing_internal_links", 0) or 0)
                incoming_internal = int(page.get("incoming_internal_links", 0) or 0)
                topical_depth = min(100.0, semantic_count * 20.0 + outgoing_internal * 3.0 + (15.0 if page.get("topic_hub") else 0.0))
                tfidf_terms = tfidf_by_url.get(src, page.get("top_terms", [])) or []
                cluster_completeness = min(100, int((semantic_count * 25) + (incoming_internal * 5)))
                orphan_topic_node = "✅" if (incoming_internal == 0 and not bool(page.get("topic_hub", False))) else "❌"
                hub_overload = "✅" if (bool(page.get("topic_hub", False)) and outgoing_internal > 150) else "❌"
                entity_consistency = round(to_float(page.get("entity_consistency_score"), 0.0), 1)
                if entity_consistency <= 0.0:
                    entity_consistency = round(min(100.0, max(0.0, to_float(page.get("lexical_diversity", 0.0), 0.0) * 120.0)), 1)
                topics_rows.append([
                    src,
                    page.get("topic_label", ""),
                    page.get("topic_hub", ""),
                    page.get("incoming_internal_links", 0),
                    outgoing_internal,
                    semantic_count,
                    "; ".join([item.get("target_url", "") for item in semantic_rows[:5]]),
                    "\n".join(semantic_summary),
                    ", ".join(tfidf_terms[:8]),
                    ", ".join((page.get("top_keywords") or [])[:8]),
                    round(topical_depth, 1),
                    delta_to_target(round(topical_depth, 1), 70.0),
                    cluster_completeness,
                    orphan_topic_node,
                    hub_overload,
                    entity_consistency,
                    page_solution("topics", page),
                    sev,
                ])
            sort_rows(topics_rows, 10, reverse=False)
            fill_sheet(
                "12_Topics_Semantics",
                topics_headers,
                topics_rows,
                severity_idx=17,
                widths=[42, 16, 10, 12, 14, 14, 42, 58, 36, 36, 14, 14, 12, 10, 10, 12, 42, 10],
                score_idx=10,
            )

            ai_headers = [
                "URL", "AI-маркеры", "Список AI-маркеров", "Пример маркеров",
                "Плотность AI /1k", "AI risk-скор", "Уровень AI-риска", "Защита от false-positive",
                "Тип страницы", "Toxicity-скор", "Доля «воды»", "Стилевые маркеры", "Disclaimer-маркеры", "Transition-маркеры", "Hedging-маркеры",
                "Уверенность false-positive", "AI-риск выше порога", "Подсказка по humanization", "Критичность",
            ]
            ai_rows = []
            for page in pages:
                sev = infer_page_severity(page)
                markers_count = int(page.get("ai_markers_count", 0) or 0)
                toxicity = to_float(page.get("toxicity_score"), 0.0)
                filler_ratio = to_float(page.get("filler_ratio"), 0.0)
                ai_density = to_float(page.get("ai_markers_density_1k"), 0.0)
                ai_risk = to_float(page.get("ai_risk_score"), min(100.0, markers_count * 6.0 + toxicity * 0.7 + filler_ratio * 0.5))
                risk_level = str(page.get("ai_risk_level") or ("high" if ai_risk >= 70 else "medium" if ai_risk >= 40 else "low"))
                marker_list = [str(m).lower() for m in (page.get("ai_markers_list") or [])]
                style_markers = sum(1 for m in marker_list if any(k in m for k in ("tone", "style", "formatted", "structured", "generic")))
                disclaimer_markers = sum(1 for m in marker_list if any(k in m for k in ("as an ai", "не могу", "cannot", "i'm unable", "модель")))
                transition_markers = sum(1 for m in marker_list if any(k in m for k in ("however", "moreover", "furthermore", "в заключение", "таким образом")))
                hedging_markers = sum(1 for m in marker_list if any(k in m for k in ("may", "might", "could", "возможно", "вероятно")))
                fp_guard = bool(page.get("ai_false_positive_guard", False))
                fp_confidence = max(0, min(100, int(70 + (15 if fp_guard else 0) - (ai_risk * 0.4))))
                ai_rows.append([
                    page.get("url", ""),
                    markers_count,
                    ", ".join((page.get("ai_markers_list") or [])[:12]) or "none",
                    page.get("ai_marker_sample", "") or "No text sample available",
                    round(ai_density, 2),
                    round(ai_risk, 1),
                    risk_level,
                    page.get("ai_false_positive_guard", False),
                    page.get("page_type", ""),
                    toxicity,
                    filler_ratio,
                    style_markers,
                    disclaimer_markers,
                    transition_markers,
                    hedging_markers,
                    fp_confidence,
                    round(max(0.0, ai_risk - 40.0), 1),
                    page_solution("content", page),
                    sev,
                ])
            sort_rows(ai_rows, 1, reverse=True)
            fill_sheet(
                "13_AI_Markers",
                ai_headers,
                ai_rows,
                severity_idx=18,
                widths=[48, 10, 48, 56, 12, 12, 12, 12, 12, 12, 10, 10, 10, 10, 10, 12, 14, 52, 10],
                score_idx=5,
            )

            crawl_budget_headers = [
                "URL", "Глубина пути", "Параметры URL", "Риск crawl budget", "Редиректы", "Статус", "Индексируемо",
                "Входящие ссылки", "Исходящие внутренние", "Близкие дубли", "Дубли обхода сверх цели", "Группа параметров", "Скор потерь обхода", "Риск глубокой индексируемости",
                "Решение по crawl budget", "Критичность",
            ]
            crawl_budget_rows = []
            for page in pages:
                sev = infer_page_severity(page)
                url_value = str(page.get("url", "") or "")
                query = ""
                if "?" in url_value:
                    query = url_value.split("?", 1)[1]
                params = [p.split("=", 1)[0].lower() for p in query.split("&") if p]
                param_group = "none"
                if any(p.startswith("utm_") for p in params):
                    param_group = "utm"
                elif any(p in ("sort", "order", "filter", "page", "p") for p in params):
                    param_group = "filter/sort"
                crawl_waste_score = min(
                    100,
                    int(
                        (int(page.get("url_params_count", 0) or 0) * 15)
                        + (int(page.get("redirect_count", 0) or 0) * 10)
                        + (int(page.get("near_duplicate_count", 0) or 0) * 8)
                    ),
                )
                deep_indexable_risk = "✅" if (bool(page.get("indexable", False)) and int(page.get("path_depth", 0) or 0) >= 4) else "❌"
                crawl_budget_rows.append(
                    [
                        url_value,
                        page.get("path_depth", 0),
                        page.get("url_params_count", 0),
                        page.get("crawl_budget_risk", ""),
                        page.get("redirect_count", 0),
                        page.get("status_code", ""),
                        page.get("indexable", ""),
                        page.get("incoming_internal_links", 0),
                        page.get("outgoing_internal_links", 0),
                        page.get("near_duplicate_count", 0),
                        max(0, int(page.get("near_duplicate_count", 0) or 0) - 0),
                        param_group,
                        crawl_waste_score,
                        deep_indexable_risk,
                        page_solution("technical", page),
                        sev,
                    ]
                )
            sort_rows(crawl_budget_rows, 9, reverse=True)
            fill_sheet(
                "CrawlBudget",
                crawl_budget_headers,
                crawl_budget_rows,
                severity_idx=15,
                widths=[52, 10, 10, 16, 10, 10, 10, 12, 12, 12, 14, 16, 12, 12, 52, 10],
                score_idx=9,
            )

            raw_issue_headers = ["Критичность", "URL", "Код", "Категория", "Первая вкладка", "Dedupe hash", "Ответственный", "Заголовок", "Детали", "Затронуто", "Рекомендация"]
            raw_issue_rows = []
            seen_raw = set()
            for issue in issues:
                if not isinstance(issue, dict):
                    continue
                severity = (issue.get("severity") or "info").lower()
                url = issue.get("url", "")
                code = issue.get("code", "")
                details = issue.get("details", "")
                details_text = str(details or "")
                fingerprint = (severity, url, code, details)
                if fingerprint in seen_raw:
                    continue
                seen_raw.add(fingerprint)
                code_text = str(code)
                category = code_text.split("_", 1)[0] if "_" in code_text else code_text
                dedupe_hash = hashlib.md5(f"{severity}|{url}|{code}|{details_text[:240]}".encode("utf-8", errors="ignore")).hexdigest()[:10]
                raw_issue_rows.append([
                    severity,
                    url,
                    code,
                    category,
                    detect_issue_tab(code),
                    dedupe_hash,
                    owner_hint_by_code(code),
                    issue.get("title", ""),
                    details_text,
                    issue.get("selector") or issue.get("path") or issue.get("field") or "",
                    issue.get("recommendation") or issue.get("fix") or issue_recommendation(issue),
                ])
            raw_issue_rows.sort(key=lambda r: (0 if str(r[0]).lower() == "critical" else 1 if str(r[0]).lower() == "warning" else 2, str(r[1]), str(r[2])))
            fill_sheet(
                "14_Issues_Raw",
                raw_issue_headers,
                raw_issue_rows,
                severity_idx=0,
                widths=[10, 56, 24, 16, 22, 14, 14, 28, 50, 24, 46],
            )

            action_plan_headers = [
                "Приоритет", "Код проблемы", "Макс. критичность", "Затронуто страниц", "Доля %", "Критично", "Предупреждение", "Инфо",
                "Импакт-скор", "Трудоемкость", "ETA", "Подсказка владельца", "Ожидаемый эффект", "Кластер первопричин", "Зависит от кодов",
                "Зависимость", "Потенциал batch-фикса", "ROI-скор", "Бакет спринта", "Репрезентативные URL", "Рекомендация",
            ]
            grouped: Dict[str, Dict[str, Any]] = {}
            for issue in issues:
                code = str(issue.get("code") or "unknown_issue")
                sev = str(issue.get("severity") or "info").lower()
                url = str(issue.get("url") or "").strip()
                node = grouped.setdefault(
                    code,
                    {
                        "critical": 0,
                        "warning": 0,
                        "info": 0,
                        "urls": set(),
                        "sample": issue,
                    },
                )
                if sev in ("critical", "warning", "info"):
                    node[sev] += 1
                else:
                    node["info"] += 1
                if url:
                    node["urls"].add(url)

            effort_map = {
                "critical": "M",
                "warning": "S",
                "info": "S",
            }
            action_rows = []
            for code, node in grouped.items():
                critical = int(node.get("critical", 0))
                warning = int(node.get("warning", 0))
                info = int(node.get("info", 0))
                if critical > 0:
                    top_severity = "critical"
                elif warning > 0:
                    top_severity = "warning"
                else:
                    top_severity = "info"
                affected_pages = len(node.get("urls", set()))
                share_pct = round((affected_pages / float(total_pages)) * 100.0, 1) if total_pages else 0.0
                impact_score = round((critical * 3.0) + (warning * 2.0) + info + (share_pct / 10.0), 1)
                if impact_score >= 40:
                    expected_lift = "High"
                elif impact_score >= 15:
                    expected_lift = "Medium"
                else:
                    expected_lift = "Low"
                eta = "1-3d" if impact_score >= 40 else ("3-7d" if impact_score >= 15 else "backlog")
                owner_hint = owner_hint_by_code(code)
                cause_cluster = root_cause_cluster(code)
                depends_on_codes = dependency_codes_for(code)
                dependency = "Dev deploy" if owner_hint in ("SEO+Dev", "Dev+Infra") else "Content approval"
                batch_fix_potential = "✅" if any(x in str(code).lower() for x in ("template", "title", "meta", "schema", "canonical", "robots")) else "❌"
                effort_weight = 1.3 if effort_map.get(top_severity, "S") == "M" else 1.0
                roi_score = round(impact_score / effort_weight, 1)
                sprint_bucket = "Сейчас" if roi_score >= 25 else ("Далее" if roi_score >= 10 else "Позже")
                rec = issue_recommendation(node.get("sample") or {"code": code, "severity": top_severity})
                if rec.startswith("Критично: "):
                    rec = rec[len("Критично: "):]
                elif rec.startswith("Предупреждение: "):
                    rec = rec[len("Предупреждение: "):]
                elif rec.startswith("Инфо: "):
                    rec = rec[len("Инфо: "):]
                action_rows.append(
                    [
                        "",  # filled after sorting
                        code,
                        top_severity,
                        affected_pages,
                        share_pct,
                        critical,
                        warning,
                        info,
                        impact_score,
                        effort_map.get(top_severity, "S"),
                        eta,
                        owner_hint,
                        expected_lift,
                        cause_cluster,
                        depends_on_codes,
                        dependency,
                        batch_fix_potential,
                        roi_score,
                        sprint_bucket,
                        ", ".join(list(node.get("urls", set()))[:5]),
                        rec,
                    ]
                )
            action_rows.sort(key=lambda row: to_float(row[8], 0.0), reverse=True)
            for idx, row in enumerate(action_rows, start=1):
                row[0] = idx
            fill_sheet(
                "15_ActionPlan",
                action_plan_headers,
                action_rows,
                severity_idx=2,
                widths=[10, 26, 12, 14, 10, 10, 10, 10, 12, 10, 10, 14, 12, 16, 22, 16, 14, 10, 12, 52, 58],
                score_idx=8,
            )

        matrix_start = 25
        health_values: List[float] = []
        weighted_health_sum = 0.0
        total_weight = 0.0
        sheet_impact_weights = {
            "2_OnPage+Structured": 1.10,
            "3_Technical": 1.25,
            "4_Content+AI": 1.15,
            "5_LinkGraph": 1.00,
            "6_Images+External": 0.95,
            "7_HierarchyErrors": 1.00,
            "8_Keywords": 1.05,
            "8b_Keywords_Summary": 0.85,
            "8c_Keywords_Insights": 0.95,
            "9_Indexability": 1.20,
            "10_StructuredData": 1.05,
            "11_Trust_EEAT": 0.90,
            "12_Topics_Semantics": 0.95,
            "13_AI_Markers": 0.90,
            "CrawlBudget": 1.05,
            "14_Issues_Raw": 1.00,
            "15_ActionPlan": 1.00,
        }
        for idx, stat in enumerate(sheet_stats, start=matrix_start):
            critical = int(stat.get("critical", 0))
            warning = int(stat.get("warning", 0))
            info = int(stat.get("info", 0))
            avg_score = stat.get("avg_score", "")
            impact_weight = float(sheet_impact_weights.get(str(stat.get("sheet", "")), 1.0))
            issue_load = round(((critical * 3.0) + (warning * 2.0) + info) / float(total_pages), 2)
            weighted_issue_load = issue_load * impact_weight
            if avg_score == "":
                health = round(max(0.0, 100.0 - (weighted_issue_load * 10.0)), 1)
            else:
                health = round(max(0.0, min(100.0, 100.0 - (weighted_issue_load * 10.0) + ((float(avg_score) - 50.0) / 2.2))), 1)
            health_values.append(health)
            weighted_health_sum += health * impact_weight
            total_weight += impact_weight
            ws.cell(row=idx, column=1, value=self._sanitize_cell_value(stat.get("sheet", "")))
            ws.cell(row=idx, column=2, value=critical)
            ws.cell(row=idx, column=3, value=warning)
            ws.cell(row=idx, column=4, value=info)
            ws.cell(row=idx, column=5, value=issue_load)
            ws.cell(row=idx, column=6, value=avg_score)
            ws.cell(row=idx, column=7, value=health)
            ws.cell(row=idx, column=8, value=impact_weight)
            for col in range(1, 9):
                self._apply_style(ws.cell(row=idx, column=col), cell_style)
            if critical > 0:
                self._apply_row_severity_fill(ws, idx, 1, 8, "critical")
        ws["A22"] = "Platform Health Index"
        ws["B22"] = round(weighted_health_sum / max(1.0, total_weight), 1) if health_values else ""
        ws["A22"].font = Font(bold=True)
        ws["B22"].font = Font(bold=True)

        filepath = os.path.join(self.reports_dir, f"{task_id}.xlsx")
        self._save_workbook(wb, filepath)
        wb.close()
        return filepath

    def generate_onpage_report(self, task_id: str, data: Dict[str, Any]) -> str:
        """Generate XLSX report for onpage_audit."""
        wb = Workbook()
        header_style = self._create_header_style()
        cell_style = self._create_cell_style()

        results = data.get("results", {}) or {}
        summary = results.get("summary", {}) or {}
        url = data.get("url", "н/д")

        ws = wb.active
        ws.title = "Сводка"
        ws["A1"] = "Отчет OnPage-аудита"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:E1")

        summary_rows = [
            ("URL", url),
            ("Движок", results.get("engine", "onpage-v1")),
            ("Оценка", results.get("score", summary.get("score", 0))),
            ("Spam-оценка", summary.get("spam_score", (results.get("scores", {}) or {}).get("spam_score", 0))),
            ("Оценка покрытия КС", summary.get("keyword_coverage_score", (results.get("scores", {}) or {}).get("keyword_coverage_score", 0))),
            ("Покрытие КС %", summary.get("keyword_coverage_pct", (results.get("keyword_coverage", {}) or {}).get("coverage_pct", 0))),
            ("AI-риск (композит)", summary.get("ai_risk_composite", (results.get("scores", {}) or {}).get("ai_risk_composite", 0))),
            ("Критично", summary.get("critical_issues", 0)),
            ("Предупреждение", summary.get("warning_issues", 0)),
            ("Инфо", summary.get("info_issues", 0)),
            ("HTTP-статус", results.get("status_code", "н/д")),
            ("Финальный URL", results.get("final_url", url)),
            ("Язык", results.get("language", "auto")),
        ]
        row = 3
        for key, value in summary_rows:
            self._apply_style(ws.cell(row=row, column=1, value=key), header_style)
            self._apply_style(ws.cell(row=row, column=2, value=value), cell_style)
            row += 1
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 120

        meta_ws = wb.create_sheet("Meta")
        meta_headers = ["Поле", "Значение"]
        for col, header in enumerate(meta_headers, 1):
            self._apply_style(meta_ws.cell(row=1, column=col, value=header), header_style)
        title_meta = results.get("title", {}) or {}
        desc_meta = results.get("description", {}) or {}
        h1_meta = results.get("h1", {}) or {}
        meta_rows = [
            ("Title", title_meta.get("text", "")),
            ("Длина title", title_meta.get("length", 0)),
            ("Description", desc_meta.get("text", "")),
            ("Длина description", desc_meta.get("length", 0)),
            ("Количество H1", h1_meta.get("count", 0)),
            ("Значения H1", ", ".join(h1_meta.get("values", []) or [])),
        ]
        for row_idx, (key, value) in enumerate(meta_rows, start=2):
            self._apply_style(meta_ws.cell(row=row_idx, column=1, value=key), cell_style)
            self._apply_style(meta_ws.cell(row=row_idx, column=2, value=value), cell_style)
        meta_ws.column_dimensions["A"].width = 24
        meta_ws.column_dimensions["B"].width = 120

        profile_ws = wb.create_sheet("Профиль контента")
        profile_headers = ["Параметр", "Значение", "Статус"]
        for col, header in enumerate(profile_headers, 1):
            self._apply_style(profile_ws.cell(row=1, column=col, value=header), header_style)
        for row_idx, item in enumerate(results.get("parameter_values", []) or [], start=2):
            vals = [item.get("parameter", ""), item.get("value", ""), str(item.get("status", "info")).upper()]
            for col, val in enumerate(vals, 1):
                self._apply_style(profile_ws.cell(row=row_idx, column=col, value=val), cell_style)
        for col, width in enumerate([34, 28, 14], 1):
            profile_ws.column_dimensions[get_column_letter(col)].width = width

        kw_ws = wb.create_sheet("Keywords")
        kw_headers = ["Ключевое слово", "Частота", "Плотность %", "В title", "В description", "В H1", "Статус"]
        for col, header in enumerate(kw_headers, 1):
            self._apply_style(kw_ws.cell(row=1, column=col, value=header), header_style)
        for row_idx, item in enumerate(results.get("keywords", []) or [], start=2):
            values = [
                item.get("keyword", ""),
                item.get("occurrences", 0),
                item.get("density_pct", 0),
                "Да" if item.get("in_title") else "Нет",
                "Да" if item.get("in_description") else "Нет",
                "Да" if item.get("in_h1") else "Нет",
                str(item.get("status", "ok")).upper(),
            ]
            for col, value in enumerate(values, 1):
                self._apply_style(kw_ws.cell(row=row_idx, column=col, value=value), cell_style)
        kw_ws.freeze_panes = "A2"
        kw_ws.auto_filter.ref = "A1:G1"
        for col, width in enumerate([30, 12, 12, 12, 15, 10, 12], 1):
            kw_ws.column_dimensions[get_column_letter(col)].width = width

        issue_ws = wb.create_sheet("Проблемы")
        issue_headers = ["Критичность", "Код", "Проблема", "Детали"]
        for col, header in enumerate(issue_headers, 1):
            self._apply_style(issue_ws.cell(row=1, column=col, value=header), header_style)
        for row_idx, issue in enumerate(results.get("issues", []) or [], start=2):
            severity = str(issue.get("severity", "info")).lower()
            values = [
                severity.upper(),
                issue.get("code", ""),
                issue.get("title", ""),
                issue.get("details", ""),
            ]
            for col, value in enumerate(values, 1):
                self._apply_style(issue_ws.cell(row=row_idx, column=col, value=value), cell_style)
            self._apply_row_severity_fill(issue_ws, row_idx, 1, len(issue_headers), severity)
        issue_ws.freeze_panes = "A2"
        issue_ws.auto_filter.ref = "A1:D1"
        for col, width in enumerate([12, 24, 40, 110], 1):
            issue_ws.column_dimensions[get_column_letter(col)].width = width

        rec_ws = wb.create_sheet("Рекомендации")
        self._apply_style(rec_ws.cell(row=1, column=1, value="Рекомендация"), header_style)
        for row_idx, rec in enumerate(results.get("recommendations", []) or [], start=2):
            self._apply_style(rec_ws.cell(row=row_idx, column=1, value=rec), cell_style)
        rec_ws.column_dimensions["A"].width = 150

        terms_ws = wb.create_sheet("Топ терминов")
        term_headers = ["Термин", "Частота", "Доля %"]
        for col, header in enumerate(term_headers, 1):
            self._apply_style(terms_ws.cell(row=1, column=col, value=header), header_style)
        for row_idx, item in enumerate(results.get("top_terms", []) or [], start=2):
            vals = [item.get("term", ""), item.get("count", 0), item.get("pct", 0)]
            for col, value in enumerate(vals, 1):
                self._apply_style(terms_ws.cell(row=row_idx, column=col, value=value), cell_style)
        for col, width in enumerate([32, 14, 14], 1):
            terms_ws.column_dimensions[get_column_letter(col)].width = width

        tech_ws = wb.create_sheet("Technical")
        tech_headers = ["Сигнал", "Значение"]
        for col, header in enumerate(tech_headers, 1):
            self._apply_style(tech_ws.cell(row=1, column=col, value=header), header_style)
        technical = results.get("technical", {}) or {}
        tech_rows = [
            ("Canonical href", technical.get("canonical_href", "")),
            ("Canonical self", "Да" if technical.get("canonical_is_self") else "Нет"),
            ("Meta robots", technical.get("robots", "")),
            ("Noindex", "Да" if technical.get("noindex") else "Нет"),
            ("Nofollow", "Да" if technical.get("nofollow") else "Нет"),
            ("Viewport", technical.get("viewport", "")),
            ("HTML lang", technical.get("lang", "")),
            ("Hreflang tags", technical.get("hreflang_count", 0)),
            ("Schema blocks", technical.get("schema_count", 0)),
        ]
        for row_idx, (key, value) in enumerate(tech_rows, start=2):
            self._apply_style(tech_ws.cell(row=row_idx, column=1, value=key), cell_style)
            self._apply_style(tech_ws.cell(row=row_idx, column=2, value=value), cell_style)
        tech_ws.column_dimensions["A"].width = 30
        tech_ws.column_dimensions["B"].width = 120

        quality_ws = wb.create_sheet("Quality")
        quality_headers = ["Метрика", "Значение"]
        for col, header in enumerate(quality_headers, 1):
            self._apply_style(quality_ws.cell(row=1, column=col, value=header), header_style)
        links = results.get("links", {}) or {}
        media = results.get("media", {}) or {}
        readability = results.get("readability", {}) or {}
        quality_rows = [
            ("Total links", links.get("links_total", 0)),
            ("Internal links", links.get("internal_links", 0)),
            ("External links", links.get("external_links", 0)),
            ("Nofollow links", links.get("nofollow_links", 0)),
            ("Empty anchor links", links.get("empty_anchor_links", 0)),
            ("Images total", media.get("images_total", 0)),
            ("Images missing alt", media.get("images_missing_alt", 0)),
            ("Sentences", readability.get("sentences_count", 0)),
            ("Avg sentence len", readability.get("avg_sentence_len", 0)),
            ("Long sentence ratio", readability.get("long_sentence_ratio", 0)),
            ("Lexical diversity", readability.get("lexical_diversity", 0)),
        ]
        spam_metrics = results.get("spam_metrics", {}) or {}
        quality_rows.extend(
            [
                ("Stopword ratio", spam_metrics.get("stopword_ratio", 0)),
                ("Content/HTML ratio", spam_metrics.get("content_html_ratio", 0)),
                ("Uppercase ratio", spam_metrics.get("uppercase_ratio", 0)),
                ("Punctuation ratio", spam_metrics.get("punctuation_ratio", 0)),
                ("Duplicate sentences", spam_metrics.get("duplicate_sentences", 0)),
                ("Duplicate sentence ratio", spam_metrics.get("duplicate_sentence_ratio", 0)),
                ("Top bigram share %", spam_metrics.get("top_bigram_pct", 0)),
                ("Top trigram share %", spam_metrics.get("top_trigram_pct", 0)),
            ]
        )
        for row_idx, (key, value) in enumerate(quality_rows, start=2):
            self._apply_style(quality_ws.cell(row=row_idx, column=1, value=key), cell_style)
            self._apply_style(quality_ws.cell(row=row_idx, column=2, value=value), cell_style)
        quality_ws.column_dimensions["A"].width = 30
        quality_ws.column_dimensions["B"].width = 36

        ngram_ws = wb.create_sheet("Ngrams")
        ngram_headers = ["Тип", "Термин", "Частота", "Доля %"]
        for col, header in enumerate(ngram_headers, 1):
            self._apply_style(ngram_ws.cell(row=1, column=col, value=header), header_style)
        ngrams = results.get("ngrams", {}) or {}
        row_idx = 2
        for item in (ngrams.get("bigrams", []) or [])[:20]:
            values = ["биграмма", item.get("term", ""), item.get("count", 0), item.get("pct", 0)]
            for col, value in enumerate(values, 1):
                self._apply_style(ngram_ws.cell(row=row_idx, column=col, value=value), cell_style)
            row_idx += 1
        for item in (ngrams.get("trigrams", []) or [])[:20]:
            values = ["триграмма", item.get("term", ""), item.get("count", 0), item.get("pct", 0)]
            for col, value in enumerate(values, 1):
                self._apply_style(ngram_ws.cell(row=row_idx, column=col, value=value), cell_style)
            row_idx += 1
        for col, width in enumerate([12, 60, 12, 12], 1):
            ngram_ws.column_dimensions[get_column_letter(col)].width = width

        schema_ws = wb.create_sheet("Schema_OG")
        schema_headers = ["Поле", "Значение"]
        for col, header in enumerate(schema_headers, 1):
            self._apply_style(schema_ws.cell(row=1, column=col, value=header), header_style)
        schema = results.get("schema", {}) or {}
        og = results.get("opengraph", {}) or {}
        schema_rows = [
            ("JSON-LD blocks", schema.get("json_ld_blocks", 0)),
            ("Valid JSON-LD", schema.get("json_ld_valid_blocks", 0)),
            ("Microdata items", schema.get("microdata_items", 0)),
            ("RDFa items", schema.get("rdfa_items", 0)),
            ("Schema types", ", ".join([x.get("type", "") for x in (schema.get("types", []) or [])[:10]])),
            ("OG tags count", og.get("tags_count", 0)),
            ("OG required present", og.get("required_present_count", 0)),
            ("OG missing", ", ".join(og.get("required_missing", []) or [])),
        ]
        for row_idx, (k, v) in enumerate(schema_rows, start=2):
            self._apply_style(schema_ws.cell(row=row_idx, column=1, value=k), cell_style)
            self._apply_style(schema_ws.cell(row=row_idx, column=2, value=v), cell_style)
        schema_ws.column_dimensions["A"].width = 32
        schema_ws.column_dimensions["B"].width = 110

        links_terms_ws = wb.create_sheet("Термины анкоров")
        self._apply_style(links_terms_ws.cell(row=1, column=1, value="Термин"), header_style)
        self._apply_style(links_terms_ws.cell(row=1, column=2, value="Частота"), header_style)
        for row_idx, item in enumerate(results.get("link_anchor_terms", []) or [], start=2):
            self._apply_style(links_terms_ws.cell(row=row_idx, column=1, value=item.get("term", "")), cell_style)
            self._apply_style(links_terms_ws.cell(row=row_idx, column=2, value=item.get("count", 0)), cell_style)
        links_terms_ws.column_dimensions["A"].width = 42
        links_terms_ws.column_dimensions["B"].width = 12

        ai_ws = wb.create_sheet("AI-сигналы")
        ai_headers = ["Сигнал", "Значение"]
        for col, header in enumerate(ai_headers, 1):
            self._apply_style(ai_ws.cell(row=1, column=col, value=header), header_style)
        ai = results.get("ai_insights", {}) or {}
        ai_rows = [
            ("AI marker density /1k", ai.get("ai_marker_density_1k", 0)),
            ("Hedging ratio", ai.get("hedging_ratio", 0)),
            ("Template repetition /1k", ai.get("template_repetition", 0)),
            ("Burstiness CV", ai.get("burstiness_cv", 0)),
            ("Perplexity proxy", ai.get("perplexity_proxy", 0)),
            ("Entity depth /1k", ai.get("entity_depth_1k", 0)),
            ("Claim specificity score", ai.get("claim_specificity_score", 0)),
            ("Author signal score", ai.get("author_signal_score", 0)),
            ("Source attribution score", ai.get("source_attribution_score", 0)),
            ("AI risk composite", ai.get("ai_risk_composite", 0)),
        ]
        for row_idx, (k, v) in enumerate(ai_rows, start=2):
            self._apply_style(ai_ws.cell(row=row_idx, column=1, value=k), cell_style)
            self._apply_style(ai_ws.cell(row=row_idx, column=2, value=v), cell_style)
        ai_ws.column_dimensions["A"].width = 34
        ai_ws.column_dimensions["B"].width = 18

        heat_ws = wb.create_sheet("Heatmap")
        heat_headers = ["Категория", "Оценка", "Проблемы", "Критично", "Предупреждение"]
        for col, h in enumerate(heat_headers, 1):
            self._apply_style(heat_ws.cell(row=1, column=col, value=h), header_style)
        for row_idx, (cat, payload) in enumerate((results.get("heatmap", {}) or {}).items(), start=2):
            vals = [cat, payload.get("score", 0), payload.get("issues", 0), payload.get("critical", 0), payload.get("warning", 0)]
            for col, v in enumerate(vals, 1):
                self._apply_style(heat_ws.cell(row=row_idx, column=col, value=v), cell_style)
        for col, width in enumerate([18, 12, 12, 12, 12], 1):
            heat_ws.column_dimensions[get_column_letter(col)].width = width

        pq_ws = wb.create_sheet("Очередь приоритетов")
        pq_headers = ["Этап", "Критичность", "Код", "Проблема", "Приоритет", "Трудозатраты"]
        for col, h in enumerate(pq_headers, 1):
            self._apply_style(pq_ws.cell(row=1, column=col, value=h), header_style)
        for row_idx, item in enumerate(results.get("priority_queue", []) or [], start=2):
            vals = [item.get("bucket", ""), item.get("severity", ""), item.get("code", ""), item.get("title", ""), item.get("priority_score", 0), item.get("effort", 0)]
            for col, v in enumerate(vals, 1):
                self._apply_style(pq_ws.cell(row=row_idx, column=col, value=v), cell_style)
        for col, width in enumerate([12, 12, 24, 68, 12, 10], 1):
            pq_ws.column_dimensions[get_column_letter(col)].width = width

        tgt_ws = wb.create_sheet("Цели")
        tgt_headers = ["Метрика", "Текущее", "Цель", "Дельта"]
        for col, h in enumerate(tgt_headers, 1):
            self._apply_style(tgt_ws.cell(row=1, column=col, value=h), header_style)
        for row_idx, item in enumerate(results.get("targets", []) or [], start=2):
            vals = [item.get("metric", ""), item.get("current", 0), item.get("target", 0), item.get("delta", 0)]
            for col, v in enumerate(vals, 1):
                self._apply_style(tgt_ws.cell(row=row_idx, column=col, value=v), cell_style)
        for col, width in enumerate([30, 16, 16, 16], 1):
            tgt_ws.column_dimensions[get_column_letter(col)].width = width

        filepath = os.path.join(self.reports_dir, f"{task_id}.xlsx")
        self._save_workbook(wb, filepath)
        wb.close()
        return filepath

    def generate_link_profile_report(self, task_id: str, data: Dict[str, Any]) -> str:
        """Generate XLSX report for link_profile_audit."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Сводка"

        header_style = self._create_header_style()
        cell_style = self._create_cell_style()

        results = data.get("results", {}) or {}
        summary = results.get("summary", {}) or {}
        tables = results.get("tables", {}) or {}
        warnings = results.get("warnings", []) or []
        errors = results.get("errors", []) or []

        ws["A1"] = "Отчет: Аудит ссылочного профиля"
        self._apply_style(ws["A1"], header_style)
        ws.merge_cells("A1:D1")
        ws["A2"] = "Домен"
        ws["B2"] = data.get("url", summary.get("our_domain", ""))
        ws["A3"] = "Дата"
        ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        summary_rows = [
            ("Строк ссылок", summary.get("rows_total", 0)),
            ("Уникальных доноров", summary.get("unique_ref_domains", 0)),
            ("Уникальных конкурентов", summary.get("unique_competitors", 0)),
            ("Ссылок на наш домен", summary.get("our_links", 0)),
            ("Dofollow", summary.get("dofollow", 0)),
            ("Nofollow", summary.get("nofollow", 0)),
            ("Unknown follow", summary.get("unknown_follow", 0)),
            ("Dofollow %", summary.get("dofollow_pct", "")),
            ("Nofollow %", summary.get("nofollow_pct", "")),
            ("Lost links %", summary.get("lost_links_pct", "")),
            ("HTTP 2xx %", summary.get("http_2xx_pct", "")),
            ("Средний DR", summary.get("avg_dr", "")),
        ]
        ws["A5"] = "Метрика"
        ws["B5"] = "Значение"
        self._apply_style(ws["A5"], header_style)
        self._apply_style(ws["B5"], header_style)
        row_idx = 6
        for k, v in summary_rows:
            ws.cell(row=row_idx, column=1, value=k)
            ws.cell(row=row_idx, column=2, value=v)
            self._apply_style(ws.cell(row=row_idx, column=1), cell_style)
            self._apply_style(ws.cell(row=row_idx, column=2), cell_style)
            row_idx += 1
        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 26
        ws.column_dimensions["C"].width = 24
        ws.column_dimensions["D"].width = 24

        def _write_sheet(title: str, rows: List[Dict[str, Any]]) -> None:
            sheet = wb.create_sheet(title[:31])
            if not rows:
                sheet["A1"] = "Нет данных"
                return
            cols = list((rows[0] or {}).keys())
            for col_idx, col_name in enumerate(cols, start=1):
                c = sheet.cell(row=1, column=col_idx, value=col_name)
                self._apply_style(c, header_style)
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(60, max(12, len(str(col_name)) + 4))
            for r_idx, item in enumerate(rows, start=2):
                for c_idx, col_name in enumerate(cols, start=1):
                    c = sheet.cell(row=r_idx, column=c_idx, value=item.get(col_name))
                    self._apply_style(c, cell_style)

        _write_sheet("Наш сайт", tables.get("our_site_overview", []) or [])
        _write_sheet("KPI сводка", tables.get("executive_kpi", []) or [])
        _write_sheet("Структура профиля", tables.get("profile_structure", []) or [])
        _write_sheet("Ссылки с главных raw", tables.get("raw_homepage_links", []) or [])
        _write_sheet("Ссылки с редиректов raw", tables.get("raw_redirect_links", []) or [])
        _write_sheet("Дубли без нас raw", tables.get("raw_duplicates_without_our", []) or [])
        _write_sheet("Конкуренты", tables.get("competitor_analysis", []) or [])
        _write_sheet("Рейтинг конкурентов", tables.get("competitor_ranking", []) or [])
        _write_sheet("Сравнение", tables.get("comparison_overview", []) or [])
        _write_sheet("Ready-buy доноры", tables.get("ready_buy_domains", []) or [])
        _write_sheet("Анкоры", tables.get("anchor_analysis", []) or [])
        _write_sheet("Слова анкоров", tables.get("anchor_word_analysis", []) or [])
        _write_sheet("Anchor mix %", tables.get("anchor_mix_pct", []) or [])
        _write_sheet("Дубликаты с нами", tables.get("duplicates_with_our_site", []) or [])
        _write_sheet("2+ конкурента без нас", tables.get("duplicates_with_two_competitors", []) or [])
        _write_sheet("Priority domains", tables.get("priority_domains", []) or [])
        _write_sheet("Priority score", tables.get("priority_score_domains", []) or [])
        _write_sheet("DR buckets", tables.get("dr_buckets", []) or [])
        _write_sheet("Доменные зоны", tables.get("zones", []) or [])
        _write_sheet("Качество конкурентов", tables.get("competitor_quality", []) or [])
        _write_sheet("Сравнение GAP", tables.get("comparison_overview", []) or [])
        _write_sheet("DR по доменам %", tables.get("dr_distribution_matrix", []) or [])
        _write_sheet("Матрица возможностей", tables.get("opportunity_domains", []) or [])
        _write_sheet("Follow mix %", tables.get("follow_mix_pct", []) or [])
        _write_sheet("Follow domains %", tables.get("follow_domain_mix_pct", []) or [])
        _write_sheet("Lost status mix", tables.get("lost_status_mix", []) or [])
        _write_sheet("HTTP class mix", tables.get("http_class_mix", []) or [])
        _write_sheet("Link type mix", tables.get("link_type_mix", []) or [])
        _write_sheet("Language mix", tables.get("language_mix", []) or [])

        notes_ws = wb.create_sheet("Warnings")
        notes_ws["A1"] = "Тип"
        notes_ws["B1"] = "Текст"
        self._apply_style(notes_ws["A1"], header_style)
        self._apply_style(notes_ws["B1"], header_style)
        cur = 2
        for item in errors:
            notes_ws.cell(row=cur, column=1, value="error")
            notes_ws.cell(row=cur, column=2, value=item)
            self._apply_style(notes_ws.cell(row=cur, column=1), cell_style)
            self._apply_style(notes_ws.cell(row=cur, column=2), cell_style)
            cur += 1
        for item in warnings:
            notes_ws.cell(row=cur, column=1, value="warning")
            notes_ws.cell(row=cur, column=2, value=item)
            self._apply_style(notes_ws.cell(row=cur, column=1), cell_style)
            self._apply_style(notes_ws.cell(row=cur, column=2), cell_style)
            cur += 1
        notes_ws.column_dimensions["A"].width = 14
        notes_ws.column_dimensions["B"].width = 120

        filepath = os.path.join(self.reports_dir, f"{task_id}.xlsx")
        self._save_workbook(wb, filepath)
        wb.close()
        return filepath

    def generate_report(self, task_id: str, task_type: str, data: Dict[str, Any]) -> str:
        """Dispatch report generation by task type."""
        generators = {
            'site_analyze': self.generate_site_analyze_report,
            'robots_check': self.generate_robots_report,
            'sitemap_validate': self.generate_sitemap_report,
            'render_audit': self.generate_render_report,
            'mobile_check': self.generate_mobile_report,
            'bot_check': self.generate_bot_report,
            'site_audit_pro': self.generate_site_audit_pro_report,
            'onpage_audit': self.generate_onpage_report,
            'link_profile_audit': self.generate_link_profile_report,
        }
        
        generator = generators.get(task_type, self.generate_site_analyze_report)
        return generator(task_id, data)


# Singleton
xlsx_generator = XLSXGenerator()
