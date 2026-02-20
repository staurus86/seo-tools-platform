import shutil
import sys
import types
import unittest
from pathlib import Path

from openpyxl import load_workbook

# Make XLSX generator import independent from runtime-only settings dependency.
if "app.config" not in sys.modules:
    fake_config = types.ModuleType("app.config")
    fake_config.settings = types.SimpleNamespace(REPORTS_DIR=".")
    sys.modules["app.config"] = fake_config

from app.reports.xlsx_generator import XLSXGenerator


class BotXlsxReportTests(unittest.TestCase):
    def test_bot_report_contains_executive_summary_and_playbooks(self):
        data = {
            "url": "https://example.com",
            "results": {
                "engine": "v2",
                "domain": "example.com",
                "summary": {
                    "total": 2,
                    "accessible": 1,
                    "unavailable": 1,
                    "with_content": 1,
                    "without_content": 1,
                    "robots_disallowed": 1,
                    "x_robots_forbidden": 0,
                    "meta_forbidden": 0,
                    "avg_response_time_ms": 321,
                    "indexable": 1,
                    "waf_cdn_detected": 1,
                },
                "bots_checked": ["Googlebot Desktop", "GPTBot"],
                "bot_rows": [
                    {
                        "bot_name": "Googlebot Desktop",
                        "category": "Google",
                        "status": 200,
                        "accessible": True,
                        "has_content": True,
                        "robots_allowed": True,
                        "response_time_ms": 200,
                        "crawlable": True,
                        "renderable": True,
                        "indexable": True,
                    },
                    {
                        "bot_name": "GPTBot",
                        "category": "AI",
                        "status": 403,
                        "accessible": False,
                        "has_content": False,
                        "robots_allowed": False,
                        "response_time_ms": 442,
                        "crawlable": False,
                        "renderable": False,
                        "indexable": False,
                        "blocked_reasons": ["status_403", "robots_disallow"],
                    },
                ],
                "category_stats": [
                    {
                        "category": "Google",
                        "total": 1,
                        "crawlable": 1,
                        "renderable": 1,
                        "indexable": 1,
                        "non_indexable": 0,
                        "indexable_pct": 100.0,
                        "sla_target_pct": 99.0,
                        "sla_met": True,
                        "priority_risk_score": 0,
                    },
                    {
                        "category": "AI",
                        "total": 1,
                        "crawlable": 0,
                        "renderable": 0,
                        "indexable": 0,
                        "non_indexable": 1,
                        "indexable_pct": 0.0,
                        "sla_target_pct": 93.0,
                        "sla_met": False,
                        "priority_risk_score": 100,
                    },
                ],
                "priority_blockers": [
                    {
                        "code": "status_403",
                        "title": "403 forbidden for bot traffic",
                        "affected_bots": 1,
                        "weighted_impact": 3.5,
                        "priority_score": 22.0,
                        "sample_bots": ["GPTBot"],
                        "details": "AI bot blocked with 403",
                    }
                ],
                "playbooks": [
                    {
                        "blocker_code": "status_403",
                        "owner": "Infra",
                        "title": "Allow known AI crawler traffic",
                        "priority_score": 22.0,
                        "actions": [
                            "Whitelist verified bot ASN/IP ranges",
                            "Disable challenge page for known bots",
                        ],
                    }
                ],
                "baseline_diff": {"has_baseline": False, "message": "No baseline found"},
                "trend": {
                    "history": [
                        {
                            "timestamp": "2026-02-20T10:00:00",
                            "indexable": 1,
                            "crawlable": 1,
                            "renderable": 1,
                            "accessible": 1,
                            "avg_response_time_ms": 250,
                            "critical_issues": 1,
                            "warning_issues": 0,
                            "waf_cdn_detected": 1,
                            "retry_profile": "standard",
                            "criticality_profile": "balanced",
                            "sla_profile": "standard",
                        }
                    ]
                },
                "issues": [],
                "recommendations": [],
            },
        }

        temp_dir = Path("tests") / ".tmp_bot_xlsx"
        if temp_dir.exists():
            shutil.rmtree(temp_dir)
        temp_dir.mkdir(parents=True, exist_ok=True)
        try:
            generator = XLSXGenerator()
            generator.reports_dir = str(temp_dir)
            report_path = generator.generate_bot_report("bot-xlsx-test", data)
            wb = load_workbook(report_path)

            self.assertIn("Summary", wb.sheetnames)
            self.assertIn("Executive Summary", wb.sheetnames)
            self.assertIn("Playbooks", wb.sheetnames)
            self.assertIn("Trend History", wb.sheetnames)

            exec_ws = wb["Executive Summary"]
            self.assertEqual(exec_ws["A1"].value, "Bot Access Check - Executive Summary")
            self.assertEqual(exec_ws["A4"].value, "Bots checked")
            self.assertEqual(exec_ws["B4"].value, 2)
            self.assertEqual(exec_ws["A12"].value, 1)
            self.assertEqual(exec_ws["B12"].value, "status_403")

            playbooks_ws = wb["Playbooks"]
            self.assertEqual(playbooks_ws["A1"].value, "Blocker code")
            self.assertEqual(playbooks_ws["A2"].value, "status_403")
            self.assertIn("Whitelist verified bot ASN/IP ranges", str(playbooks_ws["E2"].value))

            trend_ws = wb["Trend History"]
            self.assertEqual(trend_ws["A1"].value, "Run time")
            self.assertEqual(trend_ws["B2"].value, 1)
            self.assertEqual(trend_ws["F2"].value, 250)
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)


if __name__ == "__main__":
    unittest.main()
