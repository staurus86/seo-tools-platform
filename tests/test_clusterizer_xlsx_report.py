import os
import shutil
import unittest
from pathlib import Path

from openpyxl import load_workbook

from app.reports.xlsx_generator import XLSXGenerator
from app.tools.clusterizer import run_keyword_clusterizer


class ClusterizerXlsxReportTests(unittest.TestCase):
    def test_generates_clusterizer_xlsx_report(self):
        result = run_keyword_clusterizer(
            keywords=[
                "купить iphone 16",
                "iphone 16 цена",
                "samsung galaxy s25",
                "galaxy s25 купить",
                "ремонт холодильника",
            ],
            method="jaccard",
            similarity_threshold=0.35,
            min_cluster_size=2,
            clustering_mode="balanced",
        )

        temp_dir = Path("tests/.tmp_clusterizer_xlsx")
        if temp_dir.exists():
            shutil.rmtree(temp_dir, ignore_errors=True)
        temp_dir.mkdir(parents=True, exist_ok=True)

        try:
            generator = XLSXGenerator()
            generator.reports_dir = str(temp_dir)
            filepath = generator.generate_clusterizer_report(
                "clusterizer-test-1",
                {"url": "keywords://clusterizer", "results": result.get("results", {})},
            )

            self.assertTrue(os.path.exists(filepath))

            wb = load_workbook(filepath)
            try:
                self.assertIn("Сводка", wb.sheetnames)
                self.assertIn("Кластеры", wb.sheetnames)
                self.assertIn("Ключи", wb.sheetnames)
                self.assertIn("Одиночные", wb.sheetnames)
                summary_ws = wb["Сводка"]
                self.assertEqual(summary_ws["A1"].value, "Отчет кластеризатора ключевых слов")
                clusters_ws = wb["Кластеры"]
                self.assertEqual(clusters_ws["A1"].value, "Cluster ID")
            finally:
                wb.close()
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)


if __name__ == "__main__":
    unittest.main()
