import shutil
import sys
import types
import unittest
from pathlib import Path

from openpyxl import load_workbook

if "app.config" not in sys.modules:
    fake_config = types.ModuleType("app.config")
    fake_config.settings = types.SimpleNamespace(REPORTS_DIR=".")
    sys.modules["app.config"] = fake_config

from app.reports.xlsx_generator import XLSXGenerator


class RedirectCheckerXlsxReportTests(unittest.TestCase):
    def test_redirect_checker_xlsx_contains_slowest_sheet_and_durations(self):
        data = {
            "url": "https://example.com/",
            "results": {
                "checked_url": "https://example.com/",
                "selected_user_agent": {
                    "key": "googlebot_desktop",
                    "label": "Googlebot Desktop",
                },
                "summary": {
                    "total_scenarios": 3,
                    "passed": 1,
                    "warnings": 1,
                    "errors": 1,
                    "quality_score": 71,
                    "quality_grade": "C",
                    "duration_ms": 1240,
                },
                "scenarios": [
                    {
                        "id": 1,
                        "key": "http_to_https",
                        "title": "HTTP -> HTTPS",
                        "status": "error",
                        "expected": "301/308 на HTTPS",
                        "actual": "200 на http://example.com/",
                        "recommendation": "Настройте принудительный HTTPS.",
                        "test_url": "http://example.com/",
                        "response_codes": [200],
                        "duration_ms": 120,
                        "final_url": "http://example.com/",
                        "hops": 0,
                    },
                    {
                        "id": 2,
                        "key": "canonical_tag",
                        "title": "Canonical тег",
                        "status": "warning",
                        "expected": "Canonical присутствует",
                        "actual": "canonical не найден",
                        "recommendation": "Добавьте canonical в <head>.",
                        "test_url": "https://example.com/",
                        "response_codes": [200],
                        "duration_ms": 45,
                        "final_url": "https://example.com/",
                        "hops": 0,
                    },
                    {
                        "id": 3,
                        "key": "redirect_chains",
                        "title": "Цепочки редиректов",
                        "status": "passed",
                        "expected": "Не более 1 редиректа",
                        "actual": "301 -> 200",
                        "recommendation": "",
                        "test_url": "https://example.com/start",
                        "response_codes": [301, 200],
                        "duration_ms": 220,
                        "final_url": "https://example.com/final",
                        "hops": 1,
                    },
                ],
                "recommendations": [
                    "Настройте принудительный HTTPS.",
                    "Добавьте canonical в <head>.",
                ],
            },
        }

        temp_dir = Path("tests") / ".tmp_redirect_xlsx"
        if temp_dir.exists():
            shutil.rmtree(temp_dir)
        temp_dir.mkdir(parents=True, exist_ok=True)
        try:
            generator = XLSXGenerator()
            generator.reports_dir = str(temp_dir)
            report_path = generator.generate_redirect_checker_report("redirect-xlsx-test", data)
            wb = load_workbook(report_path)

            self.assertIn("Summary", wb.sheetnames)
            self.assertIn("Scenarios", wb.sheetnames)
            self.assertIn("Slowest", wb.sheetnames)
            self.assertIn("Recommendations", wb.sheetnames)

            summary_ws = wb["Summary"]
            self.assertEqual(summary_ws["A1"].value, "Redirect Checker Report")
            self.assertEqual(summary_ws["A3"].value, "URL")
            self.assertEqual(summary_ws["B11"].value, 1240)
            self.assertEqual(summary_ws["B12"].value, "auto")

            scenarios_ws = wb["Scenarios"]
            self.assertEqual(scenarios_ws["E1"].value, "Duration ms")
            self.assertEqual(scenarios_ws["B2"].value, "http_to_https")
            self.assertEqual(scenarios_ws["E2"].value, 120)

            slowest_ws = wb["Slowest"]
            self.assertEqual(slowest_ws["B2"].value, "Цепочки редиректов")
            self.assertEqual(slowest_ws["D2"].value, 220)
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)


if __name__ == "__main__":
    unittest.main()
