import shutil
import sys
import types
import unittest
from pathlib import Path

from docx import Document
from openpyxl import load_workbook

# Make report generators import independent from runtime-only settings dependency.
if "app.config" not in sys.modules:
    fake_config = types.ModuleType("app.config")
    fake_config.settings = types.SimpleNamespace(REPORTS_DIR=".")
    sys.modules["app.config"] = fake_config

from app.reports.docx_generator import DOCXGenerator
from app.reports.xlsx_generator import XLSXGenerator


def _competitor_payload() -> dict:
    return {
        "url": "https://primary.example/",
        "results": {
            "mode": "competitor",
            "strategy": "desktop",
            "source": "pagespeed_insights_api",
            "summary": {
                "total_urls": 3,
                "successful_urls": 3,
                "failed_urls": 0,
                "primary_url": "https://primary.example/",
                "primary_score": 78,
                "primary_cwv_status": "needs_improvement",
                "primary_rank": "2/3",
                "market_leader_url": "https://leader.example/",
                "market_leader_score": 92,
            },
            "benchmark": {
                "competitor_median_score": 89,
                "competitor_median_lcp_ms": 1700,
                "competitor_median_inp_ms": 90,
                "competitor_median_cls": 0.03,
            },
            "primary": {
                "url": "https://primary.example/",
                "status": "success",
                "summary": {"performance_score": 78, "core_web_vitals_status": "needs_improvement"},
                "metrics": {"lcp": {"field_value_ms": 2100}, "inp": {"field_value_ms": 110}, "cls": {"field_value": 0.07}},
            },
            "comparison_rows": [
                {
                    "url": "https://leader.example/",
                    "status": "success",
                    "cwv_status": "good",
                    "score": 92,
                    "lcp_ms": 1500,
                    "inp_ms": 70,
                    "cls": 0.02,
                    "score_delta_vs_primary": 14,
                    "lcp_delta_ms_vs_primary": -600,
                    "inp_delta_ms_vs_primary": -40,
                    "cls_delta_vs_primary": -0.05,
                    "top_focus": "Reduce unused JavaScript",
                    "error": "",
                }
            ],
            "gaps_for_primary": ["Performance score below competitor median"],
            "strengths_of_primary": ["Lower CLS volatility on key pages"],
            "common_opportunities": [{"title": "Reduce unused JavaScript", "group": "javascript", "count": 2}],
            "action_plan": [
                {
                    "priority": "P1",
                    "area": "INP",
                    "owner": "Frontend",
                    "action": "Split JS bundles",
                    "expected_impact": "Lower INP",
                }
            ],
            "recommendations": ["Cut JavaScript payload"],
        },
    }


class CoreWebVitalsReportsTests(unittest.TestCase):
    def test_docx_competitor_report_contains_comparison_sections(self):
        temp_dir = Path("tests") / ".tmp_cwv_docx"
        if temp_dir.exists():
            shutil.rmtree(temp_dir)
        temp_dir.mkdir(parents=True, exist_ok=True)
        try:
            generator = DOCXGenerator()
            generator.reports_dir = str(temp_dir)
            report_path = generator.generate_core_web_vitals_report("cwv-docx-competitor-test", _competitor_payload())

            doc = Document(report_path)
            text = "\n".join(p.text for p in doc.paragraphs)
            self.assertIn("Core Web Vitals Report", text)
            self.assertIn("2. Primary Snapshot", text)
            self.assertIn("3. Competitor Comparison", text)
            self.assertIn("4. Gap & Strength Analysis", text)
            self.assertIn("6. Action Plan (Primary)", text)
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)

    def test_xlsx_competitor_report_contains_expected_sheets(self):
        temp_dir = Path("tests") / ".tmp_cwv_xlsx"
        if temp_dir.exists():
            shutil.rmtree(temp_dir)
        temp_dir.mkdir(parents=True, exist_ok=True)
        try:
            generator = XLSXGenerator()
            generator.reports_dir = str(temp_dir)
            report_path = generator.generate_core_web_vitals_report("cwv-xlsx-competitor-test", _competitor_payload())

            wb = load_workbook(report_path)
            self.assertIn("Summary", wb.sheetnames)
            self.assertIn("Comparison", wb.sheetnames)
            self.assertIn("Gap_Analysis", wb.sheetnames)
            self.assertIn("Primary_Action_Plan", wb.sheetnames)
            self.assertIn("Recommendations", wb.sheetnames)

            summary_ws = wb["Summary"]
            self.assertEqual(summary_ws["A1"].value, "Core Web Vitals Report")
            self.assertEqual(summary_ws["A8"].value, "KPI")
            self.assertEqual(summary_ws["A13"].value, "Primary score")
            self.assertEqual(summary_ws["B13"].value, 78)

            cmp_ws = wb["Comparison"]
            self.assertEqual(cmp_ws["A1"].value, "#")
            self.assertEqual(cmp_ws["B2"].value, "https://leader.example/")
            self.assertEqual(cmp_ws["E2"].value, 92)
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)


if __name__ == "__main__":
    unittest.main()
