import unittest
from pathlib import Path
import sys
import types
import shutil

from openpyxl import load_workbook

# Make XLSX generator import independent from runtime-only settings dependency.
if "app.config" not in sys.modules:
    fake_config = types.ModuleType("app.config")
    fake_config.settings = types.SimpleNamespace(REPORTS_DIR=".")
    sys.modules["app.config"] = fake_config

from app.reports.xlsx_generator import XLSXGenerator


class SiteProXlsxLayoutTests(unittest.TestCase):
    def test_quick_report_has_8_sheets_and_unique_parameter_headers(self):
        data = {
            "url": "https://site.test",
            "results": {
                "mode": "quick",
                "summary": {
                    "total_pages": 1,
                    "internal_pages": 1,
                    "issues_total": 1,
                    "critical_issues": 0,
                    "warning_issues": 1,
                    "info_issues": 0,
                    "score": 91.0,
                },
                "pages": [
                    {
                        "url": "https://site.test",
                        "final_url": "https://site.test",
                        "status_code": 200,
                        "response_time_ms": 220,
                        "html_size_bytes": 15320,
                        "dom_nodes_count": 340,
                        "redirect_count": 0,
                        "is_https": True,
                        "compression_enabled": True,
                        "cache_enabled": True,
                        "indexable": True,
                        "health_score": 91.0,
                        "title": "Home",
                        "meta_description": "Main page",
                        "canonical": "https://site.test",
                        "meta_robots": "index,follow",
                        "schema_count": 1,
                        "hreflang_count": 0,
                        "mobile_friendly_hint": True,
                        "word_count": 330,
                        "unique_word_count": 180,
                        "lexical_diversity": 0.545,
                        "readability_score": 84.2,
                        "toxicity_score": 0.0,
                        "filler_ratio": 0.03,
                        "h1_count": 1,
                        "images_count": 8,
                        "images_without_alt": 1,
                        "external_nofollow_links": 1,
                        "external_follow_links": 2,
                        "outgoing_internal_links": 6,
                        "incoming_internal_links": 3,
                        "outgoing_external_links": 3,
                        "orphan_page": False,
                        "topic_hub": True,
                        "pagerank": 100.0,
                        "topic_label": "home",
                        "top_terms": ["home", "services", "seo"],
                        "duplicate_title_count": 1,
                        "duplicate_description_count": 1,
                        "weak_anchor_ratio": 0.08,
                        "link_quality_score": 96.5,
                        "ai_markers_count": 0,
                        "recommendation": "Maintain page quality and monitor regressions.",
                        "issues": [{"severity": "warning", "code": "thin_content", "title": "Thin", "details": ""}],
                    }
                ],
                "issues": [
                    {"severity": "warning", "url": "https://site.test", "code": "thin_content", "title": "Thin", "details": ""}
                ],
                "pipeline": {
                    "tf_idf": [{"url": "https://site.test", "top_terms": ["home", "services", "seo"]}]
                },
            },
        }

        temp_dir = Path("tests") / ".tmp_site_pro_xlsx"
        if temp_dir.exists():
            shutil.rmtree(temp_dir)
        temp_dir.mkdir(parents=True, exist_ok=True)
        try:
            generator = XLSXGenerator()
            generator.reports_dir = str(temp_dir)
            report_path = generator.generate_site_audit_pro_report("site-pro-layout-test", data)

            wb = load_workbook(report_path)
            expected_sheets = [
                "1_Executive",
                "2_OnPage+Structured",
                "3_Technical",
                "4_Content+AI",
                "5_LinkGraph",
                "6_Images+External",
                "7_HierarchyErrors",
                "8_Keywords",
            ]
            self.assertEqual(wb.sheetnames[:8], expected_sheets)

            allow_repeats = {"URL", "Severity"}
            used_headers = {}
            for sheet_name in expected_sheets[1:]:
                ws = wb[sheet_name]
                headers = [cell.value for cell in ws[1] if cell.value]
                for header in headers:
                    if header in allow_repeats:
                        continue
                    self.assertNotIn(
                        header,
                        used_headers,
                        msg=f"Header '{header}' is duplicated in '{used_headers.get(header)}' and '{sheet_name}'",
                    )
                    used_headers[header] = sheet_name

            # Smoke-check that key metric families are represented exactly once.
            required_headers = {
                "Status",
                "Response ms",
                "Schema count",
                "Word count",
                "PageRank",
                "External nofollow",
                "Code",
                "Top terms (TF-IDF)",
            }
            self.assertTrue(required_headers.issubset(set(used_headers.keys())))

            self.assertTrue(Path(report_path).exists())
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)

    def test_full_report_adds_deep_sheets(self):
        data = {
            "url": "https://site.test",
            "results": {
                "mode": "full",
                "summary": {"total_pages": 1, "issues_total": 0, "critical_issues": 0, "warning_issues": 0, "info_issues": 0},
                "pages": [{"url": "https://site.test", "topic_label": "home", "top_terms": ["home"], "issues": []}],
                "issues": [],
                "pipeline": {
                    "tf_idf": [{"url": "https://site.test", "top_terms": ["home"]}],
                    "semantic_linking_map": [
                        {"source_url": "https://site.test", "target_url": "https://site.test/a", "topic": "home", "reason": "test"}
                    ],
                    "duplicates": {"title_groups": [], "description_groups": []},
                },
            },
        }

        temp_dir = Path("tests") / ".tmp_site_pro_xlsx_full"
        if temp_dir.exists():
            shutil.rmtree(temp_dir)
        temp_dir.mkdir(parents=True, exist_ok=True)
        try:
            generator = XLSXGenerator()
            generator.reports_dir = str(temp_dir)
            report_path = generator.generate_site_audit_pro_report("site-pro-layout-full", data)
            wb = load_workbook(report_path)
            self.assertIn("9_SemanticMap", wb.sheetnames)
            self.assertIn("10_DuplicatesDeep", wb.sheetnames)
            self.assertIn("11_IssuesRaw", wb.sheetnames)
            compat_sheets = [
                "13_MainReport_Compat",
                "14_Hierarchy_Compat",
                "15_OnPage_Compat",
                "16_Content_Compat",
                "17_Technical_Compat",
                "18_EEAT_Compat",
                "19_Trust_Compat",
                "20_Health_Compat",
                "21_InternalLinks_Compat",
                "22_Images_Compat",
                "23_ExternalLinks_Compat",
                "24_Structured_Compat",
                "25_KeywordsTFIDF_Compat",
                "26_Topics_Compat",
                "27_Advanced_Compat",
                "28_LinkQuality_Compat",
                "29_AIMarkers_Compat",
            ]
            for sheet_name in compat_sheets:
                self.assertIn(sheet_name, wb.sheetnames)

            expected_headers = {
                "13_MainReport_Compat": ["URL", "Title", "Meta", "H1", "Токсичность", "Иерархия", "Health", "HTTP", "Indexable", "Canonical", "Resp ms", "Проблемы", "Solution", "Severity"],
                "14_Hierarchy_Compat": ["URL", "Статус", "Проблема", "Всего заголовков", "H1 Count", "Решение", "Severity"],
                "15_OnPage_Compat": ["URL", "Title Len", "Meta Len", "H1", "Canonical", "Canonical Status", "Mobile", "Schema", "Breadcrumbs", "Title Dup", "Meta Dup", "Solution", "Severity"],
                "16_Content_Compat": ["URL", "Words", "Unique %", "Readability", "Toxicity", "AI Markers", "AI Markers List", "Filler", "Solution", "Severity"],
                "17_Technical_Compat": ["URL", "Status", "Indexable", "Resp ms", "Size KB", "DOM", "HTML Score", "HTTPS", "Compression", "Cache", "Canonical", "Robots", "Deprecated", "Solution", "Severity"],
                "18_EEAT_Compat": ["URL", "Score", "Expertise", "Authority", "Trust", "Experience", "Solution", "Severity"],
                "19_Trust_Compat": ["URL", "Trust Score", "Contact", "Legal", "Reviews", "Badges", "Solution", "Severity"],
                "20_Health_Compat": ["URL", "Health Score", "Indexable", "Words", "Unique %", "Readability", "Title Dup", "Meta Dup", "Resp ms", "Solution", "Severity"],
                "21_InternalLinks_Compat": ["URL", "Authority", "Incoming", "Outgoing", "Is Orphan", "Solution", "Severity"],
                "22_Images_Compat": ["URL", "Total", "No Alt", "No Width", "No Lazy", "Issues", "Solution", "Severity"],
                "23_ExternalLinks_Compat": ["URL", "Total External", "Follow", "NoFollow", "Follow %", "Solution", "Severity"],
                "24_Structured_Compat": ["URL", "Total", "JSON-LD", "Microdata", "RDFa", "Hreflang", "Meta Robots", "Solution", "Severity"],
                "25_KeywordsTFIDF_Compat": ["URL", "Top Keywords", "TF-IDF 1", "TF-IDF 2", "TF-IDF 3", "Solution", "Severity"],
                "26_Topics_Compat": ["URL", "Is Hub", "Cluster", "Incoming Links", "Semantic Links", "Solution", "Severity"],
                "27_Advanced_Compat": ["URL", "Freshness Days", "Last Modified", "Status", "Indexable", "Resp ms", "Size KB", "Redirects", "Final URL", "Hidden Content", "Cloaking", "CTA Count", "List/Tables", "Solution", "Severity"],
                "28_LinkQuality_Compat": ["URL", "Linking Score", "Page Authority", "Anchor Score", "Incoming Links", "Outgoing Internal", "Orphan", "Topic Hub", "Solution", "Severity"],
                "29_AIMarkers_Compat": ["URL", "AI Markers Count", "AI Markers Found", "Text Sample with Markers", "Recommendation", "Severity"],
            }
            for sheet_name, headers_expected in expected_headers.items():
                actual = [cell.value for cell in wb[sheet_name][1] if cell.value]
                self.assertEqual(actual, headers_expected)
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)


if __name__ == "__main__":
    unittest.main()
