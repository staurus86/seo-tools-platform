"""Value-level parity checker for legacy seopro XLSX vs platform compat XLSX.

Compares mapped sheets row-by-row on shared headers with lightweight normalization.

Usage:
  python scripts/site_pro_xlsx_value_parity.py --legacy <legacy.xlsx> --current <current.xlsx> [--rows 50]
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook


LEGACY_TO_COMPAT = {
    "1. Основной отчёт": "13_MainReport_Compat",
    "2. Ошибки иерархии": "14_Hierarchy_Compat",
    "3. On-Page SEO": "15_OnPage_Compat",
    "4. Content": "16_Content_Compat",
    "5. Technical": "17_Technical_Compat",
    "6. E-E-A-T": "18_EEAT_Compat",
    "7. Trust": "19_Trust_Compat",
    "8. Health": "20_Health_Compat",
    "9. Internal Links": "21_InternalLinks_Compat",
    "10. Images": "22_Images_Compat",
    "11. External Links": "23_ExternalLinks_Compat",
    "12. Structured Data": "24_Structured_Compat",
    "13. Keywords & TF-IDF": "25_KeywordsTFIDF_Compat",
    "14. Topics": "26_Topics_Compat",
    "15. Advanced": "27_Advanced_Compat",
    "16. Link Quality": "28_LinkQuality_Compat",
    "17. AI Markers": "29_AIMarkers_Compat",
}


def normalize_cell(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r"\s+", " ", text)
    text = (
        text.replace("✅", "")
        .replace("⚠️", "")
        .replace("❌", "")
        .replace("⭐", "")
        .replace("➡️", "")
    )
    text = text.replace("Yes", "yes").replace("No", "no")
    text = text.replace(" ch", "")
    text = text.lower().strip()
    # collapse trailing .0
    if re.fullmatch(r"-?\d+\.0+", text):
        text = text.split(".")[0]
    return text


def headers_map(ws) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(1, col).value
        if raw is None:
            continue
        key = str(raw).strip()
        if key:
            mapping[key] = col
    return mapping


def extract_row_by_headers(ws, headers: List[str], row_idx: int) -> Dict[str, str]:
    hmap = headers_map(ws)
    row: Dict[str, str] = {}
    for h in headers:
        col = hmap.get(h)
        if not col:
            continue
        row[h] = normalize_cell(ws.cell(row_idx, col).value)
    return row


def compare_sheet_values(legacy_ws, compat_ws, rows_limit: int) -> Tuple[int, int, List[str]]:
    legacy_headers = [str(legacy_ws.cell(1, c).value).strip() for c in range(1, legacy_ws.max_column + 1) if legacy_ws.cell(1, c).value]
    compat_hmap = headers_map(compat_ws)
    shared = [h for h in legacy_headers if h in compat_hmap]
    if not shared:
        return 0, 0, ["no shared headers"]

    max_rows = min(rows_limit, legacy_ws.max_row - 1, compat_ws.max_row - 1)
    compared = 0
    mismatches = 0
    samples: List[str] = []
    for offset in range(1, max_rows + 1):
        lr = extract_row_by_headers(legacy_ws, shared, offset + 1)
        cr = extract_row_by_headers(compat_ws, shared, offset + 1)
        for h in shared:
            compared += 1
            lv = lr.get(h, "")
            cv = cr.get(h, "")
            if lv != cv:
                mismatches += 1
                if len(samples) < 20:
                    samples.append(f"row={offset+1} col='{h}' legacy='{lv}' current='{cv}'")
    return compared, mismatches, samples


def main() -> int:
    parser = argparse.ArgumentParser(description="Value-level parity checker for Site Pro XLSX")
    parser.add_argument("--legacy", required=True, help="Path to legacy xlsx")
    parser.add_argument("--current", required=True, help="Path to current xlsx")
    parser.add_argument("--rows", type=int, default=50, help="Rows per mapped sheet to compare")
    args = parser.parse_args()

    def out(text: str) -> None:
        try:
            print(text)
        except UnicodeEncodeError:
            enc = sys.stdout.encoding or "utf-8"
            safe = text.encode(enc, errors="replace").decode(enc, errors="replace")
            print(safe)

    legacy_path = Path(args.legacy)
    current_path = Path(args.current)
    if not legacy_path.exists():
        out(f"[error] legacy file not found: {legacy_path}")
        return 2
    if not current_path.exists():
        out(f"[error] current file not found: {current_path}")
        return 2

    legacy_wb = load_workbook(legacy_path, data_only=False)
    current_wb = load_workbook(current_path, data_only=False)

    total_compared = 0
    total_mismatches = 0
    for legacy_sheet, compat_sheet in LEGACY_TO_COMPAT.items():
        if legacy_sheet not in legacy_wb.sheetnames or compat_sheet not in current_wb.sheetnames:
            out(f"[skip] {legacy_sheet} -> {compat_sheet} (sheet missing)")
            continue
        compared, mismatches, samples = compare_sheet_values(
            legacy_wb[legacy_sheet],
            current_wb[compat_sheet],
            rows_limit=max(1, int(args.rows)),
        )
        total_compared += compared
        total_mismatches += mismatches
        mismatch_pct = (mismatches / compared * 100.0) if compared else 0.0
        out(f"[sheet] {legacy_sheet} -> {compat_sheet}: compared={compared} mismatches={mismatches} ({mismatch_pct:.2f}%)")
        for sample in samples[:3]:
            out(f"  - {sample}")

    total_pct = (total_mismatches / total_compared * 100.0) if total_compared else 0.0
    out(f"total_compared={total_compared} total_mismatches={total_mismatches} mismatch_pct={total_pct:.2f}%")
    return 1 if total_mismatches > 0 else 0


if __name__ == "__main__":
    raise SystemExit(main())
