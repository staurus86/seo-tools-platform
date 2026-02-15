"""Compare legacy seopro XLSX structure with platform Site Audit Pro XLSX.

Usage:
  python scripts/site_pro_xlsx_parity_check.py --legacy <legacy.xlsx> --current <current.xlsx>
"""
from __future__ import annotations

import argparse
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook


LEGACY_TO_COMPAT = {
    "1. Основной отчёт": "13_MainReport_Compat",
    "2. Ошибки иерархии": "14_Hierarchy_Compat",
    "3. On-Page SEO": "15_OnPage_Compat",
    "4. Content": "16_Content_Compat",
    "5. Technical": "17_Technical_Compat",
    "6. E-E-A-T": "18_EEAT_Compat",
    "7. Trust": "19_Trust_Compat",
    "8. Health": "20_Health_Compat",
    "9. Internal Links": "21_InternalLinks_Compat",
    "10. Images": "22_Images_Compat",
    "11. External Links": "23_ExternalLinks_Compat",
    "12. Structured Data": "24_Structured_Compat",
    "13. Keywords & TF-IDF": "25_KeywordsTFIDF_Compat",
    "14. Topics": "26_Topics_Compat",
    "15. Advanced": "27_Advanced_Compat",
    "16. Link Quality": "28_LinkQuality_Compat",
    "17. AI Markers": "29_AIMarkers_Compat",
}


def read_headers(path: Path) -> Dict[str, List[str]]:
    wb = load_workbook(path, data_only=False)
    data: Dict[str, List[str]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        data[sheet_name] = [str(h).strip() for h in headers if h is not None and str(h).strip()]
    return data


def compare_headers(
    legacy_headers: Dict[str, List[str]], current_headers: Dict[str, List[str]], allow_extra: bool
) -> Tuple[List[str], List[str], List[str]]:
    missing_sheets: List[str] = []
    header_mismatches: List[str] = []
    ok: List[str] = []
    for legacy_sheet, compat_sheet in LEGACY_TO_COMPAT.items():
        if legacy_sheet not in legacy_headers:
            missing_sheets.append(f"legacy sheet not found: {legacy_sheet}")
            continue
        if compat_sheet not in current_headers:
            missing_sheets.append(f"current sheet not found: {compat_sheet}")
            continue
        lhs = legacy_headers[legacy_sheet]
        rhs = current_headers[compat_sheet]
        headers_match = lhs == rhs
        if allow_extra and len(rhs) >= len(lhs):
            headers_match = rhs[: len(lhs)] == lhs
        if headers_match:
            ok.append(f"{legacy_sheet} -> {compat_sheet}")
        else:
            header_mismatches.append(
                f"{legacy_sheet} -> {compat_sheet}\n"
                f"  legacy : {lhs}\n"
                f"  current: {rhs}"
            )
    return missing_sheets, header_mismatches, ok


def main() -> int:
    parser = argparse.ArgumentParser(description="Compare legacy and current Site Pro XLSX structures")
    parser.add_argument("--legacy", required=True, help="Path to legacy seopro xlsx")
    parser.add_argument("--current", required=True, help="Path to current platform xlsx")
    parser.add_argument(
        "--allow-extra",
        action="store_true",
        help="Allow extra trailing columns in current sheet headers",
    )
    args = parser.parse_args()

    legacy_path = Path(args.legacy)
    current_path = Path(args.current)
    if not legacy_path.exists():
        print(f"[error] legacy file not found: {legacy_path}")
        return 2
    if not current_path.exists():
        print(f"[error] current file not found: {current_path}")
        return 2

    legacy = read_headers(legacy_path)
    current = read_headers(current_path)
    missing_sheets, header_mismatches, ok = compare_headers(legacy, current, allow_extra=args.allow_extra)

    print(f"legacy_sheets={len(legacy)} current_sheets={len(current)}")
    print(f"mapped_ok={len(ok)} missing={len(missing_sheets)} mismatches={len(header_mismatches)}")
    if ok:
        print("[ok]")
        for item in ok:
            print(f"- {item}")
    if missing_sheets:
        print("[missing]")
        for item in missing_sheets:
            print(f"- {item}")
    if header_mismatches:
        print("[mismatches]")
        for item in header_mismatches:
            print(f"- {item}")

    return 1 if (missing_sheets or header_mismatches) else 0


if __name__ == "__main__":
    raise SystemExit(main())
